"""
HORIZON PORTFOLIO TOOL v95
PE Secondaries & Co-Investment LP Management

Tabs:
  📂 Portfolio       add/edit/delete deals, KPIs, charts
  🔭 Pipeline        deal pipeline, stage funnel, promote → portfolio
  🧩 Pro Forma       placeholder/future deals, deployment bar
  📊 Analytics       strategy/region/vintage/sector/concentration/scatter
  📈 Segments & TWR  seed/new/money-market breakdown, TWR forecast
  💧 Dry Powder      12-month forecast, bite-size guide
  ♻️ Pacing          quarterly deployment model
  🧮 Return Calc     required IRR waterfall, Monte Carlo TWR
  ⚙️ Settings        fund params, bite sizes

Run:  python Horizon_portfolio_tool_v75.py
Open: http://localhost:8060
"""

import base64, math, os, pickle, json
import numpy as np
from datetime import datetime
from io import BytesIO
from dateutil.relativedelta import relativedelta

import dash
from dash import dcc, html, dash_table, Input, Output, State, callback_context, ALL, no_update
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# ── Persistence ───────────────────────────────────────────────────────────────
DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "horizon_v93_data.pkl")

def save_data(portfolio, pipeline, placeholders, config, next_id):
    try:
        with open(DATA_FILE, "wb") as f:
            pickle.dump({"portfolio": portfolio, "pipeline": pipeline,
                         "placeholders": placeholders, "config": config,
                         "next_id": next_id,
                         "saved_at": datetime.now().isoformat(timespec="seconds")}, f)
    except Exception as e:
        print(f"Save error: {e}")

def load_data():
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "rb") as f:
                return pickle.load(f)
        except Exception:
            return None
    return None

# ── Palette (identical to Credit Tool v9) ────────────────────────────────────
C = dict(
    bg="#07090f", panel="#0d1117", surface="#111822",
    border="#1c2a3a", border2="#243649",
    blue="#2979c8", sky="#56b4f5", teal="#2ec4b6",
    green="#2ecc71", red="#e5493a", amber="#f0a500",
    purple="#9b72cf", pink="#e05c8a",
    text="#d4e6f5", muted="#5e82a0", dim="#324d63",
    mono="'JetBrains Mono', 'Fira Mono', monospace",
    sans="'IBM Plex Sans', 'Segoe UI', sans-serif",
)

def cl(hex_c, alpha):
    r, g, b = int(hex_c[1:3],16), int(hex_c[3:5],16), int(hex_c[5:7],16)
    return f"rgba({r},{g},{b},{alpha})"

CHART = dict(
    paper_bgcolor=C["panel"], plot_bgcolor=C["surface"],
    font=dict(family=C["mono"], color=C["text"], size=11),
    margin=dict(l=52, r=20, t=40, b=40),
    xaxis=dict(gridcolor=C["border"], zerolinecolor=C["border"]),
    yaxis=dict(gridcolor=C["border"], zerolinecolor=C["border"]),
    legend=dict(bgcolor=C["panel"], bordercolor=C["border"], borderwidth=1),
)

# CHART_BASE has no axis/legend/margin keys — safe to use with **CHART_BASE alongside those kwargs
CHART_BASE = dict(
    paper_bgcolor=C["panel"], plot_bgcolor=C["surface"],
    font=dict(family=C["mono"], color=C["text"], size=11),
)

def chart(**overrides):
    """Return full chart layout with deep-merged overrides (avoids duplicate-kwarg errors)."""
    base = dict(CHART)
    for k, v in overrides.items():
        if k in base and isinstance(base[k], dict) and isinstance(v, dict):
            base[k] = {**base[k], **v}
        else:
            base[k] = v
    return base

# ── Taxonomy ──────────────────────────────────────────────────────────────────
STRATEGIES    = ["GP-Led (Multi-Asset)", "GP-Led (Single-Asset)", "Diversified LP-Led", "Co-Investment", "Primary"]
DEAL_TYPES    = ["Secondary", "Co-Investment"]
SECTORS       = ["Technology","Healthcare","Financial Services","Consumer","Industrials","Energy","Real Estate","Diversified","Other"]
REGIONS       = ["North America","Europe","Asia","Global"]
STAGES_DEAL   = ["Buyout","Growth","Venture","Liquidity"]
SEGMENTS      = ["Seed","New","MoneyMarket"]
ALLOC_STATUS  = ["Closed","Pending Close","Commitment"]
CURRENCIES    = ["USD","EUR","GBP","CHF"]
PIPE_STAGES   = ["Screening","Diligence","IC Review","Negotiation","Final Docs"]
PRIORITIES    = ["High","Medium","Low"]
VINTAGES      = list(range(2030, 2013, -1))

# ── Defaults ──────────────────────────────────────────────────────────────────
DEFAULT_CONFIG = {
    "fund_size": 1000.0, "dry_powder": 300.0,
    "target_net_twr": 0.13, "management_fee": 0.0125,
    "carry_rate": 0.125, "hurdle_rate": 0.10,
    "loss_drag": 0.01, "liquidity_reserve_pct": 0.05,
    "distribution_rate": 0.20, "cash_yield": 0.03,
    "avg_hold_period": 5.0, "deployment_years": 4.0,
    "deals_per_year": 6.0,
    "bite_min_pct": 0.005, "bite_desired_pct": 0.0275, "bite_max_pct": 0.05,
    # ── Fund Overview / Mandate ───────────────────────────────────────────────
    "fund_name": "Horizon",
    "fund_vintage": 2024,
    "fund_currency": "USD",
    "fund_strategy": "PE Secondaries & Co-Investment",
    "vehicle_type": "Evergreen",
    "domicile": "Luxembourg",
    # Strategy allocation targets (%) — must sum to ~100
    "target_secondary_pct": 70.0,
    "target_coinvest_pct":  30.0,
    # Geography targets (%)
    "target_na_pct":        50.0,
    "target_europe_pct":    35.0,
    "target_asia_pct":      10.0,
    "target_global_pct":    5.0,
    # ── Flexible restriction lists (editable in Settings) ────────────────────
    # Each entry: {label, metric_key, limit, higher_is_bad, fmt}
    # metric_key maps to computed actuals in tab_overview
    "legal_restrictions": [
        {"label": "Single Manager Concentration",  "metric_key": "max_manager_pct",   "limit": 20.0,  "higher_is_bad": True, "fmt": "%"},
        {"label": "Single Asset Concentration",    "metric_key": "max_deal_pct",       "limit": 10.0,  "higher_is_bad": True, "fmt": "%"},
        {"label": "Non-North America Exposure",    "metric_key": "non_na_pct",         "limit": 50.0,  "higher_is_bad": True, "fmt": "%"},
        {"label": "Non-Secondary Allocation",      "metric_key": "non_sec_pct",        "limit": 50.0,  "higher_is_bad": True, "fmt": "%"},
        {"label": "Listed Company Exposure",       "metric_key": "listed_pct",         "limit": 10.0,  "higher_is_bad": True, "fmt": "%"},
        {"label": "Leverage (% NAV)",              "metric_key": "leverage_pct",       "limit": 20.0,  "higher_is_bad": True, "fmt": "%"},
        {"label": "Overcommitment (% NAV)",        "metric_key": "overcommitment_pct", "limit": 120.0, "higher_is_bad": True, "fmt": "%"},
    ],
    "investment_targets": [
        {"label": "Single Vintage Year",           "metric_key": "max_vintage_pct",   "limit": 35.0, "higher_is_bad": True,  "fmt": "%"},
        {"label": "Co-Investment Allocation",      "metric_key": "ci_pct",            "limit": 40.0, "higher_is_bad": True,  "fmt": "%"},
        {"label": "Primary Fund Investments",      "metric_key": "primary_pct",       "limit": 10.0, "higher_is_bad": True,  "fmt": "%"},
        {"label": "Single Sector Concentration",   "metric_key": "max_sector_pct",    "limit": 35.0, "higher_is_bad": True,  "fmt": "%"},
        {"label": "Minimum # Positions",           "metric_key": "num_deals",         "limit": 10.0, "higher_is_bad": False, "fmt": "n"},
        {"label": "Secondaries Allocation",        "metric_key": "sec_pct",           "limit": 50.0, "higher_is_bad": False, "fmt": "%"},
    ],
    # Return targets
    "target_irr_secondary": 15.0,      # target gross IRR for secondaries (%)
    "target_irr_coinvest":  20.0,      # target gross IRR for co-investments (%)
    "target_moic":          1.8,        # target MOIC
}

SEED_PORTFOLIO = [
    dict(id=1, name="Project Alpha", manager="Hamilton Lane",
         strategy="GP-Led (Multi-Asset)", deal_type="Secondary", stage="Buyout",
         sector="Technology", region="North America", currency="USD",
         total_commitment=50.0, current_commitment=50.0, nav=58.0,
         target_irr=0.18, hold_period=4.5, moic=2.18,
         vintage=2023, segment="Seed", allocation_status="Closed",
         date_added="2023-06-01"),
    dict(id=2, name="Project Beta", manager="Coller Capital",
         strategy="Diversified LP-Led", deal_type="Secondary", stage="Buyout",
         sector="Diversified", region="Europe", currency="EUR",
         total_commitment=35.0, current_commitment=35.0, nav=40.0,
         target_irr=0.16, hold_period=3.5, moic=1.85,
         vintage=2024, segment="Seed", allocation_status="Closed",
         date_added="2024-01-15"),
    dict(id=3, name="Project Gamma", manager="Ardian",
         strategy="Co-Investment", deal_type="Co-Investment", stage="Growth",
         sector="Healthcare", region="Europe", currency="EUR",
         total_commitment=20.0, current_commitment=20.0, nav=24.0,
         target_irr=0.22, hold_period=5.0, moic=2.93,
         vintage=2024, segment="New", allocation_status="Closed",
         date_added="2024-03-20"),
    dict(id=4, name="Project Delta", manager="Blackstone",
         strategy="GP-Led (Single-Asset)", deal_type="Secondary", stage="Buyout",
         sector="Industrials", region="North America", currency="USD",
         total_commitment=45.0, current_commitment=45.0, nav=51.0,
         target_irr=0.20, hold_period=3.0, moic=1.73,
         vintage=2025, segment="New", allocation_status="Closed",
         date_added="2025-01-10"),
]

SEED_PIPELINE = [
    dict(id=10, name="Project Epsilon", strategy="GP-Led (Multi-Asset)", stage_deal="Buyout",
         sector="Technology", region="North America", size=30.0, target_irr=0.19,
         pipeline_stage="IC Review", priority="High", date_added="2025-10-01"),
    dict(id=11, name="Project Zeta", strategy="Co-Investment", stage_deal="Growth",
         sector="Healthcare", region="Europe", size=15.0, target_irr=0.24,
         pipeline_stage="Diligence", priority="Medium", date_added="2025-11-01"),
]

SEED_PLACEHOLDERS = [
    dict(id=20, name="GP-Led Placeholder 1", strategy="GP-Led (Multi-Asset)",
         deal_type="Secondary", size=35.0, expected_month=3,
         region="North America", target_irr=0.17, date_added="2025-01-01"),
    dict(id=21, name="Co-Invest Placeholder", strategy="Co-Investment",
         deal_type="Co-Investment", size=20.0, expected_month=6,
         region="Europe", target_irr=0.22, date_added="2025-01-01"),
]

loaded            = load_data()
INITIAL_PORT      = loaded.get("portfolio",    SEED_PORTFOLIO)    if loaded else SEED_PORTFOLIO
INITIAL_PIPE      = loaded.get("pipeline",     SEED_PIPELINE)     if loaded else SEED_PIPELINE
INITIAL_PH        = loaded.get("placeholders", SEED_PLACEHOLDERS) if loaded else SEED_PLACEHOLDERS
INITIAL_NEXT_ID   = loaded.get("next_id",      50)                if loaded else 50
SAVED_AT          = loaded.get("saved_at",     "Seed data")       if loaded else "Seed data"

# Always start from a full DEFAULT_CONFIG base, then overlay saved values.
# This ensures new keys (like legal_restrictions / investment_targets) are
# never stale from an older pickle that pre-dates this version.
_saved_cfg = loaded.get("config", {}) if loaded else {}
INITIAL_CFG = {**DEFAULT_CONFIG, **_saved_cfg,
               # Always reset restriction lists to current defaults unless the
               # user has already saved customised versions via the Settings tab.
               "legal_restrictions":  _saved_cfg.get("legal_restrictions",  DEFAULT_CONFIG["legal_restrictions"]),
               "investment_targets":  _saved_cfg.get("investment_targets",  DEFAULT_CONFIG["investment_targets"]),
               }

# ── Style helpers ─────────────────────────────────────────────────────────────
INP = dict(background=C["surface"], border=f"1px solid {C['border2']}",
           color=C["text"], borderRadius=6, padding="6px 10px",
           fontFamily=C["mono"], fontSize=12, outline="none", width="100%")
BTN = lambda bg, fg="#fff": dict(
    background=bg, border="none", color=fg, borderRadius=6,
    padding="7px 16px", cursor="pointer", fontWeight=600,
    fontSize=12, fontFamily=C["sans"])
TBL_CELL = dict(backgroundColor=C["panel"], color=C["text"],
                fontFamily=C["mono"], fontSize=11, padding="8px 12px",
                border=f"1px solid {C['border']}", textAlign="left")
TBL_HEAD = dict(backgroundColor=C["bg"], color=C["muted"], fontWeight=700,
                fontSize=10, letterSpacing=1.5, textTransform="uppercase",
                border=f"1px solid {C['border']}", padding="9px 12px")
TBL_ODD  = [{"if": {"row_index": "odd"}, "backgroundColor": C["surface"]}]

def card(children, style_extra=None):
    base = dict(background=C["panel"], border=f"1px solid {C['border']}", borderRadius=10, padding=18)
    if style_extra: base.update(style_extra)
    return html.Div(children, style=base)

def kpi(label, value, sub="", color=None, width=150):
    return html.Div([
        html.Div(label, style=dict(fontSize=9, letterSpacing=2, color=C["muted"],
                                   textTransform="uppercase", marginBottom=5, fontFamily=C["sans"])),
        html.Div(value, style=dict(fontSize=22, fontWeight=700, color=color or C["sky"], fontFamily=C["mono"])),
        html.Div(sub,   style=dict(fontSize=10, color=C["dim"], marginTop=3, fontFamily=C["sans"])),
    ], style=dict(background=C["surface"], border=f"1px solid {C['border']}",
                  borderRadius=8, padding="14px 18px", minWidth=width))

def section_lbl(text):
    return html.Div(text, style=dict(fontSize=9, letterSpacing=2.5, color=C["muted"],
                                      textTransform="uppercase", marginBottom=10, fontFamily=C["sans"]))

slbl = section_lbl  # alias

def _field(label, component):
    return html.Div([
        html.Label(label, style=dict(fontSize=9, color=C["muted"], display="block",
                                     marginBottom=4, letterSpacing=1, textTransform="uppercase")),
        component,
    ])

def _dd():
    return dict(backgroundColor=C["surface"], color=C["text"],
                border=f"1px solid {C['border2']}", borderRadius=6,
                fontFamily=C["mono"], fontSize=12)

def fmt_m(x):
    try: return f"${float(x or 0):,.1f}M"
    except: return "$—"

def dd_opts(lst):
    return [{"label": x, "value": x} for x in lst]

def month_options():
    names = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    return [{"label": f"{names[i%12]} {2026+(i//12)}", "value": i} for i in range(60)]

# ── Maths ─────────────────────────────────────────────────────────────────────
def w_avg(deals, key, wkey="nav"):
    total = sum(float(d.get(wkey,0) or 0) for d in deals)
    if not total: return 0.0
    return sum(float(d.get(key,0) or 0)*float(d.get(wkey,0) or 0) for d in deals)/total

def portfolio_metrics(deals):
    if not deals:
        return dict(total_nav=0, num=0, w_irr=0, top1=0, top3=0, top5=0, eff_n=0,
                    total_commit=0, curr_commit=0, total_unfunded=0)
    total = sum(d.get("nav",0) for d in deals)
    w_irr = w_avg(deals, "target_irr")
    weights = [d["nav"]/total for d in deals] if total else []
    eff_n = 1/sum(w**2 for w in weights) if weights else 0
    sd = sorted(deals, key=lambda x: x["nav"], reverse=True)
    top1 = sd[0]["nav"]/total       if sd and total else 0
    top3 = sum(d["nav"] for d in sd[:3])/total if len(sd)>=3 and total else 0
    top5 = sum(d["nav"] for d in sd[:5])/total if len(sd)>=5 and total else 0
    tc = sum(d.get("total_commitment",0) for d in deals)
    cc = sum(d.get("current_commitment",0) for d in deals)
    uf = sum(d.get("total_commitment",0)-d.get("current_commitment",0) for d in deals)
    return dict(total_nav=total, num=len(deals), w_irr=w_irr,
                top1=top1, top3=top3, top5=top5, eff_n=eff_n,
                total_commit=tc, curr_commit=cc, total_unfunded=uf)

def calc_required_irr(current_irr, nav, dry_powder, config):
    target  = float(config.get("target_net_twr", 0.13))
    fee     = float(config.get("management_fee", 0.0125))
    carry   = float(config.get("carry_rate", 0.125))
    hurdle  = float(config.get("hurdle_rate", 0.10))
    loss    = float(config.get("loss_drag", 0.01))
    liq     = float(config.get("liquidity_reserve_pct", 0.05))
    cy      = float(config.get("cash_yield", 0.03))
    invested = max(0, 1-liq)
    gross_needed = (target + fee + loss - liq*cy) / invested if invested else target
    if gross_needed > hurdle:
        gross_needed += (gross_needed-hurdle)*carry
    total = nav+dry_powder
    if not total or not dry_powder: return gross_needed
    req = (gross_needed - (nav/total)*current_irr) / (dry_powder/total)
    return max(0, min(1.5, req))

def bite_sizes(config):
    dp = float(config.get("dry_powder",300) or 300)
    mn = float(config.get("bite_min_pct",0.005))
    ds = float(config.get("bite_desired_pct",0.0275))
    mx = float(config.get("bite_max_pct",0.05))
    return {s: {"min": dp*mn, "desired": dp*ds, "max": dp*mx, "min_pct": mn, "desired_pct": ds, "max_pct": mx}
            for s in STRATEGIES}

def forecast_dp(nav, dry_powder, placeholders, config, months=12):
    rows = []
    base = datetime(2026, 1, 1)
    monthly_ret  = float(config.get("target_net_twr", 0.13))/12
    monthly_dist = float(config.get("distribution_rate", 0.20))/12
    dp = dry_powder
    n  = nav
    for m in range(months):
        d_nav  = n * monthly_ret
        dist   = n * monthly_dist
        calls  = sum(p.get("size",0) for p in (placeholders or []) if p.get("expected_month")==m)
        dp     = dp + dist - calls
        n      = n + d_nav + calls - dist
        rows.append({"month": (base+relativedelta(months=m)).strftime("%b %Y"),
                     "dry_powder": dp, "nav": n, "distributions": dist, "calls": calls})
    return rows

def build_pacing(portfolio, pipeline, placeholders, config, quarters=16):
    fund_size  = float(config.get("fund_size",1000) or 1000)
    dep_years  = float(config.get("deployment_years",4) or 4)
    deals_py   = max(float(config.get("deals_per_year",6) or 6),1)
    dist_rate  = float(config.get("distribution_rate",0.20) or 0.20)
    nav = sum(d.get("nav",0) for d in (portfolio or []))
    ph_map = {}
    for p in (placeholders or []):
        q = max(1, int(p.get("expected_month",0)/3)+1)
        ph_map[q] = ph_map.get(q,0)+p.get("size",0)
    rows = []
    for q in range(1, quarters+1):
        yr = datetime.now().year+(q-1)//4
        label = f"Q{((q-1)%4)+1} {yr}"
        base_deploy = fund_size/(dep_years*4)
        future      = ph_map.get(q,0)
        capacity    = max(fund_size-nav, 0)
        deploy      = min(base_deploy+future, capacity)
        repay       = nav*dist_rate/4 if q>4 else 0
        nav         = max(0, nav+deploy-repay)
        rows.append({"Quarter": label, "Deployment ($M)": round(deploy,1),
                     "Repayments/Dist ($M)": round(repay,1), "NAV ($M)": round(nav,1),
                     "Dry Powder ($M)": round(max(fund_size-nav,0),1),
                     "Utilisation %": round(nav/fund_size*100,1) if fund_size else 0,
                     "Deals (est)": round(deals_py/4,1)})
    return pd.DataFrame(rows)

# ── App ───────────────────────────────────────────────────────────────────────
app = dash.Dash(__name__, title="Horizon PE Secondaries", suppress_callback_exceptions=True)
server = app.server

app.layout = html.Div([
    dcc.Store(id="port-store",       data=INITIAL_PORT),
    dcc.Store(id="pipe-store",       data=INITIAL_PIPE),
    dcc.Store(id="ph-store",         data=INITIAL_PH),
    dcc.Store(id="cfg-store",        data=INITIAL_CFG),
    dcc.Store(id="next-id",          data=INITIAL_NEXT_ID),
    dcc.Store(id="edit-idx",         data=None),
    dcc.Store(id="edit-pipe-idx",    data=None),
    dcc.Store(id="edit-ph-idx",      data=None),
    dcc.Store(id="port-selected-idx",  data=None),   # persists portfolio row selection across re-renders
    dcc.Store(id="pipe-selected-idx",  data=None),   # persists pipeline row selection across re-renders
    dcc.Store(id="ph-selected-idx",    data=None),   # persists deployment plan row selection across re-renders
    dcc.Store(id="populate-port-trigger", data=None),
    dcc.Store(id="populate-pipe-trigger", data=None),
    dcc.Store(id="fund-cf-store",    data=None),   # uploaded Fund Level CF Excel
    dcc.Store(id="liquidity-store",  data=None),   # uploaded Liquidity Pull Excel
    dcc.Store(id="discount-store",   data={}),     # per-deal discount / P2NAV

    # ── Header ────────────────────────────────────────────────────────────────
    html.Div([
        html.Div([
            html.Div([
                html.Span("HORIZON", style=dict(fontSize=19, fontWeight=700, color=C["text"],
                                                 fontFamily=C["sans"], letterSpacing=2)),
                html.Span(" · ", style=dict(color=C["dim"], margin="0 6px")),
                html.Span("PE Secondaries & Co-Investment Dashboard",
                           style=dict(fontSize=16, fontWeight=300, color=C["muted"], fontFamily=C["sans"])),
            ]),
            html.Div(id="save-status-lbl",
                     style=dict(color=C["dim"], marginTop=6, fontSize=11, fontFamily=C["mono"])),
        ]),
        html.Div([
            html.Div([html.Label("Fund Size ($M)",    style=dict(fontSize=9,color=C["muted"],letterSpacing=2,display="block",marginBottom=4,textTransform="uppercase")),
                      dcc.Input(id="hdr-fund-size",  type="number", value=INITIAL_CFG.get("fund_size",1000),  style={**INP,"width":90})]),
            html.Div([html.Label("Dry Powder ($M)",   style=dict(fontSize=9,color=C["muted"],letterSpacing=2,display="block",marginBottom=4,textTransform="uppercase")),
                      dcc.Input(id="hdr-dry-powder", type="number", value=INITIAL_CFG.get("dry_powder",300),  style={**INP,"width":90})]),
            html.Div([html.Label("Target Net TWR",    style=dict(fontSize=9,color=C["muted"],letterSpacing=2,display="block",marginBottom=4,textTransform="uppercase")),
                      dcc.Input(id="hdr-twr",        type="number", value=round(INITIAL_CFG.get("target_net_twr",0.13)*100,2), step=0.25, style={**INP,"width":85})]),
            html.Div([html.Label("Mgmt Fee %",        style=dict(fontSize=9,color=C["muted"],letterSpacing=2,display="block",marginBottom=4,textTransform="uppercase")),
                      dcc.Input(id="hdr-fee",        type="number", value=round(INITIAL_CFG.get("management_fee",0.0125)*100,3), step=0.125, style={**INP,"width":80})]),
        ], style=dict(display="flex", gap=12, alignItems="flex-end", flexWrap="wrap")),
    ], style=dict(background="#070c13", borderBottom=f"1px solid {C['border']}",
                  padding="22px 36px", display="flex", justifyContent="space-between",
                  alignItems="center", flexWrap="wrap", gap=16)),

    # ── KPI strip ─────────────────────────────────────────────────────────────
    html.Div(id="kpi-strip", style=dict(padding="16px 36px", display="flex", gap=10, flexWrap="wrap")),

    # ── Tabs ──────────────────────────────────────────────────────────────────
    html.Div([
        dcc.Tabs(id="tabs", value="overview", children=[
            dcc.Tab(label="🏛 Fund Overview",     value="overview"),
            dcc.Tab(label="📂 Portfolio",         value="portfolio"),
            dcc.Tab(label="🔭 Pipeline",          value="pipeline"),
            dcc.Tab(label="🔮 Pro Forma",         value="proforma"),
            dcc.Tab(label="📅 Deployment Plan",   value="deploy"),
            dcc.Tab(label="📊 Analytics",         value="analytics"),
            dcc.Tab(label="📈 Segments & TWR",    value="segments"),
            dcc.Tab(label="💰 Deal Cashflows",    value="cashflows"),
            dcc.Tab(label="💧 Liquidity",         value="liquidity"),
            dcc.Tab(label="🏦 Dry Powder",        value="drypowder"),
            dcc.Tab(label="♻️ Pacing",            value="pacing"),
            dcc.Tab(label="🧮 Return Calc",       value="returns"),
            dcc.Tab(label="⚙️ Settings",          value="settings"),
        ], colors=dict(border=C["border"], primary=C["blue"], background=C["bg"]),
           style=dict(fontFamily=C["sans"]))
    ], style=dict(padding="0 36px", borderBottom=f"1px solid {C['border']}")),

    html.Div(id="tab-content", style=dict(padding="24px 36px 60px")),

], style=dict(background=C["bg"], minHeight="100vh", fontFamily=C["sans"], color=C["text"]))


# ── Header → config sync ──────────────────────────────────────────────────────
@app.callback(
    Output("cfg-store","data"),
    Input("hdr-fund-size","value"), Input("hdr-dry-powder","value"),
    Input("hdr-twr","value"), Input("hdr-fee","value"),
    State("cfg-store","data"),
    prevent_initial_call=True,
)
def sync_header(fund_size, dry_powder, twr, fee, config):
    # Only update the 4 header-owned keys — never touch other config keys
    # (avoids clobbering legal_restrictions, investment_targets, etc.)
    cfg = dict(config or DEFAULT_CONFIG)
    if fund_size  is not None: cfg["fund_size"]      = float(fund_size)
    if dry_powder is not None: cfg["dry_powder"]     = float(dry_powder)
    if twr        is not None: cfg["target_net_twr"] = float(twr) / 100
    if fee        is not None: cfg["management_fee"] = float(fee) / 100
    return cfg


# ── KPI strip ─────────────────────────────────────────────────────────────────
@app.callback(
    Output("kpi-strip","children"),
    Input("port-store","data"), Input("pipe-store","data"),
    Input("ph-store","data"), Input("cfg-store","data"),
)
def update_kpis(portfolio, pipeline, placeholders, config):
    m   = portfolio_metrics(portfolio or [])
    nav = m["total_nav"]
    dp  = float(config.get("dry_powder",300) or 300)
    req = calc_required_irr(m["w_irr"], nav, dp, config)
    bs  = bite_sizes(config)
    desired = bs.get("GP-Led (Multi-Asset)",{}).get("desired",0)
    sec_nav = sum(d["nav"] for d in (portfolio or []) if d.get("deal_type")=="Secondary")
    ci_nav  = sum(d["nav"] for d in (portfolio or []) if d.get("deal_type")=="Co-Investment")
    tgt_twr = float(config.get("target_net_twr",0.13))
    status  = "✅ On target" if m["w_irr"] >= tgt_twr else "⚠️ Below target"
    return [
        kpi("Investment NAV",  fmt_m(nav),             f"{m['num']} deals",                C["green"]),
        kpi("Dry Powder",      fmt_m(dp),              "Available to deploy",               C["blue"]),
        kpi("Total Fund",      fmt_m(nav+dp),          "NAV + Powder",                      C["sky"]),
        kpi("Portfolio IRR",   f"{m['w_irr']:.1%}",   status,                              C["teal"]),
        kpi("Required IRR",    f"{req:.1%}",           "On future deals",                   C["amber"]),
        kpi("Pipeline",        fmt_m(sum(p.get("size",0) for p in (pipeline or []))),
                               f"{len(pipeline or [])} deals",                              C["purple"]),
        kpi("Pro Forma",       fmt_m(sum(p.get("size",0) for p in (placeholders or []))),
                               f"{len(placeholders or [])} placeholders",                   C["pink"]),
        kpi("Secondaries",     fmt_m(sec_nav),         "In current portfolio",              C["blue"]),
        kpi("Co-Investments",  fmt_m(ci_nav),          "In current portfolio",              C["teal"]),
        kpi("Desired Bite",    fmt_m(desired),         f"{float(config.get('bite_desired_pct',0.0275))*100:.2f}% of DP", C["green"]),
    ]


# ── Tab router ────────────────────────────────────────────────────────────────
@app.callback(
    Output("tab-content","children"),
    Input("tabs","value"),
    Input("port-store","data"), Input("pipe-store","data"),
    Input("ph-store","data"), Input("cfg-store","data"),
    Input("fund-cf-store","data"), Input("liquidity-store","data"),
    Input("discount-store","data"),
)
def route(tab, portfolio, pipeline, placeholders, config, cf_data, liq_data, disc_store):
    p  = portfolio or []
    pi = pipeline  or []
    ph = placeholders or []
    def _safe(fn, *args):
        try:
            return fn(*args)
        except Exception as e:
            import traceback; traceback.print_exc()
            return html.Div([
                html.Div("⚠️ Tab render error", style=dict(color=C["amber"],fontWeight=700,fontFamily=C["mono"],fontSize=14,marginBottom=8)),
                html.Pre(str(e), style=dict(color=C["red"],fontFamily=C["mono"],fontSize=11,
                                            background=C["surface"],padding=12,borderRadius=6,overflowX="auto")),
            ], style=dict(padding=24))
    if tab == "overview":   return _safe(tab_overview, p, pi, ph, config)
    if tab == "portfolio":  return _safe(tab_portfolio, p, config)
    if tab == "pipeline":   return _safe(tab_pipeline, pi, p, config)
    if tab == "proforma":   return _safe(tab_proforma, p, pi, config)
    if tab == "deploy":     return _safe(tab_deploy, ph, config)
    if tab == "analytics":  return _safe(tab_analytics, p, pi, ph, config)
    if tab == "segments":   return _safe(tab_segments, p, pi, ph, config)
    if tab == "cashflows":  return _safe(tab_cashflows, p, cf_data, disc_store)
    if tab == "liquidity":  return _safe(tab_liquidity, liq_data)
    if tab == "drypowder":  return _safe(tab_drypowder, p, ph, config)
    if tab == "pacing":     return _safe(tab_pacing, p, pi, ph, config)
    if tab == "returns":    return _safe(tab_returns, p, pi, config)
    if tab == "settings":   return _safe(tab_settings, config)
    return html.Div()


# ══════════════════════════════════════════════════════════════════════════════
# TAB BUILDERS
# ══════════════════════════════════════════════════════════════════════════════

def _form_row(*cols, gap=8):
    """Render field columns in a flex row."""
    return html.Div(list(cols), style=dict(display="flex", gap=gap, flexWrap="wrap", alignItems="flex-end", marginBottom=10))

def _port_add_form():
    """Add-deal inline form for the Portfolio tab."""
    inp = lambda id_, **kw: dcc.Input(id=id_, **{**{"style": INP}, **kw})
    drp = lambda id_, opts, val=None: dcc.Dropdown(id=id_, options=dd_opts(opts), value=val or opts[0], style=_dd())
    return html.Div([
        section_lbl("Add / Edit Portfolio Deal"),
        _form_row(
            _field("Deal Name *",      inp("port-name",  type="text",   placeholder="Project Alpha", style={**INP,"minWidth":160})),
            _field("Manager *",        inp("port-manager",type="text",  placeholder="Hamilton Lane", style={**INP,"minWidth":150})),
            _field("Strategy *",       drp("port-strategy", STRATEGIES)),
            _field("Deal Type *",      drp("port-dtype",    DEAL_TYPES)),
            _field("Stage *",          drp("port-stage",    STAGES_DEAL, "Buyout")),
            _field("Sector",           drp("port-sector",   SECTORS, "Diversified")),
        ),
        _form_row(
            _field("Region",           drp("port-region",   REGIONS, "North America")),
            _field("Currency",         drp("port-currency", CURRENCIES, "USD")),
            _field("Total Commit $M",  inp("port-total-commit", type="number", value=30, step=0.1)),
            _field("Curr Commit $M",   inp("port-curr-commit",  type="number", value=30, step=0.1)),
            _field("Current NAV $M *", inp("port-nav",          type="number", value=30, step=0.1)),
            html.Div([
                html.Label("Unfunded $M", style=dict(fontSize=9,color=C["muted"],display="block",marginBottom=4,letterSpacing=1,textTransform="uppercase")),
                html.Div(id="unfunded-calc", style=dict(background=C["surface"], border=f"1px solid {C['border2']}", borderRadius=6,
                          padding="6px 10px", fontFamily=C["mono"], fontSize=12, color=C["amber"], minWidth=90)),
            ]),
            _field("Price / NAV",      inp("port-p2nav", type="number", value=1.0, step=0.01,
                                          placeholder="e.g. 0.88")),
        ),
        _form_row(
            _field("Target IRR % *",   inp("port-irr",   type="number", value=18, step=0.5)),
            _field("Hold Period (y)",  inp("port-hold",  type="number", value=5.0, step=0.5)),
            _field("MOIC",             inp("port-moic",  type="number", value=1.75, step=0.01)),
            _field("Vintage",          dcc.Dropdown(id="port-vintage", options=[{"label":str(y),"value":y} for y in VINTAGES], value=2025, style=_dd())),
            _field("Commitment Year",  dcc.Dropdown(id="port-commit-year", options=[{"label":str(y),"value":y} for y in VINTAGES], value=datetime.now().year, style=_dd())),
            _field("Entry Date",       inp("port-entry-date", type="text", placeholder="YYYY-MM-DD", style={**INP,"minWidth":120})),
            _field("Segment",          drp("port-segment",     SEGMENTS,     "New")),
            _field("Alloc Status",     drp("port-alloc-status",ALLOC_STATUS, "Closed")),
            html.Div([
                html.Label(" ", style=dict(display="block", marginBottom=4, fontSize=9)),
                html.Button("+ Add Deal", id="add-port-btn", style={**BTN(C["blue"]), "minWidth":100}),
            ]),
            html.Div([
                html.Label(" ", style=dict(display="block", marginBottom=4, fontSize=9)),
                html.Button("Clear",     id="clear-port-btn", style={**BTN(C["surface"], C["muted"]), "border":f"1px solid {C['border2']}","minWidth":70}),
            ]),
        ),
        html.Div(id="port-msg", style=dict(color=C["green"], fontSize=11, marginTop=4)),
    ], style=dict(background=C["surface"], border=f"1px solid {cl(C['blue'],0.4)}", borderRadius=8, padding=18, marginBottom=18))


def tab_overview(portfolio, pipeline, placeholders, config):
    """Fund Overview — identity, investment mandate, allocation targets, legal limits + live Monitor."""
    m      = portfolio_metrics(portfolio)
    nav    = m["total_nav"]
    dp     = float(config.get("dry_powder", 300) or 300)
    fs     = float(config.get("fund_size",  1000) or 1000)

    def pct(v): return float(v or 0)

    # ── Compute all possible actuals (keyed by metric_key) ───────────────────
    sec_nav  = sum(d["nav"] for d in portfolio if d.get("deal_type")=="Secondary")
    ci_nav   = sum(d["nav"] for d in portfolio if d.get("deal_type")=="Co-Investment")
    sec_pct  = sec_nav/nav*100 if nav else 0
    ci_pct   = ci_nav /nav*100 if nav else 0

    by_region  = {}
    by_sector  = {}
    by_manager = {}
    by_vintage = {}
    for d in portfolio:
        by_region [d.get("region","?")]  = by_region.get(d.get("region","?"),0)  + d["nav"]
        by_sector [d.get("sector","?")]  = by_sector.get(d.get("sector","?"),0)  + d["nav"]
        by_manager[d.get("manager","?")] = by_manager.get(d.get("manager","?"),0)+ d["nav"]
        vk = str(d.get("vintage","?"))
        by_vintage[vk] = by_vintage.get(vk, 0) + d["nav"]

    max_region_pct  = max((v/nav*100 for v in by_region.values()),  default=0) if nav else 0
    max_sector_pct  = max((v/nav*100 for v in by_sector.values()),  default=0) if nav else 0
    max_manager_pct = max((v/nav*100 for v in by_manager.values()), default=0) if nav else 0
    max_deal_pct    = m["top1"]*100
    unfunded_pct    = m["total_unfunded"]/fs*100 if fs else 0
    max_vintage_pct = max((v/nav*100 for v in by_vintage.values()), default=0) if nav else 0
    max_vintage_yr  = max(by_vintage, key=lambda k: by_vintage[k]) if by_vintage else "—"
    leverage_pct    = 0.0   # placeholder until leverage field added to deal form
    primary_nav     = sum(d["nav"] for d in portfolio if d.get("strategy","") == "Primary")
    primary_pct     = primary_nav/nav*100 if nav else 0
    non_sec_nav     = sum(d["nav"] for d in portfolio if d.get("deal_type","") != "Secondary")
    non_sec_pct     = non_sec_nav/nav*100 if nav else 0
    na_eur_nav      = by_region.get("North America",0) + by_region.get("Europe",0)
    non_naeur_pct   = (nav - na_eur_nav)/nav*100 if nav else 0
    # Non-NA only (excludes Europe from the non-NA bucket)
    na_nav          = by_region.get("North America",0)
    non_na_pct      = (nav - na_nav)/nav*100 if nav else 0
    # Listed company exposure — placeholder until deal-level flag added
    listed_pct      = 0.0
    # Overcommitment = total_commitment / NAV * 100 (reflects unfunded obligations vs portfolio value)
    overcommitment_pct = m["total_commit"]/nav*100 if nav else 0

    ACTUALS = {
        "max_deal_pct":        max_deal_pct,
        "max_manager_pct":     max_manager_pct,
        "ci_pct":              ci_pct,
        "max_sector_pct":      max_sector_pct,
        "max_region_pct":      max_region_pct,
        "unfunded_pct":        unfunded_pct,
        "num_deals":           float(m["num"]),
        "leverage_pct":        leverage_pct,
        "primary_pct":         primary_pct,
        "non_sec_pct":         non_sec_pct,
        "non_naeur_pct":       non_naeur_pct,
        "non_na_pct":          non_na_pct,
        "max_vintage_pct":     max_vintage_pct,
        "sec_pct":             sec_pct,
        "listed_pct":          listed_pct,
        "overcommitment_pct":  overcommitment_pct,
    }

    # ── Compliance row builder ────────────────────────────────────────────────
    def compliance_row(label, actual, limit, higher_is_bad=True, fmt="%"):
        ok = (actual <= limit) if higher_is_bad else (actual >= limit)
        color = C["green"] if ok else C["red"]
        icon  = "✅" if ok else "❌"
        actual_s = f"{actual:.1f}%" if fmt=="%" else f"{actual:.0f}"
        limit_s  = f"{'≤' if higher_is_bad else '≥'}{limit:.0f}%" if fmt=="%" else f"{'≥' if not higher_is_bad else '≤'}{limit:.0f}"
        return html.Tr([
            html.Td(f"{icon} {label}", style=dict(fontFamily=C["mono"],fontSize=12,padding="7px 12px",color=C["text"])),
            html.Td(actual_s,          style=dict(fontFamily=C["mono"],fontSize=12,padding="7px 12px",color=color,fontWeight=700,textAlign="right")),
            html.Td(limit_s,           style=dict(fontFamily=C["mono"],fontSize=12,padding="7px 12px",color=C["muted"],textAlign="right")),
            html.Td("Pass" if ok else "BREACH",
                    style=dict(fontFamily=C["mono"],fontSize=11,padding="7px 12px",
                               color=color,fontWeight=700,textAlign="center")),
        ])

    def make_table(restrictions, title, accent):
        rows_html = []
        n_breach  = 0
        for r in restrictions:
            mk  = r.get("metric_key","")
            lbl = r.get("label", mk)
            # append vintage year to label when relevant
            if mk == "max_vintage_pct" and max_vintage_yr != "—":
                lbl = f"{lbl} ({max_vintage_yr})"
            actual = ACTUALS.get(mk, 0.0)
            limit  = float(r.get("limit", 0))
            hib    = bool(r.get("higher_is_bad", True))
            fmt    = r.get("fmt", "%")
            ok     = (actual <= limit) if hib else (actual >= limit)
            if not ok:
                n_breach += 1
            rows_html.append(compliance_row(lbl, actual, limit, hib, fmt))

        badge_color = C["green"] if n_breach == 0 else C["red"]
        badge_text  = "All Clear" if n_breach == 0 else f"{n_breach} Breach{'es' if n_breach>1 else ''}"

        header = html.Div([
            html.Span(title, style=dict(fontSize=9,letterSpacing=2.5,color=C["muted"],
                                        textTransform="uppercase",fontFamily=C["sans"],fontWeight=700)),
            html.Span(badge_text, style=dict(
                fontSize=10, fontFamily=C["mono"], fontWeight=700, color=badge_color,
                background=cl(badge_color,0.12), border=f"1px solid {cl(badge_color,0.35)}",
                borderRadius=4, padding="2px 8px", marginLeft=10,
            )),
        ], style=dict(display="flex",alignItems="center",marginBottom=8))

        tbl = html.Table([
            html.Thead(html.Tr([
                html.Th("Restriction / Target",  style=dict(**TBL_HEAD, textAlign="left")),
                html.Th("Current",               style=dict(**TBL_HEAD, textAlign="right")),
                html.Th("Limit",                 style=dict(**TBL_HEAD, textAlign="right")),
                html.Th("Status",                style=dict(**TBL_HEAD, textAlign="center")),
            ])),
            html.Tbody(rows_html),
        ], style=dict(width="100%", borderCollapse="collapse", background=C["surface"], borderRadius=8))

        return n_breach, card([header, tbl], dict(flex=1, minWidth=340))

    # Load restriction lists from config (fall back to defaults if not yet saved)
    legal_list  = config.get("legal_restrictions",  DEFAULT_CONFIG["legal_restrictions"])
    target_list = config.get("investment_targets",  DEFAULT_CONFIG["investment_targets"])

    legal_breaches,  legal_card  = make_table(legal_list,  "Legal Restrictions",  C["red"])
    target_breaches, target_card = make_table(target_list, "Investment Targets",  C["amber"])
    n_breaches = legal_breaches + target_breaches
    def alloc_bar(label, actual, target, color):
        w_actual = min(actual, 100)
        w_target = min(target, 100)
        return html.Div([
            html.Div([
                html.Span(label, style=dict(fontSize=11,color=C["text"],fontFamily=C["mono"],flex=1)),
                html.Span(f"{actual:.1f}%", style=dict(fontSize=11,color=color,fontFamily=C["mono"],fontWeight=700)),
                html.Span(f" / {target:.0f}% target", style=dict(fontSize=10,color=C["muted"],fontFamily=C["mono"])),
            ], style=dict(display="flex",alignItems="center",marginBottom=4)),
            html.Div([
                html.Div(style=dict(width=f"{w_actual:.1f}%", height=6,
                                    background=color, borderRadius=3, position="absolute", top=0, left=0)),
                html.Div(style=dict(left=f"{w_target:.1f}%", width=2, height=14,
                                    background=C["amber"], borderRadius=1,
                                    position="absolute", top=-4)),
            ], style=dict(position="relative", height=6, background=C["border"],
                          borderRadius=3, marginBottom=12)),
        ])

    na_pct     = by_region.get("North America",0)/nav*100 if nav else 0
    eur_pct    = by_region.get("Europe",0)/nav*100        if nav else 0
    asia_pct   = by_region.get("Asia",0)/nav*100          if nav else 0

    alloc_section = card([
        slbl("Allocation vs Targets"),
        html.Div([
            html.Div([
                slbl("Strategy"),
                alloc_bar("Secondaries",    sec_pct, pct(config.get("target_secondary_pct",70)), C["blue"]),
                alloc_bar("Co-Investments", ci_pct,  pct(config.get("target_coinvest_pct",30)),  C["teal"]),
            ], style=dict(flex=1, minWidth=280)),
            html.Div([
                slbl("Geography"),
                alloc_bar("North America", na_pct,  pct(config.get("target_na_pct",50)),     C["sky"]),
                alloc_bar("Europe",        eur_pct, pct(config.get("target_europe_pct",35)), C["purple"]),
                alloc_bar("Asia",          asia_pct,pct(config.get("target_asia_pct",10)),   C["green"]),
            ], style=dict(flex=1, minWidth=280)),
        ], style=dict(display="flex", gap=32, flexWrap="wrap")),
    ], dict(marginBottom=16))

    # ── Fund identity card ────────────────────────────────────────────────────
    def info_row(label, value):
        return html.Div([
            html.Span(label, style=dict(fontSize=9,letterSpacing=2,color=C["muted"],
                                        textTransform="uppercase",fontFamily=C["sans"],minWidth=160,display="inline-block")),
            html.Span(value, style=dict(fontSize=12,color=C["text"],fontFamily=C["mono"],fontWeight=600)),
        ], style=dict(padding="6px 0", borderBottom=f"1px solid {C['border']}"))

    identity = card([
        slbl("Fund Identity"),
        info_row("Fund Name",     str(config.get("fund_name","Horizon"))),
        info_row("Strategy",      str(config.get("fund_strategy","PE Secondaries & Co-Investment"))),
        info_row("Vehicle Type",  str(config.get("vehicle_type","Evergreen"))),
        info_row("Domicile",      str(config.get("domicile","Luxembourg"))),
        info_row("Fund Currency", str(config.get("fund_currency","USD"))),
        info_row("Vintage Year",  str(config.get("fund_vintage",2024))),
        info_row("Fund Size",     fmt_m(fs)),
        info_row("Dry Powder",    fmt_m(dp)),
        info_row("Portfolio NAV", fmt_m(nav)),
        info_row("# Positions",   str(m["num"])),
    ], dict(flex=1, minWidth=280))

    # ── Return targets card ───────────────────────────────────────────────────
    tgt_sec  = pct(config.get("target_irr_secondary", 15))
    tgt_ci   = pct(config.get("target_irr_coinvest",  20))
    tgt_moic = float(config.get("target_moic", 1.8))
    tgt_twr  = float(config.get("target_net_twr", 0.13))*100

    curr_sec_irr = w_avg([d for d in portfolio if d.get("deal_type")=="Secondary"],"target_irr")*100
    curr_ci_irr  = w_avg([d for d in portfolio if d.get("deal_type")=="Co-Investment"],"target_irr")*100
    curr_moic    = w_avg(portfolio,"moic") if portfolio else 0

    def ret_row(label, actual, target, suffix="%"):
        ok = actual >= target
        return html.Div([
            html.Span(label, style=dict(fontSize=9,letterSpacing=2,color=C["muted"],
                                        textTransform="uppercase",fontFamily=C["sans"],minWidth=160,display="inline-block")),
            html.Span(f"{actual:.1f}{suffix}",
                      style=dict(fontSize=14,color=C["green"] if ok else C["amber"],fontFamily=C["mono"],fontWeight=700)),
            html.Span(f"  target {target:.1f}{suffix}",
                      style=dict(fontSize=10,color=C["dim"],fontFamily=C["mono"])),
        ], style=dict(padding="8px 0", borderBottom=f"1px solid {C['border']}"))

    returns = card([
        slbl("Return Targets"),
        ret_row("Target Net TWR",     tgt_twr,  tgt_twr),
        ret_row("Secondary IRR",      curr_sec_irr, tgt_sec),
        ret_row("Co-Invest IRR",      curr_ci_irr,  tgt_ci),
        ret_row("Portfolio MOIC",     curr_moic,    tgt_moic, suffix="x"),
    ], dict(flex=1, minWidth=280))

    # ── Summary KPIs ─────────────────────────────────────────────────────────
    compliance_color = C["green"] if n_breaches == 0 else C["red"]
    compliance_label = "All Clear" if n_breaches == 0 else f"{n_breaches} Breach{'es' if n_breaches>1 else ''}"

    top_kpis = html.Div([
        kpi("Fund NAV",          fmt_m(nav),           f"{m['num']} positions",          C["green"]),
        kpi("Dry Powder",        fmt_m(dp),            f"{dp/fs*100:.0f}% of fund",      C["blue"]),
        kpi("Secondaries",       f"{sec_pct:.1f}%",    f"Target {pct(config.get('target_secondary_pct',70)):.0f}%", C["sky"]),
        kpi("Co-Investments",    f"{ci_pct:.1f}%",     f"Target {pct(config.get('target_coinvest_pct',30)):.0f}%",  C["teal"]),
        kpi("Portfolio IRR",     f"{m['w_irr']*100:.1f}%", "NAV-weighted",              C["purple"]),
        kpi("Targets and Restrictions", compliance_label, "Mandate restrictions",       compliance_color),
    ], style=dict(display="flex",gap=10,flexWrap="wrap",marginBottom=16))

    return html.Div([
        top_kpis,
        html.Div([
            html.Div([identity, html.Div(style=dict(height=16)), returns], style=dict(flex=1, minWidth=280)),
            html.Div([
                alloc_section,
                html.Div([legal_card, target_card],
                         style=dict(display="flex", gap=16, flexWrap="wrap")),
            ], style=dict(flex=2, minWidth=400)),
        ], style=dict(display="flex", gap=16, flexWrap="wrap")),
    ])


def tab_portfolio(portfolio, config):
    m = portfolio_metrics(portfolio)
    total = m["total_nav"] or 1
    rows = []
    for i, d in enumerate(portfolio):
        uf = d.get("total_commitment",d.get("nav",0)) - d.get("current_commitment",d.get("nav",0))
        p2nav = d.get("price_to_nav", None)
        rows.append({
            "_idx": i,
            "Deal": d["name"], "Manager": d.get("manager",""),
            "Strategy": d.get("strategy",""), "Type": d.get("deal_type",""),
            "Stage": d.get("stage",""), "Sector": d.get("sector",""),
            "Region": d.get("region",""), "CCY": d.get("currency",""),
            "Total Commit": fmt_m(d.get("total_commitment",d.get("nav",0))),
            "Curr Commit":  fmt_m(d.get("current_commitment",d.get("nav",0))),
            "NAV":          fmt_m(d.get("nav",0)),
            "Unfunded":     fmt_m(uf),
            "P/NAV":        f"{p2nav:.3f}x" if p2nav is not None else "—",
            "IRR %":        f"{d.get('target_irr',0)*100:.1f}%",
            "MOIC":         f"{d.get('moic',0):.2f}x",
            "Hold (y)":     f"{d.get('hold_period',0):.1f}",
            "Weight":       f"{d.get('nav',0)/total*100:.1f}%",
            "Vintage":      str(d.get("vintage","")),
            "Commit Year":  str(d.get("commitment_year",d.get("vintage",""))),
            "Entry Date":   str(d.get("entry_date","")),
            "Segment":      d.get("segment",""),
            "Status":       d.get("allocation_status",""),
        })

    # Charts
    by_strat = {}
    by_type  = {"Secondary":0,"Co-Investment":0}
    for d in portfolio:
        k = d.get("strategy","?")
        by_strat[k] = by_strat.get(k,0)+d.get("nav",0)
        t = d.get("deal_type","Secondary")
        by_type[t]  = by_type.get(t,0)+d.get("nav",0)

    pie_colors = [C["blue"],C["purple"],C["teal"],C["green"],C["amber"]]
    fig = make_subplots(1,2,specs=[[{"type":"domain"},{"type":"domain"}]],
                        subplot_titles=["Exposure by Strategy","Secondary vs Co-Invest"])
    if any(by_strat.values()):
        fig.add_trace(go.Pie(labels=list(by_strat.keys()),values=list(by_strat.values()),hole=0.55,
                             marker_colors=pie_colors[:len(by_strat)],textfont_color=C["text"],showlegend=True),1,1)
    if any(by_type.values()):
        fig.add_trace(go.Pie(labels=list(by_type.keys()),values=list(by_type.values()),hole=0.55,
                             marker_colors=[C["sky"],C["teal"]],textfont_color=C["text"],showlegend=True),1,2)
    fig.update_layout(**chart(height=280, margin=dict(t=50,b=10,l=10,r=10)))

    # Summary KPIs for this tab
    tab_kpis = html.Div([
        kpi("Total NAV",          fmt_m(m["total_nav"]),      f"{m['num']} deals",            C["green"],  130),
        kpi("Total Commitment",   fmt_m(m["total_commit"]),   "Across portfolio",             C["blue"],   130),
        kpi("Curr Commitment",    fmt_m(m["curr_commit"]),    "Capital called",               C["sky"],    130),
        kpi("Unfunded",           fmt_m(m["total_unfunded"]), "Remaining",                    C["amber"],  120),
        kpi("Portfolio IRR",      f"{m['w_irr']*100:.1f}%",  "NAV-weighted",                 C["teal"],   120),
        kpi("Top 1 Conc",         f"{m['top1']*100:.1f}%",   "Limit 15%",                    C["red"] if m["top1"]>0.15 else C["green"], 110),
        kpi("Eff. # Positions",   f"{m['eff_n']:.1f}",       "Herfindahl",                   C["purple"], 110),
    ], style=dict(display="flex",gap=10,flexWrap="wrap",marginBottom=16))

    table = dash_table.DataTable(
        id="portfolio-table", data=rows,
        columns=[{"name":c,"id":c} for c in rows[0].keys()] if rows else [],
        hidden_columns=["_idx"],
        row_selectable="single", selected_rows=[],
        style_cell=TBL_CELL, style_header=TBL_HEAD,
        style_data_conditional=TBL_ODD+[
            {"if":{"column_id":"Type","filter_query":'{Type} = "Co-Investment"'},"color":C["teal"],"fontWeight":700},
            {"if":{"column_id":"IRR %"},"color":C["green"]},
            {"if":{"column_id":"NAV"},"color":C["sky"]},
        ],
        sort_action="native", filter_action="native",
        page_size=15, style_table={"overflowX":"auto"},
        export_format="xlsx",
    )

    actions = card([
        section_lbl("Actions"),
        html.Button("✏️ Edit Selected",   id="edit-port-btn",  style={**BTN(C["amber"]), "width":"100%", "marginBottom":8}),
        html.Button("🗑 Delete Selected", id="delete-port-btn",style={**BTN(C["red"]),   "width":"100%", "marginBottom":8}),
        html.Div(style=dict(height=4)),
        html.Div("─ NAV Sync ─", style=dict(fontSize=9,color=C["dim"],letterSpacing=2,textAlign="center",marginBottom=6,fontFamily=C["sans"])),
        html.Button("📥 Sync NAVs from CF",  id="sync-nav-cf-btn",  style={**BTN(C["blue"],  "#fff"), "width":"100%", "marginBottom":6}),
        html.Button("📥 Sync NAVs from Liq", id="sync-nav-liq-btn", style={**BTN(C["teal"],  "#fff"), "width":"100%"}),
        html.Div(id="port-action-msg", style=dict(marginTop=10, color=C["muted"], fontSize=11)),
    ], dict(minWidth=180))

    return html.Div([
        _port_add_form(),
        tab_kpis,
        html.Div([
            html.Div(table, style=dict(flex=1, minWidth=0, overflow="hidden")),
        ], style=dict(marginBottom=16)),
        html.Div([
            html.Div(card([dcc.Graph(figure=fig, config={"displayModeBar":False})]),
                     style=dict(flex=2, minWidth=300, maxWidth="60%")),
            html.Div(actions, style=dict(flex=1, minWidth=200, maxWidth=260)),
        ], style=dict(display="flex", gap=16, flexWrap="wrap", alignItems="flex-start")),
    ])


def tab_pipeline(pipeline, portfolio, config):
    rows = []
    for i, d in enumerate(pipeline):
        rows.append({
            "_idx": i,
            "Deal": d["name"], "Strategy": d.get("strategy",""),
            "Deal Stage": d.get("stage_deal",""), "Sector": d.get("sector",""),
            "Region": d.get("region",""), "Size ($M)": fmt_m(d.get("size",0)),
            "Target IRR": f"{d.get('target_irr',0)*100:.1f}%",
            "Pipeline Stage": d.get("pipeline_stage",""),
            "Priority": d.get("priority",""),
        })

    stages = PIPE_STAGES
    stage_vals = [sum(d.get("size",0) for d in pipeline if d.get("pipeline_stage")==s) for s in stages]
    fig_funnel = go.Figure(go.Funnel(y=stages, x=stage_vals, textinfo="value+percent initial",
                                     marker_color=[C["blue"],C["teal"],C["amber"],C["purple"],C["green"]],
                                     textfont=dict(color=C["text"])))
    fig_funnel.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40), xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1), height=300, title="Pipeline Funnel ($M)")

    # Add form
    form = html.Div([
        html.Div([
            section_lbl("Add / Edit Pipeline Deal"),
            html.Div(id="pipe-form-mode", style=dict(fontSize=10,color=C["amber"],fontFamily=C["mono"],marginBottom=6)),
        ]),
        _form_row(
            _field("Deal Name",       dcc.Input(id="pipe-name",       type="text",  style=INP, placeholder="Project X")),
            _field("Strategy",        dcc.Dropdown(id="pipe-strategy", options=dd_opts(STRATEGIES), value=STRATEGIES[0],  style=_dd())),
            _field("Deal Stage",      dcc.Dropdown(id="pipe-stage-deal",options=dd_opts(STAGES_DEAL),value="Buyout",       style=_dd())),
            _field("Sector",          dcc.Dropdown(id="pipe-sector",   options=dd_opts(SECTORS),    value="Technology",   style=_dd())),
            _field("Region",          dcc.Dropdown(id="pipe-region",   options=dd_opts(REGIONS),    value="North America",style=_dd())),
            _field("Size ($M)",       dcc.Input(id="pipe-size",        type="number",value=25, step=0.5, style=INP)),
            _field("Target IRR %",    dcc.Input(id="pipe-irr",         type="number",value=18, step=0.5, style=INP)),
            _field("Pipeline Stage",  dcc.Dropdown(id="pipe-pipe-stage",options=dd_opts(PIPE_STAGES),value="Screening",   style=_dd())),
            _field("Priority",        dcc.Dropdown(id="pipe-priority", options=dd_opts(PRIORITIES), value="Medium",       style=_dd())),
            html.Div([html.Label(" ",style=dict(display="block",marginBottom=4,fontSize=9)),
                      html.Button("+ Add",id="add-pipe-btn",style={**BTN(C["purple"]),"minWidth":80})]),
            html.Div([html.Label(" ",style=dict(display="block",marginBottom=4,fontSize=9)),
                      html.Button("Clear",id="clear-pipe-btn",style={**BTN(C["surface"],C["muted"]),"border":f"1px solid {C['border2']}","minWidth":60})]),
        ),
        html.Div(id="pipe-msg",style=dict(color=C["green"],fontSize=11,marginTop=4)),
    ], style=dict(background=C["surface"],border=f"1px solid {cl(C['purple'],0.4)}",borderRadius=8,padding=18,marginBottom=18))

    table = dash_table.DataTable(
        id="pipeline-table", data=rows,
        columns=[{"name":c,"id":c} for c in rows[0].keys()] if rows else [],
        hidden_columns=["_idx"],
        row_selectable="single", selected_rows=[],
        style_cell=TBL_CELL, style_header=TBL_HEAD,
        style_data_conditional=TBL_ODD+[
            {"if":{"filter_query":'{Priority} = "High"',  "column_id":"Priority"},"color":C["red"],   "fontWeight":700},
            {"if":{"filter_query":'{Priority} = "Medium"',"column_id":"Priority"},"color":C["amber"]},
            {"if":{"filter_query":'{Priority} = "Low"',   "column_id":"Priority"},"color":C["green"]},
        ],
        sort_action="native", filter_action="native",
        page_size=12, style_table={"overflowX":"auto"}, export_format="xlsx",
    )

    actions = card([
        section_lbl("Pipeline Actions"),
        html.Button("✏️ Edit Selected",       id="edit-pipe-btn",    style={**BTN(C["amber"]), "width":"100%", "marginBottom":8}),
        html.Button("✅ Promote → Portfolio", id="promote-pipe-btn", style={**BTN(C["green"]), "width":"100%", "marginBottom":8}),
        html.Button("🗑 Delete Selected",     id="delete-pipe-btn",  style={**BTN(C["red"]),   "width":"100%"}),
        html.Div(id="pipe-action-msg",style=dict(marginTop=10,color=C["muted"],fontSize=11)),
    ], dict(minWidth=200))

    return html.Div([
        form,
        html.Div([
            html.Div(table, style=dict(flex=3, minWidth=0)),
            html.Div([dcc.Graph(figure=fig_funnel,config={"displayModeBar":False}), actions],
                     style=dict(flex=1,display="flex",flexDirection="column",gap=16,minWidth=280)),
        ], style=dict(display="flex",gap=16,flexWrap="wrap")),
    ])


def tab_proforma(portfolio, pipeline, config):
    """Pro Forma = Current Portfolio + Pipeline deals (read-only view)."""
    m_curr = portfolio_metrics(portfolio)
    pipe_nav = sum(p.get("size",0) for p in pipeline)
    all_nav  = m_curr["total_nav"] + pipe_nav
    pipe_irr = w_avg(pipeline,"target_irr","size") if pipeline else 0
    blended_irr = (m_curr["total_nav"]*m_curr["w_irr"] + pipe_nav*pipe_irr)/all_nav if all_nav else 0

    # Summary KPIs
    kpi_row = html.Div([
        kpi("Current NAV",       fmt_m(m_curr["total_nav"]), f"{m_curr['num']} deals",         C["green"]),
        kpi("Pipeline (pending)",fmt_m(pipe_nav),            f"{len(pipeline)} deals",          C["purple"]),
        kpi("Pro Forma NAV",     fmt_m(all_nav),             "Current + Pipeline",              C["sky"]),
        kpi("Pro Forma IRR",     f"{blended_irr*100:.1f}%", "Blended NAV-weighted",            C["teal"]),
        kpi("Total Commit",      fmt_m(m_curr["total_commit"]+pipe_nav), "Incl. pipeline",     C["blue"]),
        kpi("Unfunded",          fmt_m(m_curr["total_unfunded"]),        "Current only",        C["amber"]),
    ], style=dict(display="flex",gap=10,flexWrap="wrap",marginBottom=16))

    # Combined table: current deals + pipeline — full portfolio column set
    total = all_nav or 1
    rows = []
    for d in portfolio:
        uf = d.get("total_commitment", d.get("nav",0)) - d.get("current_commitment", d.get("nav",0))
        p2nav = d.get("price_to_nav", None)
        rows.append({
            "Source":         "✅ Current",
            "Deal":           d["name"],
            "Manager":        d.get("manager",""),
            "Strategy":       d.get("strategy",""),
            "Type":           d.get("deal_type",""),
            "Stage":          d.get("stage",""),
            "Sector":         d.get("sector",""),
            "Region":         d.get("region",""),
            "CCY":            d.get("currency",""),
            "Total Commit":   fmt_m(d.get("total_commitment", d.get("nav",0))),
            "Curr Commit":    fmt_m(d.get("current_commitment", d.get("nav",0))),
            "NAV":            fmt_m(d.get("nav",0)),
            "Unfunded":       fmt_m(uf),
            "P/NAV":          f"{p2nav:.3f}x" if p2nav is not None else "—",
            "IRR %":          f"{d.get('target_irr',0)*100:.1f}%",
            "MOIC":           f"{d.get('moic',0):.2f}x",
            "Hold (y)":       f"{d.get('hold_period',0):.1f}",
            "Weight":         f"{d.get('nav',0)/total*100:.1f}%",
            "Vintage":        str(d.get("vintage","")),
            "Segment":        d.get("segment",""),
            "Status":         d.get("allocation_status",""),
        })
    for p in pipeline:
        size = p.get("size",0)
        rows.append({
            "Source":         "⏳ Pipeline",
            "Deal":           p["name"],
            "Manager":        "",
            "Strategy":       p.get("strategy",""),
            "Type":           p.get("deal_type","Secondary"),
            "Stage":          p.get("stage_deal",""),
            "Sector":         p.get("sector",""),
            "Region":         p.get("region",""),
            "CCY":            "—",
            "Total Commit":   fmt_m(size),
            "Curr Commit":    "—",
            "NAV":            fmt_m(size),
            "Unfunded":       "—",
            "P/NAV":          "—",
            "IRR %":          f"{p.get('target_irr',0)*100:.1f}%",
            "MOIC":           "—",
            "Hold (y)":       "—",
            "Weight":         f"{size/total*100:.1f}%",
            "Vintage":        "—",
            "Segment":        p.get("pipeline_stage",""),
            "Status":         p.get("priority",""),
        })

    tbl = dash_table.DataTable(
        id="proforma-table", data=rows,
        columns=[{"name":c,"id":c} for c in rows[0].keys()] if rows else [],
        style_cell=TBL_CELL, style_header=TBL_HEAD,
        row_selectable="single", selected_rows=[],
        style_data_conditional=TBL_ODD+[
            {"if":{"filter_query":'{Source} contains "Pipeline"'},"color":C["purple"],"fontStyle":"italic"},
            {"if":{"column_id":"Type","filter_query":'{Type} = "Co-Investment"'},"color":C["teal"],"fontWeight":700},
            {"if":{"column_id":"IRR %"},"color":C["green"]},
            {"if":{"column_id":"NAV"},"color":C["sky"]},
        ],
        sort_action="native", filter_action="native",
        page_size=20, style_table={"overflowX":"auto"}, export_format="xlsx",
    ) if rows else html.P("No deals yet.", style=dict(color=C["muted"]))

    pf_actions = card([
        section_lbl("Edit Deal"),
        html.Div("Select a row, then navigate to the Portfolio or Pipeline tab to edit it.",
                 style=dict(color=C["muted"],fontSize=11,fontFamily=C["mono"],marginBottom=8)),
        html.Button("✏️ Edit Portfolio Deal →",id="pf-goto-port-btn",style={**BTN(C["amber"]),"width":"100%","marginBottom":8}),
        html.Button("✏️ Edit Pipeline Deal →", id="pf-goto-pipe-btn",style={**BTN(C["purple"]),"width":"100%"}),
        html.Div(id="pf-action-msg",style=dict(marginTop=8,color=C["muted"],fontSize=11)),
    ], dict(minWidth=200))

    # Strategy breakdown: current vs proforma comparison
    by_strat_curr = {}
    by_strat_pf   = {}
    for d in portfolio:
        k = d.get("strategy","?")
        by_strat_curr[k] = by_strat_curr.get(k,0)+d.get("nav",0)
        by_strat_pf[k]   = by_strat_pf.get(k,0)+d.get("nav",0)
    for p in pipeline:
        k = p.get("strategy","?")
        by_strat_pf[k] = by_strat_pf.get(k,0)+p.get("size",0)

    all_strats = sorted(set(list(by_strat_curr.keys())+list(by_strat_pf.keys())))
    fig = go.Figure()
    fig.add_trace(go.Bar(name="Current", x=all_strats,
                         y=[by_strat_curr.get(s,0) for s in all_strats],
                         marker_color=C["blue"]))
    fig.add_trace(go.Bar(name="+ Pipeline", x=all_strats,
                         y=[by_strat_pf.get(s,0)-by_strat_curr.get(s,0) for s in all_strats],
                         marker_color=cl(C["purple"],0.7)))
    fig.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40), xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1), height=320, barmode="stack",
                      title="Strategy Exposure: Current vs Pro Forma",
                      yaxis_title="NAV ($M)")

    return html.Div([
        html.Div([
            html.Span("Pro Forma = Current Portfolio + Pipeline Deals",
                      style=dict(fontSize=12, color=C["muted"], fontFamily=C["mono"])),
            html.Span("  (select a row and use the edit buttons to jump to the right tab)",
                      style=dict(fontSize=11, color=C["dim"], fontFamily=C["mono"])),
        ], style=dict(marginBottom=12)),
        kpi_row,
        html.Div([
            html.Div(card([dcc.Graph(figure=fig, config={"displayModeBar":False})]), style=dict(flex=1)),
        ], style=dict(marginBottom=16)),
        html.Div([
            html.Div(card([section_lbl("Pro Forma Full Deal List"), tbl]), style=dict(flex=3,minWidth=0)),
            html.Div(pf_actions, style=dict(flex=1,minWidth=200)),
        ], style=dict(display="flex",gap=16,flexWrap="wrap")),
    ])


def tab_deploy(placeholders, config):
    """Deployment Planner — plan future commitments to pace dry powder usage."""
    mo = month_options()
    rows = []
    for i, p in enumerate(placeholders):
        midx = p.get("expected_month",0)
        mlbl = mo[midx]["label"] if 0<=midx<len(mo) else f"M{midx}"
        rows.append({
            "_idx":i, "Name":p["name"], "Strategy":p.get("strategy",""),
            "Type":p.get("deal_type",""), "Region":p.get("region",""),
            "Expected Size $M":fmt_m(p.get("size",0)),
            "Target IRR %":f"{p.get('target_irr',0)*100:.1f}%",
            "Expected":mlbl,
        })

    by_month = {}
    for p in placeholders:
        midx = p.get("expected_month",0)
        mlbl = mo[midx]["label"] if 0<=midx<len(mo) else "?"
        by_month[mlbl] = by_month.get(mlbl,0)+p.get("size",0)

    fig = go.Figure()
    if by_month:
        fig.add_trace(go.Bar(x=list(by_month.keys()), y=list(by_month.values()),
                             marker_color=C["teal"],
                             text=[fmt_m(v) for v in by_month.values()],
                             textposition="outside", textfont_color=C["text"]))
    fig.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40), xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1), height=300, title="Planned Commitment Schedule ($M)")

    form = html.Div([
        html.Div([
            section_lbl("Add / Edit Planned Future Commitment"),
            html.Div(id="ph-form-mode", style=dict(fontSize=10,color=C["amber"],fontFamily=C["mono"],marginBottom=6)),
        ]),
        html.Div([
            _field("Name",         dcc.Input(id="ph-name",     type="text",  style=INP, placeholder="GP-Led Placeholder 1")),
            _field("Strategy",     dcc.Dropdown(id="ph-strategy", options=dd_opts(STRATEGIES), value=STRATEGIES[0], style=_dd())),
            _field("Type",         dcc.Dropdown(id="ph-dtype",   options=dd_opts(DEAL_TYPES),  value="Secondary",   style=_dd())),
            _field("Region",       dcc.Dropdown(id="ph-region",  options=dd_opts(REGIONS),     value="North America",style=_dd())),
            _field("Expected $M",  dcc.Input(id="ph-size",   type="number", value=30, step=0.5, style=INP)),
            _field("Target IRR %", dcc.Input(id="ph-irr",    type="number", value=17, step=0.5, style=INP)),
            _field("Expected Month", dcc.Dropdown(id="ph-month", options=month_options(), value=3, style=_dd())),
            html.Div([html.Label(" ",style=dict(display="block",marginBottom=4,fontSize=9)),
                      html.Button("+ Add",id="add-ph-btn",style={**BTN(C["teal"]),"minWidth":80})]),
            html.Div([html.Label(" ",style=dict(display="block",marginBottom=4,fontSize=9)),
                      html.Button("Clear",id="clear-ph-btn",style={**BTN(C["surface"],C["muted"]),"border":f"1px solid {C['border2']}","minWidth":60})]),
        ], style=dict(display="flex",gap=8,flexWrap="wrap",alignItems="flex-end")),
        html.Div(id="ph-msg",style=dict(color=C["green"],fontSize=11,marginTop=4)),
    ], style=dict(background=C["surface"],border=f"1px solid {cl(C['teal'],0.4)}",borderRadius=8,padding=18,marginBottom=18))

    tbl = dash_table.DataTable(
        id="ph-table", data=rows,
        columns=[{"name":c,"id":c} for c in rows[0].keys()] if rows else [],
        hidden_columns=["_idx"],
        row_selectable="single", selected_rows=[],
        style_cell=TBL_CELL, style_header=TBL_HEAD, style_data_conditional=TBL_ODD,
        sort_action="native", page_size=12, style_table={"overflowX":"auto"}, export_format="xlsx",
    )

    actions = card([
        section_lbl("Actions"),
        html.Button("✏️ Edit Selected",   id="edit-ph-btn",   style={**BTN(C["amber"]),"width":"100%","marginBottom":8}),
        html.Button("🗑 Delete Selected", id="delete-ph-btn", style={**BTN(C["red"]),"width":"100%"}),
        html.Div(id="ph-action-msg",style=dict(marginTop=10,color=C["muted"],fontSize=11)),
    ], dict(minWidth=180))

    dp  = float(config.get("dry_powder",300) or 300)
    total_planned = sum(p.get("size",0) for p in placeholders)
    remaining = dp - total_planned

    return html.Div([
        html.Div([
            html.Span("Plan future commitments to pace dry powder deployment. ",
                      style=dict(fontSize=12, color=C["muted"], fontFamily=C["mono"])),
            html.Span("These feed into Pacing and Dry Powder forecast tabs but are NOT in the portfolio.",
                      style=dict(fontSize=11, color=C["dim"], fontFamily=C["mono"])),
        ], style=dict(marginBottom=12)),
        html.Div([
            kpi("Current Dry Powder", fmt_m(dp),           "Available",              C["blue"]),
            kpi("Total Planned",      fmt_m(total_planned), f"{len(placeholders)} commitments", C["teal"]),
            kpi("Remaining Powder",   fmt_m(remaining),     "After planned",          C["green"] if remaining>=0 else C["red"]),
        ], style=dict(display="flex",gap=10,flexWrap="wrap",marginBottom=16)),
        form,
        html.Div([
            html.Div(tbl, style=dict(flex=3, minWidth=0)),
            html.Div([dcc.Graph(figure=fig,config={"displayModeBar":False}), actions],
                     style=dict(flex=1,display="flex",flexDirection="column",gap=16,minWidth=280)),
        ], style=dict(display="flex",gap=16,flexWrap="wrap")),
    ])


# ── NAV helpers for CF/discount tabs ─────────────────────────────────────────
def _latest_nav(cf_data, deal_name):
    """Get the most recent non-zero NAV for a deal from uploaded CF data."""
    if not cf_data or not deal_name:
        return None
    dn = deal_name.strip().lower()
    for d in cf_data:
        if str(d.get("name","")).strip().lower() != dn:
            continue
        series = d.get("nav_series") or {}
        dated  = [(k,v) for k,v in series.items() if v not in (None,0,0.0)]
        if dated:
            dated.sort(key=lambda x: x[0])
            return float(dated[-1][1])
        if d.get("nav"):
            return float(d["nav"])
    return None


def tab_cashflows(portfolio, cf_data, disc_store):
    """Deal Cashflows — Excel upload, 4-section CF matrix, P/NAV discount tracker."""

    # ── Month options for start-month picker ─────────────────────────────────
    _start = datetime(2024, 9, 1)
    _end   = datetime(2041, 1, 1)
    month_opts_cf = []
    _cur = _start
    while _cur < _end:
        month_opts_cf.append({"label": _cur.strftime("%b %Y"),
                              "value": _cur.strftime("%Y-%m-%d")})
        _cur = (_cur + relativedelta(months=1)).replace(day=1)

    # Determine latest available month from CF data; fall back to now
    default_start = datetime.now().replace(day=1).strftime("%Y-%m-%d")
    if cf_data:
        try:
            _all_mk = cf_data[0].get("_all_month_keys") or sorted(
                {mk for d in cf_data for mk in (d.get("nav_series") or {}).keys() if mk})
            if _all_mk:
                # Jump to the last month that has non-zero NAV data
                _nz_months = []
                for mk in _all_mk:
                    total_nav = sum(float((d.get("nav_series") or {}).get(mk, 0) or 0) for d in cf_data)
                    if total_nav > 0:
                        _nz_months.append(mk)
                if _nz_months:
                    default_start = _nz_months[-1]
        except Exception:
            pass

    upload_section = card([
        slbl("Upload Fund Level CF Excel"),
        html.Div([
            html.Div([
                dcc.Upload(id="upload-fund-cf",
                    children=html.Button("📂 Upload Fund Level CF (.xlsx/.xlsm)",
                                         style={**BTN(C["blue"]),"width":"100%","padding":"12px 20px"}),
                    multiple=False),
                html.Small("Expected sheet: 'Fund Level CF' — Row 1: section labels, Row 2: month dates, Rows 7+: deals",
                           style=dict(color=C["muted"],fontSize=11,display="block",marginTop=6)),
            ], style=dict(flex=2)),
            html.Div(id="upload-cf-status", style=dict(flex=3,paddingLeft=16)),
        ], style=dict(display="flex",gap=12,alignItems="flex-start")),
    ], dict(marginBottom=16))

    # KPIs from upload or manual deals
    def _cf_kpis(data, portfolio):
        if data:
            tc = sum(float(d.get("commitment",0) or 0) for d in data)
            pi = sum(float(d.get("paid_in",d.get("current_commitment",0)) or 0) for d in data)
            uf = sum(float(d.get("unfunded",0) or 0) for d in data)
            cv = sum(float(d.get("nav",0) or 0) for d in data)
        else:
            tc = sum(d.get("total_commitment",d.get("nav",0)) for d in portfolio)
            pi = sum(d.get("current_commitment",d.get("nav",0)) for d in portfolio)
            uf = sum(max(d.get("total_commitment",0)-d.get("current_commitment",0),0) for d in portfolio)
            cv = sum(d.get("nav",0) for d in portfolio)
        return tc, pi, uf, cv

    tc, pi, uf, cv = _cf_kpis(cf_data, portfolio)
    cf_kpis = html.Div([
        kpi("Total Commitment",  fmt_m(tc), "Legal commitment",  C["blue"]),
        kpi("Paid In / Called",  fmt_m(pi), "Capital deployed",  C["green"]),
        kpi("Unfunded",          fmt_m(uf), "Remaining calls",   C["amber"]),
        kpi("Current NAV",       fmt_m(cv), "Latest month-end",  C["purple"]),
    ], style=dict(display="flex",gap=10,flexWrap="wrap",marginBottom=16))

    # Controls
    controls = card([
        slbl("View Controls"),
        html.Div([
            _field("Start Month", dcc.Dropdown(id="cf-start-month",
                options=month_opts_cf, value=default_start, style={**_dd(),"minWidth":160})),
            _field("Horizon", dcc.RadioItems(id="cf-horizon",
                options=[{"label":l,"value":v} for l,v in
                         [("12M",12),("18M",18),("24M",24),("36M",36),("5Y",60)]],
                value=12, inline=True,
                style=dict(fontFamily=C["mono"],fontSize=12,color=C["text"]))),
            _field("Sections", dcc.Checklist(id="cf-sections",
                options=[{"label":l,"value":v} for l,v in
                         [(" Net CF","net"),(" Calls","calls"),(" Distributions","dists"),(" NAV","nav")]],
                value=["net","calls","dists","nav"], inline=True,
                style=dict(fontFamily=C["mono"],fontSize=12,color=C["text"]))),
        ], style=dict(display="flex",gap=24,flexWrap="wrap",alignItems="flex-end")),
        html.Div(id="cf-window-badge",style=dict(marginTop=10)),
    ], dict(marginBottom=16))

    cf_table_area = html.Div([
        html.Div(id="cf-monthly-table"),
        html.Div(style=dict(height=16)),
        html.Div(id="cf-chart-area"),
    ])

    # ── Discount / P2NAV section ──────────────────────────────────────────────
    disc_rows = []
    for d in portfolio:
        name = d["name"]
        nav_val = _latest_nav(cf_data, name)
        if nav_val is None or nav_val == 0:
            nav_val = float(d.get("nav",0) or 0)
        saved = (disc_store or {}).get(name, {})
        nav_ov = saved.get("nav_override")
        if nav_ov not in (None,"",0):
            nav_val = float(nav_ov)
        disc_pct = float(saved.get("discount_pct",0) or 0)
        pp = nav_val*(1-disc_pct/100) if nav_val else 0
        p2n = pp/nav_val if nav_val else None
        disc_rows.append({
            "Deal":name, "Strategy":d.get("strategy",""),
            "NAV ($m)":round(nav_val,2),
            "Discount (%)":round(disc_pct,2),
            "Purchase Price ($m)":round(pp,2),
            "P/NAV":round(p2n,4) if p2n is not None else None,
        })

    total_nav_d  = sum(r["NAV ($m)"] for r in disc_rows)
    total_price  = sum(r["Purchase Price ($m)"] for r in disc_rows)
    wtd_disc     = sum(r["NAV ($m)"]*r["Discount (%)"] for r in disc_rows)/total_nav_d if total_nav_d else 0
    ptbnav_port  = total_price/total_nav_d if total_nav_d else None

    disc_kpis = html.Div([
        kpi("Portfolio NAV (basis)", fmt_m(total_nav_d), "Basis for discounts", C["blue"]),
        kpi("Wtd Avg Discount",      f"{wtd_disc:.1f}%", "NAV-weighted",        C["amber"]),
        kpi("Total Purchase Price",  fmt_m(total_price), "NAV × (1−disc)",      C["green"]),
        kpi("Portfolio P/NAV",       f"{ptbnav_port:.3f}x" if ptbnav_port else "—", "Weighted", C["purple"]),
    ], style=dict(display="flex",gap=10,flexWrap="wrap",marginBottom=12))

    disc_tbl = dash_table.DataTable(
        id="discount-datatable",
        data=disc_rows,
        columns=[
            {"name":"Deal","id":"Deal","editable":False},
            {"name":"Strategy","id":"Strategy","editable":False},
            {"name":"NAV ($m)","id":"NAV ($m)","editable":True,"type":"numeric"},
            {"name":"Discount (%)","id":"Discount (%)","editable":True,"type":"numeric"},
            {"name":"Purchase Price ($m)","id":"Purchase Price ($m)","editable":False,"type":"numeric"},
            {"name":"P/NAV","id":"P/NAV","editable":False,"type":"numeric"},
        ],
        editable=True, sort_action="native",
        style_table={"overflowX":"auto"},
        style_cell=TBL_CELL, style_header=TBL_HEAD,
        style_cell_conditional=[
            {"if":{"column_id":"Deal"},    "textAlign":"left","minWidth":"180px","fontWeight":"bold"},
            {"if":{"column_id":"Strategy"},"textAlign":"left","minWidth":"150px"},
        ],
        style_data_conditional=TBL_ODD+[
            {"if":{"column_id":"Discount (%)"},"backgroundColor":"#1a1600","fontWeight":"bold","border":f"2px solid {C['amber']}"},
            {"if":{"column_id":"NAV ($m)"},    "backgroundColor":"#0d1422","border":f"1px solid {C['blue']}"},
            {"if":{"filter_query":"{P/NAV} < 0.85","column_id":"P/NAV"},"color":C["green"],"fontWeight":"bold"},
            {"if":{"filter_query":"{P/NAV} >= 0.85 && {P/NAV} < 0.95","column_id":"P/NAV"},"color":C["amber"],"fontWeight":"bold"},
            {"if":{"filter_query":"{P/NAV} >= 0.95","column_id":"P/NAV"},"color":C["red"],"fontWeight":"bold"},
        ],
        tooltip_header={
            "Discount (%)":"Edit inline — enter the % discount negotiated on NAV",
            "NAV ($m)":"Override NAV if different from uploaded cashflows",
            "P/NAV":"Green <0.85 · Amber 0.85–0.95 · Red ≥0.95",
        }, tooltip_delay=300, tooltip_duration=None,
    )

    # Discount bar chart
    if disc_rows:
        sorted_disc = sorted([r for r in disc_rows if r["NAV ($m)"]>0],
                             key=lambda x: x["Discount (%)"], reverse=True)
        bar_colors = [C["green"] if d["Discount (%)"]>=15 else
                      C["amber"] if d["Discount (%)"]>=5 else C["red"]
                      for d in sorted_disc]
        fig_disc = go.Figure()
        fig_disc.add_trace(go.Bar(
            name="Discount %", y=[r["Deal"] for r in sorted_disc],
            x=[r["Discount (%)"] for r in sorted_disc],
            orientation="h", marker_color=bar_colors,
            text=[f"{r['Discount (%)']:.1f}%" for r in sorted_disc], textposition="outside",
        ))
        fig_disc.add_trace(go.Scatter(
            name="P/NAV (top axis)", y=[r["Deal"] for r in sorted_disc],
            x=[r["P/NAV"] or 1 for r in sorted_disc],
            mode="markers", xaxis="x2",
            marker=dict(symbol="diamond",size=10,color=C["purple"],line=dict(width=1,color="white")),
        ))
        fig_disc.update_layout(
            **chart(height=max(280,40*len(sorted_disc)+100),
            xaxis=dict(title="Discount (%)",gridcolor=C["border"]),
            xaxis2=dict(title="P/NAV",overlaying="x",side="top",range=[0,1.1],showgrid=False),
            yaxis=dict(autorange="reversed"),
            margin=dict(l=20,r=60,t=50,b=40), legend=dict(orientation="h",y=1.1)),
        )
    else:
        fig_disc = go.Figure()
        fig_disc.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40), xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1), height=200)

    disc_section = card([
        slbl("Discount & Price-to-NAV Tracker"),
        html.Div([
            html.Small("Edit the Discount (%) column inline. NAV auto-pulls from uploaded CF file. P/NAV auto-calculates.",
                       style=dict(color=C["muted"],fontSize=11,display="block",marginBottom=8)),
            html.Div([
                html.Button("💾 Save Discounts",    id="btn-save-disc",    style={**BTN(C["amber"]),"marginRight":8}),
                html.Button("🔄 Refresh from NAV",  id="btn-refresh-disc", style=BTN(C["surface"],C["muted"])),
            ], style=dict(marginBottom=12)),
            html.Div(id="disc-save-status",style=dict(color=C["green"],fontSize=11,marginBottom=8)),
        ]),
        disc_kpis,
        html.Div([
            html.Div(disc_tbl, style=dict(flex=2,minWidth=0)),
            html.Div(dcc.Graph(figure=fig_disc,config={"displayModeBar":False}), style=dict(flex=1,minWidth=280)),
        ], style=dict(display="flex",gap=16,flexWrap="wrap")),
    ], dict(marginTop=24))

    return html.Div([
        upload_section,
        cf_kpis,
        controls,
        cf_table_area,
        disc_section,
    ])


def tab_liquidity(liq_data):
    """Liquidity Pull — upload Excel, display waterfall, NAV projection, near-term flows."""

    upload_section = card([
        slbl("Upload Liquidity Pull Excel"),
        html.Div([
            html.Div([
                dcc.Upload(id="upload-liquidity",
                    children=html.Button("📂 Upload Liquidity Pull (.xlsx/.xlsm)",
                                         style={**BTN(C["teal"]),"width":"100%","padding":"12px 20px"}),
                    multiple=False),
                html.Small("Expected sheet: 'Liquidity Pull' — standard Horizon template structure",
                           style=dict(color=C["muted"],fontSize=11,display="block",marginTop=6)),
            ], style=dict(flex=2)),
            html.Div(id="upload-liq-status", style=dict(flex=3,paddingLeft=16)),
        ], style=dict(display="flex",gap=12,alignItems="flex-start")),
    ], dict(marginBottom=16))

    if not liq_data:
        return html.Div([
            upload_section,
            card(html.Div([
                html.Div("💧", style=dict(fontSize=48,textAlign="center",marginBottom=12)),
                html.P("Upload your Liquidity Pull Excel file above to see the waterfall, NAV projections, and near-term flows.",
                       style=dict(color=C["muted"],textAlign="center",fontFamily=C["mono"])),
            ], style=dict(padding="40px 20px"))),
        ])

    # Date cards
    date_kpis = html.Div([
        kpi("As At Date",      liq_data.get("as_at_date","—"),     "",                            C["text"]),
        kpi("Current Quarter", liq_data.get("current_quarter","—"),"",                            C["muted"]),
        kpi("Fund NAV",        fmt_m(liq_data.get("fund_nav",0)),   "Total",                      C["green"]),
        kpi("Cash Balance",    fmt_m(liq_data.get("current_cash",0)),"Row 58",                    C["blue"]),
        kpi("GLF Balance",     fmt_m(liq_data.get("glf_balance",0)), "Row 59",                    C["sky"]),
        kpi("CQS Balance",     fmt_m(liq_data.get("cqs_balance",0)), "Row 60",                    C["teal"]),
        kpi("Total Liquidity", fmt_m(liq_data.get("total_liquidity",0)),"Cash+GLF+CQS",           C["purple"]),
        kpi("Dec 2026 Dry Powder", fmt_m(liq_data.get("dec_2026_dry_powder",0)),"Row 53 Col 9",   C["amber"]),
    ], style=dict(display="flex",gap=10,flexWrap="wrap",marginBottom=16))

    # Liquidity waterfall table
    wf_items = [
        ("Fund NAV",               liq_data.get("fund_nav",0)),
        ("Cash Balance",           liq_data.get("current_cash",0)),
        ("GLF Balance",            liq_data.get("glf_balance",0)),
        ("CQS Balance",            liq_data.get("cqs_balance",0)),
        ("Total Liquidity",        liq_data.get("total_liquidity",0)),
        ("Surplus Post Buffer",    liq_data.get("surplus_liquidity_post_buffer",0)),
        ("Proj. NAV (Existing)",   liq_data.get("projected_nav_existing",0)),
        ("Proj. NAV (+ Pipeline)", liq_data.get("projected_nav_existing_pipeline",0)),
    ]
    wf_rows = [{"Item":item,"Amount ($M)":fmt_m(v)} for item,v in wf_items]
    wf_tbl = dash_table.DataTable(data=wf_rows,
                                   columns=[{"name":c,"id":c} for c in wf_rows[0]],
                                   style_cell=TBL_CELL, style_header=TBL_HEAD,
                                   style_data_conditional=TBL_ODD,
                                   style_table={"overflowX":"auto"})

    # NAV projection chart
    nav_proj = liq_data.get("nav_projections",{})
    if nav_proj:
        fig_nav = go.Figure()
        fig_nav.add_trace(go.Scatter(
            x=list(nav_proj.keys()), y=list(nav_proj.values()),
            mode="lines+markers", name="Projected NAV",
            line=dict(color=C["blue"],width=3),
            marker=dict(size=6,color=C["sky"]),
            fill="tozeroy", fillcolor=cl(C["blue"],0.12),
        ))
        fig_nav.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40), xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1), height=340, title="NAV End Projections",
                               yaxis_title="NAV ($M)", hovermode="x unified")
    else:
        fig_nav = go.Figure()
        fig_nav.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40), xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1), height=200, title="NAV Projections (no data)")

    # Near-term flows table
    ntf = liq_data.get("near_term_flows",{})
    ntf_rows = [{"Month":str(m),
                 "Subscriptions ($M)":fmt_m(v.get("subscriptions",0)),
                 "Redemptions ($M)":fmt_m(v.get("redemptions",0)),
                 "Portfolio Flows ($M)":fmt_m(v.get("portfolio_flows",0))}
                for m,v in ntf.items()]

    # Max deployable capital
    mdc = liq_data.get("max_deployable_capital",{})
    mdc_rows = [{"Month":str(m),"Max Deployable ($M)":fmt_m(v)} for m,v in mdc.items()]

    return html.Div([
        upload_section,
        date_kpis,
        html.Div([
            card([section_lbl("Liquidity Waterfall"), wf_tbl], dict(flex=1,minWidth=300)),
            card([dcc.Graph(figure=fig_nav,config={"displayModeBar":True})], dict(flex=2)),
        ], style=dict(display="flex",gap=16,flexWrap="wrap",marginBottom=16)),
        html.Div([
            card([section_lbl("Near-Term Flows (Next 12 Months)"),
                  dash_table.DataTable(data=ntf_rows,
                                       columns=[{"name":c,"id":c} for c in ntf_rows[0]] if ntf_rows else [],
                                       style_cell=TBL_CELL, style_header=TBL_HEAD,
                                       style_data_conditional=TBL_ODD,
                                       page_size=12, style_table={"overflowX":"auto"})
                  if ntf_rows else html.P("No near-term flow data.", style=dict(color=C["muted"]))
                  ], dict(flex=1)),
            card([section_lbl("Max Deployable Capital (Row 53)"),
                  dash_table.DataTable(data=mdc_rows,
                                       columns=[{"name":c,"id":c} for c in mdc_rows[0]] if mdc_rows else [],
                                       style_cell=TBL_CELL, style_header=TBL_HEAD,
                                       style_data_conditional=TBL_ODD,
                                       page_size=12, style_table={"overflowX":"auto"})
                  if mdc_rows else html.P("No deployable capital data.", style=dict(color=C["muted"]))
                  ], dict(flex=1)),
        ], style=dict(display="flex",gap=16,flexWrap="wrap")),
    ])


def tab_analytics(portfolio, pipeline, placeholders, config):
    """Analytics tab — NAV-based portfolio scope toggle:
       Current | Current + Pipeline | Current + Pipeline + Future Deals
    """
    return html.Div([
        # Scope toggle
        html.Div([
            html.Div([
                html.Span("Portfolio Scope", style=dict(
                    fontSize=9, letterSpacing=2, color=C["muted"],
                    textTransform="uppercase", fontFamily=C["sans"],
                    marginRight=12, display="inline-block"
                )),
                dcc.RadioItems(
                    id="analytics-scope",
                    options=[
                        {"label": " Current Portfolio",                              "value": "current"},
                        {"label": " + Pipeline",                                   "value": "pipeline"},
                        {"label": " + Pipeline + Deployment Plan",                 "value": "proforma"},
                    ],
                    value="current",
                    inline=True,
                    style=dict(fontFamily=C["mono"], fontSize=13, color=C["text"]),
                    inputStyle=dict(marginRight=4),
                    labelStyle=dict(marginRight=24),
                ),
            ], style=dict(
                background=C["surface"], border=f"1px solid {cl(C['blue'], 0.5)}",
                borderRadius=8, padding="14px 20px",
                display="flex", alignItems="center", flexWrap="wrap", gap=8,
            )),
        ], style=dict(marginBottom=16)),
        # KPI strip for selected scope
        html.Div(id="analytics-scope-kpis", style=dict(marginBottom=16)),
        # Charts area (rebuilt on toggle)
        html.Div(id="analytics-charts"),
    ])


def tab_segments(portfolio, pipeline, placeholders, config):
    seed_deals = [d for d in portfolio if d.get("segment","Seed")=="Seed"]
    new_deals  = [d for d in portfolio if d.get("segment","Seed")=="New"]
    mm_deals   = [d for d in portfolio if d.get("segment","Seed")=="MoneyMarket"]

    seed_nav = sum(d["nav"] for d in seed_deals)
    new_nav  = sum(d["nav"] for d in new_deals)
    mm_nav   = sum(d["nav"] for d in mm_deals)
    pipe_nav = sum(p.get("size",0) for p in pipeline)
    ph_nav   = sum(p.get("size",0) for p in placeholders)
    total_nav= seed_nav+new_nav+mm_nav

    seed_twr = w_avg(seed_deals,"target_irr") if seed_deals else 0
    new_twr  = w_avg(new_deals, "target_irr") if new_deals  else 0
    mm_twr   = 0.03
    pipe_twr = w_avg(pipeline,"target_irr","size") if pipeline else 0
    total_twr= (seed_nav*seed_twr+new_nav*new_twr+mm_nav*mm_twr)/total_nav if total_nav else 0

    kpi_row = html.Div([
        kpi("Total Portfolio",  fmt_m(total_nav), f"TWR: {total_twr:.1%}", C["blue"],  140),
        kpi("Seed Portfolio",   fmt_m(seed_nav),  f"TWR: {seed_twr:.1%}", C["green"], 140),
        kpi("New Deals",        fmt_m(new_nav),   f"TWR: {new_twr:.1%}",  C["purple"],140),
        kpi("Money Market",     fmt_m(mm_nav),    f"TWR: {mm_twr:.1%}",   C["amber"], 130),
        kpi("Pipeline",         fmt_m(pipe_nav),  f"Est IRR: {pipe_twr:.1%}", C["teal"],130),
        kpi("Pro Forma",        fmt_m(ph_nav),    f"{len(placeholders)} placeholders", C["pink"],130),
    ], style=dict(display="flex",gap=10,flexWrap="wrap",marginBottom=16))

    # ── Cashflow-based TWR forecast ──────────────────────────────────────────
    # Formula per period: sub-period return = (NAV_end - calls + dists) / NAV_start - 1
    # Chain-link across 24 months. Each deal contributes projected cashflows
    # based on its IRR, hold period, current NAV and distribution rate.
    n_months   = 24
    dist_rate  = float(config.get("distribution_rate", 0.20))  # annual
    base_dt    = datetime.now().replace(day=1)
    month_lbls = [(base_dt + relativedelta(months=i)).strftime("%b %Y") for i in range(n_months)]

    def project_deal_cashflows(deal, n):
        """Return (nav_series, calls_series, dists_series) each length n months."""
        nav0    = float(deal.get("nav", 0))
        irr     = float(deal.get("target_irr", 0.15))
        hold_y  = float(deal.get("hold_period", 5))
        hold_m  = max(1, int(hold_y * 12))
        monthly_growth = (1 + irr) ** (1/12) - 1
        monthly_dist   = dist_rate / 12
        navs   = [0.0] * n
        calls  = [0.0] * n
        dists  = [0.0] * n
        nav_t  = nav0
        for m in range(n):
            if m >= hold_m:
                # Deal has exited — NAV collapses, final distribution
                if navs[m-1] > 0 if m > 0 else nav0 > 0:
                    dists[m] = navs[m-1] if m > 0 else nav0
                navs[m] = 0.0
            else:
                d_nav  = nav_t * monthly_growth
                dist_m = nav_t * monthly_dist
                nav_t  = nav_t + d_nav - dist_m
                navs[m]  = nav_t
                dists[m] = dist_m
        return navs, calls, dists

    def compute_twr_series(deals, n):
        if not deals:
            return [0.0] * n
        # Aggregate NAV, calls, dists across all deals in this segment
        agg_nav   = [0.0] * n
        agg_calls = [0.0] * n
        agg_dists = [0.0] * n
        for d in deals:
            navs, calls, dists = project_deal_cashflows(d, n)
            for m in range(n):
                agg_nav[m]   += navs[m]
                agg_calls[m] += calls[m]
                agg_dists[m] += dists[m]
        # Starting NAV for period 0 is current portfolio NAV
        start_nav = sum(d.get("nav", 0) for d in deals)
        # Chain-link TWR
        cumulative = 1.0
        twr_series = []
        prev_nav   = start_nav
        for m in range(n):
            nav_end  = agg_nav[m]
            calls_m  = agg_calls[m]
            dists_m  = agg_dists[m]
            if prev_nav > 0:
                sub_return = (nav_end - calls_m + dists_m) / prev_nav - 1
            else:
                sub_return = 0.0
            cumulative *= (1 + sub_return)
            twr_series.append((cumulative - 1) * 100)
            prev_nav = nav_end if nav_end > 0 else prev_nav
        return twr_series

    # Money market: simple compounding
    mm_twrs = [(((1 + mm_twr) ** ((i+1)/12)) - 1) * 100 for i in range(n_months)]

    total_twrs = compute_twr_series(seed_deals + new_deals, n_months)
    seed_twrs  = compute_twr_series(seed_deals, n_months)
    new_twrs   = compute_twr_series(new_deals,  n_months)

    # TWR formula annotation
    formula_note = html.Div(
        "TWR formula: sub-period return = (NAV_end − Calls + Distributions) / NAV_start − 1, chain-linked monthly. "
        "NAV projected per deal using target IRR and hold period; distributions modelled at fund distribution rate.",
        style=dict(fontSize=10, color=C["muted"], fontFamily=C["mono"],
                   background=C["surface"], border=f"1px solid {C['border']}",
                   borderRadius=6, padding="8px 12px", marginBottom=12)
    )

    fig_twr = go.Figure()
    for name, y, color, dash in [
        ("Total Portfolio", total_twrs, C["blue"],   "solid"),
        ("Seed Portfolio",  seed_twrs,  C["green"],  "dash"),
        ("New Deals",       new_twrs,   C["purple"], "dot"),
        ("Money Market",    mm_twrs,    C["amber"],  "dashdot"),
    ]:
        fig_twr.add_trace(go.Scatter(x=month_lbls, y=y, name=name, mode="lines",
                                     line=dict(color=color, width=2, dash=dash)))
    # Target TWR reference line
    tgt = float(config.get("target_net_twr", 0.13))
    tgt_series = [(((1+tgt)**((i+1)/12))-1)*100 for i in range(n_months)]
    fig_twr.add_trace(go.Scatter(x=month_lbls, y=tgt_series, name="Target TWR",
                                  mode="lines", line=dict(color=C["red"], width=1, dash="longdash"),
                                  opacity=0.6))
    fig_twr.update_layout(**chart(height=380,
                           title="24-Month Cumulative TWR Forecast (Cashflow-Based, Chain-Linked)",
                           yaxis_title="Cumulative TWR (%)", hovermode="x unified",
                           legend=dict(orientation="h", y=1.08, x=0)))

    # Monthly sub-period returns table for Total Portfolio
    sp_returns = []
    prev_nav = sum(d.get("nav",0) for d in seed_deals+new_deals)
    all_navs = [0.0]*n_months
    all_calls= [0.0]*n_months
    all_dists= [0.0]*n_months
    for d in seed_deals+new_deals:
        navs, calls, dists = project_deal_cashflows(d, n_months)
        for m in range(n_months):
            all_navs[m]  += navs[m]
            all_calls[m] += calls[m]
            all_dists[m] += dists[m]
    for m in range(n_months):
        nav_end = all_navs[m]
        r = (nav_end - all_calls[m] + all_dists[m]) / prev_nav - 1 if prev_nav > 0 else 0
        sp_returns.append({
            "Month": month_lbls[m],
            "NAV Start ($M)": f"{prev_nav:.1f}",
            "NAV End ($M)":   f"{nav_end:.1f}",
            "Calls ($M)":     f"{all_calls[m]:.1f}",
            "Dists ($M)":     f"{all_dists[m]:.1f}",
            "Sub-period Return": f"{r*100:.2f}%",
            "Cumulative TWR":    f"{total_twrs[m]:.2f}%",
        })
        prev_nav = nav_end if nav_end > 0 else prev_nav

    sp_tbl = dash_table.DataTable(
        data=sp_returns,
        columns=[{"name":c,"id":c} for c in sp_returns[0]],
        style_cell=TBL_CELL, style_header=TBL_HEAD,
        style_data_conditional=TBL_ODD,
        page_size=12, style_table={"overflowX":"auto"},
        export_format="xlsx",
    )

    # Allocation & contribution charts
    seg_labels = ["Seed","New Deals","Money Market","Pipeline","Pro Forma"]
    seg_vals   = [seed_nav, new_nav, mm_nav, pipe_nav, ph_nav]
    seg_colors = [C["green"],C["purple"],C["amber"],C["teal"],C["pink"]]
    fig_alloc = go.Figure(go.Pie(labels=seg_labels, values=seg_vals, hole=0.5,
                                  marker_colors=seg_colors, textfont_color=C["text"]))
    fig_alloc.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40),
                             xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]),
                             yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]),
                             legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1),
                             height=300, title="Segment Allocation")

    contrib_segs = ["Seed","New Deals","MM"]
    contrib_vals = [seed_nav*seed_twr*100, new_nav*new_twr*100, mm_nav*mm_twr*100]
    fig_contrib = go.Figure(go.Bar(x=contrib_segs, y=contrib_vals,
                                    marker_color=[C["green"],C["purple"],C["amber"]],
                                    text=[f"{v:.1f}" for v in contrib_vals],
                                    textposition="outside", textfont_color=C["text"]))
    fig_contrib.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40),
                               xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]),
                               yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]),
                               legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1),
                               height=300, title="TWR Contribution (NAV × IRR)")

    def seg_table(deals, name):
        if not deals:
            return html.P(f"No {name} deals yet.", style=dict(color=C["muted"],fontSize=12))
        data = [{"Deal":d["name"],"NAV":fmt_m(d["nav"]),"IRR":f"{d.get('target_irr',0)*100:.1f}%",
                 "Vintage":str(d.get("vintage","")),"Commit Year":str(d.get("commitment_year","")),
                 "Entry Date":str(d.get("entry_date","")),"Hold(y)":f"{d.get('hold_period',0):.1f}",
                 "Manager":d.get("manager","")} for d in deals]
        return dash_table.DataTable(data=data, columns=[{"name":c,"id":c} for c in data[0]],
                                    style_cell=TBL_CELL, style_header=TBL_HEAD,
                                    style_data_conditional=TBL_ODD, page_size=8,
                                    style_table={"overflowX":"auto"})

    return html.Div([
        kpi_row,
        formula_note,
        card([dcc.Graph(figure=fig_twr, config={"displayModeBar":True})], dict(marginBottom=16)),
        card([section_lbl("Monthly TWR Detail — Total Portfolio"), sp_tbl], dict(marginBottom=16)),
        html.Div([
            card([dcc.Graph(figure=fig_alloc,  config={"displayModeBar":False})], dict(flex=1)),
            card([dcc.Graph(figure=fig_contrib,config={"displayModeBar":False})], dict(flex=1)),
        ], style=dict(display="flex",gap=16,flexWrap="wrap",marginBottom=16)),
        html.Div([
            card([section_lbl("Seed Portfolio Deals"), seg_table(seed_deals,"seed")], dict(flex=1)),
            card([section_lbl("New Deals"),            seg_table(new_deals, "new")],  dict(flex=1)),
        ], style=dict(display="flex",gap=16,flexWrap="wrap")),
    ])


def tab_drypowder(portfolio, placeholders, config):
    dp  = float(config.get("dry_powder",300) or 300)
    nav = sum(d.get("nav",0) for d in portfolio)
    fc  = forecast_dp(nav, dp, placeholders, config, 12)

    months   = [f["month"]         for f in fc]
    dp_vals  = [f["dry_powder"]    for f in fc]
    nav_vals = [f["nav"]           for f in fc]
    dists    = [f["distributions"] for f in fc]
    calls    = [f["calls"]         for f in fc]

    fig = make_subplots(specs=[[{"secondary_y":True}]])
    fig.add_trace(go.Scatter(x=months,y=dp_vals,name="Dry Powder",
                              line=dict(color=C["blue"],width=3),fill="tozeroy",
                              fillcolor=cl(C["blue"],0.15),marker=dict(size=5,color=C["sky"])),
                  secondary_y=False)
    fig.add_trace(go.Scatter(x=months,y=nav_vals,name="Portfolio NAV",
                              line=dict(color=C["green"],width=2,dash="dash"),marker=dict(size=4)),
                  secondary_y=False)
    fig.add_trace(go.Bar(x=months,y=calls,name="Capital Calls",
                          marker_color=cl(C["red"],0.65)),secondary_y=True)
    fig.add_trace(go.Bar(x=months,y=dists,name="Distributions",
                          marker_color=cl(C["green"],0.55)),secondary_y=True)
    fig.update_xaxes(gridcolor=C["border"])
    fig.update_yaxes(title_text="$M",secondary_y=False,gridcolor=C["border"])
    fig.update_yaxes(title_text="Flows $M",secondary_y=True,showgrid=False)
    fig.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40), xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1), height=400, title="12-Month Dry Powder Forecast",
                      hovermode="x unified", barmode="group")

    # Summary kpis
    kpi_row = html.Div([
        kpi("Current Dry Powder",   fmt_m(dp),                "Available now",          C["blue"]),
        kpi("Planned Calls (12M)",  fmt_m(sum(calls)),        "From placeholders",      C["amber"]),
        kpi("Forecast Distributions",fmt_m(sum(dists)),       "Over 12 months",         C["green"]),
        kpi("Forecast End DP",      fmt_m(dp_vals[-1] if dp_vals else 0), "In 12 months",C["teal"]),
    ], style=dict(display="flex",gap=10,flexWrap="wrap",marginBottom=16))

    # Bite-size table
    bs = bite_sizes(config)
    bite_rows = [{"Strategy":s,"Min $M":fmt_m(v["min"]),"Desired $M":fmt_m(v["desired"]),"Max $M":fmt_m(v["max"]),
                  "Min %":f"{v['min_pct']*100:.2f}%","Desired %":f"{v['desired_pct']*100:.2f}%","Max %":f"{v['max_pct']*100:.2f}%"}
                 for s,v in bs.items()]
    bite_tbl = dash_table.DataTable(data=bite_rows, columns=[{"name":c,"id":c} for c in bite_rows[0]],
                                     style_cell=TBL_CELL, style_header=TBL_HEAD, style_data_conditional=TBL_ODD,
                                     style_table={"overflowX":"auto"})

    dp_tbl_rows = [{"Month":f["month"],"Dry Powder $M":fmt_m(f["dry_powder"]),"NAV $M":fmt_m(f["nav"]),
                    "Distributions $M":fmt_m(f["distributions"]),"Calls $M":fmt_m(f["calls"]) if f["calls"]>0 else "—"}
                   for f in fc]
    dp_tbl = dash_table.DataTable(data=dp_tbl_rows, columns=[{"name":c,"id":c} for c in dp_tbl_rows[0]],
                                   style_cell=TBL_CELL, style_header=TBL_HEAD, style_data_conditional=TBL_ODD,
                                   page_size=12, style_table={"overflowX":"auto"})

    return html.Div([
        kpi_row,
        card([dcc.Graph(figure=fig, config={"displayModeBar":True})], dict(marginBottom=16)),
        html.Div([
            card([section_lbl("Bite Size Guide (on current dry powder)"), bite_tbl], dict(flex=1)),
            card([section_lbl("Monthly Schedule"),                         dp_tbl],  dict(flex=1)),
        ], style=dict(display="flex",gap=16,flexWrap="wrap")),
    ])


def tab_pacing(portfolio, pipeline, placeholders, config):
    df = build_pacing(portfolio, pipeline, placeholders, config)

    fig_dep = go.Figure()
    fig_dep.add_trace(go.Bar(x=df["Quarter"],y=df["Deployment ($M)"],name="New Deployment",marker_color=C["blue"]))
    fig_dep.add_trace(go.Bar(x=df["Quarter"],y=df["Repayments/Dist ($M)"],name="Repayments/Dist",marker_color=C["teal"]))
    fig_dep.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40), xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1), height=340, barmode="stack", title="Capital Deployment & Distributions")

    fig_nav = go.Figure()
    fig_nav.add_trace(go.Scatter(x=df["Quarter"],y=df["NAV ($M)"],mode="lines+markers",name="NAV",
                                  line=dict(color=C["sky"],width=3)))
    fig_nav.add_trace(go.Scatter(x=df["Quarter"],y=df["Dry Powder ($M)"],mode="lines+markers",name="Dry Powder",
                                  line=dict(color=C["amber"],width=2,dash="dash")))
    fig_nav.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40), xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1), height=340, title="NAV & Dry Powder Projection")

    fig_util = go.Figure(go.Bar(x=df["Quarter"],y=df["Utilisation %"],marker_color=C["purple"],
                                 text=[f"{v:.0f}%" for v in df["Utilisation %"]],textposition="outside",
                                 textfont_color=C["text"]))
    fig_util.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40), xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1), height=300, title="Fund Utilisation Path", yaxis_title="%")

    tbl = dash_table.DataTable(data=df.to_dict("records"),
                                columns=[{"name":c,"id":c} for c in df.columns],
                                style_cell=TBL_CELL, style_header=TBL_HEAD, style_data_conditional=TBL_ODD,
                                page_size=16, style_table={"overflowX":"auto"}, export_format="xlsx")
    return html.Div([
        html.Div([
            card([dcc.Graph(figure=fig_dep,config={"displayModeBar":False})],dict(flex=1)),
            card([dcc.Graph(figure=fig_nav,config={"displayModeBar":False})],dict(flex=1)),
        ], style=dict(display="flex",gap=16,flexWrap="wrap",marginBottom=16)),
        card([dcc.Graph(figure=fig_util,config={"displayModeBar":False})],dict(marginBottom=16)),
        card([section_lbl("Quarterly Pacing Schedule"), tbl]),
    ])


def tab_returns(portfolio, pipeline, config):
    m         = portfolio_metrics(portfolio)
    curr_irr  = m["w_irr"]
    nav       = m["total_nav"]
    dp        = float(config.get("dry_powder",300) or 300)
    tgt       = float(config.get("target_net_twr",0.13))
    fee       = float(config.get("management_fee",0.0125))
    carry     = float(config.get("carry_rate",0.125))
    hurdle    = float(config.get("hurdle_rate",0.10))
    loss      = float(config.get("loss_drag",0.01))
    liq       = float(config.get("liquidity_reserve_pct",0.05))
    cy        = float(config.get("cash_yield",0.03))
    invested  = max(0,1-liq)
    gross_needed = (tgt+fee+loss-liq*cy)/invested if invested else tgt
    if gross_needed > hurdle:
        gross_needed += (gross_needed-hurdle)*carry
    req_irr = calc_required_irr(curr_irr, nav, dp, config)

    hold_range = [2,3,4,5,6,7]
    req_moic   = [(1+gross_needed)**h for h in hold_range]
    req_rows   = [{"Hold (y)":h,"Req MOIC":f"{m:.2f}x","Gross IRR Target":f"{gross_needed:.1%}",
                   "Exit on $30M":fmt_m(30*m),"Exit on $50M":fmt_m(50*m),"Exit on $100M":fmt_m(100*m)}
                  for h,m in zip(hold_range,req_moic)]
    req_tbl = dash_table.DataTable(data=req_rows, columns=[{"name":c,"id":c} for c in req_rows[0]],
                                    style_cell=TBL_CELL, style_header=TBL_HEAD,
                                    style_data_conditional=TBL_ODD, style_table={"overflowX":"auto"})

    fig_moic = go.Figure()
    fig_moic.add_trace(go.Scatter(x=hold_range,y=req_moic,mode="lines+markers",name="Required MOIC",
                                   line=dict(color=C["amber"],width=3)))
    fig_moic.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40), xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1), height=300, title="Required MOIC by Hold Period",
                            xaxis_title="Hold (y)", yaxis_title="MOIC")

    # Waterfall: Net TWR → Required Gross
    wf_steps = ["Target Net TWR","+ Mgmt Fee","+ Loss Drag","- Cash Yield × Reserve","÷ Invested Ratio","+ Carry Drag"]
    wf_vals  = [tgt*100, fee*100, loss*100, -(liq*cy)*100, 0, 0]
    # simplify: just show gross_needed as total
    wf_display = [
        ("Target Net TWR (LPs)",    f"{tgt:.2%}"),
        ("+ Management Fee",        f"{fee:.3%}"),
        ("+ Loss Drag",             f"{loss:.2%}"),
        ("- Cash Yield on Reserves",f"-{liq*cy:.3%}"),
        ("÷ Invested Ratio",        f"÷ {invested:.0%}"),
        ("+ Carry Drag (>hurdle)",  f"+variable ({carry:.0%})"),
        ("══════════════════════",  "════════"),
        ("REQUIRED GROSS DEAL IRR", f"{gross_needed:.2%}"),
    ]
    wf_table = html.Table([
        html.Tbody([
            html.Tr([
                html.Td(step, style=dict(fontFamily=C["mono"],fontSize=13,padding="6px 12px",
                                          fontWeight="bold" if "═" in step or "REQUIRED" in step else "normal",
                                          color=C["green"] if "REQUIRED" in step else C["text"])),
                html.Td(val,  style=dict(fontFamily=C["mono"],fontSize=13,padding="6px 12px",
                                          textAlign="right", fontWeight="bold",
                                          color=C["green"] if "REQUIRED" in step else C["text"])),
            ]) for step,val in wf_display
        ])
    ], style=dict(width="100%", borderCollapse="collapse",
                  background=C["surface"], borderRadius=8, overflow="hidden"))

    # Deal scatter
    all_deals = list(portfolio)+[
        dict(name=p["name"],nav=p.get("size",0),target_irr=p.get("target_irr",0),
             hold_period=4.0, _pool="Pipeline")
        for p in pipeline
    ]
    fig_scatter = go.Figure()
    for pool, color in [("Portfolio",C["blue"]),("Pipeline",C["purple"])]:
        ds = [d for d in all_deals if d.get("_pool","Portfolio")==pool]
        if ds:
            fig_scatter.add_trace(go.Scatter(
                x=[d.get("hold_period",0) for d in ds],
                y=[d.get("target_irr",0)*100 for d in ds],
                mode="markers+text", name=pool,
                marker=dict(color=color,size=[max(d.get("nav",10)*0.25,6) for d in ds],opacity=0.85),
                text=[d.get("name","")[-10:] for d in ds],
                textposition="top center", textfont=dict(color=C["text"],size=9),
            ))
    fig_scatter.add_hline(y=gross_needed*100, line_dash="dash", line_color=C["amber"],
                           annotation_text=f"Gross target {gross_needed:.1%}")
    fig_scatter.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40), xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1), height=340, title="Deal IRR vs Hold Period",
                               xaxis_title="Hold (y)", yaxis_title="IRR (%)")

    # Monte Carlo section
    mc_section = html.Div([
        card([
            section_lbl("Monte Carlo TWR Simulation"),
            _form_row(
                _field("Current Portfolio IRR", html.Div(f"{curr_irr:.2%}",
                        style=dict(fontFamily=C["mono"],fontSize=16,color=C["green"],padding="6px 0"))),
                _field("Future Deals Mean IRR %", dcc.Input(id="mc-mean",type="number",value=25,step=0.5,style={**INP,"width":90})),
                _field("Future Deals Std Dev %",  dcc.Input(id="mc-std",  type="number",value=5,  step=0.5,style={**INP,"width":90})),
                _field("# Simulations", dcc.Dropdown(id="mc-nsims",
                        options=[{"label":"1,000","value":1000},{"label":"5,000","value":5000},{"label":"10,000","value":10000}],
                        value=5000, style={**_dd(),"width":120})),
                html.Div([html.Label(" ",style=dict(display="block",marginBottom=4,fontSize=9)),
                          html.Button("▶ Run Simulation",id="mc-run-btn",style={**BTN(C["blue"]),"minWidth":140})]),
            ),
            html.Div(id="mc-results"),
        ])
    ], style=dict(marginTop=16))

    return html.Div([
        html.Div([
            kpi("Net Target",     f"{tgt:.1%}",          "Annualised to LPs",    C["teal"]),
            kpi("Gross Target",   f"{gross_needed:.1%}", "Net + fee drag",       C["amber"]),
            kpi("Required IRR",   f"{req_irr:.1%}",      "On future deals",      C["red"] if req_irr>0.30 else C["green"]),
            kpi("MOIC @ 3y",      f"{(1+gross_needed)**3:.2f}x","Required",      C["sky"]),
            kpi("MOIC @ 5y",      f"{(1+gross_needed)**5:.2f}x","Required",      C["purple"]),
            kpi("Carry Rate",     f"{carry:.1%}",        f"Hurdle {hurdle:.1%}", C["pink"]),
        ], style=dict(display="flex",gap=10,flexWrap="wrap",marginBottom=16)),
        html.Div([
            card([section_lbl("Required Return Thresholds"), req_tbl], dict(flex=2)),
            card([section_lbl("IRR Waterfall: Gross → Net"), wf_table],dict(flex=1, minWidth=300)),
        ], style=dict(display="flex",gap=16,flexWrap="wrap",marginBottom=16)),
        html.Div([
            card([dcc.Graph(figure=fig_moic,   config={"displayModeBar":False})],dict(flex=1)),
            card([dcc.Graph(figure=fig_scatter,config={"displayModeBar":False})],dict(flex=2)),
        ], style=dict(display="flex",gap=16,flexWrap="wrap",marginBottom=16)),
        mc_section,
    ])


def tab_settings(config):
    dp   = float(config.get("dry_powder",300) or 300)
    bs   = bite_sizes(config)
    gma  = bs.get("GP-Led (Multi-Asset)",{})

    def num(id_, val, step=0.005):
        return dcc.Input(id=id_, type="number", value=round(val,6), step=step, style=INP)
    def txt(id_, val):
        return dcc.Input(id=id_, type="text", value=str(val or ""), style=INP)

    return html.Div([
        # ── Fund Identity & Mandate ───────────────────────────────────────────
        card([
            section_lbl("Fund Identity & Mandate"),
            html.Div([
                _field("Fund Name",       txt("cfg-fund-name",     config.get("fund_name","Horizon"))),
                _field("Strategy",        txt("cfg-fund-strategy", config.get("fund_strategy","PE Secondaries & Co-Investment"))),
                _field("Vehicle Type",    txt("cfg-vehicle-type",  config.get("vehicle_type","Evergreen"))),
                _field("Domicile",        txt("cfg-domicile",      config.get("domicile","Luxembourg"))),
                _field("Fund Currency",   dcc.Dropdown(id="cfg-fund-ccy", options=dd_opts(CURRENCIES),
                                              value=config.get("fund_currency","USD"), style=_dd())),
                _field("Vintage Year",    dcc.Input(id="cfg-fund-vintage", type="number",
                                              value=int(config.get("fund_vintage",2024)), step=1, style=INP)),
            ], style=dict(display="grid", gridTemplateColumns="repeat(3,1fr)", gap=12)),
        ], dict(marginBottom=16)),

        # ── Allocation Targets ────────────────────────────────────────────────
        card([
            section_lbl("Allocation Targets (%)"),
            html.Div([
                html.Div([
                    slbl("Strategy Mix"),
                    _field("Secondaries target %",    dcc.Input(id="cfg-tgt-sec",    type="number", value=float(config.get("target_secondary_pct",70)), step=1, style=INP)),
                    _field("Co-Investments target %", dcc.Input(id="cfg-tgt-ci",     type="number", value=float(config.get("target_coinvest_pct",30)),  step=1, style=INP)),
                ], style=dict(flex=1)),
                html.Div([
                    slbl("Geography"),
                    _field("North America target %", dcc.Input(id="cfg-tgt-na",     type="number", value=float(config.get("target_na_pct",50)),     step=1, style=INP)),
                    _field("Europe target %",        dcc.Input(id="cfg-tgt-eur",    type="number", value=float(config.get("target_europe_pct",35)),  step=1, style=INP)),
                    _field("Asia target %",          dcc.Input(id="cfg-tgt-asia",   type="number", value=float(config.get("target_asia_pct",10)),    step=1, style=INP)),
                    _field("Global target %",        dcc.Input(id="cfg-tgt-global", type="number", value=float(config.get("target_global_pct",5)),   step=1, style=INP)),
                ], style=dict(flex=1)),
                html.Div([
                    slbl("Return Targets"),
                    _field("Secondary Target IRR %",  dcc.Input(id="cfg-tgt-irr-sec",  type="number", value=float(config.get("target_irr_secondary",15)), step=0.5, style=INP)),
                    _field("Co-Invest Target IRR %",  dcc.Input(id="cfg-tgt-irr-ci",   type="number", value=float(config.get("target_irr_coinvest",20)),  step=0.5, style=INP)),
                    _field("Target MOIC",             dcc.Input(id="cfg-tgt-moic",     type="number", value=float(config.get("target_moic",1.8)),          step=0.05,style=INP)),
                ], style=dict(flex=1)),
            ], style=dict(display="flex", gap=24, flexWrap="wrap")),
        ], dict(marginBottom=16)),

        # ── Restrictions Editor ───────────────────────────────────────────────
        card([
            section_lbl("Legal Restrictions — Editable"),
            html.Small(
                "Edit limits directly in the table. Add rows with the button. "
                "metric_key must match one of the computed fields — see tooltip below.",
                style=dict(color=C["muted"], fontSize=10, fontFamily=C["mono"], display="block", marginBottom=10)
            ),
            dash_table.DataTable(
                id="cfg-legal-tbl",
                data=config.get("legal_restrictions", DEFAULT_CONFIG["legal_restrictions"]),
                columns=[
                    {"name": "Label",          "id": "label",          "editable": True},
                    {"name": "Metric Key",     "id": "metric_key",     "editable": True},
                    {"name": "Limit (%  or n)","id": "limit",          "editable": True, "type": "numeric"},
                    {"name": "Max (✓) / Min (✗)", "id": "higher_is_bad","editable": True},
                    {"name": "Format (% / n)", "id": "fmt",            "editable": True},
                ],
                editable=True, row_deletable=True,
                style_cell=TBL_CELL, style_header=TBL_HEAD,
                style_data_conditional=TBL_ODD,
                style_table={"overflowX": "auto"},
            ),
            html.Div([
                html.Button("+ Add Row", id="btn-add-legal-row",
                            style={**BTN(C["teal"]), "marginTop": 8, "marginRight": 8}),
                html.Button("💾 Save Restrictions", id="btn-save-restrictions",
                            style={**BTN(C["blue"]), "marginTop": 8}),
            ]),
        ], dict(marginBottom=8)),

        card([
            section_lbl("Investment Targets — Editable"),
            html.Small(
                "Same structure as Legal Restrictions above but displayed separately on the Overview page.",
                style=dict(color=C["muted"], fontSize=10, fontFamily=C["mono"], display="block", marginBottom=10)
            ),
            dash_table.DataTable(
                id="cfg-targets-tbl",
                data=config.get("investment_targets", DEFAULT_CONFIG["investment_targets"]),
                columns=[
                    {"name": "Label",          "id": "label",          "editable": True},
                    {"name": "Metric Key",     "id": "metric_key",     "editable": True},
                    {"name": "Limit (% or n)", "id": "limit",          "editable": True, "type": "numeric"},
                    {"name": "Max (✓) / Min (✗)", "id": "higher_is_bad","editable": True},
                    {"name": "Format (% / n)", "id": "fmt",            "editable": True},
                ],
                editable=True, row_deletable=True,
                style_cell=TBL_CELL, style_header=TBL_HEAD,
                style_data_conditional=TBL_ODD,
                style_table={"overflowX": "auto"},
            ),
            html.Button("+ Add Row", id="btn-add-target-row",
                        style={**BTN(C["teal"]), "marginTop": 8}),
            html.Div(id="restrictions-save-msg",
                     style=dict(marginTop=8, color=C["green"], fontSize=11, fontFamily=C["mono"])),
            html.Details([
                html.Summary("Available metric_key values", style=dict(
                    fontSize=10, color=C["muted"], fontFamily=C["mono"], cursor="pointer", marginTop=10)),
                html.Pre(
                    "max_deal_pct      — largest single asset as % NAV\n"
                    "max_manager_pct   — largest single manager as % NAV\n"
                    "ci_pct            — co-investments as % NAV\n"
                    "sec_pct           — secondaries as % NAV\n"
                    "max_sector_pct    — largest single sector as % NAV\n"
                    "max_region_pct    — largest single region as % NAV\n"
                    "non_na_pct        — non-North America exposure as % NAV\n"
                    "non_naeur_pct     — non-NA/Europe exposure as % NAV\n"
                    "non_sec_pct       — non-secondary deals as % NAV\n"
                    "unfunded_pct      — unfunded as % fund size\n"
                    "overcommitment_pct— total commitment / NAV × 100\n"
                    "leverage_pct      — leverage % NAV (placeholder)\n"
                    "listed_pct        — listed company exposure % NAV (placeholder)\n"
                    "primary_pct       — primary fund deals as % NAV\n"
                    "max_vintage_pct   — largest single vintage as % NAV\n"
                    "num_deals         — number of portfolio positions",
                    style=dict(fontSize=10, color=C["sky"], fontFamily=C["mono"],
                               background=C["bg"], padding=10, borderRadius=6, marginTop=6)
                ),
            ]),
        ], dict(marginBottom=16)),

        # ── Fund Parameters ───────────────────────────────────────────────────
        html.Div([
            card([
                section_lbl("Current Bite Sizes (on dry powder)"),
                kpi("Min Bite",     fmt_m(gma.get("min",0)),     f"{gma.get('min_pct',0)*100:.2f}% of DP", C["teal"]),
                html.Div(style=dict(height=10)),
                kpi("Desired Bite", fmt_m(gma.get("desired",0)), f"{gma.get('desired_pct',0)*100:.2f}% of DP", C["sky"]),
                html.Div(style=dict(height=10)),
                kpi("Max Bite",     fmt_m(gma.get("max",0)),     f"{gma.get('max_pct',0)*100:.2f}% of DP", C["amber"]),
            ], dict(flex=1, minWidth=220)),
            card([
                section_lbl("Fund Parameters"),
                html.Div([
                    _field("Distribution Rate (annual)",  num("cfg-dist",    float(config.get("distribution_rate",0.20)),0.01)),
                    _field("Cash Yield",                  num("cfg-cy",      float(config.get("cash_yield",0.03)),0.005)),
                    _field("Hurdle Rate",                 num("cfg-hurdle",  float(config.get("hurdle_rate",0.10)),0.005)),
                    _field("Carry Rate",                  num("cfg-carry",   float(config.get("carry_rate",0.125)),0.005)),
                    _field("Loss Drag",                   num("cfg-loss",    float(config.get("loss_drag",0.01)),0.005)),
                    _field("Liquidity Reserve",           num("cfg-liq",     float(config.get("liquidity_reserve_pct",0.05)),0.005)),
                    _field("Avg Hold Period (y)",         dcc.Input(id="cfg-hold",     type="number",value=float(config.get("avg_hold_period",5.0)),step=0.5,style=INP)),
                    _field("Deployment Years",            dcc.Input(id="cfg-dep-years",type="number",value=float(config.get("deployment_years",4.0)),step=0.5,style=INP)),
                    _field("Deals Per Year",              dcc.Input(id="cfg-deals-py", type="number",value=float(config.get("deals_per_year",6.0)),step=1,  style=INP)),
                ], style=dict(display="grid",gridTemplateColumns="1fr 1fr 1fr",gap=12)),
                html.Div(style=dict(height=14)),
                section_lbl("Bite Size Percentages (% of Dry Powder)"),
                html.Div([
                    _field("Min %",     dcc.Input(id="cfg-bite-min",     type="number",value=round(float(config.get("bite_min_pct",0.005))*100,3),step=0.1,style=INP)),
                    _field("Desired %", dcc.Input(id="cfg-bite-desired", type="number",value=round(float(config.get("bite_desired_pct",0.0275))*100,3),step=0.25,style=INP)),
                    _field("Max %",     dcc.Input(id="cfg-bite-max",     type="number",value=round(float(config.get("bite_max_pct",0.05))*100,3),step=0.25,style=INP)),
                ], style=dict(display="grid",gridTemplateColumns="1fr 1fr 1fr",gap=12)),
                html.Div(style=dict(height=12)),
                html.Button("💾 Save All Settings", id="save-cfg-btn",
                            style={**BTN(C["blue"]),"width":"100%"}),
                html.Div(id="cfg-msg",style=dict(marginTop=8,color=C["green"],fontSize=11)),
            ], dict(flex=3, minWidth=360)),
        ], style=dict(display="flex",gap=16,flexWrap="wrap")),
    ])


# ══════════════════════════════════════════════════════════════════════════════
# CALLBACKS
# ══════════════════════════════════════════════════════════════════════════════

# Auto-calc unfunded
@app.callback(
    Output("unfunded-calc","children"),
    Input("port-total-commit","value"), Input("port-curr-commit","value"),
)
def calc_unfunded(tc, cc):
    try: return fmt_m(float(tc or 0)-float(cc or 0))
    except: return "$—"


# ── Persist table row selections across re-renders ───────────────────────────
# The tab router re-renders tab content whenever any store changes, resetting
# selected_rows=[]. These callbacks capture the selection into a separate Store
# so the edit button always knows which row was last selected.

@app.callback(
    Output("port-selected-idx","data"),
    Input("portfolio-table","selected_rows"),
    Input("portfolio-table","derived_virtual_data"),
    prevent_initial_call=True,
)
def persist_port_selection(selected_rows, virtual_data):
    if not selected_rows:
        return no_update
    pos = selected_rows[0]
    vdata = virtual_data or []
    if vdata and pos < len(vdata):
        return vdata[pos].get("_idx", pos)
    return pos

@app.callback(
    Output("pipe-selected-idx","data"),
    Input("pipeline-table","selected_rows"),
    Input("pipeline-table","derived_virtual_data"),
    prevent_initial_call=True,
)
def persist_pipe_selection(selected_rows, virtual_data):
    if not selected_rows:
        return no_update
    pos = selected_rows[0]
    vdata = virtual_data or []
    if vdata and pos < len(vdata):
        return vdata[pos].get("_idx", pos)
    return pos

@app.callback(
    Output("ph-selected-idx","data"),
    Input("ph-table","selected_rows"),
    Input("ph-table","derived_virtual_data"),
    prevent_initial_call=True,
)
def persist_ph_selection(selected_rows, virtual_data):
    if not selected_rows:
        return no_update
    pos = selected_rows[0]
    vdata = virtual_data or []
    if vdata and pos < len(vdata):
        return vdata[pos].get("_idx", pos)
    return pos


# ── Portfolio: add or update ─────────────────────────────────────────────────
@app.callback(
    Output("port-store","data"),
    Output("next-id","data"),
    Output("port-msg","children"),
    Output("edit-idx","data",allow_duplicate=True),
    Output("port-selected-idx","data",allow_duplicate=True),
    Input("add-port-btn","n_clicks"),
    State("port-store","data"), State("next-id","data"), State("edit-idx","data"),
    State("port-name","value"), State("port-manager","value"),
    State("port-strategy","value"), State("port-dtype","value"),
    State("port-stage","value"), State("port-sector","value"),
    State("port-region","value"), State("port-currency","value"),
    State("port-total-commit","value"), State("port-curr-commit","value"),
    State("port-nav","value"), State("port-irr","value"),
    State("port-hold","value"), State("port-moic","value"),
    State("port-p2nav","value"),
    State("port-vintage","value"), State("port-commit-year","value"),
    State("port-entry-date","value"),
    State("port-segment","value"), State("port-alloc-status","value"),
    prevent_initial_call=True,
)
def add_or_update_portfolio(_, portfolio, nid, edit_idx,
                             name, manager, strategy, dtype, stage, sector, region,
                             currency, total_commit, curr_commit, nav, irr,
                             hold, moic, p2nav, vintage, commit_year, entry_date,
                             segment, alloc_status):
    if not name:
        return no_update, no_update, "⚠ Enter a deal name.", no_update, no_update
    p = portfolio or []
    irr_dec  = float(irr or 0)/100
    nav_val  = float(nav or 0)
    deal = dict(
        name=name, manager=manager or "", strategy=strategy, deal_type=dtype,
        stage=stage or "Buyout", sector=sector or "Diversified",
        region=region or "North America", currency=currency or "USD",
        total_commitment=float(total_commit or nav_val),
        current_commitment=float(curr_commit or nav_val),
        nav=nav_val, target_irr=irr_dec,
        hold_period=float(hold or 5), moic=float(moic or 1.0),
        price_to_nav=float(p2nav) if p2nav is not None else None,
        vintage=int(vintage or datetime.now().year),
        commitment_year=int(commit_year or vintage or datetime.now().year),
        entry_date=str(entry_date or "").strip() or datetime.now().isoformat()[:10],
        segment=segment or "New", allocation_status=alloc_status or "Closed",
        date_added=datetime.now().isoformat()[:10],
    )
    if edit_idx is not None and 0 <= edit_idx < len(p):
        deal["id"] = p[edit_idx]["id"]
        p[edit_idx] = deal
        return p, nid, f"✓ Updated: {name}", None, None
    else:
        deal["id"] = nid
        return p+[deal], nid+1, f"✓ Added: {name}", None, None


# ── Portfolio: populate edit form ────────────────────────────────────────────
@app.callback(
    Output("edit-idx","data",allow_duplicate=True),
    Output("port-name","value"), Output("port-manager","value"),
    Output("port-strategy","value"), Output("port-dtype","value"),
    Output("port-stage","value"), Output("port-sector","value"),
    Output("port-region","value"), Output("port-currency","value"),
    Output("port-total-commit","value"), Output("port-curr-commit","value"),
    Output("port-nav","value"), Output("port-irr","value"),
    Output("port-hold","value"), Output("port-moic","value"),
    Output("port-p2nav","value"),
    Output("port-vintage","value"), Output("port-commit-year","value"),
    Output("port-entry-date","value"),
    Output("port-segment","value"), Output("port-alloc-status","value"),
    Output("port-msg","children",allow_duplicate=True),
    Input("edit-port-btn","n_clicks"),
    Input("populate-port-trigger","data"),
    State("port-selected-idx","data"),
    State("port-store","data"),
    prevent_initial_call=True,
)
def populate_edit(btn_click, trigger_idx, selected_idx, portfolio):
    ctx = callback_context
    if not portfolio:
        return [no_update]*22
    triggered = ctx.triggered[0]["prop_id"] if ctx.triggered else ""

    if "populate-port-trigger" in triggered and trigger_idx is not None:
        idx = trigger_idx
    else:
        idx = selected_idx
        if idx is None:
            return [no_update]*21 + ["⚠ Select a deal row first."]

    if idx is None or idx >= len(portfolio):
        return [no_update]*21 + ["⚠ No row selected — click a row in the table first."]
    d = portfolio[idx]
    return (
        idx,
        d["name"], d.get("manager",""),
        d.get("strategy",""), d.get("deal_type","Secondary"),
        d.get("stage","Buyout"), d.get("sector","Diversified"),
        d.get("region","North America"), d.get("currency","USD"),
        d.get("total_commitment",d.get("nav",0)),
        d.get("current_commitment",d.get("nav",0)),
        d.get("nav",0), round(d.get("target_irr",0)*100,2),
        d.get("hold_period",5), d.get("moic",1.0),
        d.get("price_to_nav",None),
        d.get("vintage",2025),
        d.get("commitment_year", d.get("vintage",2025)),
        d.get("entry_date",""),
        d.get("segment","New"),
        d.get("allocation_status","Closed"),
        f"📝 Editing: {d['name']} — make changes and click '+ Add Deal' to save."
    )


# ── Portfolio: clear form ────────────────────────────────────────────────────
@app.callback(
    Output("edit-idx","data",allow_duplicate=True),
    Output("port-name","value",allow_duplicate=True),
    Output("port-manager","value",allow_duplicate=True),
    Output("port-msg","children",allow_duplicate=True),
    Output("port-selected-idx","data",allow_duplicate=True),
    Input("clear-port-btn","n_clicks"),
    prevent_initial_call=True,
)
def clear_form(_):
    return None, "", "", "", None


# ── Portfolio: delete ────────────────────────────────────────────────────────
@app.callback(
    Output("port-store","data",allow_duplicate=True),
    Output("port-action-msg","children"),
    Input("delete-port-btn","n_clicks"),
    State("port-selected-idx","data"),
    State("port-store","data"),
    prevent_initial_call=True,
)
def delete_portfolio(_, selected_idx, portfolio):
    if selected_idx is None:
        return no_update, "Select a row first."
    if not portfolio or selected_idx >= len(portfolio):
        return no_update, "Invalid selection."
    name = portfolio[selected_idx]["name"]
    return [d for i, d in enumerate(portfolio) if i != selected_idx], f"🗑 Deleted: {name}"


# ── Portfolio: Sync NAVs from uploaded Fund Level CF ─────────────────────────
@app.callback(
    Output("port-store","data",allow_duplicate=True),
    Output("port-action-msg","children",allow_duplicate=True),
    Input("sync-nav-cf-btn","n_clicks"),
    State("port-store","data"), State("fund-cf-store","data"),
    prevent_initial_call=True,
)
def sync_nav_from_cf(_, portfolio, cf_data):
    if not cf_data:
        return no_update, "⚠ Upload a Fund Level CF file first (Deal Cashflows tab)."
    if not portfolio:
        return no_update, "⚠ No portfolio deals to update."
    updated, count = list(portfolio), 0
    for i, deal in enumerate(updated):
        nav_val = _latest_nav(cf_data, deal["name"])
        if nav_val is not None and nav_val > 0:
            updated[i] = dict(deal, nav=nav_val)
            count += 1
    if count == 0:
        return no_update, "⚠ No matching deal names found in CF file — check name spelling."
    return updated, f"✅ Synced NAV for {count} deal(s) from CF file (latest available month)."


# ── Portfolio: Sync NAVs from Liquidity Pull ──────────────────────────────────
@app.callback(
    Output("port-store","data",allow_duplicate=True),
    Output("port-action-msg","children",allow_duplicate=True),
    Input("sync-nav-liq-btn","n_clicks"),
    State("port-store","data"), State("liquidity-store","data"),
    prevent_initial_call=True,
)
def sync_nav_from_liq(_, portfolio, liq_data):
    if not liq_data:
        return no_update, "⚠ Upload a Liquidity Pull file first (Liquidity tab)."
    if not portfolio:
        return no_update, "⚠ No portfolio deals to update."
    # Use fund-level NAV from liq_data to update total — or projected NAV if available
    fund_nav = float(liq_data.get("fund_nav", 0) or 0)
    proj_nav = float(liq_data.get("projected_nav_existing", 0) or 0)
    nav_to_use = proj_nav if proj_nav > 0 else fund_nav
    if nav_to_use <= 0:
        return no_update, "⚠ No fund NAV found in Liquidity Pull file."
    # Scale each deal's NAV proportionally (keep existing weights)
    total_curr = sum(float(d.get("nav",0) or 0) for d in portfolio)
    if total_curr <= 0:
        return no_update, "⚠ Portfolio NAV is zero — cannot scale proportionally."
    scale = nav_to_use / total_curr
    updated = [dict(d, nav=round(float(d.get("nav",0) or 0) * scale, 2)) for d in portfolio]
    as_at = liq_data.get("as_at_date","unknown date")
    return updated, f"✅ Scaled all NAVs by {scale:.4f}× to match Liquidity Pull fund NAV {fmt_m(nav_to_use)} (as at {as_at})."


# ── Pipeline: add ────────────────────────────────────────────────────────────
@app.callback(
    Output("pipe-store","data"),
    Output("next-id","data",allow_duplicate=True),
    Output("edit-pipe-idx","data"),
    Output("pipe-msg","children"),
    Output("pipe-selected-idx","data",allow_duplicate=True),
    Output("pipe-form-mode","children",allow_duplicate=True),
    Input("add-pipe-btn","n_clicks"),
    State("pipe-store","data"), State("next-id","data"),
    State("edit-pipe-idx","data"),
    State("pipe-name","value"), State("pipe-strategy","value"),
    State("pipe-stage-deal","value"), State("pipe-sector","value"),
    State("pipe-region","value"), State("pipe-size","value"),
    State("pipe-irr","value"), State("pipe-pipe-stage","value"),
    State("pipe-priority","value"),
    prevent_initial_call=True,
)
def add_pipeline(_, pipeline, nid, edit_pipe_idx, name, strategy, stage_deal, sector,
                  region, size, irr, pipe_stage, priority):
    if not name: return no_update, no_update, no_update, "⚠ Enter a deal name.", no_update, ""
    p = pipeline or []
    deal = dict(id=nid, name=name, strategy=strategy, stage_deal=stage_deal,
                sector=sector, region=region, size=float(size or 0),
                target_irr=float(irr or 0)/100, pipeline_stage=pipe_stage,
                priority=priority, date_added=datetime.now().isoformat()[:10])
    if edit_pipe_idx is not None and 0 <= edit_pipe_idx < len(p):
        deal["id"] = p[edit_pipe_idx]["id"]
        p[edit_pipe_idx] = deal
        return p, nid, None, f"✓ Updated: {name}", None, ""
    else:
        deal["id"] = nid
        return p+[deal], nid+1, None, f"✓ Added pipeline: {name}", None, ""


# ── Pipeline: populate edit form ─────────────────────────────────────────────
@app.callback(
    Output("edit-pipe-idx","data",allow_duplicate=True),
    Output("pipe-name","value"), Output("pipe-strategy","value"),
    Output("pipe-stage-deal","value"), Output("pipe-sector","value"),
    Output("pipe-region","value"), Output("pipe-size","value"),
    Output("pipe-irr","value"), Output("pipe-pipe-stage","value"),
    Output("pipe-priority","value"),
    Output("pipe-msg","children",allow_duplicate=True),
    Output("pipe-form-mode","children"),
    Input("edit-pipe-btn","n_clicks"),
    Input("populate-pipe-trigger","data"),
    State("pipe-selected-idx","data"),
    State("pipe-store","data"),
    prevent_initial_call=True,
)
def populate_pipe_edit(btn_click, trigger_idx, selected_idx, pipeline):
    ctx = callback_context
    if not pipeline:
        return [no_update]*12
    triggered = ctx.triggered[0]["prop_id"] if ctx.triggered else ""
    if "populate-pipe-trigger" in triggered and trigger_idx is not None:
        idx = trigger_idx
    else:
        idx = selected_idx
        if idx is None:
            return [no_update]*11 + ["⚠ Select a deal row first."]
    if idx >= len(pipeline):
        return [no_update]*12
    d = pipeline[idx]
    return (
        idx,
        d["name"], d.get("strategy", STRATEGIES[0]),
        d.get("stage_deal","Buyout"), d.get("sector","Technology"),
        d.get("region","North America"), d.get("size",0),
        round(d.get("target_irr",0)*100, 2),
        d.get("pipeline_stage","Screening"), d.get("priority","Medium"),
        f"📝 Editing: {d['name']} — make changes and click '+ Add' to save.",
        f"✏️ Edit mode — {d['name']}",
    )


# ── Pipeline: clear form ──────────────────────────────────────────────────────
@app.callback(
    Output("edit-pipe-idx","data",allow_duplicate=True),
    Output("pipe-name","value",allow_duplicate=True),
    Output("pipe-msg","children",allow_duplicate=True),
    Output("pipe-selected-idx","data",allow_duplicate=True),
    Output("pipe-form-mode","children",allow_duplicate=True),
    Input("clear-pipe-btn","n_clicks"),
    prevent_initial_call=True,
)
def clear_pipe_form(_):
    return None, "", "", None, ""


# ── Pipeline: promote → portfolio ────────────────────────────────────────────
@app.callback(
    Output("pipe-store","data",allow_duplicate=True),
    Output("port-store","data",allow_duplicate=True),
    Output("next-id","data",allow_duplicate=True),
    Output("pipe-action-msg","children"),
    Input("promote-pipe-btn","n_clicks"),
    State("pipe-selected-idx","data"),
    State("pipe-store","data"), State("port-store","data"), State("next-id","data"),
    prevent_initial_call=True,
)
def promote_pipeline(_, selected_idx, pipeline, portfolio, nid):
    if selected_idx is None: return no_update, no_update, no_update, "Select a row first."
    pipeline = pipeline or []
    if selected_idx >= len(pipeline): return no_update, no_update, no_update, "Invalid."
    p = pipeline[selected_idx]
    size = float(p.get("size",0))
    promoted = dict(id=nid, name=p["name"],
                    manager="", strategy=p.get("strategy",""),
                    deal_type="Secondary", stage=p.get("stage_deal","Buyout"),
                    sector=p.get("sector",""), region=p.get("region",""),
                    currency="USD",
                    total_commitment=size, current_commitment=size, nav=size,
                    target_irr=float(p.get("target_irr",0)),
                    hold_period=5.0, moic=1.75, vintage=datetime.now().year,
                    segment="New", allocation_status="Pending Close",
                    date_added=datetime.now().isoformat()[:10])
    new_pipe = [d for i,d in enumerate(pipeline) if i!=selected_idx]
    return new_pipe, (portfolio or [])+[promoted], nid+1, f"✅ Promoted: {p['name']}"


# ── Pipeline: delete ─────────────────────────────────────────────────────────
@app.callback(
    Output("pipe-store","data",allow_duplicate=True),
    Output("pipe-action-msg","children",allow_duplicate=True),
    Input("delete-pipe-btn","n_clicks"),
    State("pipe-selected-idx","data"),
    State("pipe-store","data"),
    prevent_initial_call=True,
)
def delete_pipeline(_, selected_idx, pipeline):
    if selected_idx is None: return no_update, "Select a row first."
    pipeline = pipeline or []
    if selected_idx >= len(pipeline): return no_update, "Invalid."
    name = pipeline[selected_idx]["name"]
    return [d for i, d in enumerate(pipeline) if i != selected_idx], f"🗑 Deleted: {name}"


# ── Placeholder: add ─────────────────────────────────────────────────────────
@app.callback(
    Output("ph-store","data"),
    Output("next-id","data",allow_duplicate=True),
    Output("edit-ph-idx","data"),
    Output("ph-msg","children"),
    Output("ph-selected-idx","data",allow_duplicate=True),
    Output("ph-form-mode","children",allow_duplicate=True),
    Input("add-ph-btn","n_clicks"),
    State("ph-store","data"), State("next-id","data"),
    State("edit-ph-idx","data"),
    State("ph-name","value"), State("ph-strategy","value"),
    State("ph-dtype","value"), State("ph-region","value"),
    State("ph-size","value"), State("ph-irr","value"),
    State("ph-month","value"),
    prevent_initial_call=True,
)
def add_placeholder(_, placeholders, nid, edit_ph_idx, name, strategy, dtype, region, size, irr, month):
    if not name: return no_update, no_update, no_update, "⚠ Enter a name.", no_update, ""
    p = placeholders or []
    ph = dict(id=nid, name=name, strategy=strategy, deal_type=dtype,
              region=region, size=float(size or 0), target_irr=float(irr or 0)/100,
              expected_month=int(month or 0), date_added=datetime.now().isoformat()[:10])
    if edit_ph_idx is not None and 0 <= edit_ph_idx < len(p):
        ph["id"] = p[edit_ph_idx]["id"]
        p[edit_ph_idx] = ph
        return p, nid, None, f"✓ Updated: {name}", None, ""
    else:
        ph["id"] = nid
        return p+[ph], nid+1, None, f"✓ Added: {name}", None, ""


# ── Placeholder: populate edit form ──────────────────────────────────────────
@app.callback(
    Output("edit-ph-idx","data",allow_duplicate=True),
    Output("ph-name","value"), Output("ph-strategy","value"),
    Output("ph-dtype","value"), Output("ph-region","value"),
    Output("ph-size","value"), Output("ph-irr","value"),
    Output("ph-month","value"),
    Output("ph-msg","children",allow_duplicate=True),
    Output("ph-form-mode","children"),
    Input("edit-ph-btn","n_clicks"),
    State("ph-selected-idx","data"), State("ph-store","data"),
    prevent_initial_call=True,
)
def populate_ph_edit(_, selected_idx, placeholders):
    if not placeholders:
        return [no_update]*10
    if selected_idx is None:
        return [no_update]*9 + ["⚠ Select a row first."]
    idx = selected_idx
    if idx >= len(placeholders):
        return [no_update]*10
    d = placeholders[idx]
    return (
        idx,
        d["name"], d.get("strategy", STRATEGIES[0]),
        d.get("deal_type","Secondary"), d.get("region","North America"),
        d.get("size",0), round(d.get("target_irr",0)*100, 2),
        d.get("expected_month",0),
        f"📝 Editing: {d['name']} — make changes and click '+ Add' to save.",
        f"✏️ Edit mode — {d['name']}",
    )


# ── Placeholder: clear form ───────────────────────────────────────────────────
@app.callback(
    Output("edit-ph-idx","data",allow_duplicate=True),
    Output("ph-name","value",allow_duplicate=True),
    Output("ph-msg","children",allow_duplicate=True),
    Output("ph-selected-idx","data",allow_duplicate=True),
    Output("ph-form-mode","children",allow_duplicate=True),
    Input("clear-ph-btn","n_clicks"),
    prevent_initial_call=True,
)
def clear_ph_form(_):
    return None, "", "", None, ""


# ── Placeholder: delete ──────────────────────────────────────────────────────
@app.callback(
    Output("ph-store","data",allow_duplicate=True),
    Output("ph-action-msg","children"),
    Input("delete-ph-btn","n_clicks"),
    State("ph-selected-idx","data"), State("ph-store","data"),
    prevent_initial_call=True,
)
def delete_placeholder(_, selected_idx, placeholders):
    if selected_idx is None: return no_update, "Select a row first."
    placeholders = placeholders or []
    if selected_idx >= len(placeholders): return no_update, "Invalid."
    name = placeholders[selected_idx]["name"]
    return [d for i,d in enumerate(placeholders) if i!=selected_idx], f"🗑 Deleted: {name}"


# ── Pro Forma: route edit to Portfolio or Pipeline tab ────────────────────────
@app.callback(
    Output("tabs","value"),
    Output("populate-port-trigger","data"),
    Output("populate-pipe-trigger","data"),
    Output("pf-action-msg","children"),
    Input("pf-goto-port-btn","n_clicks"),
    Input("pf-goto-pipe-btn","n_clicks"),
    State("proforma-table","selected_rows"),
    State("proforma-table","data"),
    State("port-store","data"),
    State("pipe-store","data"),
    prevent_initial_call=True,
)
def proforma_edit_route(port_click, pipe_click, selected, rows, portfolio, pipeline):
    if not selected or not rows:
        return no_update, no_update, no_update, "Select a deal row first."
    ctx = callback_context
    btn = ctx.triggered[0]["prop_id"].split(".")[0]
    row = rows[selected[0]]
    source = row.get("Source","")
    deal_name = row.get("Deal","")
    if btn == "pf-goto-port-btn":
        if "Pipeline" in source:
            return no_update, no_update, no_update, "⚠ That deal is in Pipeline — use the Pipeline button."
        port_idx = next((i for i,d in enumerate(portfolio or []) if d["name"]==deal_name), None)
        if port_idx is None:
            return no_update, no_update, no_update, f"⚠ '{deal_name}' not found in Portfolio."
        return "portfolio", port_idx, no_update, ""
    else:
        if "Current" in source:
            return no_update, no_update, no_update, "⚠ That deal is in Portfolio — use the Portfolio button."
        pipe_idx = next((i for i,d in enumerate(pipeline or []) if d["name"]==deal_name), None)
        if pipe_idx is None:
            return no_update, no_update, no_update, f"⚠ '{deal_name}' not found in Pipeline."
        return "pipeline", no_update, pipe_idx, ""


# ── Settings: save (fund params, identity, allocation targets only) ───────────
@app.callback(
    Output("cfg-store","data",allow_duplicate=True),
    Output("cfg-msg","children"),
    Input("save-cfg-btn","n_clicks"),
    # Fund parameters
    State("cfg-dist","value"),  State("cfg-cy","value"),
    State("cfg-hurdle","value"),State("cfg-carry","value"),
    State("cfg-loss","value"),  State("cfg-liq","value"),
    State("cfg-hold","value"),  State("cfg-dep-years","value"),
    State("cfg-deals-py","value"),
    State("cfg-bite-min","value"),State("cfg-bite-desired","value"),State("cfg-bite-max","value"),
    # Fund identity
    State("cfg-fund-name","value"), State("cfg-fund-strategy","value"),
    State("cfg-vehicle-type","value"), State("cfg-domicile","value"),
    State("cfg-fund-ccy","value"), State("cfg-fund-vintage","value"),
    # Allocation targets
    State("cfg-tgt-sec","value"), State("cfg-tgt-ci","value"),
    State("cfg-tgt-na","value"),  State("cfg-tgt-eur","value"),
    State("cfg-tgt-asia","value"),State("cfg-tgt-global","value"),
    State("cfg-tgt-irr-sec","value"), State("cfg-tgt-irr-ci","value"),
    State("cfg-tgt-moic","value"),
    State("cfg-store","data"),
    prevent_initial_call=True,
)
def save_settings(_, dist, cy, hurdle, carry, loss, liq, hold, dep_yrs, deals_py,
                   bite_min, bite_des, bite_max,
                   fund_name, fund_strategy, vehicle_type, domicile, fund_ccy, fund_vintage,
                   tgt_sec, tgt_ci, tgt_na, tgt_eur, tgt_asia, tgt_global,
                   tgt_irr_sec, tgt_irr_ci, tgt_moic,
                   config):
    cfg = dict(config or DEFAULT_CONFIG)
    def f(v, k): return float(v) if v is not None else float(cfg.get(k, 0))
    def s(v, k): return str(v) if v is not None else str(cfg.get(k, ""))
    cfg.update({
        # Fund parameters
        "distribution_rate":     f(dist,    "distribution_rate"),
        "cash_yield":            f(cy,      "cash_yield"),
        "hurdle_rate":           f(hurdle,  "hurdle_rate"),
        "carry_rate":            f(carry,   "carry_rate"),
        "loss_drag":             f(loss,    "loss_drag"),
        "liquidity_reserve_pct": f(liq,     "liquidity_reserve_pct"),
        "avg_hold_period":       f(hold,    "avg_hold_period"),
        "deployment_years":      f(dep_yrs, "deployment_years"),
        "deals_per_year":        f(deals_py,"deals_per_year"),
        "bite_min_pct":     f(bite_min,"bite_min_pct")/100 if bite_min is not None else cfg.get("bite_min_pct",0.005),
        "bite_desired_pct": f(bite_des,"bite_desired_pct")/100 if bite_des is not None else cfg.get("bite_desired_pct",0.0275),
        "bite_max_pct":     f(bite_max,"bite_max_pct")/100 if bite_max is not None else cfg.get("bite_max_pct",0.05),
        # Fund identity
        "fund_name":       s(fund_name,     "fund_name"),
        "fund_strategy":   s(fund_strategy, "fund_strategy"),
        "vehicle_type":    s(vehicle_type,  "vehicle_type"),
        "domicile":        s(domicile,      "domicile"),
        "fund_currency":   s(fund_ccy,      "fund_currency"),
        "fund_vintage":    int(fund_vintage or cfg.get("fund_vintage", 2024)),
        # Allocation targets
        "target_secondary_pct":  f(tgt_sec,    "target_secondary_pct"),
        "target_coinvest_pct":   f(tgt_ci,     "target_coinvest_pct"),
        "target_na_pct":         f(tgt_na,     "target_na_pct"),
        "target_europe_pct":     f(tgt_eur,    "target_europe_pct"),
        "target_asia_pct":       f(tgt_asia,   "target_asia_pct"),
        "target_global_pct":     f(tgt_global, "target_global_pct"),
        "target_irr_secondary":  f(tgt_irr_sec,"target_irr_secondary"),
        "target_irr_coinvest":   f(tgt_irr_ci, "target_irr_coinvest"),
        "target_moic":           f(tgt_moic,   "target_moic"),
    })
    return cfg, "✓ Settings saved"


# ── Restrictions: add blank row to Legal table ────────────────────────────────
@app.callback(
    Output("cfg-legal-tbl","data"),
    Input("btn-add-legal-row","n_clicks"),
    State("cfg-legal-tbl","data"),
    prevent_initial_call=True,
)
def add_legal_row(_, rows):
    rows = list(rows or [])
    rows.append({"label": "New Restriction", "metric_key": "max_deal_pct",
                 "limit": 100.0, "higher_is_bad": True, "fmt": "%"})
    return rows


# ── Restrictions: add blank row to Investment Targets table ───────────────────
@app.callback(
    Output("cfg-targets-tbl","data"),
    Input("btn-add-target-row","n_clicks"),
    State("cfg-targets-tbl","data"),
    prevent_initial_call=True,
)
def add_target_row(_, rows):
    rows = list(rows or [])
    rows.append({"label": "New Target", "metric_key": "max_vintage_pct",
                 "limit": 35.0, "higher_is_bad": True, "fmt": "%"})
    return rows


# ── Restrictions: save both tables to cfg-store ───────────────────────────────
@app.callback(
    Output("cfg-store","data",allow_duplicate=True),
    Output("restrictions-save-msg","children"),
    Input("btn-save-restrictions","n_clicks"),
    State("cfg-legal-tbl","data"),
    State("cfg-targets-tbl","data"),
    State("cfg-store","data"),
    prevent_initial_call=True,
)
def save_restrictions(_, legal_rows, target_rows, config):
    cfg = dict(config or DEFAULT_CONFIG)

    def clean(rows):
        out = []
        for r in (rows or []):
            try:
                out.append({
                    "label":         str(r.get("label","")).strip() or "—",
                    "metric_key":    str(r.get("metric_key","")).strip(),
                    "limit":         float(r.get("limit", 0) or 0),
                    "higher_is_bad": str(r.get("higher_is_bad","True")).strip() not in ("False","false","0"),
                    "fmt":           str(r.get("fmt","%")).strip() or "%",
                })
            except Exception:
                pass
        return out

    cfg["legal_restrictions"]  = clean(legal_rows)
    cfg["investment_targets"]  = clean(target_rows)
    n = len(cfg["legal_restrictions"]) + len(cfg["investment_targets"])
    return cfg, f"✅ Saved {n} restrictions / targets"


# ── Monte Carlo TWR ──────────────────────────────────────────────────────────
@app.callback(
    Output("mc-results","children"),
    Input("mc-run-btn","n_clicks"),
    State("port-store","data"), State("cfg-store","data"),
    State("mc-mean","value"), State("mc-std","value"), State("mc-nsims","value"),
    prevent_initial_call=True,
)
def run_mc(_, portfolio, config, mean_pct, std_pct, n_sims):
    m        = portfolio_metrics(portfolio or [])
    curr_irr = m["w_irr"]
    nav      = m["total_nav"]
    dp       = float(config.get("dry_powder",300) or 300)
    tgt      = float(config.get("target_net_twr",0.13))
    fee      = float(config.get("management_fee",0.0125))
    carry    = float(config.get("carry_rate",0.125))
    hurdle   = float(config.get("hurdle_rate",0.10))
    loss     = float(config.get("loss_drag",0.01))
    liq      = float(config.get("liquidity_reserve_pct",0.05))
    cy       = float(config.get("cash_yield",0.03))
    invested = max(0,1-liq)

    np.random.seed(42)
    future_irrs = np.random.normal(float(mean_pct or 25)/100, float(std_pct or 5)/100, int(n_sims or 5000))
    results = []
    total = nav+dp
    for fir in future_irrs:
        blended = (nav*curr_irr + dp*fir)/total if total else 0
        gross   = blended*invested + cy*liq - fee - loss
        net     = gross-(gross-hurdle)*carry if gross>hurdle else gross
        results.append(net)
    results = np.array(results)
    prob    = (results>=tgt).mean()
    p5,p95  = np.percentile(results,5), np.percentile(results,95)

    fig_dist = go.Figure()
    fig_dist.add_trace(go.Histogram(x=results*100, nbinsx=60, marker_color=C["blue"],
                                     opacity=0.8, name="Simulated Net TWR"))
    fig_dist.add_vline(x=tgt*100, line_dash="dash", line_color=C["red"],
                       annotation_text=f"Target {tgt:.1%}")
    fig_dist.add_vline(x=np.mean(results)*100, line_dash="solid", line_color=C["green"],
                       annotation_text=f"Mean {np.mean(results):.1%}")
    fig_dist.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40), xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1), height=320, title=f"Net TWR Distribution ({n_sims:,} simulations)",
                            xaxis_title="Net TWR (%)", yaxis_title="Frequency")

    # Sensitivity line
    irr_range = np.linspace(0.12,0.40,30)
    twr_sens  = []
    for fir in irr_range:
        blended = (nav*curr_irr + dp*fir)/total if total else 0
        gross   = blended*invested + cy*liq - fee - loss
        net     = gross-(gross-hurdle)*carry if gross>hurdle else gross
        twr_sens.append(net*100)
    fig_sens = go.Figure()
    fig_sens.add_trace(go.Scatter(x=irr_range*100, y=twr_sens, mode="lines+markers",
                                   line=dict(color=C["teal"],width=3), name="Net TWR"))
    fig_sens.add_hline(y=tgt*100, line_dash="dash", line_color=C["red"],
                       annotation_text=f"Target {tgt:.1%}")
    fig_sens.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40), xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]), legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1), height=300, title="Net TWR Sensitivity to Future Deal IRR",
                            xaxis_title="Future Deal IRR (%)", yaxis_title="Net TWR (%)")

    stats = html.Div([
        html.Div([
            html.Strong("Prob. of Hitting Target: "),
            html.Span(f"{prob:.1%}", style=dict(fontSize=20,fontWeight=700,fontFamily=C["mono"],
                      color=C["green"] if prob>=0.75 else C["amber"] if prob>=0.5 else C["red"])),
        ], style=dict(marginBottom=12)),
        html.Div([
            html.P([html.Strong("Mean TWR: "),       html.Span(f"{np.mean(results):.2%}",   style=dict(color=C["blue"],  fontFamily=C["mono"]))]),
            html.P([html.Strong("Median TWR: "),     html.Span(f"{np.median(results):.2%}", style=dict(color=C["green"], fontFamily=C["mono"]))]),
            html.P([html.Strong("Std Dev: "),        html.Span(f"{np.std(results):.2%}",    style=dict(color=C["muted"], fontFamily=C["mono"]))]),
            html.P([html.Strong("5th Percentile: "), html.Span(f"{p5:.2%}",                 style=dict(color=C["red"],   fontFamily=C["mono"]))]),
            html.P([html.Strong("95th Percentile: "),html.Span(f"{p95:.2%}",                style=dict(color=C["green"], fontFamily=C["mono"]))]),
        ], style=dict(fontSize=13)),
    ])

    return html.Div([
        html.Div([
            html.Div([dcc.Graph(figure=fig_dist,config={"displayModeBar":False})], style=dict(flex=2)),
            html.Div([stats], style=dict(flex=1, padding="12px")),
        ], style=dict(display="flex",gap=16,alignItems="flex-start")),
        dcc.Graph(figure=fig_sens, config={"displayModeBar":False}),
    ])


# ── Analytics: scope toggle callback ─────────────────────────────────────────
@app.callback(
    Output("analytics-scope-kpis","children"),
    Output("analytics-charts","children"),
    Input("analytics-scope","value"),
    State("port-store","data"),
    State("pipe-store","data"),
    State("ph-store","data"),
    State("cfg-store","data"),
)
def update_analytics(scope, portfolio, pipeline, placeholders, config):
    port  = portfolio    or []
    pipe  = pipeline     or []
    ph    = placeholders or []

    # Build pool list based on selected scope
    # Each deal gets a _pool tag and _nav (for display) which is the NAV basis
    all_deals = [dict(d, _pool="Current Portfolio", _nav=d.get("nav",0)) for d in port]

    if scope in ("pipeline", "proforma"):
        for p in pipe:
            all_deals.append(dict(
                name=p.get("name",""), strategy=p.get("strategy",""),
                nav=p.get("size",0), _nav=p.get("size",0),
                target_irr=p.get("target_irr",0),
                region=p.get("region","Other"), sector=p.get("sector","Other"),
                vintage=datetime.now().year, deal_type="Secondary",
                _pool="Pipeline",
            ))

    if scope == "proforma":
        for p in ph:
            all_deals.append(dict(
                name=p.get("name",""), strategy=p.get("strategy",""),
                nav=p.get("size",0), _nav=p.get("size",0),
                target_irr=p.get("target_irr",0),
                region=p.get("region","Other"), sector=p.get("sector","Other"),
                vintage=datetime.now().year,
                deal_type=p.get("deal_type","Secondary"),
                _pool="Deployment Plan",
            ))

    if not all_deals:
        empty = card(html.P("No data yet. Add deals on the Portfolio tab.", style=dict(color=C["muted"])))
        return html.Div(), empty

    total_nav = sum(d.get("nav",0) for d in all_deals)
    scope_labels = {"current":"Current Portfolio","pipeline":"+ Pipeline","proforma":"+ Pipeline + Deployment Plan"}
    scope_label  = scope_labels.get(scope,"")

    # Scope KPI row
    pool_nav = {}
    for d in all_deals:
        pk = d.get("_pool","")
        pool_nav[pk] = pool_nav.get(pk,0)+d.get("nav",0)

    pool_colors = {"Current Portfolio":C["blue"],"Pipeline":C["purple"],"Deployment Plan":C["pink"]}
    kpi_items = [
        kpi("Total Scope NAV", fmt_m(total_nav), scope_label, C["green"]),
    ] + [
        kpi(pk, fmt_m(pv), f"{pv/total_nav*100:.1f}% of scope" if total_nav else "", pool_colors.get(pk, C["muted"]))
        for pk, pv in pool_nav.items()
    ]
    kpi_strip = html.Div(kpi_items, style=dict(display="flex",gap=10,flexWrap="wrap"))

    # Group helpers
    def group_nav_scope(key_fn):
        d = {}
        for x in all_deals:
            k = key_fn(x)
            d[k] = d.get(k,0)+x.get("nav",0)
        return d

    by_strat  = group_nav_scope(lambda x: x.get("strategy","?"))
    by_region = group_nav_scope(lambda x: x.get("region","Other"))
    by_vint   = group_nav_scope(lambda x: str(x.get("vintage","?")))
    by_sector = group_nav_scope(lambda x: x.get("sector","Other"))
    by_type   = {}
    for x in all_deals:
        k = x.get("deal_type","Secondary")
        by_type[k] = by_type.get(k,0)+x.get("nav",0)
    by_pool = {}
    for x in all_deals:
        k = x.get("_pool","Portfolio")
        by_pool[k] = by_pool.get(k,0)+x.get("nav",0)

    # Chart builders
    def _bar(lbls, vals, color, title, ref_line=None, ref_label=None):
        colors = [color]*len(vals)
        if ref_line is not None:
            colors = [C["red"] if v/total_nav*100>ref_line else color for v in vals] if total_nav else colors
        f = go.Figure(go.Bar(x=lbls, y=[v/total_nav*100 if total_nav else 0 for v in vals],
                             marker_color=colors,
                             text=[f"{v/total_nav*100:.1f}%" if total_nav else "0%" for v in vals],
                             textposition="outside", textfont_color=C["text"],
                             customdata=[[fmt_m(v)] for v in vals],
                             hovertemplate="%{x}<br>%{y:.1f}% of NAV<br>NAV: %{customdata[0]}<extra></extra>"))
        if ref_line is not None:
            f.add_hline(y=ref_line, line_dash="dash", line_color=C["amber"],
                        annotation_text=f"Limit {ref_line:.0f}%", annotation_position="top right")
        f.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40),
                        xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]),
                        yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"],title="% of NAV"),
                        legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1),
                        height=300, title=title)
        return f

    def _pie(lbls, vals, colors, title):
        f = go.Figure(go.Pie(labels=lbls, values=vals, hole=0.52,
                             marker_colors=colors, textfont_color=C["text"]))
        f.update_layout(**chart(height=300, title=title, margin=dict(t=50,b=10,l=10,r=10)))
        return f

    palette = [C["blue"],C["purple"],C["teal"],C["green"],C["amber"],C["pink"],C["sky"],C["red"]]

    lim_vintage = float(config.get("limit_single_vintage_pct", 35))
    lim_region  = float(config.get("limit_single_region_pct",  60))
    lim_sector  = float(config.get("limit_single_sector_pct",  35))

    fig_strat  = _pie(list(by_strat.keys()), list(by_strat.values()), palette[:len(by_strat)], "Strategy Exposure (% of NAV)")
    fig_region = _bar(list(by_region.keys()), list(by_region.values()), C["sky"],    "Regional Exposure (% of NAV)", ref_line=lim_region)
    fig_vint   = _bar([k for k in sorted(by_vint.keys())],
                      [by_vint[k] for k in sorted(by_vint.keys())],   C["purple"],  "Vintage Exposure (% of NAV)",  ref_line=lim_vintage)
    fig_sector = _pie(list(by_sector.keys()), list(by_sector.values()), palette[:len(by_sector)], "Sector Exposure (% of NAV)")
    fig_type   = _pie(list(by_type.keys()),   list(by_type.values()),   [C["sky"],C["teal"],C["purple"],C["amber"]], "Secondary vs Co-Invest (% of NAV)")

    # Pool breakdown bar (only interesting when scope > current)
    pool_keys   = list(by_pool.keys())
    pool_colors_list = [pool_colors.get(pk, C["muted"]) for pk in pool_keys]
    fig_pool = go.Figure(go.Bar(
        x=pool_keys, y=[by_pool[k] for k in pool_keys],
        marker_color=pool_colors_list,
        text=[fmt_m(by_pool[k]) for k in pool_keys], textposition="outside", textfont_color=C["text"],
    ))
    fig_pool.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40),
                           xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]),
                           yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"],title="NAV ($M)"),
                           legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1),
                           height=260, title="NAV by Pool")

    # Concentration (based on current portfolio only for top-N metrics)
    m = portfolio_metrics(port)
    cats   = ["Top 1","Top 3","Top 5"]
    c_vals = [m["top1"]*100, m["top3"]*100, m["top5"]*100]
    c_lims = [15, 40, 60]
    fig_conc = go.Figure(go.Bar(x=cats, y=c_vals,
                                 marker_color=[C["red"] if v>l else C["green"] for v,l in zip(c_vals,c_lims)],
                                 text=[f"{v:.1f}%" for v in c_vals], textposition="outside", textfont_color=C["text"]))
    for lim in c_lims:
        fig_conc.add_hline(y=lim, line_dash="dash", line_color=C["amber"], opacity=0.6)
    fig_conc.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40),
                           xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]),
                           yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"]),
                           legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1),
                           height=300, title="Concentration Risk — Current Portfolio (% of NAV)")

    # Scatter: IRR vs NAV by pool
    tgt = float(config.get("target_net_twr",0.13))*100
    fig_scatter = go.Figure()
    pool_order = ["Current Portfolio","Pipeline","Deployment Plan"]
    p_colors   = [C["blue"],C["purple"],C["pink"]]
    for pool, pcolor in zip(pool_order, p_colors):
        ds = [d for d in all_deals if d.get("_pool")==pool]
        if ds:
            fig_scatter.add_trace(go.Scatter(
                x=[d.get("nav",0) for d in ds],
                y=[d.get("target_irr",0)*100 for d in ds],
                mode="markers+text", name=pool,
                marker=dict(color=pcolor, size=10, opacity=0.85, line=dict(color=C["border"],width=1)),
                text=[d.get("name","")[-10:] for d in ds],
                textposition="top center", textfont=dict(color=C["text"],size=9),
                hovertext=[f"<b>{d.get('name','')}</b><br>{d.get('strategy','')}<br>"
                           f"NAV: {fmt_m(d.get('nav',0))}<br>Region: {d.get('region','')}<br>"
                           f"Pool: {d.get('_pool','')}" for d in ds],
                hoverinfo="text",
            ))
    fig_scatter.add_hline(y=tgt, line_dash="dash", line_color=C["amber"],
                          annotation_text=f"Target {tgt:.1f}%", annotation_position="right")
    fig_scatter.update_layout(**CHART_BASE, margin=dict(l=52,r=20,t=40,b=40),
                               xaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"],title="NAV ($M)"),
                               yaxis=dict(gridcolor=C["border"],zerolinecolor=C["border"],title="Target IRR (%)"),
                               legend=dict(bgcolor=C["panel"],bordercolor=C["border"],borderwidth=1),
                               height=340, title="IRR vs NAV — Deal-level Scatter")

    # Wall table
    wall_rows = [{"Pool":d.get("_pool",""),"Deal":d.get("name",""),"Strategy":d.get("strategy",""),
                  "Type":d.get("deal_type",""),"Region":d.get("region",""),"Sector":d.get("sector",""),
                  "Vintage":d.get("vintage",""),
                  "NAV ($M)":fmt_m(d.get("nav",0)),
                  "% of Scope NAV": f"{d.get('nav',0)/total_nav*100:.1f}%" if total_nav else "0%",
                  "IRR":f"{d.get('target_irr',0)*100:.1f}%"}
                 for d in all_deals]
    wall_tbl = dash_table.DataTable(
        data=wall_rows, columns=[{"name":c,"id":c} for c in wall_rows[0]],
        style_cell=TBL_CELL, style_header=TBL_HEAD, style_data_conditional=TBL_ODD,
        sort_action="native", filter_action="native",
        page_size=20, style_table={"overflowX":"auto"}, export_format="xlsx",
    )

    charts = html.Div([
        html.Div([
            card([dcc.Graph(figure=fig_pool,   config={"displayModeBar":False})], dict(flex=1)),
            card([dcc.Graph(figure=fig_type,   config={"displayModeBar":False})], dict(flex=1)),
            card([dcc.Graph(figure=fig_strat,  config={"displayModeBar":False})], dict(flex=1)),
        ], style=dict(display="flex",gap=16,flexWrap="wrap",marginBottom=16)),
        html.Div([
            card([dcc.Graph(figure=fig_region, config={"displayModeBar":False})], dict(flex=1)),
            card([dcc.Graph(figure=fig_vint,   config={"displayModeBar":False})], dict(flex=1)),
            card([dcc.Graph(figure=fig_sector, config={"displayModeBar":False})], dict(flex=1)),
        ], style=dict(display="flex",gap=16,flexWrap="wrap",marginBottom=16)),
        html.Div([
            card([dcc.Graph(figure=fig_conc,   config={"displayModeBar":False})], dict(flex=1)),
            card([dcc.Graph(figure=fig_scatter,config={"displayModeBar":True})],  dict(flex=2)),
        ], style=dict(display="flex",gap=16,flexWrap="wrap",marginBottom=16)),
        card([section_lbl("Full Deal Detail — Selected Scope"), wall_tbl]),
    ])

    return kpi_strip, charts


# ── Reflect config back to header (only on settings save, not every store change) ──
@app.callback(
    Output("hdr-fund-size","value",allow_duplicate=True),
    Output("hdr-dry-powder","value",allow_duplicate=True),
    Output("hdr-twr","value",allow_duplicate=True),
    Output("hdr-fee","value",allow_duplicate=True),
    Input("save-cfg-btn","n_clicks"),
    State("cfg-store","data"),
    prevent_initial_call=True,
)
def reflect_cfg(_, config):
    if not config:
        return no_update, no_update, no_update, no_update
    return (config.get("fund_size"),
            config.get("dry_powder"),
            round(float(config.get("target_net_twr",0.13))*100,2),
            round(float(config.get("management_fee",0.0125))*100,3))


# ── Fund CF Upload ────────────────────────────────────────────────────────────
@app.callback(
    Output("fund-cf-store","data"),
    Output("upload-cf-status","children"),
    Input("upload-fund-cf","contents"),
    State("upload-fund-cf","filename"),
    prevent_initial_call=True,
)
def upload_fund_cf(contents, filename):
    if not contents:
        return no_update, ""
    try:
        from openpyxl import load_workbook
        _, content_string = contents.split(",")
        decoded = base64.b64decode(content_string)
        wb = load_workbook(BytesIO(decoded), data_only=True)
        ws = None
        for sname in ["Fund Level CF","Fund_Level_CF","FundLevelCF","Sheet1"]:
            if sname in wb.sheetnames:
                ws = wb[sname]; break
        if ws is None:
            ws = wb[wb.sheetnames[0]]

        # Detect section columns from Row 1
        SECTION_LABELS = {"Net CF":"net_cf","Called Capital":"calls","Distributions":"distributions","NAV":"nav"}
        sec_start = {}
        for col in range(1, min(ws.max_column+1, 700)):
            val = ws.cell(1,col).value
            if val in SECTION_LABELS:
                sec_start[SECTION_LABELS[val]] = col
        if not sec_start:
            sec_start = {"net_cf":14,"calls":190,"distributions":364,"nav":538}

        sec_order = ["net_cf","calls","distributions","nav"]
        sec_bounds = {}
        for i,k in enumerate(sec_order):
            s = sec_start.get(k)
            if s is None: continue
            nxt = [sec_start[j] for j in sec_order[i+1:] if j in sec_start]
            sec_bounds[k] = (s, (min(nxt)-1) if nxt else ws.max_column)

        # Month keys from Row 2 of Net CF section
        ns, ne = sec_bounds.get("net_cf",(14,189))
        all_months, col_to_mk = [], {}
        for col in range(ns, ne+1):
            v = ws.cell(2,col).value
            if v is None: continue
            try:
                mk = pd.to_datetime(v).strftime("%Y-%m-%d")
                col_to_mk[col] = mk
                if mk not in all_months: all_months.append(mk)
            except: pass

        def sec_month_map(s,e):
            r={}
            for col in range(s,e+1):
                v=ws.cell(2,col).value
                if v is None: continue
                try: r[col]=pd.to_datetime(v).strftime("%Y-%m-%d")
                except: pass
            return r

        smaps = {k: sec_month_map(s,e) for k,(s,e) in sec_bounds.items()}

        SKIP_NAMES = {"AIC Name","Investments","Investment Type","Investment Status",
                      "Commitment Year","Commitment ($m)","Total Liquidity"}
        SKIP_PREF  = ("total ","cash","liquidity")

        deals_data = []
        for row in range(7, ws.max_row+1):
            nm = ws.cell(row,5).value
            if not nm: continue
            ns2 = str(nm).strip()
            if not ns2 or ns2 in SKIP_NAMES or ns2.lower().startswith(SKIP_PREF): continue
            c_raw = ws.cell(row,11).value
            p_raw = ws.cell(row,12).value
            u_raw = ws.cell(row,13).value
            c_m = float(c_raw) if isinstance(c_raw,(int,float)) else 0.0
            p_m = float(p_raw) if isinstance(p_raw,(int,float)) else 0.0
            u_m = (float(u_raw)/1e6 if abs(float(u_raw))>1000 else float(u_raw)) if isinstance(u_raw,(int,float)) else 0.0
            dd = {"name":ns2,
                  "investments":str(ws.cell(row,6).value or ns2).strip(),
                  "portfolio_type":str(ws.cell(row,7).value or "").strip(),
                  "type":str(ws.cell(row,8).value or "").strip(),
                  "investment_status":str(ws.cell(row,9).value or "").strip(),
                  "commitment_year":ws.cell(row,10).value,
                  "commitment":c_m,"paid_in":p_m,"unfunded":u_m,
                  "current_commitment":p_m if p_m>0 else max(c_m-u_m,0),
                  "nav":0.0,
                  "net_cf_series":{},"calls_series":{},"distributions_series":{},"nav_series":{}}
            for sk,stk in [("net_cf","net_cf_series"),("calls","calls_series"),
                            ("distributions","distributions_series"),("nav","nav_series")]:
                if sk not in sec_bounds: continue
                s2,e2 = sec_bounds[sk]
                for col in range(s2,e2+1):
                    mk = smaps[sk].get(col)
                    if not mk: continue
                    cv = ws.cell(row,col).value
                    dd[stk][mk] = float(cv) if isinstance(cv,(int,float)) else 0.0
            # latest NAV
            nav_s = dd["nav_series"]
            nz = [(k,v) for k,v in sorted(nav_s.items()) if v not in (None,0,0.0)]
            dd["nav"] = float(nz[-1][1]) if nz else c_m
            deals_data.append(dd)

        if deals_data:
            deals_data[0]["_all_month_keys"] = all_months

        n_inv = sum(1 for d in deals_data if d.get("portfolio_type","").lower()!="placeholder")
        n_ph  = sum(1 for d in deals_data if d.get("portfolio_type","").lower()=="placeholder")
        rng = (f"{pd.to_datetime(all_months[0]).strftime('%b %Y')} – "
               f"{pd.to_datetime(all_months[-1]).strftime('%b %Y')}") if all_months else "?"
        msg = html.Div([
            html.Span("✅ ", style=dict(color=C["green"])),
            html.Strong(filename, style=dict(fontFamily=C["mono"],fontSize=12)),
            html.Span(f" — {n_inv} investments + {n_ph} placeholders · {len(all_months)} months ({rng})",
                      style=dict(color=C["muted"],fontFamily=C["mono"],fontSize=11)),
        ])
        return deals_data, msg
    except Exception as e:
        import traceback; traceback.print_exc()
        return no_update, html.Span(f"❌ {filename}: {e}", style=dict(color=C["red"],fontFamily=C["mono"],fontSize=11))


# ── CF window badge ───────────────────────────────────────────────────────────
@app.callback(
    Output("cf-window-badge","children"),
    Input("cf-start-month","value"), Input("cf-horizon","value"),
    Input("fund-cf-store","data"),
    prevent_initial_call=True,
)
def cf_window_badge(start_month, horizon, cf_data):
    horizon = int(horizon or 12)
    if not cf_data:
        return html.Span("Upload a Fund Level CF file to see cashflow data.",
                         style=dict(color=C["muted"],fontSize=11,fontFamily=C["mono"]))
    all_months = cf_data[0].get("_all_month_keys") or sorted(
        {mk for d in cf_data for mk in (d.get("net_cf_series") or {}).keys()})
    if not all_months:
        return html.Span("No month data found.", style=dict(color=C["amber"],fontSize=11))
    # Find start index
    sidx = 0
    if start_month:
        try:
            tgt = pd.to_datetime(start_month)
            for i,mk in enumerate(all_months):
                if pd.to_datetime(mk) >= tgt: sidx=i; break
        except: pass
    window = all_months[sidx:sidx+horizon]
    if not window:
        return html.Span("No data in selected window.", style=dict(color=C["amber"],fontSize=11))
    first = pd.to_datetime(window[0]).strftime("%b %Y")
    last  = pd.to_datetime(window[-1]).strftime("%b %Y")
    return html.Span(f"Window: {first} → {last}  ·  {len(window)} months",
                     style=dict(color=C["teal"],fontFamily=C["mono"],fontSize=12,fontWeight=600))


# ── CF monthly table ──────────────────────────────────────────────────────────
@app.callback(
    Output("cf-monthly-table","children"),
    Input("port-store","data"), Input("fund-cf-store","data"),
    Input("cf-start-month","value"), Input("cf-horizon","value"),
    Input("cf-sections","value"),
    prevent_initial_call=True,
)
def cf_monthly_table(portfolio, cf_data, start_month, horizon, sections):
    sections   = sections or ["net","calls","dists","nav"]
    num_months = int(horizon or 12)
    META = ["AIC Name","Investments","Portfolio Type","Investment Type",
            "Investment Status","Commitment Year","Commitment ($m)","Paid In","Unfunded"]

    def fmt(v):
        try: f=float(v)
        except: return "—"
        return "—" if f==0 else (f"{f:+.2f}" if f<0 else f"{f:.2f}")
    def fmt_nav(v):
        try: f=float(v)
        except: return "—"
        return "—" if f==0 else f"{f:.2f}"

    def build_section(title, key, disp_months, lbls, formatter, data):
        all_cols = META + lbls
        rows = []
        totals = {l:0.0 for l in lbls}
        for d in data:
            if d.get("name") == "_all_month_keys": continue
            s = d.get(key) or {}
            row = {"AIC Name":d.get("name",""), "Investments":d.get("investments",d.get("name","")),
                   "Portfolio Type":d.get("portfolio_type",""), "Investment Type":d.get("type",""),
                   "Investment Status":d.get("investment_status",""),
                   "Commitment Year":str(d.get("commitment_year","") or ""),
                   "Commitment ($m)":f"{float(d.get('commitment',0) or 0):.2f}",
                   "Paid In":f"{float(d.get('paid_in',d.get('current_commitment',0)) or 0):.2f}",
                   "Unfunded":f"{float(d.get('unfunded',0) or 0):.2f}"}
            for mk,lbl in zip(disp_months,lbls):
                v=float(s.get(mk) or 0); row[lbl]=formatter(v); totals[lbl]+=v
            rows.append(row)
        # totals row
        trow = {"AIC Name":"TOTAL","Investments":"","Portfolio Type":"","Investment Type":"",
                "Investment Status":"","Commitment Year":"",
                "Commitment ($m)":f"{sum(float(d.get('commitment',0) or 0) for d in data):.2f}",
                "Paid In":f"{sum(float(d.get('paid_in',d.get('current_commitment',0)) or 0) for d in data):.2f}",
                "Unfunded":f"{sum(float(d.get('unfunded',0) or 0) for d in data):.2f}"}
        for lbl in lbls: trow[lbl]=formatter(totals[lbl])
        rows.append(trow)
        n = len(rows)
        sec_colors = {"Net CF":C["blue"],"Called Capital":C["red"],"Distributions":C["green"],"NAV":C["purple"]}
        hc = sec_colors.get(title, C["text"])
        style_cond = [{"if":{"row_index":"odd"},"backgroundColor":C["surface"]},
                      {"if":{"row_index":n-1},"backgroundColor":C["surface"],"fontWeight":"bold","color":C["sky"]},
                      {"if":{"column_id":"AIC Name"},"fontWeight":"bold"}]
        if key != "nav_series":
            style_cond += [{"if":{"filter_query":f'{{{l}}} contains "-"',"column_id":l},"color":C["red"]} for l in lbls]
        return html.Div([
            html.Div([html.Span("■ ",style=dict(color=hc,fontSize=14)),
                      html.Span(title,style=dict(fontWeight="bold",fontSize=13,color=hc)),
                      html.Span(f"  ({n-1} deals · {len(lbls)} months)",
                                style=dict(fontSize=10,color=C["muted"],marginLeft=8))],
                     style=dict(padding="10px 16px",background=C["surface"],
                                borderTop=f"2px solid {hc}",borderRadius="8px 8px 0 0")),
            dash_table.DataTable(
                data=rows, columns=[{"name":c,"id":c} for c in all_cols],
                style_cell={**TBL_CELL,"textAlign":"right","minWidth":"80px","maxWidth":"140px"},
                style_cell_conditional=[{"if":{"column_id":c},"textAlign":"left","minWidth":"140px"}
                                         for c in ["AIC Name","Investments","Portfolio Type","Investment Type","Investment Status"]],
                style_header={**TBL_HEAD,"textAlign":"center","position":"sticky","top":0,"zIndex":1},
                style_data={"backgroundColor":C["panel"],"color":C["text"],"border":f"1px solid {C['border']}"},
                style_data_conditional=style_cond,
                style_table={"overflowX":"auto","maxHeight":"460px","overflowY":"auto"},
                fixed_columns={"headers":True,"data":2}, fixed_rows={"headers":True},
                page_action="none", export_format="xlsx",
            )
        ], style=dict(marginBottom=16,border=f"1px solid {C['border']}",borderRadius=8,overflow="hidden"))

    if cf_data:
        all_months = cf_data[0].get("_all_month_keys") or sorted(
            {mk for d in cf_data for mk in (d.get("net_cf_series") or {}).keys()})
        sidx = 0
        if start_month:
            try:
                tgt = pd.to_datetime(start_month)
                for i,mk in enumerate(all_months):
                    if pd.to_datetime(mk) >= tgt: sidx=i; break
            except: pass
        disp = all_months[sidx:sidx+num_months]
        lbls = [pd.to_datetime(m).strftime("%b-%y") for m in disp]
        if not disp:
            return html.Span("No data in window.", style=dict(color=C["amber"],fontSize=12))
        parts = []
        if "net"   in sections: parts.append(build_section("Net CF",        "net_cf_series",        disp,lbls,fmt,    cf_data))
        if "calls" in sections: parts.append(build_section("Called Capital","calls_series",         disp,lbls,fmt,    cf_data))
        if "dists" in sections: parts.append(build_section("Distributions", "distributions_series", disp,lbls,fmt,    cf_data))
        if "nav"   in sections: parts.append(build_section("NAV",           "nav_series",           disp,lbls,fmt_nav,cf_data))
        return html.Div(parts)

    # Fallback from manual deals
    if not portfolio:
        return html.Span("Upload a Fund Level CF file, or add deals on the Portfolio tab.",
                         style=dict(color=C["muted"],fontSize=12,fontFamily=C["mono"]))
    fb_rows = [{"AIC Name":d.get("name",""), "Investments":d.get("name",""),
                "Portfolio Type":d.get("segment",""), "Investment Type":d.get("strategy",""),
                "Investment Status":d.get("allocation_status",""),
                "Commitment Year":str(d.get("vintage","") or ""),
                "Commitment ($m)":f"{float(d.get('total_commitment',d.get('nav',0)) or 0):.2f}",
                "Paid In":f"{float(d.get('current_commitment',d.get('nav',0)) or 0):.2f}",
                "Unfunded":f"{float(max(d.get('total_commitment',0)-d.get('current_commitment',0),0)):.2f}",
                "Current NAV":f"{float(d.get('nav',0) or 0):.2f}"}
               for d in portfolio]
    cols = list(fb_rows[0].keys())
    return html.Div([
        html.Span("Showing manual deals — upload Fund Level CF for full cashflow matrix.",
                  style=dict(color=C["amber"],fontSize=11,fontFamily=C["mono"],display="block",marginBottom=8)),
        dash_table.DataTable(data=fb_rows, columns=[{"name":c,"id":c} for c in cols],
                              style_cell=TBL_CELL, style_header=TBL_HEAD,
                              style_data_conditional=TBL_ODD,
                              sort_action="native", style_table={"overflowX":"auto"}, export_format="xlsx"),
    ])


# ── CF chart ──────────────────────────────────────────────────────────────────
@app.callback(
    Output("cf-chart-area","children"),
    Input("port-store","data"), Input("fund-cf-store","data"),
    Input("cf-start-month","value"), Input("cf-horizon","value"),
    Input("cf-sections","value"),
    prevent_initial_call=True,
)
def cf_chart(portfolio, cf_data, start_month, horizon, sections):
    sections   = sections or ["net","calls","dists","nav"]
    num_months = int(horizon or 12)
    fig = go.Figure()

    if cf_data:
        all_months = cf_data[0].get("_all_month_keys") or sorted(
            {mk for d in cf_data for mk in (d.get("net_cf_series") or {}).keys()})
        sidx = 0
        if start_month:
            try:
                tgt = pd.to_datetime(start_month)
                for i,mk in enumerate(all_months):
                    if pd.to_datetime(mk) >= tgt: sidx=i; break
            except: pass
        disp = all_months[sidx:sidx+num_months]
        if not disp:
            return html.Div()
        lbls = [pd.to_datetime(m).strftime("%b %Y") for m in disp]
        calls = [sum(float((d.get("calls_series") or {}).get(mk,0) or 0) for d in cf_data) for mk in disp]
        dists = [sum(float((d.get("distributions_series") or {}).get(mk,0) or 0) for d in cf_data) for mk in disp]
        net   = [sum(float((d.get("net_cf_series") or {}).get(mk,0) or 0) for d in cf_data) for mk in disp]
        nav   = [sum(float((d.get("nav_series") or {}).get(mk,0) or 0) for d in cf_data) for mk in disp]
    else:
        lbls  = [(datetime.now().replace(day=1)+relativedelta(months=i)).strftime("%b %Y") for i in range(num_months)]
        total_nav = sum(d.get("nav",0) for d in (portfolio or []))
        calls,dists,net,nav = [0]*num_months,[0]*num_months,[0]*num_months,[total_nav]*num_months

    if "calls" in sections:
        fig.add_trace(go.Bar(x=lbls,y=calls,name="Called Capital",marker_color=C["red"],opacity=0.85))
    if "dists" in sections:
        fig.add_trace(go.Bar(x=lbls,y=dists,name="Distributions", marker_color=C["green"],opacity=0.85))
    if "net"   in sections:
        fig.add_trace(go.Scatter(x=lbls,y=net,name="Net CF",mode="lines+markers",
                                  line=dict(color=C["amber"],width=2,dash="dot"),marker=dict(size=4)))
    if "nav"   in sections:
        fig.add_trace(go.Scatter(x=lbls,y=nav,name="Ending NAV",mode="lines+markers",
                                  line=dict(color=C["blue"],width=3),marker=dict(size=4,color=C["sky"]),
                                  yaxis="y2"))
    fig.update_layout(**chart(barmode="group", height=400,
                      yaxis=dict(title="Cashflows ($M)",gridcolor=C["border"],zeroline=True,zerolinecolor=C["border2"]),
                      yaxis2=dict(title="NAV ($M)",overlaying="y",side="right",showgrid=False),
                      hovermode="x unified",
                      xaxis=dict(tickangle=-45,tickfont=dict(size=10,family=C["mono"])),
                      legend=dict(orientation="h",y=1.08,x=0)))
    return dcc.Graph(figure=fig, config={"displayModeBar":True})


# ── Liquidity Upload ──────────────────────────────────────────────────────────
@app.callback(
    Output("liquidity-store","data"),
    Output("upload-liq-status","children"),
    Input("upload-liquidity","contents"),
    State("upload-liquidity","filename"),
    prevent_initial_call=True,
)
def upload_liquidity(contents, filename):
    if not contents:
        return no_update, ""
    try:
        from openpyxl import load_workbook
        _, cs = contents.split(",")
        decoded = base64.b64decode(cs)
        wb = load_workbook(BytesIO(decoded), data_only=True)
        ws = None
        for sname in ["Liquidity Pull","Liquidity_Pull","LiquidityPull"]:
            if sname in wb.sheetnames: ws=wb[sname]; break
        if ws is None: ws = wb[wb.sheetnames[0]]

        def cell(r,c):
            v = ws.cell(r,c).value
            return float(v)/1e6 if isinstance(v,(int,float)) else None

        liq = {
            "as_at_date":     str(ws.cell(4,2).value or ""),
            "current_quarter":str(ws.cell(5,2).value or ""),
            "current_month":  str(ws.cell(6,2).value or ""),
            "fund_nav":       cell(10,3) or 0,
            "current_cash":   cell(58,3) or 0,
            "glf_balance":    cell(59,3) or 0,
            "cqs_balance":    cell(60,3) or 0,
            "total_liquidity":cell(16,3) or 0,
            "surplus_liquidity_post_buffer": cell(63,3) or 0,
            "projected_nav_existing":        cell(71,3) or 0,
            "projected_nav_existing_pipeline":cell(72,3) or 0,
            "dec_2026_dry_powder": cell(53,9) or 0,
            "max_deployable_capital": {},
            "near_term_flows": {},
            "nav_projections": {},
        }
        for col in range(1,25):
            try:
                lbl = ws.cell(28,col).value
                md  = ws.cell(53,col).value
                if lbl and md and isinstance(md,(int,float)):
                    liq["max_deployable_capital"][str(lbl)] = float(md)/1e6
                nav_end = ws.cell(39,col).value
                if lbl and nav_end and isinstance(nav_end,(int,float)):
                    liq["nav_projections"][str(lbl)] = float(nav_end)/1e6
                subs  = ws.cell(32,col).value or 0
                reds  = ws.cell(33,col).value or 0
                flows = ws.cell(34,col).value or 0
                if lbl:
                    liq["near_term_flows"][str(lbl)] = {
                        "subscriptions":  float(subs)/1e6  if isinstance(subs,(int,float))  else 0,
                        "redemptions":    float(reds)/1e6  if isinstance(reds,(int,float))  else 0,
                        "portfolio_flows":float(flows)/1e6 if isinstance(flows,(int,float)) else 0,
                    }
            except: pass

        msg = html.Div([
            html.Span("✅ ", style=dict(color=C["green"])),
            html.Strong(filename, style=dict(fontFamily=C["mono"],fontSize=12)),
            html.Span(f" — NAV: {liq['fund_nav']:.1f}M · Cash: {liq['current_cash']:.1f}M · "
                      f"Dec-26 DP: {liq['dec_2026_dry_powder']:.1f}M",
                      style=dict(color=C["muted"],fontFamily=C["mono"],fontSize=11)),
        ])
        return liq, msg
    except Exception as e:
        return no_update, html.Span(f"❌ {filename}: {e}", style=dict(color=C["red"],fontFamily=C["mono"],fontSize=11))


# ── Discount: recalc on edit ──────────────────────────────────────────────────
@app.callback(
    Output("discount-datatable","data"),
    Input("discount-datatable","data_timestamp"),
    State("discount-datatable","data"),
    prevent_initial_call=True,
)
def recalc_discount(_, rows):
    if not rows: return rows
    out = []
    for row in rows:
        try:
            nav = float(row.get("NAV ($m)",0) or 0)
            disc = float(row.get("Discount (%)",0) or 0)
            pp = nav*(1-disc/100) if nav else 0
            p2n = pp/nav if nav else None
            row["Purchase Price ($m)"] = round(pp,2)
            row["P/NAV"] = round(p2n,4) if p2n is not None else None
        except: pass
        out.append(row)
    return out


# ── Discount: save ────────────────────────────────────────────────────────────
@app.callback(
    Output("discount-store","data"),
    Output("disc-save-status","children"),
    Input("btn-save-disc","n_clicks"),
    State("discount-datatable","data"),
    State("discount-store","data"),
    prevent_initial_call=True,
)
def save_discounts(_, rows, store):
    if not rows:
        return store or {}, ""
    s = dict(store or {})
    for r in rows:
        n = r.get("Deal","")
        if not n: continue
        try: s[n] = {"discount_pct": float(r.get("Discount (%)",0) or 0),
                     "nav_override":  float(r.get("NAV ($m)",0) or 0)}
        except: pass
    return s, f"✅ Saved {len(s)} deal discount(s)"


# ── Discount: refresh (clear nav overrides) ───────────────────────────────────
@app.callback(
    Output("discount-store","data",allow_duplicate=True),
    Output("disc-save-status","children",allow_duplicate=True),
    Input("btn-refresh-disc","n_clicks"),
    State("discount-store","data"),
    prevent_initial_call=True,
)
def refresh_discounts(_, store):
    if not store: return {}, "Cleared"
    refreshed = {n: {"discount_pct": v.get("discount_pct",0), "nav_override": 0}
                 for n,v in (store or {}).items()}
    return refreshed, "🔄 NAV overrides cleared — will re-pull from uploaded CF"


# ── Auto-persist ─────────────────────────────────────────────────────────────
@app.callback(
    Output("save-status-lbl","children"),
    Input("port-store","data"), Input("pipe-store","data"),
    Input("ph-store","data"), Input("cfg-store","data"),
    Input("next-id","data"),
)
def persist(portfolio, pipeline, placeholders, config, next_id):
    try:
        save_data(portfolio, pipeline, placeholders, config, next_id)
        n_deals = len(portfolio or [])
        return html.Span([
            html.Span("💾 ", style=dict(color=C["green"])),
            html.Span(f"Saved {n_deals} deal{'s' if n_deals!=1 else ''} · {datetime.now().strftime('%H:%M:%S')}",
                      style=dict(color=C["muted"], fontSize=11, fontFamily=C["mono"])),
            html.Span(f" → {DATA_FILE}", style=dict(color=C["dim"], fontSize=10, fontFamily=C["mono"])),
        ])
    except Exception as e:
        return html.Span(f"❌ Save failed: {e}",
                         style=dict(color=C["red"], fontSize=11, fontFamily=C["mono"]))


if __name__ == "__main__":
    print("\n" + "="*70)
    print("HORIZON PORTFOLIO TOOL v88 — PE Secondaries & Co-Investment")
    print("="*70)
    print(f"\n✅  http://localhost:8060")
    print("\nTabs: Portfolio | Pipeline | Pro Forma | Analytics | Segments & TWR |")
    print("      Dry Powder | Pacing | Return Calc | Settings")
    print("\nPress CTRL+C to stop\n")
    app.run(debug=True, host="0.0.0.0", port=8060)