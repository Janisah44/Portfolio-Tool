"""
HORIZON PORTFOLIO TOOL - INSTITUTIONAL GRADE LP MANAGEMENT
Comprehensive evergreen fund portfolio management and analytics platform

Features:
✓ Current Portfolio Management (add/edit/delete deals)
✓ Total Commitment / Current Commitment / Unfunded tracking
✓ NAV auto-pulled from latest available month in Fund Level CF upload
✓ Pro Forma Portfolio (placeholder/future deals)
✓ Dry Powder Forecasting & Bite Sizing
✓ Pipeline Management
✓ Return Calculator with Waterfall
✓ Pacing Model with Multi-Year Forecasts
✓ Portfolio Construction Analytics
✓ Deal Sizing Recommendations (Min/Desired/Max)

Run: python ULTIMATE_Evergreen_Fund_Dashboard.py
Open: http://localhost:8050
"""

import dash
from dash import dcc, html, dash_table, Input, Output, State, callback_context, ALL
import dash_bootstrap_components as dbc
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import pandas as pd
import numpy as np
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import json
import base64
from io import BytesIO
import pickle
import os

# ==================== DATA PERSISTENCE ====================

DATA_FILE = 'portfolio_data.pkl'


def save_data(deals, pipeline, placeholders, config):
    data = {
        'deals': deals,
        'pipeline': pipeline,
        'placeholders': placeholders,
        'config': config,
        'saved_at': datetime.now().isoformat()
    }
    try:
        with open(DATA_FILE, 'wb') as f:
            pickle.dump(data, f)
        print(f"✅ Data saved at {datetime.now().strftime('%H:%M:%S')}")
    except Exception as e:
        print(f"⚠️ Error saving: {e}")


def load_data():
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'rb') as f:
                data = pickle.load(f)
            print(f"✅ Data loaded from {data.get('saved_at', 'unknown')}")
            return data
        except Exception as e:
            print(f"⚠️ Error loading: {e}")
    return None


# ==================== COLORS ====================

C = dict(
    bg="#07090f", panel="#0d1117", surface="#111822",
    border="#1c2a3a", border2="#243649",
    blue="#2979c8", sky="#56b4f5", teal="#2ec4b6",
    green="#2ecc71", red="#e5493a", amber="#f0a500",
    purple="#9b72cf", pink="#e05c8a",
    text="#d4e6f5", muted="#5e82a0", dim="#324d63",
    mono="'JetBrains Mono', 'Fira Mono', monospace",
    sans="'IBM Plex Sans', 'Segoe UI', sans-serif",
)


def rgba(hex_color, alpha):
    r, g, b = int(hex_color[1:3], 16), int(hex_color[3:5], 16), int(hex_color[5:7], 16)
    return f"rgba({r},{g},{b},{alpha})"


CHART_BASE = dict(
    paper_bgcolor=C["panel"], plot_bgcolor=C["surface"],
    font=dict(family=C["mono"], color=C["text"], size=11),
    margin=dict(l=52, r=24, t=40, b=40),
    xaxis=dict(gridcolor=C["border"], zerolinecolor=C["border"]),
    yaxis=dict(gridcolor=C["border"], zerolinecolor=C["border"]),
    legend=dict(bgcolor=C["panel"], bordercolor=C["border"], borderwidth=1),
)

COLORS = {
    'primary': C['blue'], 'secondary': C['purple'], 'accent': C['amber'],
    'success': C['green'], 'danger': C['red'], 'dark': C['bg'],
    'light': C['panel'], 'warning': C['amber'], 'info': C['sky']
}


# ==================== HELPER: NAV FROM CASHFLOWS ====================

def get_latest_nav_from_cashflows(deal_name, fund_cf_data):
    """
    Search fund_cf_data for deal_name and return the latest non-zero NAV value.
    Returns (nav_value, month_label) or (None, None) if not found.

    fund_cf_data entries look like:
    {
      'name': 'Deal X',
      'nav': float,              # pre-computed latest NAV
      'monthly_navs': {0: val, 1: val, ...},  # month_index -> NAV
      'monthly_cfs': {0: val, ...},
      ...
    }
    """
    if not fund_cf_data:
        return None, None

    # Normalise name for fuzzy matching
    search_name = deal_name.strip().lower()

    for cf_deal in fund_cf_data:
        cf_name = cf_deal.get('name', '').strip().lower()
        # Exact or substring match
        if cf_name == search_name or search_name in cf_name or cf_name in search_name:
            # Prefer monthly_navs if present (scan from highest index downward)
            monthly_navs = cf_deal.get('monthly_navs', {})
            if monthly_navs:
                for idx in sorted(monthly_navs.keys(), reverse=True):
                    val = monthly_navs[idx]
                    if val and abs(val) > 0.001:
                        base_date = datetime(2026, 1, 1)
                        try:
                            month_label = (base_date + relativedelta(months=int(idx))).strftime('%b %Y')
                        except Exception:
                            month_label = f"Month {idx}"
                        return float(val), month_label

            # Fall back to top-level 'nav'
            nav_val = cf_deal.get('nav', 0)
            if nav_val and abs(nav_val) > 0.001:
                return float(nav_val), "Latest available"

    return None, None


# ==================== HELPER FUNCTIONS ====================

def generate_month_options():
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    options = []
    for i in range(180):
        options.append({"label": f"{month_names[i % 12]} {2026 + (i // 12)}", "value": i})
    return options


def calculate_required_future_irr(current_portfolio_irr, current_nav, dry_powder, config):
    target_twr = config['fund_parameters']['target_net_twr']
    mgmt_fee = config['fund_parameters']['management_fee']
    carry_rate = config['fund_parameters']['carry_rate']
    hurdle = config['fund_parameters']['hurdle_rate']
    loss_drag = config['fund_parameters']['loss_drag']
    cash_reserve = config['fund_parameters']['liquidity_reserve_pct']
    cash_yield = config['fund_parameters'].get('cash_yield', 0.03)

    total_fund = current_nav + dry_powder
    if total_fund == 0:
        return 0.25

    invested_ratio = max(0, 1 - cash_reserve)
    idle_ratio = cash_reserve

    gross_needed = (target_twr + mgmt_fee + loss_drag - (idle_ratio * cash_yield)) / invested_ratio

    if gross_needed > hurdle:
        carry_drag = (gross_needed - hurdle) * carry_rate
        gross_needed += carry_drag

    if dry_powder == 0:
        return gross_needed

    current_weight = current_nav / total_fund
    future_weight = dry_powder / total_fund
    required_future = (gross_needed - (current_weight * current_portfolio_irr)) / future_weight

    return max(0, min(1.0, required_future))


def calculate_portfolio_metrics(deals):
    if not deals:
        return {
            'total_nav': 0, 'total_commitment': 0, 'total_current_commitment': 0,
            'total_unfunded': 0, 'num_deals': 0, 'weighted_irr': 0,
            'by_strategy': {}, 'by_vintage': {}, 'by_sector': {},
            'concentration_top1': 0, 'concentration_top3': 0, 'concentration_top5': 0,
            'effective_n': 0
        }

    total_nav = sum(d.get('nav', d.get('size', 0)) for d in deals)
    total_commitment = sum(d.get('total_commitment', d.get('commitment', d.get('size', 0))) for d in deals)
    total_current_commitment = sum(d.get('current_commitment', d.get('commitment', d.get('size', 0))) for d in deals)
    total_unfunded = total_commitment - total_current_commitment
    num_deals = len(deals)
    weighted_irr = sum(
        d.get('nav', d.get('size', 0)) * d['target_irr'] for d in deals) / total_nav if total_nav > 0 else 0

    weights = [d.get('nav', d.get('size', 0)) / total_nav for d in deals] if total_nav > 0 else []
    effective_n = 1 / sum(w ** 2 for w in weights) if weights else 0

    by_strategy = {}
    for deal in deals:
        strategy = deal['strategy']
        if strategy not in by_strategy:
            by_strategy[strategy] = {'nav': 0, 'count': 0, 'deals': []}
        by_strategy[strategy]['nav'] += deal.get('nav', deal.get('size', 0))
        by_strategy[strategy]['count'] += 1
        by_strategy[strategy]['deals'].append(deal)

    for strategy in by_strategy:
        nav = by_strategy[strategy]['nav']
        by_strategy[strategy]['weighted_irr'] = sum(
            d.get('nav', d.get('size', 0)) * d['target_irr']
            for d in by_strategy[strategy]['deals']
        ) / nav if nav > 0 else 0
        by_strategy[strategy]['allocation'] = nav / total_nav if total_nav > 0 else 0

    by_vintage = {}
    for deal in deals:
        vintage = deal.get('vintage', 2024)
        if vintage not in by_vintage:
            by_vintage[vintage] = {'nav': 0, 'count': 0}
        by_vintage[vintage]['nav'] += deal.get('nav', deal.get('size', 0))
        by_vintage[vintage]['count'] += 1

    by_sector = {}
    for deal in deals:
        sector = deal.get('sector', 'Other')
        if sector not in by_sector:
            by_sector[sector] = {'nav': 0, 'count': 0}
        by_sector[sector]['nav'] += deal.get('nav', deal.get('size', 0))
        by_sector[sector]['count'] += 1

    sorted_deals = sorted(deals, key=lambda x: x.get('nav', x.get('size', 0)), reverse=True)
    concentration_top1 = sorted_deals[0].get('nav', sorted_deals[0].get('size', 0)) / total_nav \
        if len(sorted_deals) > 0 and total_nav > 0 else 0
    concentration_top3 = sum(d.get('nav', d.get('size', 0)) for d in sorted_deals[:3]) / total_nav \
        if len(sorted_deals) >= 3 and total_nav > 0 else 0
    concentration_top5 = sum(d.get('nav', d.get('size', 0)) for d in sorted_deals[:5]) / total_nav \
        if len(sorted_deals) >= 5 and total_nav > 0 else 0

    return {
        'total_nav': total_nav,
        'total_commitment': total_commitment,
        'total_current_commitment': total_current_commitment,
        'total_unfunded': total_unfunded,
        'num_deals': num_deals,
        'weighted_irr': weighted_irr,
        'by_strategy': by_strategy,
        'by_vintage': by_vintage,
        'by_sector': by_sector,
        'concentration_top1': concentration_top1,
        'concentration_top3': concentration_top3,
        'concentration_top5': concentration_top5,
        'effective_n': effective_n
    }


def calculate_bite_sizes(dry_powder, config):
    bite_sizes = {}
    for strategy_name in ['GP-Led (Multi-Asset)', 'GP-Led (Single-Asset)', 'Diversified LP-Led', 'Co-Investments']:
        if 'Multi-Asset' in strategy_name:
            min_pct, desired_pct, max_pct = 0.005, 0.0275, 0.05
        elif 'Single-Asset' in strategy_name:
            min_pct, desired_pct, max_pct = 0.005, 0.0225, 0.04
        elif 'LP-Led' in strategy_name:
            min_pct, desired_pct, max_pct = 0.005, 0.0275, 0.05
        else:
            min_pct, desired_pct, max_pct = 0.005, 0.0175, 0.03
        bite_sizes[strategy_name] = {
            'min': dry_powder * min_pct, 'desired': dry_powder * desired_pct, 'max': dry_powder * max_pct,
            'min_pct': min_pct, 'desired_pct': desired_pct, 'max_pct': max_pct
        }
    return bite_sizes


def forecast_dry_powder(current_nav, dry_powder, deals, placeholder_deals, config, months=12):
    forecast = []
    nav = current_nav
    powder = dry_powder
    base_date = datetime(2026, 1, 1)

    for month in range(months):
        month_date = base_date + relativedelta(months=month)
        monthly_return = config['fund_parameters']['target_net_twr'] / 12
        nav_growth = nav * monthly_return
        monthly_dist = nav * (config['fund_parameters']['distribution_rate'] / 12)
        calls_this_month = sum(pd['size'] for pd in placeholder_deals if pd.get('expected_month') == month)
        powder = powder + monthly_dist - calls_this_month
        nav = nav + nav_growth + calls_this_month - monthly_dist
        forecast.append({
            'month': month_date.strftime('%b %Y'), 'dry_powder': powder,
            'nav': nav, 'distributions': monthly_dist, 'calls': calls_this_month
        })
    return forecast


# ==================== DASH APP ====================

app = dash.Dash(__name__,
                external_stylesheets=[dbc.themes.BOOTSTRAP, dbc.icons.FONT_AWESOME],
                suppress_callback_exceptions=True)
app.title = "Horizon Portfolio Tool"
server = app.server

app.index_string = '''
<!DOCTYPE html>
<html>
    <head>
        {%metas%}
        <title>{%title%}</title>
        {%favicon%}
        {%css%}
        <style>
            @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500;600&display=swap');
            body { background-color: ''' + C['bg'] + ''' !important; color: ''' + C[
    'text'] + ''' !important; font-family: ''' + C['sans'] + ''' !important; }
            .card { background-color: ''' + C['panel'] + ''' !important; border: 1px solid ''' + C[
                       'border'] + ''' !important; color: ''' + C['text'] + ''' !important; }
            .card-header { background-color: ''' + C['surface'] + ''' !important; border-bottom: 1px solid ''' + C[
                       'border'] + ''' !important; color: ''' + C['text'] + ''' !important; }
            .card-body { background-color: ''' + C['panel'] + ''' !important; }
            .nav-pills .nav-link { color: ''' + C[
                       'muted'] + ''' !important; background-color: transparent !important; border-radius: 8px !important; font-family: ''' + \
                   C['sans'] + ''' !important; }
            .nav-pills .nav-link:hover { background-color: ''' + C['surface'] + ''' !important; color: ''' + C['text'] + ''' !important; }
            .nav-pills .nav-link.active { background: linear-gradient(135deg, ''' + C['blue'] + ''' 0%, ''' + C['sky'] + ''' 100%) !important; color: white !important; font-weight: 600 !important; }
            .navbar { background: linear-gradient(135deg, ''' + C['bg'] + ''' 0%, ''' + C[
                       'surface'] + ''' 100%) !important; border-bottom: 2px solid ''' + C['border'] + ''' !important; }
            .btn-primary { background: linear-gradient(135deg, ''' + C['blue'] + ''' 0%, ''' + C['sky'] + ''' 100%) !important; border: none !important; font-weight: 600 !important; }
            .btn-primary:hover { background: linear-gradient(135deg, ''' + C['sky'] + ''' 0%, ''' + C['teal'] + ''' 100%) !important; transform: translateY(-1px); }
            .btn-success { background: linear-gradient(135deg, ''' + C['green'] + ''' 0%, ''' + C['teal'] + ''' 100%) !important; border: none !important; }
            .btn-danger { background-color: ''' + C['red'] + ''' !important; border: none !important; }
            .btn-secondary { background-color: ''' + C['surface'] + ''' !important; border: 1px solid ''' + C[
                       'border'] + ''' !important; color: ''' + C['text'] + ''' !important; }
            .btn-warning { background-color: ''' + C['amber'] + ''' !important; border: none !important; color: #000 !important; }
            .btn-info { background: linear-gradient(135deg, ''' + C['purple'] + ''' 0%, ''' + C['pink'] + ''' 100%) !important; border: none !important; }
            .modal-content { background-color: ''' + C['panel'] + ''' !important; border: 1px solid ''' + C['border'] + ''' !important; }
            .modal-header { background-color: ''' + C['surface'] + ''' !important; border-bottom: 1px solid ''' + C[
                       'border'] + ''' !important; color: ''' + C['text'] + ''' !important; }
            .modal-body { background-color: ''' + C['panel'] + ''' !important; color: ''' + C['text'] + ''' !important; }
            .modal-footer { background-color: ''' + C['surface'] + ''' !important; border-top: 1px solid ''' + C[
                       'border'] + ''' !important; }
            .form-control, .form-select { background-color: ''' + C['surface'] + ''' !important; border: 1px solid ''' + \
                   C['border'] + ''' !important; color: ''' + C['text'] + ''' !important; font-family: ''' + C['mono'] + ''' !important; }
            .form-control:focus, .form-select:focus { background-color: ''' + C[
                       'surface'] + ''' !important; border-color: ''' + C['blue'] + ''' !important; color: ''' + C[
                       'text'] + ''' !important; box-shadow: 0 0 0 0.2rem ''' + rgba(C['blue'], 0.25) + ''' !important; }
            .form-control:disabled { background-color: ''' + C['bg'] + ''' !important; color: ''' + C['muted'] + ''' !important; }
            .form-label { color: ''' + C['text'] + ''' !important; font-weight: 600 !important; font-family: ''' + C[
                       'sans'] + ''' !important; }
            .text-muted { color: ''' + C['muted'] + ''' !important; }
            .alert-info { background-color: ''' + rgba(C['blue'], 0.1) + ''' !important; border-color: ''' + C[
                       'blue'] + ''' !important; color: ''' + C['text'] + ''' !important; }
            .alert-warning { background-color: ''' + rgba(C['amber'], 0.1) + ''' !important; border-color: ''' + C[
                       'amber'] + ''' !important; color: ''' + C['text'] + ''' !important; }
            h1,h2,h3,h4,h5,h6 { color: ''' + C['text'] + ''' !important; font-family: ''' + C['sans'] + ''' !important; }
            small { color: ''' + C['muted'] + ''' !important; }
            .dash-spreadsheet td, .dash-spreadsheet th { background-color: ''' + C[
                       'panel'] + ''' !important; color: ''' + C['text'] + ''' !important; border: 1px solid ''' + C[
                       'border'] + ''' !important; }
            .dash-spreadsheet th { background-color: ''' + C['surface'] + ''' !important; }
            .nav-tabs .nav-link { color: ''' + C['muted'] + ''' !important; background-color: transparent !important; border: none !important; border-bottom: 2px solid transparent !important; }
            .nav-tabs .nav-link.active { color: ''' + C['blue'] + ''' !important; border-bottom: 2px solid ''' + C[
                       'blue'] + ''' !important; background-color: transparent !important; }
            ::-webkit-scrollbar { width: 10px; height: 10px; }
            ::-webkit-scrollbar-track { background: ''' + C['surface'] + '''; }
            ::-webkit-scrollbar-thumb { background: ''' + C['border2'] + '''; border-radius: 5px; }
            ::-webkit-scrollbar-thumb:hover { background: ''' + C['muted'] + '''; }
            .commitment-badge { display: inline-block; padding: 2px 8px; border-radius: 4px; font-family: ''' + C[
                       'mono'] + '''; font-size: 12px; font-weight: 600; margin: 2px 0; }
            .nav-pulled { color: ''' + C['green'] + ''' !important; font-size: 10px; font-style: italic; }
            .nav-manual { color: ''' + C['amber'] + ''' !important; font-size: 10px; font-style: italic; }
            .deal-card { transition: border-color 0.2s; }
            .deal-card:hover { border-color: ''' + C['blue'] + ''' !important; }
        </style>
    </head>
    <body>
        {%app_entry%}
        <footer>
            {%config%}
            {%scripts%}
            {%renderer%}
        </footer>
    </body>
</html>
'''

DEFAULT_CONFIG = {
    'fund_parameters': {
        'dry_powder': 450,
        'target_net_twr': 0.13,
        'management_fee': 0.0125,
        'carry_rate': 0.125,
        'hurdle_rate': 0.10,
        'loss_drag': 0.01,
        'liquidity_reserve_pct': 0.05,
        'distribution_rate': 0.20,
        'cash_yield': 0.03,
        'avg_hold_period': 5.0
    },
    'bite_sizing': {'enabled': True}
}

# ==================== LAYOUT ====================

navbar = dbc.Navbar(
    dbc.Container([
        html.Div([
            html.I(className="fas fa-chart-line me-3", style={'fontSize': '32px', 'color': C['blue']}),
            dbc.NavbarBrand("Horizon Portfolio Tool", style={
                'fontSize': '24px', 'fontWeight': 'bold', 'fontFamily': C['sans'],
                'background': f'linear-gradient(135deg, {C["blue"]} 0%, {C["sky"]} 100%)',
                '-webkit-background-clip': 'text', '-webkit-text-fill-color': 'transparent'
            })
        ]),
        html.Div(id='live-clock', style={'fontSize': '14px', 'color': C['muted'], 'fontFamily': C['mono']})
    ], fluid=True),
    style={'background': f'linear-gradient(135deg, {C["bg"]} 0%, {C["surface"]} 100%)',
           'borderBottom': f'2px solid {C["border"]}', 'padding': '1.2rem 2rem'},
    dark=True, className="mb-4"
)

sidebar = dbc.Card([
    dbc.CardBody([
        html.H5("📊 Navigation", className="mb-4", style={'fontWeight': 'bold', 'color': C['text']}),
        dbc.Nav([
            dbc.NavLink([html.I(className="fas fa-home me-2"), "Dashboard"], href="/", active="exact"),
            dbc.NavLink([html.I(className="fas fa-briefcase me-2"), "Current Portfolio"], href="/portfolio",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-layer-group me-2"), "Portfolio Segments & TWR"], href="/segmentation",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-calendar-plus me-2"), "Future Deals"], href="/future",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-water me-2"), "Dry Powder Forecast"], href="/drypowder",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-tint me-2"), "Liquidity Assumptions"], href="/liquidity",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-bullseye me-2"), "Return Calculator"], href="/calculator",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-chart-area me-2"), "TWR Forecaster"], href="/twr", active="exact"),
            dbc.NavLink([html.I(className="fas fa-dollar-sign me-2"), "Deal Cashflows"], href="/cashflows",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-magic me-2"), "Pro Forma"], href="/proforma", active="exact"),
            dbc.NavLink([html.I(className="fas fa-funnel-dollar me-2"), "Pipeline"], href="/pipeline", active="exact"),
            dbc.NavLink([html.I(className="fas fa-chart-bar me-2"), "Analytics"], href="/analytics", active="exact"),
            dbc.NavLink([html.I(className="fas fa-cog me-2"), "Settings"], href="/settings", active="exact"),
        ], vertical=True, pills=True),
        html.Hr(style={'borderColor': C['border']}),
        html.H6("📈 Quick Stats", className="mb-3", style={'fontWeight': 'bold', 'color': C['text']}),
        html.Div(id='sidebar-stats'),
        html.Hr(style={'borderColor': C['border']}),
        dbc.Button("📥 Export All", id="btn-export", color="secondary", size="sm", className="w-100"),
        dcc.Download(id="download-csv"),
    ], style={'backgroundColor': C['panel']})
], style={'position': 'sticky', 'top': '20px', 'backgroundColor': C['panel'], 'border': f'1px solid {C["border"]}'})

LOADED_DATA = load_data()

app.layout = dbc.Container([
    dcc.Location(id='url', refresh=False),
    html.Div(id='save-status', style={'display': 'none'}),
    dcc.Store(id='deals-store', data=LOADED_DATA.get('deals', []) if LOADED_DATA else []),
    dcc.Store(id='placeholder-deals-store', data=LOADED_DATA.get('placeholders', []) if LOADED_DATA else []),
    dcc.Store(id='pipeline-store', data=LOADED_DATA.get('pipeline', []) if LOADED_DATA else []),
    dcc.Store(id='cashflows-store', data=[]),
    dcc.Store(id='fund-cf-data-store', data=None),
    dcc.Store(id='liquidity-data-store', data=None),
    dcc.Store(id='proforma-scenario-store', data=[]),
    dcc.Store(id='liquidity-assumptions-store', data={
        'GP-Led (Multi-Asset)': {'annual_dist_rate': 0.20, 'call_pattern': 'immediate'},
        'GP-Led (Single-Asset)': {'annual_dist_rate': 0.25, 'call_pattern': 'staged'},
        'Diversified LP-Led': {'annual_dist_rate': 0.18, 'call_pattern': 'staged'},
        'Co-Investments': {'annual_dist_rate': 0.15, 'call_pattern': 'delayed'},
    }),
    dcc.Store(id='config-store', data=LOADED_DATA.get('config', DEFAULT_CONFIG) if LOADED_DATA else DEFAULT_CONFIG),
    dcc.Interval(id='clock-interval', interval=1000),
    navbar,
    dbc.Row([
        dbc.Col(sidebar, width=2),
        dbc.Col(html.Div(id="page-content"), width=10)
    ])
], fluid=True, style={'backgroundColor': C['bg'], 'minHeight': '100vh'})


# ==================== PAGE LAYOUTS ====================

def dashboard_page():
    return html.Div([
        html.H2("📊 Fund Dashboard", className="mb-4",
                style={'fontWeight': 'bold', 'fontFamily': C['sans'], 'color': C['text']}),
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("💰 Current NAV", className="text-muted mb-2"),
                html.H3(id="dash-nav", style={'color': C['green'], 'fontWeight': 'bold', 'fontFamily': C['mono']}),
                html.Small(id="dash-num-deals", className="text-muted")
            ])], style={'background': f'linear-gradient(135deg, {rgba(C["green"], 0.1)} 0%, {C["panel"]} 100%)'}),
                width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("💵 Dry Powder", className="text-muted mb-2"),
                html.H3(id="dash-dry-powder",
                        style={'color': C['blue'], 'fontWeight': 'bold', 'fontFamily': C['mono']}),
                html.Small("Available", className="text-muted")
            ])], style={'background': f'linear-gradient(135deg, {rgba(C["blue"], 0.1)} 0%, {C["panel"]} 100%)'}),
                width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("📊 Total Fund", className="text-muted mb-2"),
                html.H3(id="dash-total", style={'color': C['sky'], 'fontWeight': 'bold', 'fontFamily': C['mono']}),
                html.Small("NAV + Powder", className="text-muted")
            ])], style={'background': f'linear-gradient(135deg, {rgba(C["sky"], 0.1)} 0%, {C["panel"]} 100%)'}),
                width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("🎯 Portfolio IRR", className="text-muted mb-2"),
                html.H3(id="dash-current-irr",
                        style={'color': C['teal'], 'fontWeight': 'bold', 'fontFamily': C['mono']}),
                html.Small("Weighted", className="text-muted")
            ])], style={'background': f'linear-gradient(135deg, {rgba(C["teal"], 0.1)} 0%, {C["panel"]} 100%)'}),
                width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("🚀 Required IRR", className="text-muted mb-2"),
                html.H3(id="dash-req-irr", style={'color': C['amber'], 'fontWeight': 'bold', 'fontFamily': C['mono']}),
                html.Small("Future Deals", className="text-muted")
            ])], style={'background': f'linear-gradient(135deg, {rgba(C["amber"], 0.1)} 0%, {C["panel"]} 100%)'}),
                width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("📅 Placeholders", className="text-muted mb-2"),
                html.H3(id="dash-placeholders",
                        style={'color': C['purple'], 'fontWeight': 'bold', 'fontFamily': C['mono']}),
                html.Small(id="dash-placeholder-value", className="text-muted")
            ])], style={'background': f'linear-gradient(135deg, {rgba(C["purple"], 0.1)} 0%, {C["panel"]} 100%)'}),
                width=2),
        ], className="mb-4"),
        dbc.Row([
            dbc.Col([dbc.Card([
                dbc.CardHeader("Portfolio Allocation", style={'fontWeight': 'bold'}),
                dbc.CardBody([dcc.Graph(id='dash-allocation-chart', config={'displayModeBar': False})])
            ])], width=4),
            dbc.Col([dbc.Card([
                dbc.CardHeader("Dry Powder Forecast (12 Months)", style={'fontWeight': 'bold'}),
                dbc.CardBody([dcc.Graph(id='dash-forecast-chart', config={'displayModeBar': False})])
            ])], width=4),
            dbc.Col([dbc.Card([
                dbc.CardHeader("Deal Bite Sizing Guide", style={'fontWeight': 'bold'}),
                dbc.CardBody([html.Div(id='dash-bite-sizing')])
            ])], width=4),
        ], className="mb-4"),
        dbc.Row([
            dbc.Col([dbc.Card([
                dbc.CardHeader("Recent Portfolio Activity", style={'fontWeight': 'bold'}),
                dbc.CardBody([html.Div(id='dash-recent-activity')])
            ])], width=6),
            dbc.Col([dbc.Card([
                dbc.CardHeader("Pipeline Status", style={'fontWeight': 'bold'}),
                dbc.CardBody([html.Div(id='dash-pipeline-status')])
            ])], width=6),
        ])
    ])


def portfolio_page():
    return html.Div([
        dbc.Row([
            dbc.Col(html.H2("💼 Current Portfolio", style={'fontWeight': 'bold'}), width=8),
            dbc.Col(dbc.Button("➕ Add Deal", id="btn-open-deal", color="primary", className="float-end", size="lg"),
                    width=4)
        ], className="mb-3"),

        # Enhanced Fund Status Cards — now shows commitment breakdown
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("💰 Current NAV", className="text-muted"),
                html.H4(id="port-nav", style={'color': C['green'], 'fontFamily': C['mono']}),
                html.Small("From CF upload or manual entry", className="text-muted")
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("📋 Total Commitment", className="text-muted"),
                html.H4(id="port-total-commitment", style={'color': C['blue'], 'fontFamily': C['mono']}),
                html.Small("Agreed at signing", className="text-muted")
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("💸 Current Commitment", className="text-muted"),
                html.H4(id="port-current-commitment", style={'color': C['purple'], 'fontFamily': C['mono']}),
                html.Small("Capital called to date", className="text-muted")
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("⚠️ Unfunded", className="text-muted"),
                html.H4(id="port-unfunded", style={'color': C['amber'], 'fontFamily': C['mono']}),
                html.Small("Total – Current Commitment", className="text-muted")
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("💵 Dry Powder", className="text-muted"),
                html.H4(id="port-dry-powder", style={'color': C['sky'], 'fontFamily': C['mono']}),
                html.Small("Available to deploy", className="text-muted")
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("🎯 Required IRR", className="text-muted"),
                html.H4(id="port-target-return", style={'color': C['red'], 'fontFamily': C['mono']}),
                html.Small("Future deal target", className="text-muted")
            ])], className="shadow-sm"), width=2),
        ], className="mb-4"),

        # CF data indicator banner
        html.Div(id='port-cf-indicator'),

        # Add Deal Modal
        dbc.Modal([
            dbc.ModalHeader(dbc.ModalTitle("➕ Add Deal to Current Portfolio")),
            dbc.ModalBody([
                dbc.Alert([
                    html.I(className="fas fa-info-circle me-2"),
                    html.Strong("Current Portfolio: "),
                    "Enter actual deal data. NAV is auto-populated from the Fund Level CF upload if a match is found."
                ], color="info", className="mb-3"),

                dbc.Row([
                    dbc.Col([dbc.Label("Deal Name *"),
                             dbc.Input(id="in-name", type="text", placeholder="e.g., Coller VII")], width=6),
                    dbc.Col([dbc.Label("Fund Manager *"),
                             dbc.Input(id="in-manager", type="text", placeholder="e.g., Coller Capital")], width=6)
                ], className="mb-3"),
                dbc.Row([
                    dbc.Col([dbc.Label("Strategy Type *"),
                             dbc.Select(id="in-strategy", options=[
                                 {"label": "GP-Led (Multi-Asset)", "value": "GP-Led (Multi-Asset)"},
                                 {"label": "GP-Led (Single-Asset)", "value": "GP-Led (Single-Asset)"},
                                 {"label": "Diversified LP-Led", "value": "Diversified LP-Led"},
                                 {"label": "Co-Investments", "value": "Co-Investments"},
                             ])], width=6),
                    dbc.Col([dbc.Label("Stage *"),
                             dbc.Select(id="in-stage", options=[
                                 {"label": "Buyout", "value": "Buyout"}, {"label": "Venture", "value": "Venture"},
                                 {"label": "Growth", "value": "Growth"}, {"label": "Liquidity", "value": "Liquidity"},
                             ], value="Buyout")], width=6)
                ], className="mb-3"),

                # Commitment block
                html.Div([
                    html.H6("📋 Commitment Details",
                            style={'color': C['sky'], 'fontFamily': C['sans'],
                                   'borderBottom': f'1px solid {C["border"]}',
                                   'paddingBottom': '6px', 'marginBottom': '12px'}),
                    dbc.Row([
                        dbc.Col([
                            dbc.Label("Total Commitment ($mm) *"),
                            dbc.Input(id="in-total-commitment", type="number", step=0.1, placeholder="70.0"),
                            html.Small("Full amount agreed at signing", className="text-muted")
                        ], width=4),
                        dbc.Col([
                            dbc.Label("Current Commitment ($mm) *"),
                            dbc.Input(id="in-current-commitment", type="number", step=0.1, placeholder="52.0"),
                            html.Small("Capital called/drawn to date", className="text-muted")
                        ], width=4),
                        dbc.Col([
                            dbc.Label("Unfunded ($mm)"),
                            html.Div(id="unfunded-display", style={
                                'marginTop': '8px', 'padding': '10px',
                                'backgroundColor': rgba(C['amber'], 0.12),
                                'border': f'1px solid {C["amber"]}',
                                'borderRadius': '6px', 'fontFamily': C['mono'],
                                'fontSize': '16px', 'fontWeight': 'bold', 'color': C['amber']
                            }),
                            html.Small("Auto: Total – Current", className="text-muted")
                        ], width=4),
                    ], className="mb-3"),
                ], style={'backgroundColor': rgba(C['sky'], 0.04), 'border': f'1px solid {C["border"]}',
                          'borderRadius': '8px', 'padding': '12px', 'marginBottom': '12px'}),

                # NAV block
                html.Div([
                    html.H6("📈 Current NAV",
                            style={'color': C['green'], 'fontFamily': C['sans'],
                                   'borderBottom': f'1px solid {C["border"]}',
                                   'paddingBottom': '6px', 'marginBottom': '12px'}),
                    dbc.Row([
                        dbc.Col([
                            dbc.Label("Current NAV ($mm) *"),
                            dbc.Input(id="in-nav", type="number", step=0.1, placeholder="48.0"),
                            html.Small(id="in-nav-cf-hint",
                                       children="Enter manually or load CF file for auto-population",
                                       className="text-muted")
                        ], width=6),
                        dbc.Col([
                            dbc.Label("NAV Source"),
                            html.Div(id="in-nav-source-display", style={
                                'marginTop': '8px', 'padding': '10px',
                                'backgroundColor': rgba(C['green'], 0.08),
                                'border': f'1px solid {C["border"]}',
                                'borderRadius': '6px', 'fontFamily': C['mono'],
                                'fontSize': '13px', 'color': C['muted']
                            }, children="Manual entry")
                        ], width=6),
                    ]),
                ], style={'backgroundColor': rgba(C['green'], 0.04), 'border': f'1px solid {C["border"]}',
                          'borderRadius': '8px', 'padding': '12px', 'marginBottom': '12px'}),

                dbc.Row([
                    dbc.Col([dbc.Label("Target Gross IRR (%) *"),
                             dbc.Input(id="in-irr", type="number", step=0.5, placeholder="22.0")], width=4),
                    dbc.Col(
                        [dbc.Label("Hold Period (years)"), dbc.Input(id="in-hold", type="number", value=5, step=0.5)],
                        width=4),
                    dbc.Col([dbc.Label("Currency"),
                             dbc.Select(id="in-currency", options=[
                                 {"label": "USD", "value": "USD"}, {"label": "EUR", "value": "EUR"},
                                 {"label": "GBP", "value": "GBP"},
                             ], value="USD")], width=4),
                ], className="mb-3"),
                dbc.Row([
                    dbc.Col([dbc.Label("Vintage Year"),
                             dbc.Select(id="in-vintage",
                                        options=[{"label": str(y), "value": y} for y in range(2026, 2014, -1)],
                                        value=2024)], width=4),
                    dbc.Col([dbc.Label("Portfolio Segment *"),
                             dbc.Select(id="in-segment", options=[
                                 {"label": "Seed Portfolio", "value": "Seed"},
                                 {"label": "New Deals", "value": "New"},
                                 {"label": "Money Market", "value": "MoneyMarket"},
                             ], value="Seed"),
                             html.Small("For TWR tracking", className="text-muted")], width=4),
                    dbc.Col([dbc.Label("Allocation Status *"),
                             dbc.Select(id="in-allocation-status", options=[
                                 {"label": "Closed", "value": "Closed"},
                                 {"label": "Pending Close", "value": "Pending Close"},
                                 {"label": "Pending Allocation", "value": "Pending Allocation"},
                             ], value="Closed")], width=4),
                ], className="mb-3"),
                dbc.Row([
                    dbc.Col([dbc.Label("Sector"),
                             dbc.Select(id="in-sector", options=[
                                 {"label": "Diversified", "value": "Diversified"},
                                 {"label": "Technology", "value": "Technology"},
                                 {"label": "Healthcare", "value": "Healthcare"},
                                 {"label": "Consumer", "value": "Consumer"},
                                 {"label": "Industrials", "value": "Industrials"},
                                 {"label": "Financials", "value": "Financials"},
                                 {"label": "Other", "value": "Other"}
                             ], value="Diversified")], width=6),
                    dbc.Col([dbc.Label("Geography"),
                             dbc.Select(id="in-geo", options=[
                                 {"label": "North America", "value": "North America"},
                                 {"label": "Europe", "value": "Europe"},
                                 {"label": "Asia", "value": "Asia"},
                                 {"label": "Global", "value": "Global"}
                             ], value="North America")], width=6),
                ])
            ]),
            dbc.ModalFooter([
                dbc.Button("Cancel", id="btn-cancel", color="secondary"),
                dbc.Button("✅ Add Deal", id="btn-submit", color="primary")
            ])
        ], id="modal-deal", size="lg", is_open=False),

        # Edit Deal Modal — ENHANCED with CF NAV pull
        dbc.Modal([
            dbc.ModalHeader(dbc.ModalTitle("✏️ Edit Deal")),
            dbc.ModalBody([
                dcc.Store(id='edit-deal-index', data=None),

                # CF NAV pull banner
                html.Div(id="edit-cf-pull-banner", className="mb-3"),

                dbc.Row([
                    dbc.Col([dbc.Label("Deal Name *"), dbc.Input(id="edit-name", type="text")], width=6),
                    dbc.Col([dbc.Label("Fund Manager *"), dbc.Input(id="edit-manager", type="text")], width=6)
                ], className="mb-3"),
                dbc.Row([
                    dbc.Col([dbc.Label("Strategy Type *"),
                             dbc.Select(id="edit-strategy", options=[
                                 {"label": "GP-Led (Multi-Asset)", "value": "GP-Led (Multi-Asset)"},
                                 {"label": "GP-Led (Single-Asset)", "value": "GP-Led (Single-Asset)"},
                                 {"label": "Diversified LP-Led", "value": "Diversified LP-Led"},
                                 {"label": "Co-Investments", "value": "Co-Investments"},
                             ])], width=6),
                    dbc.Col([dbc.Label("Stage *"),
                             dbc.Select(id="edit-stage", options=[
                                 {"label": "Buyout", "value": "Buyout"}, {"label": "Venture", "value": "Venture"},
                                 {"label": "Growth", "value": "Growth"}, {"label": "Liquidity", "value": "Liquidity"},
                             ])], width=6)
                ], className="mb-3"),

                # Commitment block (edit)
                html.Div([
                    html.H6("📋 Commitment Details",
                            style={'color': C['sky'], 'fontFamily': C['sans'],
                                   'borderBottom': f'1px solid {C["border"]}', 'paddingBottom': '6px',
                                   'marginBottom': '12px'}),
                    dbc.Row([
                        dbc.Col([
                            dbc.Label("Total Commitment ($mm) *"),
                            dbc.Input(id="edit-total-commitment", type="number", step=0.1),
                            html.Small("Full amount agreed at signing", className="text-muted")
                        ], width=4),
                        dbc.Col([
                            dbc.Label("Current Commitment ($mm) *"),
                            dbc.Input(id="edit-current-commitment", type="number", step=0.1),
                            html.Small("Capital called to date", className="text-muted")
                        ], width=4),
                        dbc.Col([
                            dbc.Label("Unfunded ($mm)"),
                            html.Div(id="edit-unfunded-display", style={
                                'marginTop': '8px', 'padding': '10px',
                                'backgroundColor': rgba(C['amber'], 0.12),
                                'border': f'1px solid {C["amber"]}',
                                'borderRadius': '6px', 'fontFamily': C['mono'],
                                'fontSize': '16px', 'fontWeight': 'bold', 'color': C['amber']
                            }),
                            html.Small("Auto: Total – Current", className="text-muted")
                        ], width=4),
                    ]),
                ], style={'backgroundColor': rgba(C['sky'], 0.04), 'border': f'1px solid {C["border"]}',
                          'borderRadius': '8px', 'padding': '12px', 'marginBottom': '12px'}),

                # NAV block (edit) — always editable, shows pull source
                html.Div([
                    html.H6("📈 Current NAV",
                            style={'color': C['green'], 'fontFamily': C['sans'],
                                   'borderBottom': f'1px solid {C["border"]}', 'paddingBottom': '6px',
                                   'marginBottom': '12px'}),
                    dbc.Row([
                        dbc.Col([
                            dbc.Label("Current NAV ($mm) *"),
                            dbc.Input(id="edit-nav", type="number", step=0.1),
                            html.Small(id="edit-nav-source",
                                       children="Enter or let the system pull from CF data",
                                       className="text-muted")
                        ], width=6),
                        dbc.Col([
                            dbc.Label("Latest CF Month"),
                            html.Div(id="edit-nav-month-display", style={
                                'marginTop': '8px', 'padding': '10px',
                                'backgroundColor': rgba(C['green'], 0.08),
                                'border': f'1px solid {C["border"]}',
                                'borderRadius': '6px', 'fontFamily': C['mono'],
                                'fontSize': '13px', 'color': C['muted']
                            }, children="—")
                        ], width=6),
                    ]),
                ], style={'backgroundColor': rgba(C['green'], 0.04), 'border': f'1px solid {C["border"]}',
                          'borderRadius': '8px', 'padding': '12px', 'marginBottom': '12px'}),

                dbc.Row([
                    dbc.Col([dbc.Label("Target Gross IRR (%) *"), dbc.Input(id="edit-irr", type="number", step=0.5)],
                            width=6),
                    dbc.Col([dbc.Label("Hold Period (years)"), dbc.Input(id="edit-hold", type="number", step=0.5)],
                            width=6),
                ]),
            ]),
            dbc.ModalFooter([
                dbc.Button("Cancel", id="btn-cancel-edit", color="secondary"),
                dbc.Button("💾 Save Changes", id="btn-save-edit", color="warning")
            ])
        ], id="modal-edit-deal", size="lg", is_open=False),

        html.Div(id='portfolio-table')
    ])


def future_deals_page():
    return html.Div([
        dbc.Row([
            dbc.Col(html.H2("📅 Pro Forma Portfolio (Future/Placeholder Deals)", style={'fontWeight': 'bold'}), width=8),
            dbc.Col(dbc.Button("➕ Add Placeholder", id="btn-open-placeholder", color="success", className="float-end",
                               size="lg"), width=4)
        ], className="mb-3"),
        dbc.Alert([html.I(className="fas fa-lightbulb me-2"),
                   "Plan future deals to forecast dry powder usage. These are NOT in the current portfolio."],
                  color="info", className="mb-4"),
        dbc.Modal([
            dbc.ModalHeader("Add Placeholder Deal (Future Commitment)"),
            dbc.ModalBody([
                dbc.Row([
                    dbc.Col([dbc.Label("Deal Name *"),
                             dbc.Input(id="in-ph-name", type="text", placeholder="GP-Led (Multi-Asset) 1")], width=6),
                    dbc.Col([dbc.Label("Strategy Type *"),
                             dbc.Select(id="in-ph-strategy", options=[
                                 {"label": "GP-Led (Multi-Asset)", "value": "GP-Led (Multi-Asset)"},
                                 {"label": "GP-Led (Single-Asset)", "value": "GP-Led (Single-Asset)"},
                                 {"label": "Diversified LP-Led", "value": "Diversified LP-Led"},
                                 {"label": "Co-Investments", "value": "Co-Investments"},
                             ])], width=6)
                ], className="mb-3"),
                dbc.Row([
                    dbc.Col([dbc.Label("Expected Size ($mm) *"),
                             dbc.Input(id="in-ph-size", type="number", step=0.1, placeholder="15.0"),
                             html.Small(id="ph-bite-warning", className="text-muted")], width=4),
                    dbc.Col([dbc.Label("Custom Bite Size ($mm)"),
                             dbc.Input(id="in-ph-bite", type="number", step=0.1, placeholder="Auto"),
                             html.Small("Override default bite size", className="text-muted")], width=4),
                    dbc.Col([dbc.Label("Expected Month *"),
                             dbc.Select(id="in-ph-month", options=generate_month_options(), value=0,
                                        style={'fontFamily': C['mono']})], width=4),
                ], className="mb-3"),
                dbc.Row([
                    dbc.Col([dbc.Label("Target Weight (% of Portfolio)"),
                             dbc.Input(id="in-ph-weight", type="number", step=0.5, placeholder="5.0")], width=4),
                    dbc.Col([dbc.Label("Region"),
                             dbc.Select(id="in-ph-region", options=[
                                 {"label": "North America", "value": "North America"},
                                 {"label": "Europe", "value": "Europe"},
                                 {"label": "Asia", "value": "Asia"}, {"label": "Global", "value": "Global"},
                             ], value="North America")], width=4),
                    dbc.Col([dbc.Label("Deal Type"),
                             dbc.Select(id="in-ph-type", options=[
                                 {"label": "Secondary", "value": "Secondary"},
                                 {"label": "Co-Investment", "value": "Co-Investment"},
                             ], value="Secondary")], width=4),
                ])
            ]),
            dbc.ModalFooter([
                dbc.Button("Cancel", id="btn-cancel-ph", color="secondary"),
                dbc.Button("Add Placeholder", id="btn-submit-ph", color="success")
            ])
        ], id="modal-placeholder", size="lg", is_open=False),
        html.Div(id='placeholder-table')
    ])


def drypowder_page():
    return html.Div([
        html.H2("💧 Dry Powder Forecaster", className="mb-4", style={'fontWeight': 'bold'}),
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([html.H6("Current Dry Powder", className="text-muted"),
                                            html.H3(id="dp-current", style={'color': COLORS['primary']})])]), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([html.H6("Planned Deployments", className="text-muted"),
                                            html.H3(id="dp-planned", style={'color': COLORS['warning']})])]), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([html.H6("Expected Distributions", className="text-muted"),
                                            html.H3(id="dp-distributions", style={'color': COLORS['success']})])]),
                    width=3),
            dbc.Col(dbc.Card([dbc.CardBody([html.H6("12-Month Forecast", className="text-muted"),
                                            html.H3(id="dp-forecast-12m", style={'color': COLORS['info']})])]),
                    width=3),
        ], className="mb-4"),
        dbc.Card([
            dbc.CardHeader("12-Month Dry Powder Forecast", style={'fontWeight': 'bold'}),
            dbc.CardBody([dcc.Graph(id='dp-forecast-chart', config={'displayModeBar': True})])
        ], className="mb-4"),
        dbc.Card([
            dbc.CardHeader("Monthly Breakdown", style={'fontWeight': 'bold'}),
            dbc.CardBody([html.Div(id='dp-monthly-table')])
        ])
    ])


def calculator_page():
    return html.Div([
        html.H2("🧮 Return Calculator & Waterfall", className="mb-4", style={'fontWeight': 'bold'}),
        dbc.Card([dbc.CardBody([
            html.H5("Calculate required deal IRR to achieve target net TWR", className="text-muted mb-4"),
            html.Hr(),
            dbc.Row([
                dbc.Col([html.H6("Current NAV:", className="text-muted"),
                         html.H3(id="calc-current-nav", style={'color': COLORS['success']})], width=4),
                dbc.Col([html.H6("Weighted IRR:", className="text-muted"),
                         html.H3(id="calc-current-irr", style={'color': COLORS['success']})], width=4),
                dbc.Col([html.H6("Number of Deals:", className="text-muted"),
                         html.H3(id="calc-num-deals", style={'color': COLORS['success']})], width=4),
            ], className="mb-4"),
            html.Hr(),
            dbc.Row([
                dbc.Col([html.H6("Total Fund Size:", className="text-muted"),
                         html.H3(id="calc-total-fund", style={'color': COLORS['primary']}),
                         html.Small("NAV + Dry Powder")], width=4),
                dbc.Col([html.H6("Target Net TWR:", className="text-muted"),
                         html.H3(id="calc-target-twr", style={'color': COLORS['warning']})], width=4),
                dbc.Col([html.H6("Dry Powder:", className="text-muted"),
                         html.H3(id="calc-dry-powder", style={'color': COLORS['info']}),
                         html.Small("For Pipeline/Future Only")], width=4),
            ], className="mb-4"),
            html.Hr(),
            html.Div([
                html.H3("✅ REQUIRED FUTURE DEAL IRR", className="text-center mb-3",
                        style={'color': COLORS['success'], 'fontWeight': 'bold'}),
                html.H1(id="calc-required-irr", className="text-center mb-2",
                        style={'fontSize': '80px', 'fontWeight': 'bold', 'color': COLORS['success']}),
                html.P(id="calc-explanation", className="text-center text-muted mb-4", style={'fontSize': '18px'}),
            ], style={'backgroundColor': rgba(C['green'], 0.08), 'padding': '2rem', 'borderRadius': '10px',
                      'border': f'2px solid {C["green"]}'}),
            html.Hr(className="mt-4"),
            html.H5("💡 Return Waterfall Breakdown", className="mb-3", style={'fontWeight': 'bold'}),
            html.Div(id="calc-waterfall")
        ])])
    ])


def twr_forecaster_page():
    return html.Div([
        html.H2("📈 TWR Returns Forecaster", className="mb-4", style={'fontWeight': 'bold'}),
        dbc.Row([dbc.Col([dbc.Card([
            dbc.CardHeader("Monte Carlo Simulation Parameters", style={'fontWeight': 'bold'}),
            dbc.CardBody([
                dbc.Row([
                    dbc.Col([dbc.Label("Current Portfolio IRR"),
                             html.H4(id="twr-current-irr", style={'color': C['green']})], width=4),
                    dbc.Col([dbc.Label("Future Deals Mean IRR (%)"),
                             dbc.Input(id="twr-future-mean", type="number", value=25, step=0.5)], width=4),
                    dbc.Col([dbc.Label("Future Deals Std Dev (%)"),
                             dbc.Input(id="twr-future-std", type="number", value=5, step=0.5)], width=4),
                ], className="mb-3"),
                dbc.Row([
                    dbc.Col([dbc.Label("# Simulations"),
                             dbc.Select(id="twr-n-sims", options=[
                                 {"label": "1,000", "value": 1000}, {"label": "5,000", "value": 5000},
                                 {"label": "10,000", "value": 10000},
                             ], value=5000)], width=6),
                    dbc.Col([dbc.Button("🎲 Run Simulation", id="btn-run-twr", color="primary", size="lg",
                                        className="w-100")], width=6)
                ])
            ])
        ])], width=12)]),
        dbc.Row([
            dbc.Col([dbc.Card(
                [dbc.CardHeader("Probability Distribution"), dbc.CardBody([dcc.Graph(id='twr-distribution-chart')])])],
                    width=8),
            dbc.Col([dbc.Card([dbc.CardHeader("Key Statistics"), dbc.CardBody([html.Div(id='twr-statistics')])])],
                    width=4),
        ], className="mb-4"),
        dbc.Row([dbc.Col([dbc.Card(
            [dbc.CardHeader("Sensitivity Analysis"), dbc.CardBody([dcc.Graph(id='twr-sensitivity-chart')])])])])
    ])


def cashflows_page():
    return html.Div([
        html.H2("💰 Fund Level Cashflows", className="mb-4", style={'fontWeight': 'bold'}),
        dbc.Alert([html.I(className="fas fa-info-circle me-2"),
                   "Upload your Excel Fund Level CF file. NAV values are auto-pulled into the Current Portfolio tab."],
                  color="info", className="mb-4"),
        dbc.Card([
            dbc.CardHeader("📥 Upload Fund Level CF Excel File", style={'fontWeight': 'bold'}),
            dbc.CardBody([dbc.Row([
                dbc.Col([
                    dcc.Upload(id='upload-fund-cf',
                               children=dbc.Button(
                                   [html.I(className="fas fa-file-excel me-2"), 'Upload Fund Level CF Excel File'],
                                   color="primary", size="lg", className="w-100"), multiple=False),
                    html.Small("Upload the 'Fund Level CF' tab (.xlsx, .xlsm)", className="text-muted d-block mt-2")
                ], width=6),
                dbc.Col([html.Div(id='upload-fund-cf-status', className="mt-2")], width=6)
            ])])
        ], className="mb-4"),
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody(
                [html.H6("Total Commitment"), html.H4(id="cf-total-commitment", style={'color': C['blue']})])]),
                    width=3),
            dbc.Col(dbc.Card([dbc.CardBody(
                [html.H6("Total Paid In"), html.H4(id="cf-total-paid-in", style={'color': C['green']})])]), width=3),
            dbc.Col(dbc.Card([dbc.CardBody(
                [html.H6("Total Unfunded"), html.H4(id="cf-total-unfunded", style={'color': C['amber']})])]), width=3),
            dbc.Col(dbc.Card([dbc.CardBody(
                [html.H6("Current Total NAV"), html.H4(id="cf-current-nav", style={'color': C['purple']})])]), width=3),
        ], className="mb-4"),
        dbc.Row([dbc.Col([dbc.Card([dbc.CardBody([dbc.Row([
            dbc.Col([dbc.Label("View Period"), dbc.Select(id="cf-view-period", options=[
                {"label": "Next 12 Months", "value": 12}, {"label": "Next 24 Months", "value": 24},
                {"label": "Next 36 Months", "value": 36}, {"label": "5 Years", "value": 60},
            ], value=12)], width=4),
            dbc.Col([dbc.Label("Cashflow Type"), dbc.Select(id="cf-type-view", options=[
                {"label": "All (Calls + Dists + NAV)", "value": "all"}, {"label": "Net Cashflows Only", "value": "net"},
                {"label": "NAV Only", "value": "nav"},
            ], value="all")], width=4),
            dbc.Col([dbc.Button("📥 Export to CSV", id="btn-export-cf", color="success", className="w-100 mt-4")],
                    width=4)
        ])])])])], className="mb-4"),
        dbc.Card([dbc.CardHeader("Fund Level Cashflows (Deal-by-Deal)"),
                  dbc.CardBody([html.Div(id='cf-monthly-table', style={'overflowX': 'auto'})])], className="mb-4"),
        dbc.Card([dbc.CardHeader("Monthly Cashflow Totals"), dbc.CardBody([dcc.Graph(id='cf-monthly-chart')])]),
    ])


def proforma_page():
    return html.Div([
        html.H2("🔮 Pro Forma Analyzer", className="mb-4", style={'fontWeight': 'bold'}),
        dbc.Row([dbc.Col([dbc.Card([
            dbc.CardHeader("Pro Forma Settings", style={'fontWeight': 'bold'}),
            dbc.CardBody([dbc.Row([
                dbc.Col([dbc.Label("Target Month for NAV Calculation *"),
                         dbc.Select(id="pf-target-month", options=generate_month_options(), value=12),
                         html.Small("Calculate pro forma NAV at this month", className="text-muted")], width=6),
                dbc.Col([dbc.Button("📊 Calculate Pro Forma NAV", id="btn-calc-pf-nav", color="primary", size="lg",
                                    className="w-100 mt-4")], width=6)
            ])])
        ], className="mb-4")])]),
        dbc.Tabs([
            dbc.Tab(label="📋 Complete Portfolio", tab_id="tab-pf-portfolio"),
            dbc.Tab(label="📊 Metrics Comparison", tab_id="tab-pf-metrics"),
            dbc.Tab(label="📈 Impact Charts", tab_id="tab-pf-charts"),
        ], id="pf-tabs", active_tab="tab-pf-portfolio"),
        html.Div(id='pf-tab-content', className="mt-4")
    ])


def liquidity_assumptions_page():
    return html.Div([
        html.H2("💧 Liquidity Pull", className="mb-4", style={'fontWeight': 'bold'}),
        dbc.Alert([html.I(className="fas fa-database me-2"), html.Strong("Data Pull From Liquidity Model — "),
                   "Upload your Excel Liquidity Pull file or view current liquidity position."], color="info",
                  className="mb-4"),
        dbc.Card([
            dbc.CardHeader("📥 Upload Liquidity Pull Excel File", style={'fontWeight': 'bold'}),
            dbc.CardBody([dbc.Row([
                dbc.Col([dcc.Upload(id='upload-liquidity-pull',
                                    children=dbc.Button([html.I(className="fas fa-file-excel me-2"),
                                                         'Upload Liquidity Pull Excel File'],
                                                        color="info", size="lg", className="w-100"), multiple=False),
                         html.Small("Upload the 'Liquidity Pull' tab (.xlsx, .xlsm)",
                                    className="text-muted d-block mt-2")], width=6),
                dbc.Col([html.Div(id='upload-liquidity-status', className="mt-2")], width=6)
            ])])
        ], className="mb-4"),
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([html.H6("Today's Date"), html.H5(id="liq-today-date")])]), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([html.H6("As At"), html.H5(id="liq-as-at-date")])]), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([html.H6("Current Quarter"), html.H5(id="liq-current-quarter")])]), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([html.H6("Current Month"), html.H5(id="liq-current-month")])]), width=3),
        ], className="mb-4"),
        dbc.Card([dbc.CardHeader("Liquidity Waterfall"), dbc.CardBody([html.Div(id='liq-waterfall-table')])],
                 className="mb-4"),
        dbc.Card(
            [dbc.CardHeader("Near Term Flows (Next 12 Months)"), dbc.CardBody([html.Div(id='liq-near-term-flows')])],
            className="mb-4"),
        dbc.Card([dbc.CardHeader("NAV End Projections"), dbc.CardBody([dcc.Graph(id='liq-nav-projection-chart')])],
                 className="mb-4"),
        dbc.Tabs([
            dbc.Tab(label="Subscriptions & Redemptions", tab_id="tab-liq-subs"),
            dbc.Tab(label="Portfolio Net Flows", tab_id="tab-liq-flows"),
            dbc.Tab(label="Unfunded Commitments", tab_id="tab-liq-unfunded"),
        ], id="liq-tabs", active_tab="tab-liq-subs"),
        html.Div(id='liq-tab-content', className="mt-4"),
    ])


def segmentation_page():
    return html.Div([
        html.H2("📊 Portfolio Segmentation & TWR Analysis", className="mb-4", style={'fontWeight': 'bold'}),
        dbc.Alert([html.I(className="fas fa-layer-group me-2"),
                   "Track TWR performance across portfolio segments: Seed Portfolio, New Deals, Money Market, Pipeline, and Future Deals"],
                  color="info", className="mb-4"),
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody(
                [html.H6("Total Portfolio"), html.H4(id="seg-total-nav", style={'color': C['blue']}),
                 html.Small(id="seg-total-twr")])]), width=2),
            dbc.Col(dbc.Card([dbc.CardBody(
                [html.H6("Seed Portfolio"), html.H4(id="seg-seed-nav", style={'color': C['green']}),
                 html.Small(id="seg-seed-twr")])]), width=2),
            dbc.Col(dbc.Card([dbc.CardBody(
                [html.H6("New Deals"), html.H4(id="seg-new-nav", style={'color': C['purple']}),
                 html.Small(id="seg-new-twr")])]), width=2),
            dbc.Col(dbc.Card([dbc.CardBody(
                [html.H6("Money Market"), html.H4(id="seg-mm-nav", style={'color': C['amber']}),
                 html.Small(id="seg-mm-twr")])]), width=2),
            dbc.Col(dbc.Card([dbc.CardBody(
                [html.H6("Pipeline"), html.H4(id="seg-pipeline-nav", style={'color': C['teal']}),
                 html.Small(id="seg-pipeline-twr")])]), width=2),
            dbc.Col(dbc.Card([dbc.CardBody(
                [html.H6("Future/Placeholder"), html.H4(id="seg-future-nav", style={'color': C['pink']}),
                 html.Small(id="seg-future-twr")])]), width=2),
        ], className="mb-4"),
        dbc.Card([dbc.CardHeader("TWR Forecast by Segment"), dbc.CardBody([dcc.Graph(id='seg-twr-forecast-chart')])],
                 className="mb-4"),
        dbc.Row([
            dbc.Col([dbc.Card([dbc.CardHeader("Seed Portfolio Deals"), dbc.CardBody([html.Div(id='seg-seed-table')])])],
                    width=6),
            dbc.Col([dbc.Card([dbc.CardHeader("New Deals"), dbc.CardBody([html.Div(id='seg-new-table')])])], width=6),
        ], className="mb-4"),
        dbc.Row([
            dbc.Col([dbc.Card([dbc.CardHeader("Segment Allocation"), dbc.CardBody(
                [dcc.Graph(id='seg-allocation-chart', config={'displayModeBar': False})])])], width=6),
            dbc.Col([dbc.Card([dbc.CardHeader("TWR Contribution by Segment"), dbc.CardBody(
                [dcc.Graph(id='seg-contribution-chart', config={'displayModeBar': False})])])], width=6),
        ])
    ])


def analytics_page():
    return html.Div([
        html.H2("📊 Portfolio Analytics", className="mb-4", style={'fontWeight': 'bold'}),
        dbc.Row([dbc.Col([dbc.Card([dbc.CardBody([dbc.Row([
            dbc.Col([html.H6("Portfolio View:", style={'marginBottom': 0})], width=2),
            dbc.Col([dbc.RadioItems(id="analytics-view-toggle", options=[
                {"label": " Current Portfolio", "value": "current"},
                {"label": " Current + Pipeline", "value": "current_pipeline"},
                {"label": " Current + Pipeline + Placeholder", "value": "full_proforma"},
            ], value="current", inline=True)], width=10)
        ])])])])], className="mb-4"),
        dbc.Row([
            dbc.Col(dbc.Card(
                [dbc.CardBody([html.H6("Total NAV"), html.H4(id="analytics-total-nav", style={'color': C['blue']})])]),
                    width=3),
            dbc.Col(dbc.Card([dbc.CardBody(
                [html.H6("Number of Deals"), html.H4(id="analytics-num-deals", style={'color': C['green']})])]),
                    width=3),
            dbc.Col(dbc.Card([dbc.CardBody(
                [html.H6("Weighted IRR"), html.H4(id="analytics-weighted-irr", style={'color': C['purple']})])]),
                    width=3),
            dbc.Col(dbc.Card([dbc.CardBody(
                [html.H6("Top 1 Concentration"), html.H4(id="analytics-top1-conc", style={'color': C['amber']})])]),
                    width=3),
        ], className="mb-4"),
        dbc.Row([
            dbc.Col([dbc.Card([dbc.CardHeader("Strategy Exposure"),
                               dbc.CardBody([dcc.Graph(id='analytics-strategy', config={'displayModeBar': False})])])],
                    width=6),
            dbc.Col([dbc.Card([dbc.CardHeader("Regional Exposure"),
                               dbc.CardBody([dcc.Graph(id='analytics-region', config={'displayModeBar': False})])])],
                    width=6),
        ], className="mb-4"),
        dbc.Row([
            dbc.Col([dbc.Card([dbc.CardHeader("Vintage Year Exposure"),
                               dbc.CardBody([dcc.Graph(id='analytics-vintage', config={'displayModeBar': False})])])],
                    width=6),
            dbc.Col([dbc.Card([dbc.CardHeader("Secondary vs Co-Investment"),
                               dbc.CardBody([dcc.Graph(id='analytics-dealtype', config={'displayModeBar': False})])])],
                    width=6),
        ], className="mb-4"),
        dbc.Row([
            dbc.Col([dbc.Card([dbc.CardHeader("Sector Exposure"),
                               dbc.CardBody([dcc.Graph(id='analytics-sector', config={'displayModeBar': False})])])],
                    width=6),
            dbc.Col([dbc.Card([dbc.CardHeader("Concentration Risk"), dbc.CardBody(
                [dcc.Graph(id='analytics-concentration', config={'displayModeBar': False})])])], width=6),
        ])
    ])


def pipeline_page():
    return html.Div([
        dbc.Row([
            dbc.Col(html.H2("🔍 Deal Pipeline", style={'fontWeight': 'bold'}), width=8),
            dbc.Col(
                dbc.Button("➕ Add to Pipeline", id="btn-open-pipeline", color="info", className="float-end", size="lg"),
                width=4)
        ], className="mb-3"),
        dbc.Modal([
            dbc.ModalHeader("Add Deal to Pipeline"),
            dbc.ModalBody([
                dbc.Row([
                    dbc.Col([dbc.Label("Deal Name *"),
                             dbc.Input(id="in-pipe-name", type="text", placeholder="Project Sigil")], width=6),
                    dbc.Col([dbc.Label("Deal Type *"),
                             dbc.Select(id="in-pipe-type", options=[
                                 {"label": "GP-Led (Multi-Asset)", "value": "GP-Led (Multi-Asset)"},
                                 {"label": "GP-Led (Single-Asset)", "value": "GP-Led (Single-Asset)"},
                                 {"label": "Diversified LP-Led", "value": "Diversified LP-Led"},
                                 {"label": "Co-Investments", "value": "Co-Investments"},
                             ])], width=6)
                ], className="mb-3"),
                dbc.Row([
                    dbc.Col([dbc.Label("Stage"),
                             dbc.Select(id="in-pipe-stage", options=[
                                 {"label": "Screening", "value": "Screening"},
                                 {"label": "Due Diligence", "value": "Due Diligence"},
                                 {"label": "Term Sheet", "value": "Term Sheet"},
                                 {"label": "Final Docs", "value": "Final Docs"},
                             ], value="Screening")], width=4),
                    dbc.Col([dbc.Label("Target Size ($mm)"),
                             dbc.Input(id="in-pipe-size", type="number", step=0.1, placeholder="7.5")], width=4),
                    dbc.Col([dbc.Label("Target IRR (%)"),
                             dbc.Input(id="in-pipe-irr", type="number", step=0.5, placeholder="16")], width=4),
                ])
            ]),
            dbc.ModalFooter([
                dbc.Button("Cancel", id="btn-cancel-pipe", color="secondary"),
                dbc.Button("Add to Pipeline", id="btn-submit-pipe", color="info")
            ])
        ], id="modal-pipeline", size="lg", is_open=False),
        html.Div(id='pipeline-table')
    ])


def settings_page():
    return html.Div([
        html.H2("⚙️ Fund Settings", className="mb-4", style={'fontWeight': 'bold'}),
        dbc.Card([dbc.CardBody([
            html.H5("💰 Fund Parameters", className="mb-4", style={'fontWeight': 'bold'}),
            dbc.Row([
                dbc.Col(
                    [dbc.Label("💵 Dry Powder ($mm)"), dbc.Input(id="set-dry-powder", type="number", value=450, step=10),
                     html.Small("Available capital to deploy")], width=4),
                dbc.Col([dbc.Label("🎯 Target Net TWR (%)"), dbc.Input(id="set-twr", type="number", value=13, step=0.5),
                         html.Small("Net return to LPs")], width=4),
                dbc.Col([dbc.Label("📊 Average Hold Period (years)"),
                         dbc.Input(id="set-hold", type="number", value=5.0, step=0.5),
                         html.Small("For MOIC translation")], width=4),
            ], className="mb-4"),
            html.Hr(),
            html.H5("💼 Fee Structure", className="mb-4", style={'fontWeight': 'bold'}),
            dbc.Row([
                dbc.Col([dbc.Label("Management Fee (% p.a.)"),
                         dbc.Input(id="set-fee", type="number", value=1.25, step=0.05),
                         html.Small("Annual fee on NAV")], width=4),
                dbc.Col([dbc.Label("Carry Rate (%)"), dbc.Input(id="set-carry", type="number", value=12.5, step=0.5),
                         html.Small("Performance fee")], width=4),
                dbc.Col([dbc.Label("Hurdle Rate (%)"), dbc.Input(id="set-hurdle", type="number", value=10, step=0.5),
                         html.Small("Return before carry")], width=4),
            ], className="mb-4"),
            html.Hr(),
            html.H5("🔧 Portfolio Assumptions", className="mb-4", style={'fontWeight': 'bold'}),
            dbc.Row([
                dbc.Col([dbc.Label("Liquidity Reserve (%)"), dbc.Input(id="set-liq", type="number", value=5, step=1),
                         html.Small("Cash buffer")], width=4),
                dbc.Col([dbc.Label("Loss Drag (%)"), dbc.Input(id="set-loss", type="number", value=1, step=0.5),
                         html.Small("Expected annual impairment")], width=4),
                dbc.Col([dbc.Label("Cash Yield (%)"), dbc.Input(id="set-cash-yield", type="number", value=3, step=0.5),
                         html.Small("Return on uninvested cash")], width=4),
            ], className="mb-4"),
            html.Hr(),
            dbc.Row([dbc.Col([
                dbc.Button("💾 Save Settings", id="btn-save-settings", color="primary", size="lg", className="me-2"),
                dbc.Button("🔄 Reset to Default", id="btn-reset-settings", color="secondary", size="lg")
            ])])
        ])])
    ])


# ==================== CALLBACKS ====================

@app.callback(Output('live-clock', 'children'), Input('clock-interval', 'n_intervals'))
def update_clock(n):
    return datetime.now().strftime('%H:%M:%S')


@app.callback(Output('sidebar-stats', 'children'),
              [Input('deals-store', 'data'), Input('config-store', 'data')])
def update_sidebar(deals, config):
    m = calculate_portfolio_metrics(deals)
    dry_powder = config['fund_parameters']['dry_powder']
    return [
        html.Small("💰 NAV", className="text-muted"),
        html.H6(f"${m['total_nav']:.0f}M", className="mb-1", style={'fontWeight': 'bold', 'fontFamily': C['mono']}),
        html.Small("💵 Dry Powder", className="text-muted"),
        html.H6(f"${dry_powder:.0f}M", className="mb-1", style={'fontWeight': 'bold', 'fontFamily': C['mono']}),
        html.Small("📊 Deals", className="text-muted"),
        html.H6(str(m['num_deals']), className="mb-1", style={'fontWeight': 'bold', 'fontFamily': C['mono']}),
        html.Small("📈 Avg IRR", className="text-muted"),
        html.H6(f"{m['weighted_irr']:.1%}", style={'fontWeight': 'bold', 'fontFamily': C['mono']})
    ]


@app.callback(Output('page-content', 'children'), Input('url', 'pathname'))
def display_page(pathname):
    pages = {
        '/portfolio': portfolio_page, '/segmentation': segmentation_page,
        '/future': future_deals_page, '/drypowder': drypowder_page,
        '/liquidity': liquidity_assumptions_page, '/calculator': calculator_page,
        '/twr': twr_forecaster_page, '/cashflows': cashflows_page,
        '/proforma': proforma_page, '/pipeline': pipeline_page,
        '/analytics': analytics_page, '/settings': settings_page,
    }
    return pages.get(pathname, dashboard_page)()


# Modal toggles
@app.callback(Output('modal-deal', 'is_open'),
              [Input('btn-open-deal', 'n_clicks'), Input('btn-cancel', 'n_clicks'), Input('btn-submit', 'n_clicks')],
              State('modal-deal', 'is_open'), prevent_initial_call=True)
def toggle_deal_modal(o, c, s, is_open):
    return not is_open


@app.callback(Output('modal-placeholder', 'is_open'),
              [Input('btn-open-placeholder', 'n_clicks'), Input('btn-cancel-ph', 'n_clicks'),
               Input('btn-submit-ph', 'n_clicks')],
              State('modal-placeholder', 'is_open'), prevent_initial_call=True)
def toggle_placeholder_modal(o, c, s, is_open):
    return not is_open


@app.callback(Output('modal-pipeline', 'is_open'),
              [Input('btn-open-pipeline', 'n_clicks'), Input('btn-cancel-pipe', 'n_clicks'),
               Input('btn-submit-pipe', 'n_clicks')],
              State('modal-pipeline', 'is_open'), prevent_initial_call=True)
def toggle_pipeline_modal(o, c, s, is_open):
    return not is_open


@app.callback(Output('modal-edit-deal', 'is_open'),
              [Input({'type': 'edit-deal', 'index': ALL}, 'n_clicks'),
               Input('btn-cancel-edit', 'n_clicks'), Input('btn-save-edit', 'n_clicks')],
              State('modal-edit-deal', 'is_open'), prevent_initial_call=True)
def toggle_edit_modal(edit_clicks, cancel, save, is_open):
    return not is_open


# ── Populate Edit Modal with CF NAV pull ──────────────────────────────────────
@app.callback(
    [Output('edit-deal-index', 'data'),
     Output('edit-name', 'value'), Output('edit-manager', 'value'),
     Output('edit-strategy', 'value'), Output('edit-stage', 'value'),
     Output('edit-total-commitment', 'value'), Output('edit-current-commitment', 'value'),
     Output('edit-nav', 'value'), Output('edit-irr', 'value'), Output('edit-hold', 'value'),
     Output('edit-nav-source', 'children'), Output('edit-nav-month-display', 'children'),
     Output('edit-cf-pull-banner', 'children')],
    Input({'type': 'edit-deal', 'index': ALL}, 'n_clicks'),
    [State('deals-store', 'data'), State('fund-cf-data-store', 'data')],
    prevent_initial_call=True
)
def populate_edit_modal(n_clicks, deals, fund_cf_data):
    if not any(n_clicks) or not deals:
        return None, "", "", "", "", 0, 0, 0, 0, 5, "—", "—", ""

    ctx = callback_context
    if not ctx.triggered:
        return None, "", "", "", "", 0, 0, 0, 0, 5, "—", "—", ""

    button_id = ctx.triggered[0]['prop_id'].split('.')[0]
    idx = json.loads(button_id)['index']

    if idx < 0 or idx >= len(deals):
        return None, "", "", "", "", 0, 0, 0, 0, 5, "—", "—", ""

    deal = deals[idx]

    # Existing stored NAV
    stored_nav = deal.get('nav', deal.get('size', 0))
    total_commitment = deal.get('total_commitment', deal.get('commitment', stored_nav))
    current_commitment = deal.get('current_commitment', deal.get('commitment', stored_nav))

    # Try CF pull
    cf_nav, cf_month = get_latest_nav_from_cashflows(deal['name'], fund_cf_data)

    if cf_nav is not None:
        nav_value = cf_nav
        nav_source = f"✅ Auto-pulled from CF upload"
        month_display = cf_month
        banner = dbc.Alert([
            html.I(className="fas fa-check-circle me-2"),
            html.Strong("NAV auto-populated from Fund Level CF upload: "),
            f"${cf_nav:.2f}M as of {cf_month}. You may override below if needed."
        ], color="success", className="mb-0", style={'fontSize': '13px'})
    else:
        nav_value = stored_nav
        nav_source = "⚠️ Manual entry (no CF data matched)"
        month_display = "No CF data available"
        banner = dbc.Alert([
            html.I(className="fas fa-exclamation-circle me-2"),
            "No matching deal found in CF upload. NAV shows last saved value. "
            "Upload a Fund Level CF file on the Deal Cashflows page to enable auto-pull."
        ], color="warning", className="mb-0", style={'fontSize': '13px'}) if not fund_cf_data else \
            dbc.Alert([
                html.I(className="fas fa-search me-2"),
                f"No match found for '{deal['name']}' in CF data. Check deal names match between portfolio and CF file."
            ], color="warning", className="mb-0", style={'fontSize': '13px'})

    return (
        idx,
        deal['name'], deal.get('manager', ''),
        deal.get('strategy', ''), deal.get('stage', 'Buyout'),
        total_commitment, current_commitment,
        nav_value,
        deal.get('target_irr', 0) * 100,
        deal.get('hold_period', 5),
        nav_source, month_display, banner
    )


# ── Auto-calc unfunded (Add modal) ────────────────────────────────────────────
@app.callback(
    Output('unfunded-display', 'children'),
    [Input('in-total-commitment', 'value'), Input('in-current-commitment', 'value')]
)
def calculate_unfunded(total, current):
    if total is None or current is None:
        return "$0.0M"
    unfunded = float(total) - float(current)
    color = C['red'] if unfunded < 0 else C['amber']
    return html.Span(f"${unfunded:.1f}M", style={'color': color})


# ── Auto-calc unfunded (Edit modal) ───────────────────────────────────────────
@app.callback(
    Output('edit-unfunded-display', 'children'),
    [Input('edit-total-commitment', 'value'), Input('edit-current-commitment', 'value')]
)
def calculate_edit_unfunded(total, current):
    if total is None or current is None:
        return "$0.0M"
    unfunded = float(total) - float(current)
    color = C['red'] if unfunded < 0 else C['amber']
    return html.Span(f"${unfunded:.1f}M", style={'color': color})


# ── Save edited deal ──────────────────────────────────────────────────────────
@app.callback(
    Output('deals-store', 'data', allow_duplicate=True),
    Input('btn-save-edit', 'n_clicks'),
    [State('edit-deal-index', 'data'),
     State('edit-name', 'value'), State('edit-manager', 'value'),
     State('edit-strategy', 'value'), State('edit-stage', 'value'),
     State('edit-total-commitment', 'value'), State('edit-current-commitment', 'value'),
     State('edit-nav', 'value'), State('edit-irr', 'value'), State('edit-hold', 'value'),
     State('deals-store', 'data')],
    prevent_initial_call='initial_duplicate'
)
def save_edited_deal(n, idx, name, manager, strategy, stage, total_comm, current_comm, nav, irr, hold, deals):
    if not n or idx is None or not deals:
        return deals or []
    if idx < 0 or idx >= len(deals):
        return deals
    try:
        unfunded = float(total_comm) - float(current_comm)
        deals[idx].update({
            'name': name, 'manager': manager, 'strategy': strategy, 'stage': stage,
            'total_commitment': float(total_comm), 'current_commitment': float(current_comm),
            'nav': float(nav), 'size': float(nav),
            'commitment': float(total_comm),
            'unfunded': unfunded,
            'target_irr': float(irr) / 100,
            'hold_period': float(hold),
            'moic': (1 + float(irr) / 100) ** float(hold),
        })
        print(f"✅ Deal saved: {name} | NAV=${nav}M | Unfunded=${unfunded}M")
        return deals
    except Exception as e:
        print(f"ERROR saving deal: {e}")
        return deals


# ── Add Deal ──────────────────────────────────────────────────────────────────
@app.callback(
    Output('deals-store', 'data', allow_duplicate=True),
    Input('btn-submit', 'n_clicks'),
    [State('in-name', 'value'), State('in-manager', 'value'), State('in-strategy', 'value'),
     State('in-stage', 'value'), State('in-total-commitment', 'value'), State('in-current-commitment', 'value'),
     State('in-nav', 'value'), State('in-irr', 'value'), State('in-hold', 'value'), State('in-currency', 'value'),
     State('in-vintage', 'value'), State('in-segment', 'value'), State('in-allocation-status', 'value'),
     State('in-sector', 'value'), State('in-geo', 'value'),
     State('deals-store', 'data'), State('config-store', 'data')],
    prevent_initial_call='initial_duplicate'
)
def add_deal(n, name, manager, strat, stage, total_commitment, current_commitment,
             nav, irr, hold, currency, vint, segment, alloc_status, sec, geo, deals, config):
    if not n or not name or not manager or not strat:
        return deals or []
    if total_commitment is None or current_commitment is None or nav is None or irr is None:
        return deals or []
    try:
        unfunded = float(total_commitment) - float(current_commitment)
        new_deal = {
            'name': name, 'manager': manager, 'strategy': strat, 'stage': stage or 'Buyout',
            'total_commitment': float(total_commitment),
            'current_commitment': float(current_commitment),
            'nav': float(nav), 'size': float(nav),
            'commitment': float(total_commitment),
            'unfunded': unfunded,
            'target_irr': float(irr) / 100,
            'hold_period': float(hold) if hold else 5.0,
            'moic': (1 + float(irr) / 100) ** (float(hold) if hold else 5.0),
            'currency': currency or 'USD',
            'vintage': int(vint) if vint else 2024,
            'segment': segment or 'Seed',
            'allocation_status': alloc_status or 'Closed',
            'sector': sec or 'Diversified',
            'geography': geo or 'North America',
            'date_added': datetime.now().isoformat(),
            'is_actual': True
        }
        print(
            f"✅ Added: {name} | Total={total_commitment} | Current={current_commitment} | NAV={nav} | Unfunded={unfunded}")
        return (deals or []) + [new_deal]
    except Exception as e:
        print(f"ERROR adding deal: {e}")
        return deals or []


# ── Add Placeholder ───────────────────────────────────────────────────────────
@app.callback(
    Output('placeholder-deals-store', 'data', allow_duplicate=True),
    Input('btn-submit-ph', 'n_clicks'),
    [State('in-ph-name', 'value'), State('in-ph-strategy', 'value'),
     State('in-ph-size', 'value'), State('in-ph-bite', 'value'),
     State('in-ph-weight', 'value'), State('in-ph-month', 'value'),
     State('in-ph-region', 'value'), State('in-ph-type', 'value'),
     State('placeholder-deals-store', 'data')],
    prevent_initial_call='initial_duplicate'
)
def add_placeholder(n, name, strat, size, bite, weight, month, region, deal_type, placeholders):
    if not n or not all([name, strat, size is not None, month is not None]):
        return placeholders or []
    return (placeholders or []) + [{
        'name': name, 'strategy': strat, 'size': float(size),
        'custom_bite_size': float(bite) if bite else None,
        'target_weight': float(weight) if weight else None,
        'expected_month': int(month), 'region': region or 'North America',
        'deal_type': deal_type or 'Secondary', 'date_added': datetime.now().isoformat()
    }]


# ── Add Pipeline ──────────────────────────────────────────────────────────────
@app.callback(
    Output('pipeline-store', 'data', allow_duplicate=True),
    Input('btn-submit-pipe', 'n_clicks'),
    [State('in-pipe-name', 'value'), State('in-pipe-type', 'value'),
     State('in-pipe-stage', 'value'), State('in-pipe-size', 'value'),
     State('in-pipe-irr', 'value'), State('pipeline-store', 'data')],
    prevent_initial_call='initial_duplicate'
)
def add_pipeline(n, name, ptype, stage, size, irr, pipeline):
    if not all([name, ptype]):
        return pipeline
    return (pipeline or []) + [{
        'name': name, 'type': ptype, 'stage': stage or 'Screening',
        'size': float(size) if size else 0, 'target_irr': float(irr) / 100 if irr else 0,
        'date_added': datetime.now().isoformat()
    }]


# ── Delete callbacks ──────────────────────────────────────────────────────────
@app.callback(Output('deals-store', 'data', allow_duplicate=True),
              Input({'type': 'delete-deal', 'index': ALL}, 'n_clicks'),
              State('deals-store', 'data'), prevent_initial_call='initial_duplicate')
def delete_deal(n_clicks, deals):
    if not any(n_clicks): return deals
    ctx = callback_context
    if not ctx.triggered: return deals
    idx = json.loads(ctx.triggered[0]['prop_id'].split('.')[0])['index']
    if 0 <= idx < len(deals): deals.pop(idx)
    return deals


@app.callback(Output('placeholder-deals-store', 'data', allow_duplicate=True),
              Input({'type': 'delete-placeholder', 'index': ALL}, 'n_clicks'),
              State('placeholder-deals-store', 'data'), prevent_initial_call='initial_duplicate')
def delete_placeholder(n_clicks, placeholders):
    if not any(n_clicks): return placeholders
    ctx = callback_context
    if not ctx.triggered: return placeholders
    idx = json.loads(ctx.triggered[0]['prop_id'].split('.')[0])['index']
    if 0 <= idx < len(placeholders): placeholders.pop(idx)
    return placeholders


@app.callback(Output('pipeline-store', 'data', allow_duplicate=True),
              Input({'type': 'delete-pipeline', 'index': ALL}, 'n_clicks'),
              State('pipeline-store', 'data'), prevent_initial_call='initial_duplicate')
def delete_pipeline(n_clicks, pipeline):
    if not any(n_clicks): return pipeline
    ctx = callback_context
    if not ctx.triggered: return pipeline
    idx = json.loads(ctx.triggered[0]['prop_id'].split('.')[0])['index']
    if 0 <= idx < len(pipeline): pipeline.pop(idx)
    return pipeline


# ── PORTFOLIO TABLE — enhanced commitment display ─────────────────────────────
@app.callback(Output('portfolio-table', 'children'),
              [Input('deals-store', 'data'), Input('fund-cf-data-store', 'data')])
def update_portfolio_table(deals, fund_cf_data):
    if not deals:
        return dbc.Alert("📝 No deals yet. Click 'Add Deal' to start building your portfolio.", color="info")

    rows = []
    for idx, d in enumerate(deals):
        total_commitment = d.get('total_commitment', d.get('commitment', d.get('size', 0)))
        current_commitment = d.get('current_commitment', d.get('commitment', d.get('size', 0)))
        unfunded = d.get('unfunded', total_commitment - current_commitment)

        # NAV: try CF pull first, fall back to stored
        cf_nav, cf_month = get_latest_nav_from_cashflows(d['name'], fund_cf_data)
        if cf_nav is not None:
            display_nav = cf_nav
            nav_tag = html.Span(f"📡 CF: {cf_month}", className="nav-pulled")
        else:
            display_nav = d.get('nav', d.get('size', 0))
            nav_tag = html.Span("✏️ Manual", className="nav-manual")

        # Utilisation %
        utilisation = (current_commitment / total_commitment * 100) if total_commitment > 0 else 0

        # Status badge
        status = d.get('allocation_status', 'Closed')
        status_color = {'Closed': 'success', 'Pending Close': 'warning', 'Pending Allocation': 'info'}.get(status,
                                                                                                           'secondary')

        rows.append(dbc.Card([dbc.CardBody([
            dbc.Row([
                # Col 1: Deal identity
                dbc.Col([
                    html.Div([
                        html.H5([
                            d['name'],
                            dbc.Badge(status, color=status_color, className="ms-2", style={'fontSize': '10px'})
                        ], className="mb-1", style={'fontFamily': C['sans'], 'color': C['text']}),
                        html.Small([
                            html.Strong(d.get('manager', 'N/A'), style={'color': C['blue']}),
                            f" • {d['strategy']} • {d.get('stage', 'Buyout')} • {d.get('sector', '—')} • {d.get('geography', '—')}"
                        ], className="text-muted", style={'fontFamily': C['mono'], 'fontSize': '11px'}),
                        html.Br(),
                        html.Small([
                            f"Vintage {d.get('vintage', '—')} • Segment: ",
                            html.Strong(d.get('segment', 'Seed'), style={'color': C['purple']})
                        ], style={'fontFamily': C['mono'], 'fontSize': '11px', 'color': C['muted']})
                    ])
                ], width=3),

                # Col 2: Commitment breakdown — the key new section
                dbc.Col([
                    html.Div([
                        # Header
                        html.Div("📋 Commitment",
                                 style={'fontSize': '10px', 'fontWeight': '700', 'color': C['sky'],
                                        'textTransform': 'uppercase', 'letterSpacing': '0.05em',
                                        'marginBottom': '6px', 'fontFamily': C['sans']}),
                        # Total Commitment
                        dbc.Row([
                            dbc.Col(html.Small("Total:", style={'color': C['muted'], 'fontFamily': C['sans']}),
                                    width=5),
                            dbc.Col(html.Span(f"${total_commitment:.1f}M",
                                              style={'color': C['blue'], 'fontFamily': C['mono'], 'fontWeight': '600',
                                                     'fontSize': '13px'}), width=7),
                        ], className="mb-1"),
                        # Current Commitment
                        dbc.Row([
                            dbc.Col(html.Small("Called:", style={'color': C['muted'], 'fontFamily': C['sans']}),
                                    width=5),
                            dbc.Col(html.Span(f"${current_commitment:.1f}M",
                                              style={'color': C['purple'], 'fontFamily': C['mono'], 'fontWeight': '600',
                                                     'fontSize': '13px'}), width=7),
                        ], className="mb-1"),
                        # Unfunded
                        dbc.Row([
                            dbc.Col(html.Small("Unfunded:", style={'color': C['muted'], 'fontFamily': C['sans']}),
                                    width=5),
                            dbc.Col(html.Span(f"${unfunded:.1f}M",
                                              style={'color': C['amber'], 'fontFamily': C['mono'], 'fontWeight': '600',
                                                     'fontSize': '13px'}), width=7),
                        ], className="mb-1"),
                        # Utilisation bar
                        html.Div([
                            html.Div(style={
                                'height': '4px', 'borderRadius': '2px',
                                'backgroundColor': C['border2'], 'marginTop': '6px'
                            }, children=[
                                html.Div(style={
                                    'height': '4px', 'borderRadius': '2px',
                                    'width': f'{min(utilisation, 100):.0f}%',
                                    'backgroundColor': C['purple'] if utilisation < 80 else C[
                                        'amber'] if utilisation < 100 else C['red'],
                                })
                            ]),
                            html.Small(f"Called: {utilisation:.0f}%",
                                       style={'color': C['muted'], 'fontFamily': C['mono'], 'fontSize': '10px'})
                        ])
                    ], style={
                        'backgroundColor': rgba(C['sky'], 0.05), 'border': f'1px solid {C["border"]}',
                        'borderRadius': '6px', 'padding': '10px'
                    })
                ], width=3),

                # Col 3: NAV — with CF pull indicator
                dbc.Col([
                    html.Div([
                        html.Div("📈 Current NAV",
                                 style={'fontSize': '10px', 'fontWeight': '700', 'color': C['green'],
                                        'textTransform': 'uppercase', 'letterSpacing': '0.05em',
                                        'marginBottom': '6px', 'fontFamily': C['sans']}),
                        html.H4(f"${display_nav:.1f}M",
                                style={'color': C['green'], 'fontFamily': C['mono'], 'fontWeight': '700',
                                       'marginBottom': '4px'}),
                        nav_tag,
                        html.Br(),
                        # MOIC on NAV vs commitment
                        html.Small(
                            f"TVPI: {display_nav / current_commitment:.2f}x" if current_commitment > 0 else "TVPI: —",
                            style={'color': C['muted'], 'fontFamily': C['mono'], 'fontSize': '11px'}
                        )
                    ], style={
                        'backgroundColor': rgba(C['green'], 0.06), 'border': f'1px solid {C["border"]}',
                        'borderRadius': '6px', 'padding': '10px'
                    })
                ], width=2),

                # Col 4: Return metrics
                dbc.Col([
                    html.Div([
                        html.Div("📊 Returns",
                                 style={'fontSize': '10px', 'fontWeight': '700', 'color': C['teal'],
                                        'textTransform': 'uppercase', 'letterSpacing': '0.05em',
                                        'marginBottom': '6px', 'fontFamily': C['sans']}),
                        dbc.Row([
                            dbc.Col(html.Small("IRR:", style={'color': C['muted']}), width=5),
                            dbc.Col(html.Span(f"{d['target_irr']:.1%}",
                                              style={'color': C['text'], 'fontFamily': C['mono'], 'fontWeight': '600'}),
                                    width=7),
                        ], className="mb-1"),
                        dbc.Row([
                            dbc.Col(html.Small("MOIC:", style={'color': C['muted']}), width=5),
                            dbc.Col(html.Span(f"{d.get('moic', 0):.2f}x",
                                              style={'color': C['text'], 'fontFamily': C['mono'], 'fontWeight': '600'}),
                                    width=7),
                        ], className="mb-1"),
                        dbc.Row([
                            dbc.Col(html.Small("Hold:", style={'color': C['muted']}), width=5),
                            dbc.Col(html.Span(f"{d.get('hold_period', 5):.1f}y",
                                              style={'color': C['text'], 'fontFamily': C['mono']}), width=7),
                        ]),
                    ], style={
                        'backgroundColor': rgba(C['teal'], 0.05), 'border': f'1px solid {C["border"]}',
                        'borderRadius': '6px', 'padding': '10px'
                    })
                ], width=2),

                # Col 5: Actions
                dbc.Col([
                    html.Div([
                        dbc.Button("✏️ Edit", id={'type': 'edit-deal', 'index': idx},
                                   color="warning", size="sm", className="w-100 mb-2"),
                        dbc.Button("🗑️ Delete", id={'type': 'delete-deal', 'index': idx},
                                   color="danger", size="sm", outline=True, className="w-100")
                    ])
                ], width=2, className="d-flex align-items-center")
            ])
        ])], className="mb-2 shadow-sm deal-card",
            style={'backgroundColor': C['panel'], 'border': f'1px solid {C["border"]}'}))

    return html.Div(rows)


# ── Portfolio page summary metrics ────────────────────────────────────────────
@app.callback(
    [Output('port-nav', 'children'), Output('port-total-commitment', 'children'),
     Output('port-current-commitment', 'children'), Output('port-unfunded', 'children'),
     Output('port-dry-powder', 'children'), Output('port-target-return', 'children'),
     Output('port-cf-indicator', 'children')],
    [Input('deals-store', 'data'), Input('config-store', 'data'), Input('fund-cf-data-store', 'data')]
)
def update_portfolio_metrics(deals, config, fund_cf_data):
    m = calculate_portfolio_metrics(deals)
    dry_powder = config['fund_parameters']['dry_powder']
    req_irr = calculate_required_future_irr(m['weighted_irr'], m['total_nav'], dry_powder, config)

    # CF indicator
    if fund_cf_data:
        cf_banner = dbc.Alert([
            html.I(className="fas fa-satellite-dish me-2"),
            html.Strong("Fund Level CF data loaded — "),
            f"{len(fund_cf_data)} deals. NAV values in portfolio cards are auto-pulled from the latest available CF month. "
            "Click ✏️ Edit on any deal to see the source month."
        ], color="success", className="mb-3", style={'fontSize': '13px'})
    else:
        cf_banner = dbc.Alert([
            html.I(className="fas fa-info-circle me-2"),
            "No CF file loaded. NAV values are manually entered. Upload a Fund Level CF file on the ",
            dcc.Link("Deal Cashflows page", href="/cashflows", style={'color': C['blue']}),
            " to enable auto-population."
        ], color="info", className="mb-3", style={'fontSize': '13px'})

    return (
        f"${m['total_nav']:.1f}M",
        f"${m['total_commitment']:.1f}M",
        f"${m['total_current_commitment']:.1f}M",
        f"${m['total_unfunded']:.1f}M",
        f"${dry_powder:.0f}M",
        f"{req_irr:.1%}",
        cf_banner
    )


# ── Placeholder Table ─────────────────────────────────────────────────────────
@app.callback(Output('placeholder-table', 'children'), Input('placeholder-deals-store', 'data'))
def update_placeholder_table(placeholders):
    if not placeholders:
        return dbc.Alert("📝 No placeholder deals yet. Add future deals to forecast dry powder.", color="info")
    rows = []
    for idx, pd in enumerate(placeholders):
        base_date = datetime(2026, 1, 1)
        month_name = (base_date + relativedelta(months=pd['expected_month'])).strftime('%b %Y')
        rows.append(dbc.Card([dbc.CardBody([dbc.Row([
            dbc.Col([
                html.H5(pd['name'], className="mb-1", style={'color': C['text']}),
                html.Small(f"{pd['strategy']} • Expected: {month_name}", className="text-muted")
            ], width=8),
            dbc.Col([html.H5(f"${pd['size']:.1f}M", className="text-end", style={'color': C['green']})], width=2),
            dbc.Col([dbc.ButtonGroup([
                dbc.Button("✏️", id={'type': 'edit-placeholder', 'index': idx}, color="warning", size="sm",
                           outline=True, className="me-1"),
                dbc.Button("🗑️", id={'type': 'delete-placeholder', 'index': idx}, color="danger", size="sm",
                           outline=True)
            ])], width=2, className="text-end")
        ])])], className="mb-2 shadow-sm",
            style={'backgroundColor': C['panel'], 'border': f'1px solid {C["border"]}'}))
    return html.Div(rows)


# ── Pipeline Table ────────────────────────────────────────────────────────────
@app.callback(Output('pipeline-table', 'children'), Input('pipeline-store', 'data'))
def update_pipeline_table(pipeline):
    if not pipeline:
        return dbc.Alert("📝 No pipeline deals yet.", color="info")
    rows = []
    for idx, p in enumerate(pipeline):
        rows.append(dbc.Card([dbc.CardBody([dbc.Row([
            dbc.Col([
                html.H5(p['name'], className="mb-1", style={'color': C['text']}),
                html.Small(f"{p['type']} • {p['stage']}", className="text-muted")
            ], width=6),
            dbc.Col([html.Small("Size: "),
                     html.Strong(f"${p['size']:.1f}M" if p['size'] > 0 else "TBD", style={'color': C['green']})],
                    width=2),
            dbc.Col([html.Small("IRR: "), html.Strong(f"{p['target_irr']:.1%}" if p['target_irr'] > 0 else "TBD")],
                    width=2),
            dbc.Col([dbc.ButtonGroup([
                dbc.Button("✏️", id={'type': 'edit-pipeline', 'index': idx}, color="warning", size="sm", outline=True,
                           className="me-1"),
                dbc.Button("🗑️", id={'type': 'delete-pipeline', 'index': idx}, color="danger", size="sm", outline=True)
            ])], width=2, className="text-end")
        ])])], className="mb-2 shadow-sm",
            style={'backgroundColor': C['panel'], 'border': f'1px solid {C["border"]}'}))
    return html.Div(rows)


# ── Dashboard KPIs ─────────────────────────────────────────────────────────────
@app.callback(
    [Output('dash-nav', 'children'), Output('dash-num-deals', 'children'),
     Output('dash-dry-powder', 'children'), Output('dash-total', 'children'),
     Output('dash-current-irr', 'children'), Output('dash-req-irr', 'children'),
     Output('dash-placeholders', 'children'), Output('dash-placeholder-value', 'children')],
    [Input('deals-store', 'data'), Input('placeholder-deals-store', 'data'),
     Input('config-store', 'data'), Input('liquidity-data-store', 'data'), Input('fund-cf-data-store', 'data')]
)
def update_dashboard_kpis(deals, placeholders, config, liquidity_data, fund_cf_data):
    m = calculate_portfolio_metrics(deals)
    investment_nav = m['total_nav']

    cash = glf = cqs = 0
    if liquidity_data:
        cash = liquidity_data.get('current_cash', 0)
        glf = liquidity_data.get('glf_balance', 0)
        cqs = liquidity_data.get('cqs_balance', 0)

    if fund_cf_data and len(fund_cf_data) > 0:
        investment_nav = sum(deal.get('nav', deal.get('size', 0)) for deal in fund_cf_data)

    total_nav = investment_nav + cash + glf + cqs
    dry_powder = config['fund_parameters']['dry_powder']
    req_irr = calculate_required_future_irr(m['weighted_irr'], investment_nav, dry_powder, config)
    total_placeholder_value = sum(p['size'] for p in placeholders)

    return (
        f"${total_nav:.1f}M", f"{m['num_deals']} deals",
        f"${dry_powder:.0f}M", f"${total_nav + dry_powder:.0f}M",
        f"{m['weighted_irr']:.1%}", f"{req_irr:.1%}",
        str(len(placeholders)), f"${total_placeholder_value:.0f}M"
    )


# ── Dashboard Charts ───────────────────────────────────────────────────────────
@app.callback(
    [Output('dash-allocation-chart', 'figure'), Output('dash-forecast-chart', 'figure'),
     Output('dash-bite-sizing', 'children')],
    [Input('deals-store', 'data'), Input('placeholder-deals-store', 'data'), Input('config-store', 'data')]
)
def update_dashboard_charts(deals, placeholders, config):
    m = calculate_portfolio_metrics(deals)
    dry_powder = config['fund_parameters']['dry_powder']

    if m['by_strategy']:
        labels = list(m['by_strategy'].keys())
        values = [m['by_strategy'][s]['nav'] for s in labels]
        fig_alloc = go.Figure(data=[go.Pie(
            labels=labels, values=values, hole=0.4,
            marker=dict(colors=[C['blue'], C['purple'], C['teal'], C['green']][:len(labels)],
                        line=dict(color=C['border'], width=1)),
            textfont=dict(color=C['text'], family=C['mono'])
        )])
        fig_alloc.update_layout(**CHART_BASE, height=300, margin=dict(t=20, b=0, l=0, r=0), showlegend=False)
    else:
        fig_alloc = go.Figure()
        fig_alloc.update_layout(**CHART_BASE, height=300)

    forecast = forecast_dry_powder(m['total_nav'], dry_powder, deals, placeholders, config, 12)
    fig_forecast = go.Figure()
    fig_forecast.add_trace(go.Scatter(
        x=[f['month'] for f in forecast], y=[f['dry_powder'] for f in forecast],
        mode='lines+markers', line=dict(color=C['blue'], width=3),
        fill='tozeroy', fillcolor=rgba(C['blue'], 0.2),
        marker=dict(size=6, color=C['sky'])
    ))
    fig_forecast.update_layout(**CHART_BASE, height=300, xaxis_title="Month", yaxis_title="$mm", hovermode='x unified')

    bite_sizes = calculate_bite_sizes(dry_powder, config)
    bite_cards = [html.Div([
        html.H6(s, style={'fontSize': '12px', 'fontWeight': 'bold'}),
        html.Small(f"Min: ${v['min']:.1f}M ({v['min_pct']:.1%})", className="d-block text-muted"),
        html.Small(f"Desired: ${v['desired']:.1f}M ({v['desired_pct']:.1%})", className="d-block",
                   style={'color': C['green']}),
        html.Small(f"Max: ${v['max']:.1f}M ({v['max_pct']:.1%})", className="d-block", style={'color': C['red']}),
        html.Hr(className="my-2")
    ]) for s, v in bite_sizes.items()]

    return fig_alloc, fig_forecast, html.Div(bite_cards)


# ── Calculator ────────────────────────────────────────────────────────────────
@app.callback(
    [Output('calc-current-nav', 'children'), Output('calc-current-irr', 'children'),
     Output('calc-num-deals', 'children'), Output('calc-total-fund', 'children'),
     Output('calc-target-twr', 'children'), Output('calc-dry-powder', 'children'),
     Output('calc-required-irr', 'children'), Output('calc-explanation', 'children'),
     Output('calc-waterfall', 'children')],
    [Input('deals-store', 'data'), Input('config-store', 'data')]
)
def update_calculator(deals, config):
    m = calculate_portfolio_metrics(deals)
    dry_powder = config['fund_parameters']['dry_powder']
    target_twr = config['fund_parameters']['target_net_twr']
    req_irr = calculate_required_future_irr(m['weighted_irr'], m['total_nav'], dry_powder, config)
    gap = req_irr - m['weighted_irr'] if m['total_nav'] > 0 else 0
    explanation = f"{'Higher' if gap > 0 else 'Lower'} than current portfolio by {abs(gap):.1%}" if m[
                                                                                                        'total_nav'] > 0 else "Add deals to calculate"

    invested_ratio = max(0, 1 - config['fund_parameters']['liquidity_reserve_pct'])
    idle_ratio = config['fund_parameters']['liquidity_reserve_pct']
    waterfall_steps = [
        ("1. Target Net TWR (to LPs)", f"{target_twr:.1%}"),
        ("2. + Management Fee", f"{config['fund_parameters']['management_fee']:.2%}"),
        ("3. + Loss Drag (impairments)", f"{config['fund_parameters']['loss_drag']:.1%}"),
        ("4. - Cash Yield on reserves", f"-{idle_ratio * config['fund_parameters']['cash_yield']:.2%}"),
        ("5. ÷ Invested Ratio", f"÷ {invested_ratio:.1%}"),
        ("6. + Carry Drag (if > hurdle)", f"+Variable ({config['fund_parameters']['carry_rate']:.0%})"),
        ("═══════════════════════", "═══════════"),
        ("REQUIRED GROSS DEAL RETURN", f"{req_irr:.1%}")
    ]
    waterfall_div = dbc.Table([html.Tbody([
        html.Tr([
            html.Td(step, style={'fontSize': '14px', 'fontWeight': 'bold' if '═' in step else 'normal'}),
            html.Td(val, className="text-end", style={'fontSize': '14px', 'fontWeight': 'bold',
                                                      'color': C['green'] if '═' in step else C['text']})
        ]) for step, val in waterfall_steps
    ])], bordered=True, striped=True, hover=True)

    return (f"${m['total_nav']:.1f}M", f"{m['weighted_irr']:.1%}", str(m['num_deals']),
            f"${m['total_nav'] + dry_powder:.0f}M", f"{target_twr:.1%}", f"${dry_powder:.0f}M",
            f"{req_irr:.1%}", explanation, waterfall_div)


# ── Dry Powder Page ───────────────────────────────────────────────────────────
@app.callback(
    [Output('dp-current', 'children'), Output('dp-planned', 'children'),
     Output('dp-distributions', 'children'), Output('dp-forecast-12m', 'children'),
     Output('dp-forecast-chart', 'figure'), Output('dp-monthly-table', 'children')],
    [Input('deals-store', 'data'), Input('placeholder-deals-store', 'data'), Input('config-store', 'data')]
)
def update_drypowder_page(deals, placeholders, config):
    m = calculate_portfolio_metrics(deals)
    dry_powder = config['fund_parameters']['dry_powder']
    total_planned = sum(p['size'] for p in placeholders)
    annual_dist = m['total_nav'] * config['fund_parameters']['distribution_rate']
    forecast = forecast_dry_powder(m['total_nav'], dry_powder, deals, placeholders, config, 12)
    forecast_12m = forecast[-1]['dry_powder']

    months = [f['month'] for f in forecast]
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(go.Scatter(x=months, y=[f['dry_powder'] for f in forecast], name='Dry Powder',
                             line=dict(color=C['blue'], width=3), fill='tozeroy', fillcolor=rgba(C['blue'], 0.2)),
                  secondary_y=False)
    fig.add_trace(go.Bar(x=months, y=[f['calls'] for f in forecast], name='Capital Calls',
                         marker_color=C['red'], opacity=0.7), secondary_y=True)
    fig.add_trace(go.Bar(x=months, y=[f['distributions'] for f in forecast], name='Distributions',
                         marker_color=C['green'], opacity=0.7), secondary_y=True)
    fig.update_xaxes(gridcolor=C['border'])
    fig.update_yaxes(title_text="Dry Powder ($mm)", secondary_y=False, gridcolor=C['border'])
    fig.update_yaxes(title_text="Flows ($mm)", secondary_y=True, gridcolor=C['border'])
    fig.update_layout(**CHART_BASE, height=400, hovermode='x unified')

    table_data = [{'Month': f['month'], 'Dry Powder': f"${f['dry_powder']:.1f}M",
                   'NAV': f"${f['nav']:.1f}M", 'Distributions': f"${f['distributions']:.1f}M",
                   'Calls': f"${f['calls']:.1f}M" if f['calls'] > 0 else "-"} for f in forecast]
    table = dash_table.DataTable(data=table_data, columns=[{"name": c, "id": c} for c in table_data[0].keys()],
                                 style_cell={'textAlign': 'left', 'padding': '10px', 'fontFamily': C['mono'],
                                             'fontSize': '11px'},
                                 style_header={'backgroundColor': C['surface'], 'color': C['text'],
                                               'fontWeight': 'bold'},
                                 style_data={'backgroundColor': C['panel'], 'color': C['text']},
                                 style_data_conditional=[{'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']}],
                                 page_size=12)

    return (f"${dry_powder:.0f}M", f"${total_planned:.0f}M", f"${annual_dist:.0f}M", f"${forecast_12m:.0f}M", fig,
            table)


# ── TWR Forecaster ────────────────────────────────────────────────────────────
@app.callback(Output('twr-current-irr', 'children'), Input('deals-store', 'data'))
def update_twr_current_irr(deals):
    return f"{calculate_portfolio_metrics(deals)['weighted_irr']:.1%}"


@app.callback(
    [Output('twr-distribution-chart', 'figure'), Output('twr-statistics', 'children'),
     Output('twr-sensitivity-chart', 'figure')],
    Input('btn-run-twr', 'n_clicks'),
    [State('deals-store', 'data'), State('config-store', 'data'),
     State('twr-future-mean', 'value'), State('twr-future-std', 'value'), State('twr-n-sims', 'value')],
    prevent_initial_call=True
)
def run_twr_simulation(n, deals, config, future_mean, future_std, n_sims):
    m = calculate_portfolio_metrics(deals)
    current_irr = m['weighted_irr']
    current_nav = m['total_nav']
    dry_powder = config['fund_parameters']['dry_powder']
    target_twr = config['fund_parameters']['target_net_twr']
    mgmt_fee = config['fund_parameters']['management_fee']
    loss_drag = config['fund_parameters']['loss_drag']
    liq_reserve = config['fund_parameters']['liquidity_reserve_pct']
    cash_yield = config['fund_parameters']['cash_yield']
    carry_rate = config['fund_parameters']['carry_rate']
    hurdle = config['fund_parameters']['hurdle_rate']

    np.random.seed(42)
    future_irrs = np.random.normal(future_mean / 100, future_std / 100, n_sims)
    results = []
    for fir in future_irrs:
        total = current_nav + dry_powder
        blended = (current_nav * current_irr + dry_powder * fir) / total if total > 0 else 0
        gross = (blended * (1 - liq_reserve)) + (cash_yield * liq_reserve) - mgmt_fee - loss_drag
        net = gross - (gross - hurdle) * carry_rate if gross > hurdle else gross
        results.append(net)
    results = np.array(results)

    fig_dist = go.Figure()
    fig_dist.add_trace(go.Histogram(x=results * 100, nbinsx=50, marker_color=C['blue'], opacity=0.7))
    fig_dist.add_vline(x=target_twr * 100, line_color=C['red'], line_width=3, line_dash='dash',
                       annotation_text=f"Target: {target_twr:.1%}")
    fig_dist.add_vline(x=np.median(results) * 100, line_color=C['green'], line_width=3,
                       annotation_text=f"Median: {np.median(results):.1%}")
    fig_dist.update_layout(**CHART_BASE, xaxis_title="Net TWR (%)", yaxis_title="Frequency", height=400)

    prob = (results >= target_twr).sum() / len(results)
    stats_div = html.Div([
        html.Div([html.Strong("P(Hit Target):"), html.H3(f"{prob:.1%}",
                                                         style={'color': C['green'] if prob >= 0.75 else C[
                                                             'amber'] if prob >= 0.5 else C['red']})],
                 className="mb-3"),
        html.Hr(style={'borderColor': C['border']}),
        html.P([html.Strong("Mean: "),
                html.Span(f"{np.mean(results):.2%}", style={'color': C['blue'], 'fontFamily': C['mono']})]),
        html.P([html.Strong("Median: "),
                html.Span(f"{np.median(results):.2%}", style={'color': C['green'], 'fontFamily': C['mono']})]),
        html.P([html.Strong("Std Dev: "),
                html.Span(f"{np.std(results):.2%}", style={'color': C['muted'], 'fontFamily': C['mono']})]),
        html.P([html.Strong("5th %ile: "),
                html.Span(f"{np.percentile(results, 5):.2%}", style={'color': C['red'], 'fontFamily': C['mono']})]),
        html.P([html.Strong("95th %ile: "),
                html.Span(f"{np.percentile(results, 95):.2%}", style={'color': C['green'], 'fontFamily': C['mono']})]),
    ])

    fir_range = np.linspace(0.15, 0.35, 20)
    twr_sens = []
    for fir in fir_range:
        total = current_nav + dry_powder
        blended = (current_nav * current_irr + dry_powder * fir) / total if total > 0 else 0
        gross = (blended * (1 - liq_reserve)) + (cash_yield * liq_reserve) - mgmt_fee - loss_drag
        net = gross - (gross - hurdle) * carry_rate if gross > hurdle else gross
        twr_sens.append(net)

    fig_sens = go.Figure()
    fig_sens.add_trace(go.Scatter(x=fir_range * 100, y=np.array(twr_sens) * 100, mode='lines+markers',
                                  line=dict(color=C['blue'], width=3), marker=dict(size=6, color=C['sky'])))
    fig_sens.add_hline(y=target_twr * 100, line_color=C['red'], line_dash='dash',
                       annotation_text=f"Target: {target_twr:.1%}")
    fig_sens.update_layout(**CHART_BASE, xaxis_title="Future Deal IRR (%)", yaxis_title="Net TWR (%)", height=400)
    return fig_dist, stats_div, fig_sens


# ── Segmentation callbacks ────────────────────────────────────────────────────
@app.callback(
    [Output('seg-total-nav', 'children'), Output('seg-total-twr', 'children'),
     Output('seg-seed-nav', 'children'), Output('seg-seed-twr', 'children'),
     Output('seg-new-nav', 'children'), Output('seg-new-twr', 'children'),
     Output('seg-mm-nav', 'children'), Output('seg-mm-twr', 'children'),
     Output('seg-pipeline-nav', 'children'), Output('seg-pipeline-twr', 'children'),
     Output('seg-future-nav', 'children'), Output('seg-future-twr', 'children')],
    [Input('deals-store', 'data'), Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')]
)
def update_segment_summary(deals, pipeline, placeholders):
    seed_deals = [d for d in deals if d.get('segment', 'Seed') == 'Seed'] if deals else []
    new_deals = [d for d in deals if d.get('segment', 'Seed') == 'New'] if deals else []
    mm_deals = [d for d in deals if d.get('segment', 'Seed') == 'MoneyMarket'] if deals else []
    seed_nav = sum(d.get('nav', d.get('size', 0)) for d in seed_deals)
    new_nav = sum(d.get('nav', d.get('size', 0)) for d in new_deals)
    mm_nav = sum(d.get('nav', d.get('size', 0)) for d in mm_deals)
    pipeline_nav = sum(p['size'] for p in pipeline) if pipeline else 0
    future_nav = sum(p['size'] for p in placeholders) if placeholders else 0
    total_nav = seed_nav + new_nav + mm_nav
    seed_twr = sum(
        d['target_irr'] * d.get('nav', d.get('size', 0)) for d in seed_deals) / seed_nav if seed_nav > 0 else 0
    new_twr = sum(d['target_irr'] * d.get('nav', d.get('size', 0)) for d in new_deals) / new_nav if new_nav > 0 else 0
    mm_twr = 0.03
    pipeline_twr = sum(
        p['target_irr'] * p['size'] for p in pipeline) / pipeline_nav if pipeline_nav > 0 and pipeline else 0
    total_twr = (seed_nav * seed_twr + new_nav * new_twr + mm_nav * mm_twr) / total_nav if total_nav > 0 else 0
    return (f"${total_nav:.1f}M", f"TWR: {total_twr:.1%}", f"${seed_nav:.1f}M", f"TWR: {seed_twr:.1%}",
            f"${new_nav:.1f}M", f"TWR: {new_twr:.1%}", f"${mm_nav:.1f}M", f"TWR: {mm_twr:.1%}",
            f"${pipeline_nav:.1f}M", f"Est TWR: {pipeline_twr:.1%}", f"${future_nav:.1f}M", "Target TWR: 25.0%")


@app.callback(Output('seg-twr-forecast-chart', 'figure'),
              [Input('deals-store', 'data'), Input('config-store', 'data')])
def generate_twr_forecast_by_segment(deals, config):
    base_date = datetime(2026, 1, 1)
    months = [(base_date + relativedelta(months=i)).strftime('%b %Y') for i in range(12)]
    target_twr = config['fund_parameters']['target_net_twr']
    fig = go.Figure()
    for name, mult, color, dash in [
        ('Total Portfolio', 1.0, C['blue'], 'solid'),
        ('Seed Portfolio', 0.95, C['green'], 'dash'),
        ('New Deals', 1.05, C['purple'], 'dot'),
        ('Money Market', None, C['amber'], 'dashdot'),
    ]:
        if mult:
            vals = [((1 + target_twr) ** (i / 12) * mult - 1) * 100 for i in range(12)]
        else:
            vals = [(((1 + 0.03) ** (i / 12)) - 1) * 100 for i in range(12)]
        fig.add_trace(go.Scatter(x=months, y=vals, name=name, mode='lines',
                                 line=dict(color=color, width=2 if mult != 1.0 else 3, dash=dash)))
    fig.update_layout(**CHART_BASE, yaxis_title='Cumulative TWR (%)', height=400, hovermode='x unified')
    return fig


@app.callback([Output('seg-seed-table', 'children'), Output('seg-new-table', 'children')],
              Input('deals-store', 'data'))
def update_segment_tables(deals):
    def make_table(seg_deals, label):
        if not seg_deals:
            return html.P(f"No {label} yet", className="text-muted")
        data = [{'Deal': d['name'], 'NAV': f"${d.get('nav', d.get('size', 0)):.1f}M",
                 'IRR': f"{d['target_irr']:.1%}", 'Vintage': d.get('vintage', 'N/A')} for d in seg_deals]
        return dash_table.DataTable(data=data, columns=[{"name": c, "id": c} for c in data[0].keys()],
                                    style_cell={'textAlign': 'left', 'padding': '10px', 'fontFamily': C['mono'],
                                                'fontSize': '11px'},
                                    style_header={'backgroundColor': C['surface'], 'color': C['text'],
                                                  'fontWeight': 'bold'},
                                    style_data={'backgroundColor': C['panel'], 'color': C['text']},
                                    style_data_conditional=[
                                        {'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']}], page_size=10)

    seed_deals = [d for d in deals if d.get('segment', 'Seed') == 'Seed'] if deals else []
    new_deals = [d for d in deals if d.get('segment', 'Seed') == 'New'] if deals else []
    return make_table(seed_deals, 'Seed'), make_table(new_deals, 'New')


@app.callback(Output('seg-allocation-chart', 'figure'),
              [Input('deals-store', 'data'), Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')])
def generate_allocation_chart(deals, pipeline, placeholders):
    seg_data = {
        'Seed Portfolio': (
            sum(d.get('nav', d.get('size', 0)) for d in deals if d.get('segment', 'Seed') == 'Seed') if deals else 0,
            C['green']),
        'New Deals': (
            sum(d.get('nav', d.get('size', 0)) for d in deals if d.get('segment', 'Seed') == 'New') if deals else 0,
            C['purple']),
        'Money Market': (sum(
            d.get('nav', d.get('size', 0)) for d in deals if d.get('segment', 'Seed') == 'MoneyMarket') if deals else 0,
                         C['amber']),
        'Pipeline': (sum(p['size'] for p in pipeline) if pipeline else 0, C['teal']),
        'Future': (sum(p['size'] for p in placeholders) if placeholders else 0, C['pink']),
    }
    labels, values, colors = zip(*[(k, v, c) for k, (v, c) in seg_data.items() if v > 0]) if any(
        v for v, _ in seg_data.values()) else ([], [], [])
    fig = go.Figure(data=[go.Pie(labels=list(labels), values=list(values), hole=0.4,
                                 marker=dict(colors=list(colors)), textfont=dict(color=C['text']))])
    fig.update_layout(**CHART_BASE, height=350, margin=dict(t=20, b=0, l=0, r=0))
    return fig


@app.callback(Output('seg-contribution-chart', 'figure'), Input('deals-store', 'data'))
def generate_contribution_chart(deals):
    if not deals:
        fig = go.Figure();
        fig.update_layout(**CHART_BASE);
        return fig
    seed_c = sum(d['target_irr'] * d.get('nav', d.get('size', 0)) for d in deals if d.get('segment', 'Seed') == 'Seed')
    new_c = sum(d['target_irr'] * d.get('nav', d.get('size', 0)) for d in deals if d.get('segment', 'Seed') == 'New')
    mm_c = sum(0.03 * d.get('nav', d.get('size', 0)) for d in deals if d.get('segment', 'Seed') == 'MoneyMarket')
    segs = [('Seed', seed_c, C['green']), ('New', new_c, C['purple']), ('MM', mm_c, C['amber'])]
    segs = [(n, v * 100, c) for n, v, c in segs if v > 0]
    if not segs: fig = go.Figure(); fig.update_layout(**CHART_BASE); return fig
    names, vals, colors = zip(*segs)
    fig = go.Figure(data=[go.Bar(x=list(names), y=list(vals), marker_color=list(colors),
                                 text=[f"{v:.1f}" for v in vals], textposition='outside')])
    fig.update_layout(**CHART_BASE, yaxis_title="TWR Contribution", height=350)
    return fig


# ── Cashflows Upload ──────────────────────────────────────────────────────────
@app.callback(
    [Output('fund-cf-data-store', 'data'), Output('upload-fund-cf-status', 'children')],
    Input('upload-fund-cf', 'contents'),
    State('upload-fund-cf', 'filename')
)
def upload_fund_cf_file(contents, filename):
    if contents is None:
        return None, ""
    try:
        content_type, content_string = contents.split(',')
        decoded = base64.b64decode(content_string)
        deals_data = []

        if filename.lower().endswith('.csv'):
            import csv
            from io import StringIO
            rows = list(csv.reader(StringIO(decoded.decode('utf-8'))))
            header_row_idx = next((i for i, r in enumerate(rows)
                                   if any('Investment' in str(c) or 'AIC Name' in str(c) for c in r)), None)
            if header_row_idx is not None:
                for row in rows[header_row_idx + 1:]:
                    if len(row) > 5 and row[0] and 'Export' not in str(row[0]):
                        d = {'name': row[0].strip(), 'type': row[1].strip() if len(row) > 1 else '',
                             'commitment': 0, 'nav': 0, 'monthly_navs': {}, 'monthly_cfs': {}}
                        try:
                            d['commitment'] = float(str(row[2]).replace(',', '').replace('$', ''))
                        except:
                            pass
                        try:
                            d['nav'] = float(str(row[3]).replace(',', '').replace('$', ''))
                        except:
                            pass
                        deals_data.append(d)
        else:
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(decoded), data_only=True)
            ws = next((wb[n] for n in ['Fund Level CF', 'Fund_Level_CF', 'FundLevelCF'] if n in wb.sheetnames),
                      wb[wb.sheetnames[0]])

            # Find header row
            header_row = None
            for row in range(1, min(20, ws.max_row + 1)):
                for col in range(1, min(10, ws.max_column + 1)):
                    val = ws.cell(row, col).value
                    if val and ('AIC Name' in str(val) or 'Investment' in str(val) or 'Deal' in str(val)):
                        header_row = row
                        break
                if header_row:
                    break

            if header_row:
                headers = {str(ws.cell(header_row, col).value).strip(): col
                           for col in range(1, min(ws.max_column + 1, 600))
                           if ws.cell(header_row, col).value}

                name_col = headers.get('AIC Name') or headers.get('Investments') or headers.get('Investment')

                for row in range(header_row + 1, min(header_row + 200, ws.max_row + 1)):
                    if not name_col:
                        break
                    val = ws.cell(row, name_col).value
                    if not val or not str(val).strip() or 'Export' in str(val) or 'Total' in str(val):
                        continue
                    deal_name = str(val).strip()

                    commitment = 0
                    if 'Commitment ($m)' in headers:
                        try:
                            commitment = float(ws.cell(row, headers['Commitment ($m)']).value or 0)
                        except:
                            pass

                    paid_in = 0
                    if 'Paid In' in headers:
                        try:
                            paid_in = float(ws.cell(row, headers['Paid In']).value or 0)
                        except:
                            pass

                    unfunded = 0
                    if 'Unfunded' in headers:
                        try:
                            unfunded = float(ws.cell(row, headers['Unfunded']).value or 0)
                        except:
                            pass

                    # Scan for latest non-zero NAV (scan all columns from right)
                    nav = 0
                    nav_col_start = max(headers.values()) - 200 if headers else 337
                    for col in range(min(ws.max_column, nav_col_start + 300), max(1, nav_col_start - 1), -1):
                        try:
                            v = ws.cell(row, col).value
                            if v and isinstance(v, (int, float)) and abs(v) > 0.001:
                                nav = float(v) / 1_000_000 if abs(v) > 100 else float(v)
                                break
                        except:
                            continue

                    if nav == 0:
                        nav = commitment

                    # Build monthly NAV dict from last N columns (up to 60 months back from max)
                    monthly_navs = {}
                    nav_section_start = max(1, ws.max_column - 60)
                    month_idx = 0
                    for col in range(nav_section_start, ws.max_column + 1):
                        try:
                            v = ws.cell(row, col).value
                            if v and isinstance(v, (int, float)) and abs(v) > 0.001:
                                monthly_navs[month_idx] = float(v) / 1_000_000 if abs(v) > 100 else float(v)
                        except:
                            pass
                        month_idx += 1

                    deals_data.append({
                        'name': deal_name,
                        'type': str(ws.cell(row, headers.get('Investment Type', 1)).value or ''),
                        'commitment': commitment,
                        'total_commitment': commitment,
                        'current_commitment': paid_in if paid_in > 0 else commitment,
                        'nav': nav,
                        'unfunded': unfunded / 1_000_000 if unfunded > 1000 else unfunded,
                        'monthly_navs': monthly_navs,
                        'monthly_cfs': {}
                    })

        return deals_data, dbc.Alert([html.I(className="fas fa-check-circle me-2"),
                                      f"✅ {filename} uploaded — {len(deals_data)} deals found. NAV auto-pull is now active."],
                                     color="success", className="mt-2")

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return None, dbc.Alert([html.I(className="fas fa-exclamation-triangle me-2"),
                                f"❌ Error: {str(e)}"], color="danger", className="mt-2")


@app.callback(
    [Output('cf-total-commitment', 'children'), Output('cf-total-paid-in', 'children'),
     Output('cf-total-unfunded', 'children'), Output('cf-current-nav', 'children')],
    [Input('deals-store', 'data'), Input('fund-cf-data-store', 'data')]
)
def update_cf_summary(deals, uploaded_data):
    if uploaded_data:
        tc = sum(d.get('commitment', 0) for d in uploaded_data)
        pi = sum(d.get('current_commitment', tc * 0.75) for d in uploaded_data)
        uf = tc - pi
        nav = sum(d.get('nav', 0) for d in uploaded_data)
    elif deals:
        tc = sum(d.get('total_commitment', d.get('commitment', d.get('size', 0))) for d in deals)
        pi = sum(d.get('current_commitment', d.get('size', 0)) for d in deals)
        uf = tc - pi
        nav = sum(d.get('nav', d.get('size', 0)) for d in deals)
    else:
        return "$0.0M", "$0.0M", "$0.0M", "$0.0M"
    return f"${tc:.1f}M", f"${pi:.1f}M", f"${uf:.1f}M", f"${nav:.1f}M"


@app.callback(Output('cf-monthly-table', 'children'),
              [Input('deals-store', 'data'), Input('fund-cf-data-store', 'data'),
               Input('cf-view-period', 'value'), Input('cf-type-view', 'value')])
def generate_monthly_cf_table(deals, uploaded_data, num_months, cf_type):
    data_source = uploaded_data if uploaded_data else deals
    if not data_source:
        return html.P("Upload Fund Level CF Excel file or add deals.", className="text-muted")

    base_date = datetime(2026, 1, 1)
    months = [(base_date + relativedelta(months=i)).strftime('%b-%y') for i in range(num_months)]

    rows = []
    for deal in data_source:
        row = {'Deal': deal.get('name', ''), 'Type': deal.get('type', deal.get('strategy', '')),
               'Commitment': f"${deal.get('commitment', deal.get('total_commitment', 0)):.1f}M",
               'NAV': f"${deal.get('nav', 0):.1f}M"}
        for i, month in enumerate(months):
            cf_value = deal.get('monthly_navs', deal.get('monthly_cfs', {})).get(i, 0)
            row[month] = f"${cf_value:.2f}M" if cf_value != 0 else "—"
        rows.append(row)

    if not rows:
        return html.P("No data", className="text-muted")

    cols = ['Deal', 'Type', 'Commitment', 'NAV'] + months
    return dash_table.DataTable(data=rows, columns=[{"name": c, "id": c} for c in cols],
                                style_cell={'textAlign': 'left', 'padding': '8px', 'fontFamily': C['mono'],
                                            'fontSize': '11px', 'minWidth': '80px'},
                                style_header={'backgroundColor': C['surface'], 'color': C['text'], 'fontWeight': 'bold',
                                              'position': 'sticky', 'top': 0},
                                style_data={'backgroundColor': C['panel'], 'color': C['text']},
                                style_data_conditional=[{'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']},
                                                        {'if': {'column_id': ['Deal', 'Type']}, 'fontWeight': 'bold'}],
                                style_table={'overflowX': 'auto', 'maxHeight': '600px', 'overflowY': 'auto'},
                                fixed_columns={'headers': True, 'data': 2}, fixed_rows={'headers': True},
                                export_format='xlsx', export_headers='display')


@app.callback(Output('cf-monthly-chart', 'figure'),
              [Input('deals-store', 'data'), Input('cf-view-period', 'value')])
def generate_monthly_cf_chart(deals, num_months):
    base_date = datetime(2026, 1, 1)
    months = [(base_date + relativedelta(months=i)).strftime('%b %Y') for i in range(num_months)]
    nav = [sum(d.get('nav', d.get('size', 0)) for d in deals) if deals else 0] * num_months
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=months, y=nav, name='Ending NAV', line=dict(color=C['blue'], width=3)))
    fig.update_layout(**CHART_BASE, height=400, yaxis_title='NAV ($mm)', hovermode='x unified')
    return fig


# ── Pro Forma ─────────────────────────────────────────────────────────────────
@app.callback(Output('pf-tab-content', 'children'),
              [Input('pf-tabs', 'active_tab'), Input('deals-store', 'data'),
               Input('pipeline-store', 'data'), Input('pf-target-month', 'value'),
               Input('cashflows-store', 'data')])
def render_pf_tab_content(active_tab, deals, pipeline, target_month, cashflows):
    if active_tab == "tab-pf-portfolio":
        return render_complete_portfolio(deals, pipeline, target_month, cashflows)
    elif active_tab == "tab-pf-metrics":
        return render_metrics_comparison(deals, pipeline)
    elif active_tab == "tab-pf-charts":
        return render_impact_charts(deals, pipeline)
    return html.Div()


def render_complete_portfolio(deals, pipeline, target_month, cashflows):
    month_names = generate_month_options()
    target_month_name = month_names[target_month]['label'] if target_month < len(month_names) else "Unknown"
    all_deals = []
    for deal in (deals or []):
        dc = deal.copy();
        dc['source'] = 'Current Portfolio';
        dc['status'] = '✓ Active'
        dc['nav_at_target'] = deal.get('nav', deal.get('size', 0))
        all_deals.append(dc)
    for p in (pipeline or []):
        all_deals.append({'name': p['name'], 'strategy': p['type'], 'size': p['size'],
                          'target_irr': p['target_irr'], 'hold_period': 5.0, 'moic': (1 + p['target_irr']) ** 5,
                          'vintage': 2026, 'sector': 'TBD', 'geography': 'Global',
                          'source': 'Pipeline', 'status': '⏳ Pending', 'nav_at_target': p['size']})
    if not all_deals:
        return dbc.Alert("No deals to display.", color="info")

    table_data = [{'Deal': d['name'], 'Strategy': d['strategy'], 'Size': f"${d['size']:.1f}M",
                   'IRR': f"{d['target_irr']:.1%}", 'MOIC': f"{d.get('moic', 0):.2f}x",
                   'Vintage': str(d.get('vintage', 'N/A')), 'Source': d['source'],
                   f'NAV @ {target_month_name}': f"${d['nav_at_target']:.1f}M"} for d in all_deals]

    summary = dbc.Row([
        dbc.Col(dbc.Card([dbc.CardBody([html.H6("Current NAV"),
                                        html.H4(f"${sum(d.get('nav', d.get('size', 0)) for d in (deals or [])):.1f}M",
                                                style={'color': C['green']})])]), width=3),
        dbc.Col(dbc.Card([dbc.CardBody([html.H6("Pipeline NAV"),
                                        html.H4(f"${sum(p['size'] for p in (pipeline or [])):.1f}M",
                                                style={'color': C['blue']})])]), width=3),
        dbc.Col(dbc.Card([dbc.CardBody([html.H6(f"Pro Forma @ {target_month_name}"),
                                        html.H4(f"${sum(d['nav_at_target'] for d in all_deals):.1f}M",
                                                style={'color': C['amber']})])]), width=3),
        dbc.Col(dbc.Card(
            [dbc.CardBody([html.H6("Total Deals"), html.H4(str(len(all_deals)), style={'color': C['purple']})])]),
                width=3),
    ], className="mb-4")

    return html.Div([summary, dbc.Card([
        dbc.CardHeader(f"Complete Pro Forma Portfolio ({len(all_deals)} deals)"),
        dbc.CardBody([dash_table.DataTable(data=table_data,
                                           columns=[{"name": c, "id": c} for c in
                                                    table_data[0].keys()] if table_data else [],
                                           style_cell={'textAlign': 'left', 'padding': '12px', 'fontFamily': C['mono'],
                                                       'fontSize': '12px'},
                                           style_header={'backgroundColor': C['surface'], 'color': C['text'],
                                                         'fontWeight': 'bold'},
                                           style_data={'backgroundColor': C['panel'], 'color': C['text']},
                                           style_data_conditional=[
                                               {'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']},
                                               {'if': {'column_id': 'Source', 'filter_query': '{Source} = "Pipeline"'},
                                                'color': C['blue'], 'fontWeight': 'bold'},
                                               {'if': {'column_id': 'Source',
                                                       'filter_query': '{Source} = "Current Portfolio"'},
                                                'color': C['green'], 'fontWeight': 'bold'},
                                           ], page_size=100, sort_action='native', filter_action='native')])
    ])])


def render_metrics_comparison(deals, pipeline):
    current_m = calculate_portfolio_metrics(deals)
    if pipeline:
        pf_m = calculate_portfolio_metrics((deals or []) + [
            {'name': p['name'], 'strategy': p['type'], 'size': p['size'], 'nav': p['size'],
             'target_irr': p['target_irr'], 'hold_period': 5.0, 'moic': (1 + p['target_irr']) ** 5,
             'vintage': 2026, 'sector': 'TBD', 'geography': 'Global'} for p in pipeline])
    else:
        pf_m = current_m
    rows = [
        {"Metric": "Total NAV ($mm)", "Current": f"${current_m['total_nav']:.1f}",
         "Pro Forma": f"${pf_m['total_nav']:.1f}", "Change": f"+${pf_m['total_nav'] - current_m['total_nav']:.1f}"},
        {"Metric": "Total Commitment ($mm)", "Current": f"${current_m['total_commitment']:.1f}",
         "Pro Forma": f"${pf_m['total_commitment']:.1f}", "Change": "—"},
        {"Metric": "Total Unfunded ($mm)", "Current": f"${current_m['total_unfunded']:.1f}",
         "Pro Forma": f"${pf_m['total_unfunded']:.1f}", "Change": "—"},
        {"Metric": "Weighted IRR", "Current": f"{current_m['weighted_irr']:.2%}",
         "Pro Forma": f"{pf_m['weighted_irr']:.2%}",
         "Change": f"{pf_m['weighted_irr'] - current_m['weighted_irr']:+.2%}"},
        {"Metric": "Number of Deals", "Current": str(current_m['num_deals']), "Pro Forma": str(pf_m['num_deals']),
         "Change": f"+{pf_m['num_deals'] - current_m['num_deals']}"},
        {"Metric": "Top 1 Concentration", "Current": f"{current_m['concentration_top1']:.1%}",
         "Pro Forma": f"{pf_m['concentration_top1']:.1%}",
         "Change": f"{pf_m['concentration_top1'] - current_m['concentration_top1']:+.1%}"},
    ]
    return dbc.Card([dbc.CardHeader("Before vs After Comparison"),
                     dbc.CardBody([dash_table.DataTable(data=rows, columns=[{"name": c, "id": c} for c in
                                                                            ["Metric", "Current", "Pro Forma",
                                                                             "Change"]],
                                                        style_cell={'textAlign': 'left', 'padding': '12px',
                                                                    'fontFamily': C['mono'], 'fontSize': '13px'},
                                                        style_header={'backgroundColor': C['surface'],
                                                                      'color': C['text'], 'fontWeight': 'bold'},
                                                        style_data={'backgroundColor': C['panel'], 'color': C['text']},
                                                        style_data_conditional=[
                                                            {'if': {'column_id': 'Change'}, 'fontWeight': 'bold',
                                                             'color': C['blue']}])])])


def render_impact_charts(deals, pipeline):
    cm = calculate_portfolio_metrics(deals)
    pf_deals = (deals or []) + [{'name': p['name'], 'strategy': p['type'], 'size': p['size'], 'nav': p['size'],
                                 'target_irr': p['target_irr'], 'hold_period': 5.0, 'moic': (1 + p['target_irr']) ** 5,
                                 'vintage': 2026, 'sector': 'TBD', 'geography': 'Global'} for p in (pipeline or [])]
    pm = calculate_portfolio_metrics(pf_deals)
    fig_s = go.Figure()
    if cm['by_strategy']:
        fig_s.add_trace(go.Bar(name='Current', x=list(cm['by_strategy'].keys()),
                               y=[cm['by_strategy'][s]['nav'] for s in cm['by_strategy']], marker_color=C['blue']))
    if pm['by_strategy']:
        fig_s.add_trace(go.Bar(name='Pro Forma', x=list(pm['by_strategy'].keys()),
                               y=[pm['by_strategy'][s]['nav'] for s in pm['by_strategy']], marker_color=C['green']))
    fig_s.update_layout(**CHART_BASE, barmode='group', yaxis_title="NAV ($mm)", height=400)
    cats = ['Top 1', 'Top 3', 'Top 5']
    fig_c = go.Figure()
    fig_c.add_trace(go.Bar(name='Current', x=cats, y=[cm['concentration_top1'] * 100, cm['concentration_top3'] * 100,
                                                      cm['concentration_top5'] * 100], marker_color=C['blue']))
    fig_c.add_trace(go.Bar(name='Pro Forma', x=cats, y=[pm['concentration_top1'] * 100, pm['concentration_top3'] * 100,
                                                        pm['concentration_top5'] * 100], marker_color=C['green']))
    fig_c.update_layout(**CHART_BASE, barmode='group', yaxis_title="% of NAV", height=400)
    return dbc.Row([
        dbc.Col([dbc.Card([dbc.CardHeader("Strategy Allocation"),
                           dbc.CardBody([dcc.Graph(figure=fig_s, config={'displayModeBar': False})])])], width=6),
        dbc.Col([dbc.Card([dbc.CardHeader("Concentration Risk"),
                           dbc.CardBody([dcc.Graph(figure=fig_c, config={'displayModeBar': False})])])], width=6),
    ])


# ── Liquidity callbacks ───────────────────────────────────────────────────────
@app.callback(
    [Output('liquidity-data-store', 'data'), Output('upload-liquidity-status', 'children')],
    Input('upload-liquidity-pull', 'contents'),
    State('upload-liquidity-pull', 'filename')
)
def upload_liquidity_file(contents, filename):
    if contents is None:
        return None, ""
    try:
        content_type, content_string = contents.split(',')
        decoded = base64.b64decode(content_string)
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(decoded), data_only=True)
        ws = next((wb[n] for n in ['Liquidity Pull', 'Liquidity_Pull', 'LiquidityPull'] if n in wb.sheetnames),
                  wb[wb.sheetnames[0]])
        liq_data = {'as_at_date': None, 'current_quarter': None, 'current_month': None,
                    'fund_nav': 0, 'current_cash': 0, 'glf_balance': 0, 'cqs_balance': 0,
                    'total_liquidity': 0, 'near_term_flows': {}, 'nav_projections': {}}
        for k, (r, c) in [('as_at_date', (4, 2)), ('current_quarter', (5, 2)), ('current_month', (6, 2))]:
            v = ws.cell(r, c).value
            if v: liq_data[k] = str(v)
        for k, (r, c) in [('fund_nav', (10, 3)), ('current_cash', (58, 3)), ('glf_balance', (59, 3)),
                          ('cqs_balance', (60, 3)), ('total_liquidity', (16, 3))]:
            v = ws.cell(r, c).value
            if v and isinstance(v, (int, float)):
                liq_data[k] = float(v) / 1_000_000
        return liq_data, dbc.Alert([html.I(className="fas fa-check-circle me-2"), f"✅ {filename} uploaded"],
                                   color="success", className="mt-2")
    except Exception as e:
        return None, dbc.Alert(f"❌ Error: {str(e)}", color="danger", className="mt-2")


@app.callback(
    [Output('liq-today-date', 'children'), Output('liq-as-at-date', 'children'),
     Output('liq-current-quarter', 'children'), Output('liq-current-month', 'children')],
    Input('liquidity-data-store', 'data')
)
def update_liq_dates(data):
    today = datetime.now()
    if data:
        return today.strftime('%Y-%m-%d'), data.get('as_at_date', 'N/A'), data.get('current_quarter', 'N/A'), data.get(
            'current_month', 'N/A')
    as_at = (today.replace(day=1) - relativedelta(days=1))
    return today.strftime('%Y-%m-%d'), as_at.strftime('%Y-%m-%d'), 'Q' + str(
        (today.month - 1) // 3 + 1), today.strftime('%b %Y')


@app.callback(Output('liq-waterfall-table', 'children'),
              [Input('deals-store', 'data'), Input('liquidity-data-store', 'data')])
def generate_liquidity_waterfall(deals, data):
    nav = data.get('fund_nav', 0) if data else sum(d.get('nav', d.get('size', 0)) for d in deals) if deals else 0
    cash = data.get('current_cash', 0) if data else 0
    glf = data.get('glf_balance', 0) if data else 0
    cqs = data.get('cqs_balance', 0) if data else 0
    total = cash + glf + cqs
    rows = [
        {'Source': 'Fund NAV', 'Item': 'Total Current NAV', 'Amount ($mm)': f'{nav:.2f}'},
        {'Source': 'Cash', 'Item': 'Cash Balance', 'Amount ($mm)': f'{cash:.2f}'},
        {'Source': 'GLF', 'Item': 'GLF Balance', 'Amount ($mm)': f'{glf:.2f}'},
        {'Source': 'CQS', 'Item': 'CQS Balance', 'Amount ($mm)': f'{cqs:.2f}'},
        {'Source': 'Total Liquidity', 'Item': 'Total Liquidity Balance', 'Amount ($mm)': f'{total:.2f}'},
    ]
    return dash_table.DataTable(data=rows, columns=[{"name": c, "id": c} for c in rows[0].keys()],
                                style_cell={'textAlign': 'left', 'padding': '12px', 'fontFamily': C['mono'],
                                            'fontSize': '12px'},
                                style_header={'backgroundColor': C['surface'], 'color': C['text'],
                                              'fontWeight': 'bold'},
                                style_data={'backgroundColor': C['panel'], 'color': C['text']},
                                style_data_conditional=[
                                    {'if': {'column_id': 'Source', 'filter_query': '{Source} = "Total Liquidity"'},
                                     'fontWeight': 'bold', 'color': C['green']}])


@app.callback(Output('liq-near-term-flows', 'children'),
              [Input('deals-store', 'data'), Input('liquidity-data-store', 'data')])
def generate_near_term_flows(deals, data):
    base = datetime(2026, 3, 31)
    months = [(base + relativedelta(months=i)).strftime('%b-%y') for i in range(12)]
    rows = [
        {'Flow Type': 'Subscriptions', **{m: '$50.0M' if i < 3 else '—' for i, m in enumerate(months)}},
        {'Flow Type': 'Redemptions', **{m: '—' for m in months}},
        {'Flow Type': 'Portfolio Net Flows', **{m: f'${2.6 + i * 0.5:.1f}M' for i, m in enumerate(months)}},
    ]
    return dash_table.DataTable(data=rows, columns=[{"name": c, "id": c} for c in ['Flow Type'] + months],
                                style_cell={'textAlign': 'left', 'padding': '10px', 'fontFamily': C['mono'],
                                            'fontSize': '11px', 'minWidth': '100px'},
                                style_header={'backgroundColor': C['surface'], 'color': C['text'],
                                              'fontWeight': 'bold'},
                                style_data={'backgroundColor': C['panel'], 'color': C['text']},
                                style_data_conditional=[{'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']},
                                                        {'if': {'column_id': 'Flow Type'}, 'fontWeight': 'bold'}],
                                style_table={'overflowX': 'auto'}, fixed_columns={'headers': True, 'data': 1})


@app.callback(Output('liq-nav-projection-chart', 'figure'),
              [Input('deals-store', 'data'), Input('liquidity-data-store', 'data')])
def generate_nav_projection_chart(deals, data):
    base = datetime(2026, 3, 31)
    base_nav = sum(d.get('nav', d.get('size', 0)) for d in deals) if deals else 100
    months = [(base + relativedelta(months=i)).strftime('%b %Y') for i in range(12)]
    navs = [base_nav * (1 + 0.13 / 12) ** i for i in range(12)]
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=months, y=navs, mode='lines+markers', line=dict(color=C['blue'], width=3),
                             fill='tozeroy', fillcolor=rgba(C['blue'], 0.2)))
    fig.update_layout(**CHART_BASE, height=400, yaxis_title='NAV ($mm)', hovermode='x unified')
    return fig


@app.callback(Output('liq-tab-content', 'children'), Input('liq-tabs', 'active_tab'))
def render_liq_tab_content(active_tab):
    contents = {
        "tab-liq-subs": [html.H6("Subscriptions & Redemptions"), html.P("Expected subscriptions: $150M (confirmed)"),
                         html.P("Expected redemptions: $0M")],
        "tab-liq-flows": [html.H6("Portfolio Net Flows"), html.P("Seed Portfolio Net Flows from deal-level cashflows")],
        "tab-liq-unfunded": [html.H6("Unfunded Commitments"), html.P("Current unfunded commitments tracked per deal")],
    }
    return dbc.Card([dbc.CardBody(contents.get(active_tab, []))])


# ── Analytics ─────────────────────────────────────────────────────────────────
def get_portfolio_for_analytics_view(view, deals, pipeline, placeholders):
    base = deals if deals else []
    if view == 'current':
        return base
    extra_pipe = [{'name': p['name'], 'size': p['size'], 'nav': p['size'], 'strategy': p.get('type', 'GP-Led'),
                   'geography': p.get('region', 'North America'), 'sector': 'Technology',
                   'vintage': 2026, 'target_irr': p.get('target_irr', 0), 'deal_type': p.get('deal_type', 'Secondary')}
                  for p in (pipeline if pipeline else [])]
    if view == 'current_pipeline':
        return base + extra_pipe
    extra_ph = [{'name': p['name'], 'size': p['size'], 'nav': p['size'], 'strategy': p.get('strategy', 'GP-Led'),
                 'geography': p.get('region', 'North America'), 'sector': 'Technology',
                 'vintage': 2026, 'target_irr': 0.25, 'deal_type': p.get('deal_type', 'Secondary')}
                for p in (placeholders if placeholders else [])]
    return base + extra_pipe + extra_ph


@app.callback(
    [Output('analytics-total-nav', 'children'), Output('analytics-num-deals', 'children'),
     Output('analytics-weighted-irr', 'children'), Output('analytics-top1-conc', 'children')],
    [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
     Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')]
)
def update_analytics_summary(view, deals, pipeline, placeholders):
    pd = get_portfolio_for_analytics_view(view, deals, pipeline, placeholders)
    if not pd: return "$0.0M", "0", "0.0%", "0.0%"
    tnav = sum(d.get('nav', d.get('size', 0)) for d in pd)
    wirr = sum(d.get('target_irr', 0) * d.get('nav', d.get('size', 0)) for d in pd) / tnav if tnav > 0 else 0
    top1 = max(d.get('nav', d.get('size', 0)) for d in pd) / tnav if tnav > 0 else 0
    return f"${tnav:.1f}M", str(len(pd)), f"{wirr:.1%}", f"{top1:.1%}"


def make_pie(by_key, pd_list, colors_list):
    by_k = {}
    for d in pd_list:
        k = d.get(by_key, 'Other')
        by_k[k] = by_k.get(k, 0) + d.get('nav', d.get('size', 0))
    if not by_k:
        fig = go.Figure();
        fig.update_layout(**CHART_BASE, height=350);
        return fig
    labels, values = zip(*by_k.items())
    fig = go.Figure(data=[go.Pie(labels=list(labels), values=list(values), hole=0.4,
                                 marker=dict(colors=colors_list[:len(labels)]), textfont=dict(color=C['text']))])
    fig.update_layout(**CHART_BASE, height=350, margin=dict(t=20, b=0, l=0, r=0))
    return fig


@app.callback(Output('analytics-strategy', 'figure'),
              [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
               Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')])
def update_strategy_chart(v, d, p, ph):
    return make_pie('strategy', get_portfolio_for_analytics_view(v, d, p, ph),
                    [C['blue'], C['purple'], C['teal'], C['green']])


@app.callback(Output('analytics-region', 'figure'),
              [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
               Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')])
def update_region_chart(v, d, p, ph):
    pd = get_portfolio_for_analytics_view(v, d, p, ph)
    if not pd: fig = go.Figure(); fig.update_layout(**CHART_BASE); return fig
    by_r = {}
    for deal in pd:
        r = deal.get('geography', deal.get('region', 'North America'))
        by_r[r] = by_r.get(r, 0) + deal.get('nav', deal.get('size', 0))
    fig = go.Figure(data=[go.Bar(x=list(by_r.keys()), y=list(by_r.values()), marker_color=C['blue'],
                                 text=[f"${v:.1f}M" for v in by_r.values()], textposition='outside')])
    fig.update_layout(**CHART_BASE, yaxis_title="NAV ($mm)", height=350)
    return fig


@app.callback(Output('analytics-vintage', 'figure'),
              [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
               Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')])
def update_vintage_chart(v, d, p, ph):
    pd = get_portfolio_for_analytics_view(v, d, p, ph)
    if not pd: fig = go.Figure(); fig.update_layout(**CHART_BASE); return fig
    by_v = {}
    for deal in pd:
        vt = deal.get('vintage', 2024)
        by_v[vt] = by_v.get(vt, 0) + deal.get('nav', deal.get('size', 0))
    labels = [str(k) for k in sorted(by_v.keys())]
    fig = go.Figure(data=[go.Bar(x=labels, y=[by_v[int(k)] for k in labels], marker_color=C['purple'],
                                 text=[f"${by_v[int(k)]:.1f}M" for k in labels], textposition='outside')])
    fig.update_layout(**CHART_BASE, yaxis_title="NAV ($mm)", height=350)
    return fig


@app.callback(Output('analytics-dealtype', 'figure'),
              [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
               Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')])
def update_dealtype_chart(v, d, p, ph):
    pd = get_portfolio_for_analytics_view(v, d, p, ph)
    sec = sum(deal.get('nav', deal.get('size', 0)) for deal in pd if deal.get('deal_type', 'Secondary') == 'Secondary')
    co = sum(
        deal.get('nav', deal.get('size', 0)) for deal in pd if deal.get('deal_type', 'Secondary') == 'Co-Investment')
    fig = go.Figure(data=[go.Pie(labels=['Secondary', 'Co-Investment'], values=[sec, co], hole=0.4,
                                 marker=dict(colors=[C['blue'], C['green']]), textfont=dict(color=C['text']))])
    fig.update_layout(**CHART_BASE, height=350, margin=dict(t=20, b=0, l=0, r=0))
    return fig


@app.callback(Output('analytics-sector', 'figure'),
              [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
               Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')])
def update_sector_chart(v, d, p, ph):
    return make_pie('sector', get_portfolio_for_analytics_view(v, d, p, ph),
                    [C['teal'], C['purple'], C['amber'], C['green'], C['pink']])


@app.callback(Output('analytics-concentration', 'figure'),
              [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
               Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')])
def update_concentration_chart(v, d, p, ph):
    pd = get_portfolio_for_analytics_view(v, d, p, ph)
    if not pd: fig = go.Figure(); fig.update_layout(**CHART_BASE); return fig
    tnav = sum(deal.get('nav', deal.get('size', 0)) for deal in pd)
    sd = sorted(pd, key=lambda x: x.get('nav', x.get('size', 0)), reverse=True)
    top1 = sd[0].get('nav', sd[0].get('size', 0)) / tnav * 100 if sd and tnav > 0 else 0
    top3 = sum(d.get('nav', d.get('size', 0)) for d in sd[:3]) / tnav * 100 if len(sd) >= 3 and tnav > 0 else 0
    top5 = sum(d.get('nav', d.get('size', 0)) for d in sd[:5]) / tnav * 100 if len(sd) >= 5 and tnav > 0 else 0
    vals = [top1, top3, top5];
    limits = [15, 40, 60]
    fig = go.Figure(data=[go.Bar(x=['Top 1', 'Top 3', 'Top 5'], y=vals,
                                 marker_color=[C['red'] if v > l else C['green'] for v, l in zip(vals, limits)],
                                 text=[f"{v:.1f}%" for v in vals], textposition='outside')])
    for limit in limits:
        fig.add_hline(y=limit, line_dash='dash', line_color=C['red'], opacity=0.5)
    fig.update_layout(**CHART_BASE, yaxis_title="% of NAV", height=350)
    return fig


# ── Dashboard activity & pipeline ────────────────────────────────────────────
@app.callback([Output('dash-recent-activity', 'children'), Output('dash-pipeline-status', 'children')],
              [Input('deals-store', 'data'), Input('pipeline-store', 'data')])
def update_dashboard_activity(deals, pipeline):
    if not deals:
        recent = html.P("No deals yet", className="text-muted")
    else:
        items = []
        for d in sorted(deals, key=lambda x: x.get('date_added', ''), reverse=True)[:5]:
            items.append(html.Div([
                html.Strong(d['name'], style={'color': C['blue'], 'fontFamily': C['mono']}), html.Br(),
                html.Small(f"${d.get('nav', d.get('size', 0)):.1f}M NAV • {d['target_irr']:.1%} IRR",
                           style={'color': C['muted'], 'fontFamily': C['mono']}),
                html.Hr(style={'borderColor': C['border'], 'margin': '0.5rem 0'})
            ]))
        recent = html.Div(items)

    if not pipeline:
        pipe_status = html.P("No pipeline deals", className="text-muted")
    else:
        stage_counts = {}
        for p in pipeline:
            stage_counts[p.get('stage', 'Screening')] = stage_counts.get(p.get('stage', 'Screening'), 0) + 1
        stage_colors = {'Screening': C['muted'], 'Due Diligence': C['amber'], 'Term Sheet': C['blue'],
                        'Final Docs': C['green']}
        pipe_status = html.Div([html.Div([
            html.Span("●", style={'color': stage_colors.get(s, C['muted']), 'marginRight': '0.5rem'}),
            html.Strong(f"{s}: "), html.Span(f"{c} deals", style={'color': C['muted']})
        ], style={'marginBottom': '0.5rem'}) for s, c in stage_counts.items()])

    return recent, pipe_status


# ── Export ─────────────────────────────────────────────────────────────────────
@app.callback(Output("download-csv", "data"),
              Input("btn-export", "n_clicks"),
              [State("deals-store", "data"), State("placeholder-deals-store", "data"), State("pipeline-store", "data")],
              prevent_initial_call=True)
def export_all(n, deals, placeholders, pipeline):
    if deals:
        return dcc.send_data_frame(pd.DataFrame(deals).to_csv, f"portfolio_{date.today()}.csv", index=False)


# ── Auto-save ─────────────────────────────────────────────────────────────────
@app.callback(Output('save-status', 'children'),
              [Input('deals-store', 'data'), Input('pipeline-store', 'data'),
               Input('placeholder-deals-store', 'data'), Input('config-store', 'data')],
              prevent_initial_call=True)
def auto_save(deals, pipeline, placeholders, config):
    save_data(deals or [], pipeline or [], placeholders or [], config or DEFAULT_CONFIG)
    return ""


# ── Settings ──────────────────────────────────────────────────────────────────
@app.callback(Output('config-store', 'data', allow_duplicate=True),
              Input('btn-save-settings', 'n_clicks'),
              [State('set-dry-powder', 'value'), State('set-twr', 'value'), State('set-hold', 'value'),
               State('set-fee', 'value'), State('set-carry', 'value'), State('set-hurdle', 'value'),
               State('set-liq', 'value'), State('set-loss', 'value'), State('set-cash-yield', 'value'),
               State('config-store', 'data')],
              prevent_initial_call='initial_duplicate')
def save_settings(n, dp, twr, hold, fee, carry, hurdle, liq, loss, cash_yield, config):
    if not n: return config
    config['fund_parameters'].update({
        'dry_powder': float(dp or 450), 'target_net_twr': float(twr or 13) / 100,
        'avg_hold_period': float(hold or 5), 'management_fee': float(fee or 1.25) / 100,
        'carry_rate': float(carry or 12.5) / 100, 'hurdle_rate': float(hurdle or 10) / 100,
        'liquidity_reserve_pct': float(liq or 5) / 100, 'loss_drag': float(loss or 1) / 100,
        'cash_yield': float(cash_yield or 3) / 100,
    })
    return config


server = app.server

if __name__ == '__main__':
    print("\n" + "=" * 80)
    print("HORIZON PORTFOLIO TOOL — Enhanced v2")
    print("=" * 80)
    print("\n✅ Starting on http://localhost:8050")
    print("\n🆕 Key Enhancements:")
    print("   • Total Commitment / Current Commitment / Unfunded tracked per deal")
    print("   • Called % utilisation bar on each portfolio card")
    print("   • TVPI = NAV / Current Commitment displayed per deal")
    print("   • NAV auto-pulled from latest CF month when Fund Level CF is uploaded")
    print("   • CF pull shows source month in Edit modal with green/amber banner")
    print("   • Portfolio summary shows total commitments, unfunded exposure")
    print("   • CF data upload banner with link in portfolio page header")
    print("\nPress CTRL+C to stop\n")
    app.run(debug=True, host='0.0.0.0', port=8050)
