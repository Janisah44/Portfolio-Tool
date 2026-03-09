"""
HORIZON PORTFOLIO TOOL - INSTITUTIONAL GRADE LP MANAGEMENT
Comprehensive evergreen fund portfolio management and analytics platform

Features:
✓ Current Portfolio Management (add/delete deals)
✓ Pro Forma Portfolio (placeholder/future deals)
✓ Dry Powder Forecasting & Bite Sizing
✓ Pipeline Management
✓ Return Calculator with Waterfall
✓ Pacing Model with Multi-Year Forecasts
✓ Portfolio Construction Analytics
✓ Deal Sizing Recommendations (Min/Desired/Max)

Run: python ULTIMATE_Evergreen_Fund_Dashboard.py
Open: http://localhost:8050
"""

import dash
from dash import dcc, html, dash_table, Input, Output, State, callback_context, ALL
import dash_bootstrap_components as dbc
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import pandas as pd
import numpy as np
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import json
import base64
from io import BytesIO
import pickle
import os

# ==================== DATA PERSISTENCE ====================

DATA_FILE = 'portfolio_data.pkl'


def save_data(deals, pipeline, placeholders, config):
    """Save all data to disk"""
    data = {
        'deals': deals,
        'pipeline': pipeline,
        'placeholders': placeholders,
        'config': config,
        'saved_at': datetime.now().isoformat()
    }
    try:
        with open(DATA_FILE, 'wb') as f:
            pickle.dump(data, f)
        print(f"✅ Data saved successfully at {datetime.now().strftime('%H:%M:%S')}")
    except Exception as e:
        print(f"⚠️ Error saving data: {e}")


def load_data():
    """Load data from disk"""
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'rb') as f:
                data = pickle.load(f)
            print(f"✅ Data loaded from {data.get('saved_at', 'unknown time')}")
            return data
        except Exception as e:
            print(f"⚠️ Error loading data: {e}")
            return None
    return None


# ==================== COLORS ====================

C = dict(
    bg="#07090f", panel="#0d1117", surface="#111822",
    border="#1c2a3a", border2="#243649",
    blue="#2979c8", sky="#56b4f5", teal="#2ec4b6",
    green="#2ecc71", red="#e5493a", amber="#f0a500",
    purple="#9b72cf", pink="#e05c8a",
    text="#d4e6f5", muted="#5e82a0", dim="#324d63",
    mono="'JetBrains Mono', 'Fira Mono', monospace",
    sans="'IBM Plex Sans', 'Segoe UI', sans-serif",
)


def rgba(hex_color, alpha):
    r, g, b = int(hex_color[1:3], 16), int(hex_color[3:5], 16), int(hex_color[5:7], 16)
    return f"rgba({r},{g},{b},{alpha})"


CHART_BASE = dict(
    paper_bgcolor=C["panel"], plot_bgcolor=C["surface"],
    font=dict(family=C["mono"], color=C["text"], size=11),
    margin=dict(l=52, r=24, t=40, b=40),
    xaxis=dict(gridcolor=C["border"], zerolinecolor=C["border"]),
    yaxis=dict(gridcolor=C["border"], zerolinecolor=C["border"]),
    legend=dict(bgcolor=C["panel"], bordercolor=C["border"], borderwidth=1),
)

# Legacy color mappings for compatibility
COLORS = {
    'primary': C['blue'],
    'secondary': C['purple'],
    'accent': C['amber'],
    'success': C['green'],
    'danger': C['red'],
    'dark': C['bg'],
    'light': C['panel'],
    'warning': C['amber'],
    'info': C['sky']
}


# ==================== HELPER FUNCTIONS ====================

# ==================== HELPER FUNCTIONS ====================

def generate_month_options():
    """Generate month options from Jan 2026 to Dec 2040 (180 months)"""
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    options = []
    for i in range(180):  # 15 years * 12 months
        month_name = month_names[i % 12]
        year = 2026 + (i // 12)
        options.append({"label": f"{month_name} {year}", "value": i})
    return options


def calculate_required_future_irr(current_portfolio_irr, current_nav, dry_powder, config):
    """Calculate IRR needed on future deals to hit target net TWR"""
    target_twr = config['fund_parameters']['target_net_twr']
    mgmt_fee = config['fund_parameters']['management_fee']
    carry_rate = config['fund_parameters']['carry_rate']
    hurdle = config['fund_parameters']['hurdle_rate']
    loss_drag = config['fund_parameters']['loss_drag']
    cash_reserve = config['fund_parameters']['liquidity_reserve_pct']
    cash_yield = config['fund_parameters'].get('cash_yield', 0.03)

    # Calculate invested ratio
    total_fund = current_nav + dry_powder
    if total_fund == 0:
        return 0.25

    invested_ratio = max(0, 1 - cash_reserve)
    idle_ratio = cash_reserve

    # Gross return needed before carry
    gross_needed = (target_twr + mgmt_fee + loss_drag - (idle_ratio * cash_yield)) / invested_ratio

    # Add carry drag if above hurdle
    if gross_needed > hurdle:
        carry_drag = (gross_needed - hurdle) * carry_rate
        gross_needed += carry_drag

    # Calculate weighted average
    if dry_powder == 0:
        return gross_needed

    current_weight = current_nav / total_fund
    future_weight = dry_powder / total_fund

    required_future = (gross_needed - (current_weight * current_portfolio_irr)) / future_weight

    return max(0, min(1.0, required_future))


def calculate_portfolio_metrics(deals):
    """Calculate comprehensive portfolio metrics"""
    if not deals:
        return {
            'total_nav': 0,
            'num_deals': 0,
            'weighted_irr': 0,
            'by_strategy': {},
            'by_vintage': {},
            'by_sector': {},
            'concentration_top1': 0,
            'concentration_top3': 0,
            'concentration_top5': 0,
            'effective_n': 0
        }

    total_nav = sum(d['size'] for d in deals)
    num_deals = len(deals)
    weighted_irr = sum(d['size'] * d['target_irr'] for d in deals) / total_nav if total_nav > 0 else 0

    # Effective number of positions (Herfindahl)
    weights = [d['size'] / total_nav for d in deals] if total_nav > 0 else []
    effective_n = 1 / sum(w ** 2 for w in weights) if weights else 0

    # By strategy
    by_strategy = {}
    for deal in deals:
        strategy = deal['strategy']
        if strategy not in by_strategy:
            by_strategy[strategy] = {'nav': 0, 'count': 0, 'deals': []}
        by_strategy[strategy]['nav'] += deal['size']
        by_strategy[strategy]['count'] += 1
        by_strategy[strategy]['deals'].append(deal)

    for strategy in by_strategy:
        nav = by_strategy[strategy]['nav']
        by_strategy[strategy]['weighted_irr'] = sum(
            d['size'] * d['target_irr'] for d in by_strategy[strategy]['deals']) / nav if nav > 0 else 0
        by_strategy[strategy]['allocation'] = nav / total_nav if total_nav > 0 else 0

    # By vintage
    by_vintage = {}
    for deal in deals:
        vintage = deal.get('vintage', 2024)
        if vintage not in by_vintage:
            by_vintage[vintage] = {'nav': 0, 'count': 0}
        by_vintage[vintage]['nav'] += deal['size']
        by_vintage[vintage]['count'] += 1

    # By sector
    by_sector = {}
    for deal in deals:
        sector = deal.get('sector', 'Other')
        if sector not in by_sector:
            by_sector[sector] = {'nav': 0, 'count': 0}
        by_sector[sector]['nav'] += deal['size']
        by_sector[sector]['count'] += 1

    # Concentration
    sorted_deals = sorted(deals, key=lambda x: x['size'], reverse=True)
    concentration_top1 = sorted_deals[0]['size'] / total_nav if len(sorted_deals) > 0 and total_nav > 0 else 0
    concentration_top3 = sum(d['size'] for d in sorted_deals[:3]) / total_nav if len(
        sorted_deals) >= 3 and total_nav > 0 else 0
    concentration_top5 = sum(d['size'] for d in sorted_deals[:5]) / total_nav if len(
        sorted_deals) >= 5 and total_nav > 0 else 0

    return {
        'total_nav': total_nav,
        'num_deals': num_deals,
        'weighted_irr': weighted_irr,
        'by_strategy': by_strategy,
        'by_vintage': by_vintage,
        'by_sector': by_sector,
        'concentration_top1': concentration_top1,
        'concentration_top3': concentration_top3,
        'concentration_top5': concentration_top5,
        'effective_n': effective_n
    }


def calculate_bite_sizes(dry_powder, config):
    """Calculate recommended deal sizes (min/desired/max) for each strategy"""
    bite_sizes = {}

    for strategy_name in ['GP-Led (Multi-Asset)', 'GP-Led (Single-Asset)', 'Diversified LP-Led', 'Co-Investments']:
        if 'Multi-Asset' in strategy_name:
            min_pct, desired_pct, max_pct = 0.005, 0.0275, 0.05
        elif 'Single-Asset' in strategy_name:
            min_pct, desired_pct, max_pct = 0.005, 0.0225, 0.04
        elif 'LP-Led' in strategy_name:
            min_pct, desired_pct, max_pct = 0.005, 0.0275, 0.05
        else:  # Co-Investments
            min_pct, desired_pct, max_pct = 0.005, 0.0175, 0.03

        bite_sizes[strategy_name] = {
            'min': dry_powder * min_pct,
            'desired': dry_powder * desired_pct,
            'max': dry_powder * max_pct,
            'min_pct': min_pct,
            'desired_pct': desired_pct,
            'max_pct': max_pct
        }

    return bite_sizes


def forecast_dry_powder(current_nav, dry_powder, deals, placeholder_deals, config, months=12):
    """Forecast dry powder availability over next N months starting from Jan 2026"""
    forecast = []
    nav = current_nav
    powder = dry_powder

    base_date = datetime(2026, 1, 1)

    for month in range(months):
        month_date = base_date + relativedelta(months=month)

        # Simple monthly growth on NAV
        monthly_return = config['fund_parameters']['target_net_twr'] / 12
        nav_growth = nav * monthly_return

        # Distributions (20% annually = 1.67% monthly)
        monthly_dist = nav * (config['fund_parameters']['distribution_rate'] / 12)

        # Capital calls from placeholder deals in this month
        calls_this_month = 0
        for pd_deal in placeholder_deals:
            if pd_deal.get('expected_month') == month:
                calls_this_month += pd_deal['size']

        # Update
        powder = powder + monthly_dist - calls_this_month
        nav = nav + nav_growth + calls_this_month - monthly_dist

        forecast.append({
            'month': month_date.strftime('%b %Y'),
            'dry_powder': powder,
            'nav': nav,
            'distributions': monthly_dist,
            'calls': calls_this_month
        })

    return forecast


# ==================== DASH APP ====================

app = dash.Dash(__name__,
                external_stylesheets=[dbc.themes.BOOTSTRAP, dbc.icons.FONT_AWESOME],
                suppress_callback_exceptions=True)

app.title = "Ultimate Evergreen Fund Manager"
server = app.server

# Custom CSS for dark theme
app.index_string = '''
<!DOCTYPE html>
<html>
    <head>
        {%metas%}
        <title>{%title%}</title>
        {%favicon%}
        {%css%}
        <style>
            @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500;600&display=swap');

            body {
                background-color: ''' + C['bg'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
                font-family: ''' + C['sans'] + ''' !important;
            }

            .card {
                background-color: ''' + C['panel'] + ''' !important;
                border: 1px solid ''' + C['border'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
            }

            .card-header {
                background-color: ''' + C['surface'] + ''' !important;
                border-bottom: 1px solid ''' + C['border'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
            }

            .card-body {
                background-color: ''' + C['panel'] + ''' !important;
            }

            .nav-pills .nav-link {
                color: ''' + C['muted'] + ''' !important;
                background-color: transparent !important;
                border-radius: 8px !important;
                font-family: ''' + C['sans'] + ''' !important;
            }

            .nav-pills .nav-link:hover {
                background-color: ''' + C['surface'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
            }

            .nav-pills .nav-link.active {
                background: linear-gradient(135deg, ''' + C['blue'] + ''' 0%, ''' + C['sky'] + ''' 100%) !important;
                color: white !important;
                font-weight: 600 !important;
            }

            .navbar {
                background: linear-gradient(135deg, ''' + C['bg'] + ''' 0%, ''' + C['surface'] + ''' 100%) !important;
                border-bottom: 2px solid ''' + C['border'] + ''' !important;
            }

            .btn-primary {
                background: linear-gradient(135deg, ''' + C['blue'] + ''' 0%, ''' + C['sky'] + ''' 100%) !important;
                border: none !important;
                font-weight: 600 !important;
                font-family: ''' + C['sans'] + ''' !important;
            }

            .btn-primary:hover {
                background: linear-gradient(135deg, ''' + C['sky'] + ''' 0%, ''' + C['teal'] + ''' 100%) !important;
                transform: translateY(-1px);
                box-shadow: 0 4px 12px ''' + rgba(C['blue'], 0.3) + ''' !important;
            }

            .btn-success {
                background: linear-gradient(135deg, ''' + C['green'] + ''' 0%, ''' + C['teal'] + ''' 100%) !important;
                border: none !important;
            }

            .btn-danger {
                background-color: ''' + C['red'] + ''' !important;
                border: none !important;
            }

            .btn-secondary {
                background-color: ''' + C['surface'] + ''' !important;
                border: 1px solid ''' + C['border'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
            }

            .btn-info {
                background: linear-gradient(135deg, ''' + C['purple'] + ''' 0%, ''' + C['pink'] + ''' 100%) !important;
                border: none !important;
            }

            .modal-content {
                background-color: ''' + C['panel'] + ''' !important;
                border: 1px solid ''' + C['border'] + ''' !important;
            }

            .modal-header {
                background-color: ''' + C['surface'] + ''' !important;
                border-bottom: 1px solid ''' + C['border'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
            }

            .modal-body {
                background-color: ''' + C['panel'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
            }

            .modal-footer {
                background-color: ''' + C['surface'] + ''' !important;
                border-top: 1px solid ''' + C['border'] + ''' !important;
            }

            .form-control, .form-select {
                background-color: ''' + C['surface'] + ''' !important;
                border: 1px solid ''' + C['border'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
                font-family: ''' + C['mono'] + ''' !important;
            }

            .form-control:focus, .form-select:focus {
                background-color: ''' + C['surface'] + ''' !important;
                border-color: ''' + C['blue'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
                box-shadow: 0 0 0 0.2rem ''' + rgba(C['blue'], 0.25) + ''' !important;
            }

            .form-label {
                color: ''' + C['text'] + ''' !important;
                font-weight: 600 !important;
                font-family: ''' + C['sans'] + ''' !important;
            }

            .text-muted {
                color: ''' + C['muted'] + ''' !important;
            }

            .alert-info {
                background-color: ''' + rgba(C['blue'], 0.1) + ''' !important;
                border-color: ''' + C['blue'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
            }

            h1, h2, h3, h4, h5, h6 {
                color: ''' + C['text'] + ''' !important;
                font-family: ''' + C['sans'] + ''' !important;
            }

            .shadow-sm {
                box-shadow: 0 2px 8px ''' + rgba(C['bg'], 0.4) + ''' !important;
            }

            .border-0 {
                border: 1px solid ''' + C['border'] + ''' !important;
            }

            small {
                color: ''' + C['muted'] + ''' !important;
            }

            /* Dash DataTable */
            .dash-table-container {
                font-family: ''' + C['mono'] + ''' !important;
            }

            .dash-spreadsheet {
                background-color: ''' + C['panel'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
            }

            .dash-spreadsheet td {
                background-color: ''' + C['panel'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
                border: 1px solid ''' + C['border'] + ''' !important;
            }

            .dash-spreadsheet th {
                background-color: ''' + C['surface'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
                border: 1px solid ''' + C['border'] + ''' !important;
            }

            /* Tab styling */
            .nav-tabs .nav-link {
                color: ''' + C['muted'] + ''' !important;
                background-color: transparent !important;
                border: none !important;
                border-bottom: 2px solid transparent !important;
            }

            .nav-tabs .nav-link.active {
                color: ''' + C['blue'] + ''' !important;
                border-bottom: 2px solid ''' + C['blue'] + ''' !important;
                background-color: transparent !important;
            }

            /* Scrollbar */
            ::-webkit-scrollbar {
                width: 10px;
                height: 10px;
            }

            ::-webkit-scrollbar-track {
                background: ''' + C['surface'] + ''';
            }

            ::-webkit-scrollbar-thumb {
                background: ''' + C['border2'] + ''';
                border-radius: 5px;
            }

            ::-webkit-scrollbar-thumb:hover {
                background: ''' + C['muted'] + ''';
            }
        </style>
    </head>
    <body>
        {%app_entry%}
        <footer>
            {%config%}
            {%scripts%}
            {%renderer%}
        </footer>
    </body>
</html>
'''

DEFAULT_CONFIG = {
    'fund_parameters': {
        'dry_powder': 450,
        'target_net_twr': 0.13,
        'management_fee': 0.0125,
        'carry_rate': 0.125,
        'hurdle_rate': 0.10,
        'loss_drag': 0.01,
        'liquidity_reserve_pct': 0.05,
        'distribution_rate': 0.20,
        'cash_yield': 0.03,
        'avg_hold_period': 5.0
    },
    'bite_sizing': {
        'enabled': True
    }
}

# ==================== LAYOUT ====================

navbar = dbc.Navbar(
    dbc.Container([
        html.Div([
            html.I(className="fas fa-chart-line me-3", style={'fontSize': '32px', 'color': C['blue']}),
            dbc.NavbarBrand("Horizon Portfolio Tool", style={
                'fontSize': '24px',
                'fontWeight': 'bold',
                'fontFamily': C['sans'],
                'background': f'linear-gradient(135deg, {C["blue"]} 0%, {C["sky"]} 100%)',
                '-webkit-background-clip': 'text',
                '-webkit-text-fill-color': 'transparent'
            })
        ]),
        html.Div(id='live-clock', style={
            'fontSize': '14px',
            'color': C['muted'],
            'fontFamily': C['mono']
        })
    ], fluid=True),
    style={
        'background': f'linear-gradient(135deg, {C["bg"]} 0%, {C["surface"]} 100%)',
        'borderBottom': f'2px solid {C["border"]}',
        'padding': '1.2rem 2rem'
    },
    dark=True,
    className="mb-4"
)

sidebar = dbc.Card([
    dbc.CardBody([
        html.H5("📊 Navigation", className="mb-4", style={
            'fontWeight': 'bold',
            'color': C['text'],
            'fontFamily': C['sans']
        }),
        dbc.Nav([
            dbc.NavLink([html.I(className="fas fa-home me-2"), "Dashboard"], href="/", active="exact"),
            dbc.NavLink([html.I(className="fas fa-briefcase me-2"), "Current Portfolio"], href="/portfolio",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-layer-group me-2"), "Portfolio Segments & TWR"], href="/segmentation",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-calendar-plus me-2"), "Future Deals"], href="/future",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-water me-2"), "Dry Powder Forecast"], href="/drypowder",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-tint me-2"), "Liquidity Assumptions"], href="/liquidity",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-bullseye me-2"), "Return Calculator"], href="/calculator",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-chart-area me-2"), "TWR Forecaster"], href="/twr", active="exact"),
            dbc.NavLink([html.I(className="fas fa-dollar-sign me-2"), "Deal Cashflows"], href="/cashflows",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-magic me-2"), "Pro Forma"], href="/proforma", active="exact"),
            dbc.NavLink([html.I(className="fas fa-funnel-dollar me-2"), "Pipeline"], href="/pipeline", active="exact"),
            dbc.NavLink([html.I(className="fas fa-chart-bar me-2"), "Analytics"], href="/analytics", active="exact"),
            dbc.NavLink([html.I(className="fas fa-cog me-2"), "Settings"], href="/settings", active="exact"),
        ], vertical=True, pills=True),
        html.Hr(style={'borderColor': C['border']}),
        html.H6("📈 Quick Stats", className="mb-3", style={
            'fontWeight': 'bold',
            'color': C['text'],
            'fontFamily': C['sans']
        }),
        html.Div(id='sidebar-stats'),
        html.Hr(style={'borderColor': C['border']}),
        dbc.Button("📥 Export All", id="btn-export", color="secondary", size="sm", className="w-100"),
        dcc.Download(id="download-csv"),
    ], style={'backgroundColor': C['panel']})
], style={
    'position': 'sticky',
    'top': '20px',
    'backgroundColor': C['panel'],
    'border': f'1px solid {C["border"]}'
})

# Load saved data at startup
LOADED_DATA = load_data()

app.layout = dbc.Container([
    dcc.Location(id='url', refresh=False),
    html.Div(id='save-status', style={'display': 'none'}),  # Hidden save status
    dcc.Store(id='deals-store', data=LOADED_DATA.get('deals', []) if LOADED_DATA else []),
    dcc.Store(id='placeholder-deals-store', data=LOADED_DATA.get('placeholders', []) if LOADED_DATA else []),
    dcc.Store(id='pipeline-store', data=LOADED_DATA.get('pipeline', []) if LOADED_DATA else []),
    dcc.Store(id='cashflows-store', data=[]),
    dcc.Store(id='fund-cf-data-store', data=None),  # Fund Level CF upload
    dcc.Store(id='liquidity-data-store', data=None),  # Liquidity Pull upload
    dcc.Store(id='proforma-scenario-store', data=[]),
    dcc.Store(id='liquidity-assumptions-store', data={
        'GP-Led (Multi-Asset)': {'annual_dist_rate': 0.20, 'call_pattern': 'immediate'},
        'GP-Led (Single-Asset)': {'annual_dist_rate': 0.25, 'call_pattern': 'staged'},
        'Diversified LP-Led': {'annual_dist_rate': 0.18, 'call_pattern': 'staged'},
        'Co-Investments': {'annual_dist_rate': 0.15, 'call_pattern': 'delayed'},
    }),
    dcc.Store(id='config-store', data=LOADED_DATA.get('config', DEFAULT_CONFIG) if LOADED_DATA else DEFAULT_CONFIG),
    dcc.Interval(id='clock-interval', interval=1000),
    navbar,
    dbc.Row([
        dbc.Col(sidebar, width=2),
        dbc.Col(html.Div(id="page-content"), width=10)
    ])
], fluid=True, style={'backgroundColor': C['bg'], 'minHeight': '100vh'})


# ==================== PAGE LAYOUTS ====================

def dashboard_page():
    return html.Div([
        html.H2("📊 Fund Dashboard", className="mb-4", style={
            'fontWeight': 'bold',
            'fontFamily': C['sans'],
            'color': C['text']
        }),

        # Top KPIs
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("💰 Current NAV", className="text-muted mb-2"),
                html.H3(id="dash-nav", style={'color': C['green'], 'fontWeight': 'bold', 'fontFamily': C['mono']}),
                html.Small(id="dash-num-deals", className="text-muted")
            ])], className="shadow-sm border-0",
                style={'background': f'linear-gradient(135deg, {rgba(C["green"], 0.1)} 0%, {C["panel"]} 100%)'}),
                width=2),

            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("💵 Dry Powder", className="text-muted mb-2"),
                html.H3(id="dash-dry-powder",
                        style={'color': C['blue'], 'fontWeight': 'bold', 'fontFamily': C['mono']}),
                html.Small("Available", className="text-muted")
            ])], className="shadow-sm border-0",
                style={'background': f'linear-gradient(135deg, {rgba(C["blue"], 0.1)} 0%, {C["panel"]} 100%)'}),
                width=2),

            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("📊 Total Fund", className="text-muted mb-2"),
                html.H3(id="dash-total", style={'color': C['sky'], 'fontWeight': 'bold', 'fontFamily': C['mono']}),
                html.Small("NAV + Powder", className="text-muted")
            ])], className="shadow-sm border-0",
                style={'background': f'linear-gradient(135deg, {rgba(C["sky"], 0.1)} 0%, {C["panel"]} 100%)'}),
                width=2),

            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("🎯 Portfolio IRR", className="text-muted mb-2"),
                html.H3(id="dash-current-irr",
                        style={'color': C['teal'], 'fontWeight': 'bold', 'fontFamily': C['mono']}),
                html.Small("Weighted", className="text-muted")
            ])], className="shadow-sm border-0",
                style={'background': f'linear-gradient(135deg, {rgba(C["teal"], 0.1)} 0%, {C["panel"]} 100%)'}),
                width=2),

            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("🚀 Required IRR", className="text-muted mb-2"),
                html.H3(id="dash-req-irr", style={'color': C['amber'], 'fontWeight': 'bold', 'fontFamily': C['mono']}),
                html.Small("Future Deals", className="text-muted")
            ])], className="shadow-sm border-0",
                style={'background': f'linear-gradient(135deg, {rgba(C["amber"], 0.1)} 0%, {C["panel"]} 100%)'}),
                width=2),

            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("📅 Placeholders", className="text-muted mb-2"),
                html.H3(id="dash-placeholders",
                        style={'color': C['purple'], 'fontWeight': 'bold', 'fontFamily': C['mono']}),
                html.Small(id="dash-placeholder-value", className="text-muted")
            ])], className="shadow-sm border-0",
                style={'background': f'linear-gradient(135deg, {rgba(C["purple"], 0.1)} 0%, {C["panel"]} 100%)'}),
                width=2),
        ], className="mb-4"),

        # Charts Row
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Portfolio Allocation", style={
                        'fontWeight': 'bold',
                        'backgroundColor': C['surface'],
                        'borderBottom': f'1px solid {C["border"]}'
                    }),
                    dbc.CardBody([dcc.Graph(id='dash-allocation-chart', config={'displayModeBar': False})])
                ], className="shadow-sm border-0")
            ], width=4),

            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Dry Powder Forecast (12 Months)", style={
                        'fontWeight': 'bold',
                        'backgroundColor': C['surface'],
                        'borderBottom': f'1px solid {C["border"]}'
                    }),
                    dbc.CardBody([dcc.Graph(id='dash-forecast-chart', config={'displayModeBar': False})])
                ], className="shadow-sm border-0")
            ], width=4),

            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Deal Bite Sizing Guide", style={
                        'fontWeight': 'bold',
                        'backgroundColor': C['surface'],
                        'borderBottom': f'1px solid {C["border"]}'
                    }),
                    dbc.CardBody([html.Div(id='dash-bite-sizing')])
                ], className="shadow-sm border-0")
            ], width=4),
        ], className="mb-4"),

        # Recent Activity
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Recent Portfolio Activity", style={
                        'fontWeight': 'bold',
                        'backgroundColor': C['surface'],
                        'borderBottom': f'1px solid {C["border"]}'
                    }),
                    dbc.CardBody([html.Div(id='dash-recent-activity')])
                ], className="shadow-sm border-0")
            ], width=6),

            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Pipeline Status", style={
                        'fontWeight': 'bold',
                        'backgroundColor': C['surface'],
                        'borderBottom': f'1px solid {C["border"]}'
                    }),
                    dbc.CardBody([html.Div(id='dash-pipeline-status')])
                ], className="shadow-sm border-0")
            ], width=6),
        ])
    ])


def portfolio_page():
    return html.Div([
        dbc.Row([
            dbc.Col(html.H2("💼 Current Portfolio", style={'fontWeight': 'bold'}), width=8),
            dbc.Col(dbc.Button("➕ Add Deal", id="btn-open-deal", color="primary", className="float-end", size="lg"),
                    width=4)
        ], className="mb-3"),

        # Fund Status Cards
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("💰 Current NAV", className="text-muted"),
                html.H4(id="port-nav", style={'color': COLORS['success']}),
                html.Small("Invested Capital", className="text-muted")
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("💵 Dry Powder", className="text-muted"),
                html.H4(id="port-dry-powder", style={'color': COLORS['primary']}),
                html.Small("For Pipeline/Future Only", className="text-muted")
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("📊 Total Fund Size", className="text-muted"),
                html.H4(id="port-total-fund", style={'color': COLORS['dark']}),
                html.Small("NAV + Dry Powder", className="text-muted")
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("🎯 Target Return", className="text-muted"),
                html.H4(id="port-target-return", style={'color': COLORS['warning']}),
                html.Small("Required Deal IRR", className="text-muted")
            ])], className="shadow-sm"), width=3),
        ], className="mb-4"),

        # Add Deal Modal
        dbc.Modal([
            dbc.ModalHeader("Add Deal to Current Portfolio"),
            dbc.ModalBody([
                dbc.Alert([
                    html.I(className="fas fa-info-circle me-2"),
                    html.Strong("Current Portfolio: "),
                    "Enter ACTUAL NAV of deals already in the portfolio. These do NOT reduce dry powder."
                ], color="info", className="mb-3", style={'fontSize': '13px'}),

                # Row 1: Deal Name, Fund Manager
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Deal Name *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-name", type="text", placeholder="e.g., Coller VII")
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Fund Manager *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-manager", type="text", placeholder="e.g., Coller Capital")
                    ], width=6)
                ], className="mb-3"),

                # Row 2: Strategy, Stage
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Strategy Type *", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-strategy", options=[
                            {"label": "GP-Led (Multi-Asset)", "value": "GP-Led (Multi-Asset)"},
                            {"label": "GP-Led (Single-Asset)", "value": "GP-Led (Single-Asset)"},
                            {"label": "Diversified LP-Led", "value": "Diversified LP-Led"},
                            {"label": "Co-Investments", "value": "Co-Investments"},
                        ])
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Stage *", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-stage", options=[
                            {"label": "Buyout", "value": "Buyout"},
                            {"label": "Venture", "value": "Venture"},
                            {"label": "Growth", "value": "Growth"},
                            {"label": "Liquidity", "value": "Liquidity"},
                        ], value="Buyout")
                    ], width=6)
                ], className="mb-3"),

                # Row 3: Commitment, NAV, Unfunded
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Total Commitment ($mm) *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-commitment", type="number", step=0.1, placeholder="50.0"),
                        html.Small("Total committed capital", className="text-muted")
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Current NAV ($mm) *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-size", type="number", step=0.1, placeholder="35.0"),
                        html.Small("Current Net Asset Value", className="text-muted")
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Unfunded ($mm)", style={'fontWeight': 'bold'}),
                        html.Div(id="unfunded-display", style={
                            'marginTop': '8px',
                            'padding': '8px',
                            'backgroundColor': C['surface'],
                            'borderRadius': '4px',
                            'fontFamily': C['mono'],
                            'fontSize': '16px',
                            'fontWeight': 'bold',
                            'color': C['amber']
                        }),
                        html.Small("Auto: Commitment - NAV", className="text-muted")
                    ], width=4),
                ], className="mb-3"),

                # Row 4: IRR, Hold Period, Currency
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Target Gross IRR (%) *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-irr", type="number", step=0.5, placeholder="22.0")
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Hold Period (years)", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-hold", type="number", value=5, step=0.5)
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Currency", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-currency", options=[
                            {"label": "USD", "value": "USD"},
                            {"label": "EUR", "value": "EUR"},
                            {"label": "GBP", "value": "GBP"},
                        ], value="USD")
                    ], width=4),
                ], className="mb-3"),

                # Row 5: Vintage, Segment, Allocation Status
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Vintage Year", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-vintage", options=[
                            {"label": str(y), "value": y} for y in range(2026, 2014, -1)
                        ], value=2024)
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Portfolio Segment *", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-segment", options=[
                            {"label": "Seed Portfolio", "value": "Seed"},
                            {"label": "New Deals", "value": "New"},
                            {"label": "Money Market", "value": "MoneyMarket"},
                        ], value="Seed"),
                        html.Small("For TWR tracking", className="text-muted")
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Allocation Status *", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-allocation-status", options=[
                            {"label": "Closed", "value": "Closed"},
                            {"label": "Pending Close", "value": "Pending Close"},
                            {"label": "Pending Allocation", "value": "Pending Allocation"},
                        ], value="Closed")
                    ], width=4),
                ], className="mb-3"),

                # Row 6: Sector, Geography
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Sector", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-sector", options=[
                            {"label": "Diversified", "value": "Diversified"},
                            {"label": "Technology", "value": "Technology"},
                            {"label": "Healthcare", "value": "Healthcare"},
                            {"label": "Consumer", "value": "Consumer"},
                            {"label": "Industrials", "value": "Industrials"},
                            {"label": "Financials", "value": "Financials"},
                            {"label": "Other", "value": "Other"}
                        ], value="Diversified")
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Geography", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-geo", options=[
                            {"label": "North America", "value": "North America"},
                            {"label": "Europe", "value": "Europe"},
                            {"label": "Asia", "value": "Asia"},
                            {"label": "Global", "value": "Global"}
                        ], value="North America")
                    ], width=6),
                ])
            ]),
            dbc.ModalFooter([
                dbc.Button("Cancel", id="btn-cancel", color="secondary"),
                dbc.Button("Add Deal", id="btn-submit", color="primary")
            ])
        ], id="modal-deal", size="lg", is_open=False),

        html.Div(id='portfolio-table')
    ])


def future_deals_page():
    return html.Div([
        dbc.Row([
            dbc.Col(html.H2("📅 Pro Forma Portfolio (Future/Placeholder Deals)", style={'fontWeight': 'bold'}), width=8),
            dbc.Col(dbc.Button("➕ Add Placeholder", id="btn-open-placeholder", color="success", className="float-end",
                               size="lg"), width=4)
        ], className="mb-3"),

        dbc.Alert([
            html.I(className="fas fa-lightbulb me-2"),
            "Plan future deals to forecast dry powder usage and ensure you stay within bite size limits. These are NOT in your current portfolio."
        ], color="info", className="mb-4"),

        # Add Placeholder Modal
        dbc.Modal([
            dbc.ModalHeader("Add Placeholder Deal (Future Commitment)"),
            dbc.ModalBody([
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Deal Name *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-ph-name", type="text", placeholder="e.g., GP-Led (Multi-Asset) 1")
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Strategy Type *", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-ph-strategy", options=[
                            {"label": "GP-Led (Multi-Asset)", "value": "GP-Led (Multi-Asset)"},
                            {"label": "GP-Led (Single-Asset)", "value": "GP-Led (Single-Asset)"},
                            {"label": "Diversified LP-Led", "value": "Diversified LP-Led"},
                            {"label": "Co-Investments", "value": "Co-Investments"},
                        ])
                    ], width=6)
                ], className="mb-3"),
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Expected Size ($mm) *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-ph-size", type="number", step=0.1, placeholder="15.0"),
                        html.Small(id="ph-bite-warning", className="text-muted")
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Custom Bite Size ($mm)", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-ph-bite", type="number", step=0.1, placeholder="Auto-calculate"),
                        html.Small("Override default bite size limits", className="text-muted")
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Expected Month *", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-ph-month", options=generate_month_options(), value=0,
                                   style={'fontFamily': C['mono']})
                    ], width=4),
                ], className="mb-3"),
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Target Weight (% of Portfolio)", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-ph-weight", type="number", step=0.5, placeholder="5.0"),
                        html.Small("Desired portfolio allocation %", className="text-muted")
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Region", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-ph-region", options=[
                            {"label": "North America", "value": "North America"},
                            {"label": "Europe", "value": "Europe"},
                            {"label": "Asia", "value": "Asia"},
                            {"label": "Global", "value": "Global"},
                        ], value="North America")
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Deal Type", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-ph-type", options=[
                            {"label": "Secondary", "value": "Secondary"},
                            {"label": "Co-Investment", "value": "Co-Investment"},
                        ], value="Secondary")
                    ], width=4),
                ])
            ]),
            dbc.ModalFooter([
                dbc.Button("Cancel", id="btn-cancel-ph", color="secondary"),
                dbc.Button("Add Placeholder", id="btn-submit-ph", color="success")
            ])
        ], id="modal-placeholder", size="lg", is_open=False),

        html.Div(id='placeholder-table')
    ])


def drypowder_page():
    return html.Div([
        html.H2("💧 Dry Powder Forecaster", className="mb-4", style={'fontWeight': 'bold'}),

        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Current Dry Powder", className="text-muted"),
                html.H3(id="dp-current", style={'color': COLORS['primary']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Planned Deployments", className="text-muted"),
                html.H3(id="dp-planned", style={'color': COLORS['warning']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Expected Distributions", className="text-muted"),
                html.H3(id="dp-distributions", style={'color': COLORS['success']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("12-Month Forecast", className="text-muted"),
                html.H3(id="dp-forecast-12m", style={'color': COLORS['info']})
            ])], className="shadow-sm"), width=3),
        ], className="mb-4"),

        dbc.Card([
            dbc.CardHeader("12-Month Dry Powder Forecast",
                           style={'fontWeight': 'bold', 'backgroundColor': COLORS['light']}),
            dbc.CardBody([
                dcc.Graph(id='dp-forecast-chart', config={'displayModeBar': True})
            ])
        ], className="shadow-sm mb-4"),

        dbc.Card([
            dbc.CardHeader("Monthly Breakdown", style={'fontWeight': 'bold', 'backgroundColor': COLORS['light']}),
            dbc.CardBody([
                html.Div(id='dp-monthly-table')
            ])
        ], className="shadow-sm")
    ])


def calculator_page():
    return html.Div([
        html.H2("🧮 Return Calculator & Waterfall", className="mb-4", style={'fontWeight': 'bold'}),

        dbc.Card([dbc.CardBody([
            html.H5("Calculate required deal IRR to achieve target net TWR", className="text-muted mb-4"),

            html.Hr(),
            html.H4("📊 Current Portfolio", className="mt-4", style={'color': COLORS['dark'], 'fontWeight': 'bold'}),
            dbc.Row([
                dbc.Col([
                    html.H6("Current NAV:", className="text-muted"),
                    html.H3(id="calc-current-nav", style={'color': COLORS['success']})
                ], width=4),
                dbc.Col([
                    html.H6("Weighted IRR:", className="text-muted"),
                    html.H3(id="calc-current-irr", style={'color': COLORS['success']})
                ], width=4),
                dbc.Col([
                    html.H6("Number of Deals:", className="text-muted"),
                    html.H3(id="calc-num-deals", style={'color': COLORS['success']})
                ], width=4),
            ], className="mb-4"),

            html.Hr(),
            html.H4("🎯 Evergreen Fund Structure", className="mt-4",
                    style={'color': COLORS['dark'], 'fontWeight': 'bold'}),
            dbc.Row([
                dbc.Col([
                    html.H6("Total Fund Size:", className="text-muted"),
                    html.H3(id="calc-total-fund", style={'color': COLORS['primary']}),
                    html.Small("NAV + Dry Powder", className="text-muted")
                ], width=4),
                dbc.Col([
                    html.H6("Target Net TWR:", className="text-muted"),
                    html.H3(id="calc-target-twr", style={'color': COLORS['warning']})
                ], width=4),
                dbc.Col([
                    html.H6("Dry Powder:", className="text-muted"),
                    html.H3(id="calc-dry-powder", style={'color': COLORS['info']}),
                    html.Small("For Pipeline/Future Only", className="text-muted")
                ], width=4),
            ], className="mb-4"),

            html.Hr(),
            html.Div([
                html.H3("✅ REQUIRED FUTURE DEAL IRR", className="text-center mb-3",
                        style={'color': COLORS['success'], 'fontWeight': 'bold'}),
                html.H1(id="calc-required-irr", className="text-center mb-2",
                        style={'fontSize': '80px', 'fontWeight': 'bold', 'color': COLORS['success']}),
                html.P(id="calc-explanation", className="text-center text-muted mb-4", style={'fontSize': '18px'}),
            ], style={'backgroundColor': '#F0F8F0', 'padding': '2rem', 'borderRadius': '10px',
                      'border': f'3px solid {COLORS["success"]}'}),

            html.Hr(className="mt-4"),
            html.H5("💡 Return Waterfall Breakdown", className="mb-3",
                    style={'color': COLORS['dark'], 'fontWeight': 'bold'}),
            html.Div(id="calc-waterfall")
        ])], className="shadow-sm")
    ])


def twr_forecaster_page():
    return html.Div([
        html.H2("📈 TWR Returns Forecaster", className="mb-4",
                style={'fontWeight': 'bold', 'fontFamily': C['sans'], 'color': C['text']}),

        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Monte Carlo Simulation Parameters",
                                   style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([
                        dbc.Row([
                            dbc.Col([
                                dbc.Label("Current Portfolio IRR",
                                          style={'fontWeight': 'bold', 'fontFamily': C['sans']}),
                                html.H4(id="twr-current-irr", style={'color': C['green'], 'fontFamily': C['mono']})
                            ], width=4),
                            dbc.Col([
                                dbc.Label("Future Deals Mean IRR (%)", style={'fontWeight': 'bold'}),
                                dbc.Input(id="twr-future-mean", type="number", value=25, step=0.5,
                                          style={'fontFamily': C['mono']})
                            ], width=4),
                            dbc.Col([
                                dbc.Label("Future Deals Std Dev (%)", style={'fontWeight': 'bold'}),
                                dbc.Input(id="twr-future-std", type="number", value=5, step=0.5,
                                          style={'fontFamily': C['mono']})
                            ], width=4),
                        ], className="mb-3"),
                        dbc.Row([
                            dbc.Col([
                                dbc.Label("# Simulations", style={'fontWeight': 'bold'}),
                                dbc.Select(id="twr-n-sims", options=[
                                    {"label": "1,000", "value": 1000},
                                    {"label": "5,000", "value": 5000},
                                    {"label": "10,000", "value": 10000},
                                ], value=5000, style={'fontFamily': C['mono']})
                            ], width=6),
                            dbc.Col([
                                dbc.Button("🎲 Run Simulation", id="btn-run-twr", color="primary", size="lg",
                                           className="w-100")
                            ], width=6)
                        ])
                    ])
                ], className="shadow-sm mb-4")
            ], width=12),
        ]),

        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Probability Distribution",
                                   style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='twr-distribution-chart', config={'displayModeBar': True})])
                ], className="shadow-sm")
            ], width=8),
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Key Statistics", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([html.Div(id='twr-statistics')])
                ], className="shadow-sm")
            ], width=4),
        ], className="mb-4"),

        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Sensitivity Analysis",
                                   style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='twr-sensitivity-chart', config={'displayModeBar': True})])
                ], className="shadow-sm")
            ], width=12),
        ])
    ])


def cashflows_page():
    return html.Div([
        html.H2("💰 Fund Level Cashflows", className="mb-4",
                style={'fontWeight': 'bold', 'fontFamily': C['sans'], 'color': C['text']}),

        dbc.Alert([
            html.I(className="fas fa-info-circle me-2"),
            "Upload your Excel Fund Level CF file or view monthly cashflows by deal. Each row = one deal. Columns = months."
        ], color="info", className="mb-4"),

        # File Upload Section
        dbc.Card([
            dbc.CardHeader("📥 Upload Fund Level CF Excel File",
                           style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
            dbc.CardBody([
                dbc.Row([
                    dbc.Col([
                        dcc.Upload(
                            id='upload-fund-cf',
                            children=dbc.Button([
                                html.I(className="fas fa-file-excel me-2"),
                                'Upload Fund Level CF Excel File'
                            ], color="primary", size="lg", className="w-100"),
                            multiple=False
                        ),
                        html.Small("Upload the 'Fund Level CF' tab from your Excel file (.xlsx, .xlsm)",
                                   className="text-muted d-block mt-2")
                    ], width=6),
                    dbc.Col([
                        html.Div(id='upload-fund-cf-status', className="mt-2")
                    ], width=6)
                ])
            ])
        ], className="shadow-sm mb-4"),

        # Summary Cards
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Total Commitment", className="text-muted"),
                html.H4(id="cf-total-commitment", style={'color': C['blue'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Total Paid In", className="text-muted"),
                html.H4(id="cf-total-paid-in", style={'color': C['green'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Total Unfunded", className="text-muted"),
                html.H4(id="cf-total-unfunded", style={'color': C['amber'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Current Total NAV", className="text-muted"),
                html.H4(id="cf-current-nav", style={'color': C['purple'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
        ], className="mb-4"),

        # Controls
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardBody([
                        dbc.Row([
                            dbc.Col([
                                dbc.Label("View Period", style={'fontWeight': 'bold'}),
                                dbc.Select(id="cf-view-period", options=[
                                    {"label": "Next 12 Months", "value": 12},
                                    {"label": "Next 24 Months", "value": 24},
                                    {"label": "Next 36 Months", "value": 36},
                                    {"label": "5 Years", "value": 60},
                                ], value=12, style={'fontFamily': C['mono']})
                            ], width=4),
                            dbc.Col([
                                dbc.Label("Cashflow Type", style={'fontWeight': 'bold'}),
                                dbc.Select(id="cf-type-view", options=[
                                    {"label": "All (Calls + Dists + NAV)", "value": "all"},
                                    {"label": "Net Cashflows Only", "value": "net"},
                                    {"label": "NAV Only", "value": "nav"},
                                ], value="all", style={'fontFamily': C['mono']})
                            ], width=4),
                            dbc.Col([
                                dbc.Button("📥 Export to CSV", id="btn-export-cf", color="success",
                                           className="w-100 mt-4")
                            ], width=4)
                        ])
                    ])
                ], className="shadow-sm")
            ], width=12)
        ], className="mb-4"),

        # Main Cashflow Table - Deal by Deal with Monthly Columns
        dbc.Card([
            dbc.CardHeader("Fund Level Cashflows (Deal-by-Deal Monthly View)",
                           style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
            dbc.CardBody([
                html.Div(id='cf-monthly-table', style={'overflowX': 'auto'})
            ])
        ], className="shadow-sm mb-4"),

        # Monthly Totals Chart
        dbc.Card([
            dbc.CardHeader("Monthly Cashflow Totals", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
            dbc.CardBody([
                dcc.Graph(id='cf-monthly-chart', config={'displayModeBar': True})
            ])
        ], className="shadow-sm"),

        # Hidden store for uploaded data
        dcc.Store(id='fund-cf-data-store', data=None)
    ])


def proforma_page():
    return html.Div([
        html.H2("🔮 Pro Forma Analyzer", className="mb-4",
                style={'fontWeight': 'bold', 'fontFamily': C['sans'], 'color': C['text']}),

        # Month selector for NAV calculation
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Pro Forma Settings", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([
                        dbc.Row([
                            dbc.Col([
                                dbc.Label("Target Month for NAV Calculation *", style={'fontWeight': 'bold'}),
                                dbc.Select(id="pf-target-month", options=generate_month_options(), value=12,
                                           style={'fontFamily': C['mono']}),
                                html.Small("Calculate pro forma NAV at this month using actual cashflows",
                                           className="text-muted")
                            ], width=6),
                            dbc.Col([
                                dbc.Button("📊 Calculate Pro Forma NAV", id="btn-calc-pf-nav",
                                           color="primary", size="lg", className="w-100 mt-4")
                            ], width=6)
                        ])
                    ])
                ], className="shadow-sm mb-4")
            ], width=12)
        ]),

        # Tabs
        dbc.Tabs([
            dbc.Tab(label="📋 Complete Portfolio", tab_id="tab-pf-portfolio"),
            dbc.Tab(label="📊 Metrics Comparison", tab_id="tab-pf-metrics"),
            dbc.Tab(label="📈 Impact Charts", tab_id="tab-pf-charts"),
        ], id="pf-tabs", active_tab="tab-pf-portfolio"),

        html.Div(id='pf-tab-content', className="mt-4")
    ])


def liquidity_assumptions_page():
    return html.Div([
        html.H2("💧 Liquidity Pull", className="mb-4",
                style={'fontWeight': 'bold', 'fontFamily': C['sans'], 'color': C['text']}),

        dbc.Alert([
            html.I(className="fas fa-database me-2"),
            html.Strong("Data Pull From Liquidity Model - "),
            "Upload your Excel Liquidity Pull file or view current liquidity position and projections"
        ], color="info", className="mb-4"),

        # File Upload Section
        dbc.Card([
            dbc.CardHeader("📥 Upload Liquidity Pull Excel File",
                           style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
            dbc.CardBody([
                dbc.Row([
                    dbc.Col([
                        dcc.Upload(
                            id='upload-liquidity-pull',
                            children=dbc.Button([
                                html.I(className="fas fa-file-excel me-2"),
                                'Upload Liquidity Pull Excel File'
                            ], color="info", size="lg", className="w-100"),
                            multiple=False
                        ),
                        html.Small("Upload the 'Liquidity Pull' tab from your Excel file (.xlsx, .xlsm)",
                                   className="text-muted d-block mt-2")
                    ], width=6),
                    dbc.Col([
                        html.Div(id='upload-liquidity-status', className="mt-2")
                    ], width=6)
                ])
            ])
        ], className="shadow-sm mb-4"),

        # Current Date Info
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Today's Date", className="text-muted"),
                html.H5(id="liq-today-date", style={'color': C['text'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("As At", className="text-muted"),
                html.H5(id="liq-as-at-date", style={'color': C['text'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Current Quarter", className="text-muted"),
                html.H5(id="liq-current-quarter", style={'color': C['text'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Current Month", className="text-muted"),
                html.H5(id="liq-current-month", style={'color': C['text'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
        ], className="mb-4"),

        # Liquidity Waterfall
        dbc.Card([
            dbc.CardHeader("Liquidity Waterfall", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
            dbc.CardBody([
                html.Div(id='liq-waterfall-table')
            ])
        ], className="shadow-sm mb-4"),

        # Near Term Flows (Monthly Projections)
        dbc.Card([
            dbc.CardHeader("Near Term Flows (Next 12 Months)",
                           style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
            dbc.CardBody([
                html.Div(id='liq-near-term-flows')
            ])
        ], className="shadow-sm mb-4"),

        # NAV Projections
        dbc.Card([
            dbc.CardHeader("NAV End Projections", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
            dbc.CardBody([
                dcc.Graph(id='liq-nav-projection-chart', config={'displayModeBar': True})
            ])
        ], className="shadow-sm mb-4"),

        # Detailed Breakdown Tabs
        dbc.Tabs([
            dbc.Tab(label="Subscriptions & Redemptions", tab_id="tab-liq-subs"),
            dbc.Tab(label="Portfolio Net Flows", tab_id="tab-liq-flows"),
            dbc.Tab(label="Unfunded Commitments", tab_id="tab-liq-unfunded"),
        ], id="liq-tabs", active_tab="tab-liq-subs"),

        html.Div(id='liq-tab-content', className="mt-4"),

        # Hidden store for uploaded data
        dcc.Store(id='liquidity-data-store', data=None)
    ])


def segmentation_page():
    return html.Div([
        html.H2("📊 Portfolio Segmentation & TWR Analysis", className="mb-4",
                style={'fontWeight': 'bold', 'fontFamily': C['sans'], 'color': C['text']}),

        dbc.Alert([
            html.I(className="fas fa-layer-group me-2"),
            "Track TWR performance across portfolio segments: Seed Portfolio, New Deals, Money Market, Pipeline, and Future Deals"
        ], color="info", className="mb-4"),

        # Segment Summary Cards
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Total Portfolio", className="text-muted"),
                html.H4(id="seg-total-nav", style={'color': C['blue'], 'fontFamily': C['mono']}),
                html.Small(id="seg-total-twr", className="text-muted", style={'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Seed Portfolio", className="text-muted"),
                html.H4(id="seg-seed-nav", style={'color': C['green'], 'fontFamily': C['mono']}),
                html.Small(id="seg-seed-twr", className="text-muted", style={'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("New Deals", className="text-muted"),
                html.H4(id="seg-new-nav", style={'color': C['purple'], 'fontFamily': C['mono']}),
                html.Small(id="seg-new-twr", className="text-muted", style={'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Money Market", className="text-muted"),
                html.H4(id="seg-mm-nav", style={'color': C['amber'], 'fontFamily': C['mono']}),
                html.Small(id="seg-mm-twr", className="text-muted", style={'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Pipeline", className="text-muted"),
                html.H4(id="seg-pipeline-nav", style={'color': C['teal'], 'fontFamily': C['mono']}),
                html.Small(id="seg-pipeline-twr", className="text-muted", style={'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Future/Placeholder", className="text-muted"),
                html.H4(id="seg-future-nav", style={'color': C['pink'], 'fontFamily': C['mono']}),
                html.Small(id="seg-future-twr", className="text-muted", style={'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=2),
        ], className="mb-4"),

        # TWR Forecast by Segment
        dbc.Card([
            dbc.CardHeader("TWR Forecast by Segment (Next 12 Months)",
                           style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
            dbc.CardBody([
                dcc.Graph(id='seg-twr-forecast-chart', config={'displayModeBar': True})
            ])
        ], className="shadow-sm mb-4"),

        # Segment Breakdown Tables
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Seed Portfolio Deals",
                                   style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([html.Div(id='seg-seed-table')])
                ], className="shadow-sm")
            ], width=6),
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("New Deals", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([html.Div(id='seg-new-table')])
                ], className="shadow-sm")
            ], width=6),
        ], className="mb-4"),

        # Allocation and Contribution Charts
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Segment Allocation", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='seg-allocation-chart', config={'displayModeBar': False})])
                ], className="shadow-sm")
            ], width=6),
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("TWR Contribution by Segment",
                                   style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='seg-contribution-chart', config={'displayModeBar': False})])
                ], className="shadow-sm")
            ], width=6),
        ])
    ])


def analytics_page():
    return html.Div([
        html.H2("📊 Portfolio Analytics", className="mb-4",
                style={'fontWeight': 'bold', 'fontFamily': C['sans'], 'color': C['text']}),

        # Toggle between Current and Pro Forma
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardBody([
                        dbc.Row([
                            dbc.Col([
                                html.H6("Portfolio View:",
                                        style={'color': C['text'], 'fontFamily': C['sans'], 'marginBottom': 0}),
                            ], width=2),
                            dbc.Col([
                                dbc.RadioItems(
                                    id="analytics-view-toggle",
                                    options=[
                                        {"label": " Current Portfolio", "value": "current"},
                                        {"label": " Current + Pipeline", "value": "current_pipeline"},
                                        {"label": " Current + Pipeline + Placeholder", "value": "full_proforma"},
                                    ],
                                    value="current",
                                    inline=True,
                                    style={'fontFamily': C['sans'], 'fontSize': '13px'}
                                )
                            ], width=10)
                        ])
                    ])
                ], className="shadow-sm")
            ], width=12)
        ], className="mb-4"),

        # Summary Cards
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Total NAV", className="text-muted"),
                html.H4(id="analytics-total-nav", style={'color': C['blue'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Number of Deals", className="text-muted"),
                html.H4(id="analytics-num-deals", style={'color': C['green'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Weighted IRR", className="text-muted"),
                html.H4(id="analytics-weighted-irr", style={'color': C['purple'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Top 1 Concentration", className="text-muted"),
                html.H4(id="analytics-top1-conc", style={'color': C['amber'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
        ], className="mb-4"),

        # Exposure Charts Row 1
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Strategy Exposure", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='analytics-strategy', config={'displayModeBar': False})])
                ], className="shadow-sm")
            ], width=6),
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Regional Exposure", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='analytics-region', config={'displayModeBar': False})])
                ], className="shadow-sm")
            ], width=6),
        ], className="mb-4"),

        # Exposure Charts Row 2
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Vintage Year Exposure",
                                   style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='analytics-vintage', config={'displayModeBar': False})])
                ], className="shadow-sm")
            ], width=6),
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Secondary vs Co-Investment",
                                   style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='analytics-dealtype', config={'displayModeBar': False})])
                ], className="shadow-sm")
            ], width=6),
        ], className="mb-4"),

        # Exposure Charts Row 3
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Sector Exposure", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='analytics-sector', config={'displayModeBar': False})])
                ], className="shadow-sm")
            ], width=6),
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Concentration Risk", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='analytics-concentration', config={'displayModeBar': False})])
                ], className="shadow-sm")
            ], width=6),
        ])
    ])


def pipeline_page():
    return html.Div([
        dbc.Row([
            dbc.Col(html.H2("🔍 Deal Pipeline", style={'fontWeight': 'bold'}), width=8),
            dbc.Col(
                dbc.Button("➕ Add to Pipeline", id="btn-open-pipeline", color="info", className="float-end", size="lg"),
                width=4)
        ], className="mb-3"),

        # Add Pipeline Modal
        dbc.Modal([
            dbc.ModalHeader("Add Deal to Pipeline"),
            dbc.ModalBody([
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Deal Name *"),
                        dbc.Input(id="in-pipe-name", type="text", placeholder="Project Sigil")
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Deal Type *"),
                        dbc.Select(id="in-pipe-type", options=[
                            {"label": "GP-Led (Multi-Asset)", "value": "GP-Led (Multi-Asset)"},
                            {"label": "GP-Led (Single-Asset)", "value": "GP-Led (Single-Asset)"},
                            {"label": "Diversified LP-Led", "value": "Diversified LP-Led"},
                            {"label": "Co-Investments", "value": "Co-Investments"},
                        ])
                    ], width=6)
                ], className="mb-3"),
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Stage"),
                        dbc.Select(id="in-pipe-stage", options=[
                            {"label": "Screening", "value": "Screening"},
                            {"label": "Due Diligence", "value": "Due Diligence"},
                            {"label": "Term Sheet", "value": "Term Sheet"},
                            {"label": "Final Docs", "value": "Final Docs"},
                        ], value="Screening")
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Target Size ($mm)"),
                        dbc.Input(id="in-pipe-size", type="number", step=0.1, placeholder="7.5")
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Target IRR (%)"),
                        dbc.Input(id="in-pipe-irr", type="number", step=0.5, placeholder="16")
                    ], width=4),
                ])
            ]),
            dbc.ModalFooter([
                dbc.Button("Cancel", id="btn-cancel-pipe", color="secondary"),
                dbc.Button("Add to Pipeline", id="btn-submit-pipe", color="info")
            ])
        ], id="modal-pipeline", size="lg", is_open=False),

        html.Div(id='pipeline-table')
    ])


def settings_page():
    return html.Div([
        html.H2("⚙️ Fund Settings", className="mb-4", style={'fontWeight': 'bold'}),

        dbc.Card([dbc.CardBody([
            html.H5("💰 Fund Parameters", className="mb-4", style={'color': COLORS['dark'], 'fontWeight': 'bold'}),
            dbc.Row([
                dbc.Col([
                    dbc.Label("💵 Dry Powder ($mm)", style={'fontWeight': 'bold'}),
                    dbc.Input(id="set-dry-powder", type="number", value=450, step=10),
                    html.Small("Available capital to deploy", className="text-muted")
                ], width=4),
                dbc.Col([
                    dbc.Label("🎯 Target Net TWR (%)", style={'fontWeight': 'bold'}),
                    dbc.Input(id="set-twr", type="number", value=13, step=0.5),
                    html.Small("Net return to LPs", className="text-muted")
                ], width=4),
                dbc.Col([
                    dbc.Label("📊 Average Hold Period (years)", style={'fontWeight': 'bold'}),
                    dbc.Input(id="set-hold", type="number", value=5.0, step=0.5),
                    html.Small("For MOIC translation", className="text-muted")
                ], width=4),
            ], className="mb-4"),

            html.Hr(),
            html.H5("💼 Fee Structure", className="mb-4", style={'color': COLORS['dark'], 'fontWeight': 'bold'}),
            dbc.Row([
                dbc.Col([
                    dbc.Label("Management Fee (% p.a.)", style={'fontWeight': 'bold'}),
                    dbc.Input(id="set-fee", type="number", value=1.25, step=0.05),
                    html.Small("Annual fee on NAV", className="text-muted")
                ], width=4),
                dbc.Col([
                    dbc.Label("Carry Rate (%)", style={'fontWeight': 'bold'}),
                    dbc.Input(id="set-carry", type="number", value=12.5, step=0.5),
                    html.Small("Performance fee", className="text-muted")
                ], width=4),
                dbc.Col([
                    dbc.Label("Hurdle Rate (%)", style={'fontWeight': 'bold'}),
                    dbc.Input(id="set-hurdle", type="number", value=10, step=0.5),
                    html.Small("Return before carry", className="text-muted")
                ], width=4),
            ], className="mb-4"),

            html.Hr(),
            html.H5("🔧 Portfolio Assumptions", className="mb-4", style={'color': COLORS['dark'], 'fontWeight': 'bold'}),
            dbc.Row([
                dbc.Col([
                    dbc.Label("Liquidity Reserve (%)", style={'fontWeight': 'bold'}),
                    dbc.Input(id="set-liq", type="number", value=5, step=1),
                    html.Small("Cash buffer", className="text-muted")
                ], width=4),
                dbc.Col([
                    dbc.Label("Loss Drag (%)", style={'fontWeight': 'bold'}),
                    dbc.Input(id="set-loss", type="number", value=1, step=0.5),
                    html.Small("Expected annual impairment", className="text-muted")
                ], width=4),
                dbc.Col([
                    dbc.Label("Cash Yield (%)", style={'fontWeight': 'bold'}),
                    dbc.Input(id="set-cash-yield", type="number", value=3, step=0.5),
                    html.Small("Return on uninvested cash", className="text-muted")
                ], width=4),
            ], className="mb-4"),

            html.Hr(),
            dbc.Row([
                dbc.Col([
                    dbc.Button("💾 Save Settings", id="btn-save-settings", color="primary", size="lg", className="me-2"),
                    dbc.Button("🔄 Reset to Default", id="btn-reset-settings", color="secondary", size="lg")
                ])
            ])
        ])], className="shadow-sm")
    ])


# ==================== CALLBACKS ====================

@app.callback(Output('live-clock', 'children'), Input('clock-interval', 'n_intervals'))
def update_clock(n):
    return datetime.now().strftime('%H:%M:%S')


@app.callback(Output('sidebar-stats', 'children'),
              [Input('deals-store', 'data'), Input('config-store', 'data')])
def update_sidebar(deals, config):
    m = calculate_portfolio_metrics(deals)
    dry_powder = config['fund_parameters']['dry_powder']
    return [
        html.Small("💰 NAV", className="text-muted"),
        html.H6(f"${m['total_nav']:.0f}M", className="mb-2", style={'fontWeight': 'bold'}),
        html.Small("💵 Dry Powder", className="text-muted"),
        html.H6(f"${dry_powder:.0f}M", className="mb-2", style={'fontWeight': 'bold'}),
        html.Small("📊 Deals", className="text-muted"),
        html.H6(str(m['num_deals']), className="mb-2", style={'fontWeight': 'bold'}),
        html.Small("📈 Avg IRR", className="text-muted"),
        html.H6(f"{m['weighted_irr']:.1%}", style={'fontWeight': 'bold'})
    ]


@app.callback(Output('page-content', 'children'), Input('url', 'pathname'))
def display_page(pathname):
    if pathname == '/portfolio':
        return portfolio_page()
    elif pathname == '/segmentation':
        return segmentation_page()
    elif pathname == '/future':
        return future_deals_page()
    elif pathname == '/drypowder':
        return drypowder_page()
    elif pathname == '/liquidity':
        return liquidity_assumptions_page()
    elif pathname == '/calculator':
        return calculator_page()
    elif pathname == '/twr':
        return twr_forecaster_page()
    elif pathname == '/cashflows':
        return cashflows_page()
    elif pathname == '/proforma':
        return proforma_page()
    elif pathname == '/pipeline':
        return pipeline_page()
    elif pathname == '/analytics':
        return analytics_page()
    elif pathname == '/settings':
        return settings_page()
    else:
        return dashboard_page()


# Deal Modal
@app.callback(
    Output('modal-deal', 'is_open'),
    [Input('btn-open-deal', 'n_clicks'), Input('btn-cancel', 'n_clicks'), Input('btn-submit', 'n_clicks')],
    State('modal-deal', 'is_open'), prevent_initial_call=True
)
def toggle_deal_modal(o, c, s, is_open):
    return not is_open


# Placeholder Modal
@app.callback(
    Output('modal-placeholder', 'is_open'),
    [Input('btn-open-placeholder', 'n_clicks'), Input('btn-cancel-ph', 'n_clicks'), Input('btn-submit-ph', 'n_clicks')],
    State('modal-placeholder', 'is_open'), prevent_initial_call=True
)
def toggle_placeholder_modal(o, c, s, is_open):
    return not is_open


# Pipeline Modal
@app.callback(
    Output('modal-pipeline', 'is_open'),
    [Input('btn-open-pipeline', 'n_clicks'), Input('btn-cancel-pipe', 'n_clicks'),
     Input('btn-submit-pipe', 'n_clicks')],
    State('modal-pipeline', 'is_open'), prevent_initial_call=True
)
def toggle_pipeline_modal(o, c, s, is_open):
    return not is_open


# Auto-calculate Unfunded
@app.callback(
    Output('unfunded-display', 'children'),
    [Input('in-commitment', 'value'), Input('in-size', 'value')]
)
def calculate_unfunded(commitment, nav):
    if commitment is None or nav is None:
        return "$0.0M"
    unfunded = float(commitment) - float(nav)
    return f"${unfunded:.1f}M"


# Add Deal
@app.callback(
    Output('deals-store', 'data', allow_duplicate=True),
    Input('btn-submit', 'n_clicks'),
    [State('in-name', 'value'), State('in-manager', 'value'), State('in-strategy', 'value'),
     State('in-stage', 'value'), State('in-commitment', 'value'), State('in-size', 'value'),
     State('in-irr', 'value'), State('in-hold', 'value'), State('in-currency', 'value'),
     State('in-vintage', 'value'), State('in-segment', 'value'), State('in-allocation-status', 'value'),
     State('in-sector', 'value'), State('in-geo', 'value'),
     State('deals-store', 'data'), State('config-store', 'data')],
    prevent_initial_call=True
)
def add_deal(n, name, manager, strat, stage, commitment, size, irr, hold, currency, vint, segment,
             alloc_status, sec, geo, deals, config):
    if not all([name, manager, strat, commitment, size, irr]):
        return deals

    # NOTE: Current portfolio deals have ACTUAL NAV and do NOT reduce dry powder
    # Dry powder is only reduced by pipeline/future deals

    unfunded = float(commitment) - float(size)

    return deals + [{
        'name': name,
        'manager': manager,
        'strategy': strat,
        'stage': stage or 'Buyout',
        'commitment': float(commitment),  # Total commitment
        'size': float(size),  # Current NAV
        'unfunded': unfunded,  # Auto-calculated
        'target_irr': float(irr) / 100,
        'hold_period': float(hold) if hold else 5.0,
        'moic': (1 + float(irr) / 100) ** (float(hold) if hold else 5.0),
        'currency': currency or 'USD',
        'vintage': int(vint) if vint else 2024,
        'segment': segment or 'Seed',
        'allocation_status': alloc_status or 'Closed',
        'sector': sec or 'Diversified',
        'geography': geo or 'North America',
        'date_added': datetime.now().isoformat(),
        'is_actual': True
    }]


# Add Placeholder
@app.callback(
    Output('placeholder-deals-store', 'data', allow_duplicate=True),
    Input('btn-submit-ph', 'n_clicks'),
    [State('in-ph-name', 'value'), State('in-ph-strategy', 'value'),
     State('in-ph-size', 'value'), State('in-ph-bite', 'value'),
     State('in-ph-weight', 'value'), State('in-ph-month', 'value'),
     State('in-ph-region', 'value'), State('in-ph-type', 'value'),
     State('placeholder-deals-store', 'data')],
    prevent_initial_call=True
)
def add_placeholder(n, name, strat, size, bite, weight, month, region, deal_type, placeholders):
    if not all([name, strat, size is not None, month is not None]):
        return placeholders

    return placeholders + [{
        'name': name, 'strategy': strat, 'size': float(size),
        'custom_bite_size': float(bite) if bite else None,
        'target_weight': float(weight) if weight else None,
        'expected_month': int(month),
        'region': region or 'North America',
        'deal_type': deal_type or 'Secondary',
        'date_added': datetime.now().isoformat()
    }]


# Add Pipeline Deal
@app.callback(
    Output('pipeline-store', 'data', allow_duplicate=True),
    Input('btn-submit-pipe', 'n_clicks'),
    [State('in-pipe-name', 'value'), State('in-pipe-type', 'value'),
     State('in-pipe-stage', 'value'), State('in-pipe-size', 'value'),
     State('in-pipe-irr', 'value'), State('pipeline-store', 'data')],
    prevent_initial_call=True
)
def add_pipeline(n, name, ptype, stage, size, irr, pipeline):
    if not all([name, ptype]):
        return pipeline

    return pipeline + [{
        'name': name, 'type': ptype, 'stage': stage or 'Screening',
        'size': float(size) if size else 0, 'target_irr': float(irr) / 100 if irr else 0,
        'date_added': datetime.now().isoformat()
    }]


# Delete Deal
@app.callback(
    Output('deals-store', 'data', allow_duplicate=True),
    Input({'type': 'delete-deal', 'index': ALL}, 'n_clicks'),
    State('deals-store', 'data'),
    prevent_initial_call=True
)
def delete_deal(n_clicks, deals):
    if not any(n_clicks):
        return deals
    ctx = callback_context
    if not ctx.triggered:
        return deals
    button_id = ctx.triggered[0]['prop_id'].split('.')[0]
    if button_id == '':
        return deals
    button_info = json.loads(button_id)
    delete_idx = button_info['index']
    if 0 <= delete_idx < len(deals):
        deals.pop(delete_idx)
    return deals


# Delete Placeholder
@app.callback(
    Output('placeholder-deals-store', 'data', allow_duplicate=True),
    Input({'type': 'delete-placeholder', 'index': ALL}, 'n_clicks'),
    State('placeholder-deals-store', 'data'),
    prevent_initial_call=True
)
def delete_placeholder(n_clicks, placeholders):
    if not any(n_clicks):
        return placeholders
    ctx = callback_context
    if not ctx.triggered:
        return placeholders
    button_id = ctx.triggered[0]['prop_id'].split('.')[0]
    if button_id == '':
        return placeholders
    button_info = json.loads(button_id)
    delete_idx = button_info['index']
    if 0 <= delete_idx < len(placeholders):
        placeholders.pop(delete_idx)
    return placeholders


# Portfolio Table
@app.callback(Output('portfolio-table', 'children'), Input('deals-store', 'data'))
def update_portfolio_table(deals):
    if not deals:
        return dbc.Alert("📝 No deals yet. Click 'Add Deal' to start.", color="info")
    rows = []
    for idx, d in enumerate(deals):
        commitment = d.get('commitment', d['size'])
        nav = d['size']
        unfunded = d.get('unfunded', commitment - nav)

        # Allocation status badge color
        status = d.get('allocation_status', 'Closed')
        if status == 'Closed':
            status_color = 'success'
        elif status == 'Pending Close':
            status_color = 'warning'
        else:
            status_color = 'info'

        rows.append(dbc.Card([dbc.CardBody([
            dbc.Row([
                dbc.Col([
                    html.H5([
                        d['name'],
                        dbc.Badge(status, color=status_color, className="ms-2", style={'fontSize': '10px'})
                    ], className="mb-1", style={'fontFamily': C['sans'], 'color': C['text']}),
                    html.Small([
                        html.Strong(d.get('manager', 'N/A'), style={'color': C['blue']}),
                        f" • {d['strategy']} • {d.get('stage', 'Buyout')} • {d['sector']} • {d['geography']}"
                    ], className="text-muted", style={'fontFamily': C['mono'], 'fontSize': '12px'})
                ], width=4),
                dbc.Col([
                    html.Div([
                        html.Strong("Commitment: ",
                                    style={'fontFamily': C['sans'], 'color': C['muted'], 'fontSize': '11px'}),
                        html.Span(f"${commitment:.1f}M {d.get('currency', 'USD')}",
                                  style={'fontFamily': C['mono'], 'color': C['blue'], 'fontSize': '13px'}),
                        html.Br(),
                        html.Strong("Current NAV: ",
                                    style={'fontFamily': C['sans'], 'color': C['muted'], 'fontSize': '11px'}),
                        html.Span(f"${nav:.1f}M {d.get('currency', 'USD')}",
                                  style={'fontFamily': C['mono'], 'color': C['green'], 'fontSize': '13px'}),
                        html.Br(),
                        html.Strong("Unfunded: ",
                                    style={'fontFamily': C['sans'], 'color': C['muted'], 'fontSize': '11px'}),
                        html.Span(f"${unfunded:.1f}M",
                                  style={'fontFamily': C['mono'], 'color': C['amber'], 'fontSize': '13px'})
                    ])
                ], width=3),
                dbc.Col([
                    html.Div([
                        html.Strong("IRR: ", style={'fontFamily': C['sans'], 'color': C['muted'], 'fontSize': '11px'}),
                        html.Span(f"{d['target_irr']:.1%}",
                                  style={'fontFamily': C['mono'], 'color': C['text'], 'fontSize': '13px'}),
                        html.Br(),
                        html.Strong("MOIC: ", style={'fontFamily': C['sans'], 'color': C['muted'], 'fontSize': '11px'}),
                        html.Span(f"{d.get('moic', 0):.2f}x",
                                  style={'fontFamily': C['mono'], 'color': C['text'], 'fontSize': '13px'}),
                        html.Br(),
                        html.Strong("Hold: ", style={'fontFamily': C['sans'], 'color': C['muted'], 'fontSize': '11px'}),
                        html.Span(f"{d.get('hold_period', 5):.1f}y",
                                  style={'fontFamily': C['mono'], 'color': C['text'], 'fontSize': '13px'}),
                    ])
                ], width=2),
                dbc.Col([
                    html.Div([
                        html.Strong("Vintage: ",
                                    style={'fontFamily': C['sans'], 'color': C['muted'], 'fontSize': '11px'}),
                        html.Span(str(d.get('vintage', 2024)),
                                  style={'fontFamily': C['mono'], 'color': C['text'], 'fontSize': '13px'}),
                        html.Br(),
                        html.Strong("Segment: ",
                                    style={'fontFamily': C['sans'], 'color': C['muted'], 'fontSize': '11px'}),
                        html.Span(d.get('segment', 'Seed'),
                                  style={'fontFamily': C['mono'], 'color': C['purple'], 'fontSize': '11px'}),
                        html.Br(),
                        html.Strong("Added: ",
                                    style={'fontFamily': C['sans'], 'color': C['muted'], 'fontSize': '11px'}),
                        html.Span(d.get('date_added', '')[:10] if d.get('date_added') else 'N/A',
                                  style={'fontFamily': C['mono'], 'color': C['text'], 'fontSize': '11px'})
                    ])
                ], width=2),
                dbc.Col([
                    dbc.Button("🗑️", id={'type': 'delete-deal', 'index': idx},
                               color="danger", size="sm", outline=True)
                ], width=1, className="text-end")
            ])
        ])], className="mb-2 shadow-sm"))
    return html.Div(rows)


# Placeholder Table
@app.callback(Output('placeholder-table', 'children'), Input('placeholder-deals-store', 'data'))
def update_placeholder_table(placeholders):
    if not placeholders:
        return dbc.Alert("📝 No placeholder deals yet. Add future deals to forecast dry powder.", color="info")
    rows = []
    for idx, pd in enumerate(placeholders):
        # Month index starts from Jan 2026
        base_date = datetime(2026, 1, 1)
        month_date = base_date + relativedelta(months=pd['expected_month'])
        month_name = month_date.strftime('%b %Y')
        rows.append(dbc.Card([dbc.CardBody([
            dbc.Row([
                dbc.Col([
                    html.H5(pd['name'], className="mb-1", style={'fontFamily': C['sans'], 'color': C['text']}),
                    html.Small(f"{pd['strategy']} • Expected: {month_name}",
                               className="text-muted", style={'fontFamily': C['mono']})
                ], width=8),
                dbc.Col([
                    html.H5(f"${pd['size']:.1f}M", className="text-end mb-1",
                            style={'color': C['green'], 'fontFamily': C['mono']})
                ], width=2),
                dbc.Col([
                    dbc.Button("🗑️", id={'type': 'delete-placeholder', 'index': idx},
                               color="danger", size="sm", outline=True)
                ], width=2, className="text-end")
            ])
        ])], className="mb-2 shadow-sm", style={'backgroundColor': C['panel'], 'border': f'1px solid {C["border"]}'}))
    return html.Div(rows)


# Pipeline Table
@app.callback(Output('pipeline-table', 'children'), Input('pipeline-store', 'data'))
def update_pipeline_table(pipeline):
    if not pipeline:
        return dbc.Alert("📝 No pipeline deals yet.", color="info")

    data = [{
        'Deal': p['name'],
        'Type': p['type'],
        'Stage': p['stage'],
        'Size': f"${p['size']:.1f}M" if p['size'] > 0 else "TBD",
        'IRR': f"{p['target_irr']:.1%}" if p['target_irr'] > 0 else "TBD"
    } for p in pipeline]

    return dash_table.DataTable(
        data=data,
        columns=[{"name": c, "id": c} for c in data[0].keys()] if data else [],
        style_cell={'textAlign': 'left', 'padding': '12px', 'fontFamily': C['mono'], 'fontSize': '11px'},
        style_header={
            'backgroundColor': C['surface'],
            'color': C['text'],
            'fontWeight': 'bold',
            'border': f'1px solid {C["border"]}'
        },
        style_data={
            'backgroundColor': C['panel'],
            'color': C['text'],
            'border': f'1px solid {C["border"]}'
        },
        style_data_conditional=[{
            'if': {'row_index': 'odd'},
            'backgroundColor': C['surface']
        }]
    )


# Dashboard KPIs
@app.callback(
    [Output('dash-nav', 'children'), Output('dash-num-deals', 'children'),
     Output('dash-dry-powder', 'children'), Output('dash-total', 'children'),
     Output('dash-current-irr', 'children'), Output('dash-req-irr', 'children'),
     Output('dash-placeholders', 'children'), Output('dash-placeholder-value', 'children')],
    [Input('deals-store', 'data'), Input('placeholder-deals-store', 'data'),
     Input('config-store', 'data'), Input('liquidity-data-store', 'data'),
     Input('fund-cf-data-store', 'data')]
)
def update_dashboard_kpis(deals, placeholders, config, liquidity_data, fund_cf_data):
    """
    Total NAV = Investment NAV + Cash + GLF + CQS

    Sources:
    - Investment NAV: From fund_cf_data (monthly) OR deals manually entered
    - Cash, GLF, CQS: From liquidity_data (Rows 58, 59, 60)
    """
    # Calculate investment NAV from deals
    m = calculate_portfolio_metrics(deals)
    investment_nav = m['total_nav']  # NAV from manually entered deals

    # Get Cash, GLF, CQS from liquidity upload (if available)
    cash = 0
    glf = 0
    cqs = 0

    if liquidity_data:
        cash = liquidity_data.get('current_cash', 0)
        glf = liquidity_data.get('glf_balance', 0)
        cqs = liquidity_data.get('cqs_balance', 0)

    # If fund_cf_data is available, use it for investment NAV
    # (This would override manually entered deals if both exist)
    if fund_cf_data and len(fund_cf_data) > 0:
        # Sum up latest NAV from fund CF data
        # This represents the most recent monthly snapshot
        fund_cf_nav = sum(deal.get('nav', deal.get('size', 0)) for deal in fund_cf_data)
        investment_nav = fund_cf_nav

    # TOTAL NAV = Investment NAV + Cash + GLF + CQS
    total_nav = investment_nav + cash + glf + cqs

    # Dry powder calculation
    dry_powder = config['fund_parameters']['dry_powder']

    # Required IRR calculation (based on investment NAV only, not cash/GLF/CQS)
    req_irr = calculate_required_future_irr(m['weighted_irr'], investment_nav, dry_powder, config)

    # Placeholder stats
    num_placeholders = len(placeholders)
    total_placeholder_value = sum(p['size'] for p in placeholders)

    return (
        f"${total_nav:.1f}M",  # Shows Total NAV (investments + cash + GLF + CQS)
        f"{m['num_deals']} deals",
        f"${dry_powder:.0f}M",
        f"${total_nav + dry_powder:.0f}M",  # Total Fund = Total NAV + Dry Powder
        f"{m['weighted_irr']:.1%}",
        f"{req_irr:.1%}",
        str(num_placeholders),
        f"${total_placeholder_value:.0f}M"
    )


# Dashboard Charts
@app.callback(
    [Output('dash-allocation-chart', 'figure'), Output('dash-forecast-chart', 'figure'),
     Output('dash-bite-sizing', 'children')],
    [Input('deals-store', 'data'), Input('placeholder-deals-store', 'data'), Input('config-store', 'data')]
)
def update_dashboard_charts(deals, placeholders, config):
    m = calculate_portfolio_metrics(deals)
    dry_powder = config['fund_parameters']['dry_powder']

    # Allocation Pie
    if m['by_strategy']:
        labels = list(m['by_strategy'].keys())
        values = [m['by_strategy'][s]['nav'] for s in labels]
        colors_pie = [C['blue'], C['purple'], C['teal'], C['green']][:len(labels)]

        fig_alloc = go.Figure(data=[go.Pie(
            labels=labels, values=values, hole=0.4,
            marker=dict(colors=colors_pie, line=dict(color=C['border'], width=1)),
            textfont=dict(color=C['text'], family=C['mono'])
        )])
        fig_alloc.update_layout(**CHART_BASE, height=300, margin=dict(t=20, b=0, l=0, r=0), showlegend=False)
    else:
        fig_alloc = go.Figure()
        fig_alloc.update_layout(**CHART_BASE)

    # Forecast
    forecast = forecast_dry_powder(m['total_nav'], dry_powder, deals, placeholders, config, 12)
    months = [f['month'] for f in forecast]
    dp_values = [f['dry_powder'] for f in forecast]

    fig_forecast = go.Figure()
    fig_forecast.add_trace(go.Scatter(
        x=months, y=dp_values, mode='lines+markers', name='Dry Powder',
        line=dict(color=C['blue'], width=3),
        fill='tozeroy',
        fillcolor=rgba(C['blue'], 0.2),
        marker=dict(size=6, color=C['sky'], line=dict(color=C['blue'], width=2))
    ))
    fig_forecast.update_layout(
        **CHART_BASE,
        height=300,
        xaxis_title="Month",
        yaxis_title="Dry Powder ($mm)",
        hovermode='x unified'
    )

    # Bite Sizing
    bite_sizes = calculate_bite_sizes(dry_powder, config)
    bite_cards = []
    for strategy, sizes in bite_sizes.items():
        bite_cards.append(html.Div([
            html.H6(strategy, style={'fontSize': '12px', 'fontWeight': 'bold'}),
            html.Small(f"Min: ${sizes['min']:.1f}M ({sizes['min_pct']:.1%})", className="d-block text-muted"),
            html.Small(f"Desired: ${sizes['desired']:.1f}M ({sizes['desired_pct']:.1%})",
                       className="d-block text-success"),
            html.Small(f"Max: ${sizes['max']:.1f}M ({sizes['max_pct']:.1%})", className="d-block text-danger"),
            html.Hr(className="my-2")
        ]))

    return fig_alloc, fig_forecast, html.Div(bite_cards)


# Portfolio Page Metrics
@app.callback(
    [Output('port-nav', 'children'), Output('port-dry-powder', 'children'),
     Output('port-total-fund', 'children'), Output('port-target-return', 'children')],
    [Input('deals-store', 'data'), Input('config-store', 'data')]
)
def update_portfolio_metrics(deals, config):
    m = calculate_portfolio_metrics(deals)
    dry_powder = config['fund_parameters']['dry_powder']
    req_irr = calculate_required_future_irr(m['weighted_irr'], m['total_nav'], dry_powder, config)
    return (
        f"${m['total_nav']:.1f}M",
        f"${dry_powder:.0f}M",
        f"${m['total_nav'] + dry_powder:.0f}M",
        f"{req_irr:.1%}"
    )


# Calculator Page
@app.callback(
    [Output('calc-current-nav', 'children'), Output('calc-current-irr', 'children'),
     Output('calc-num-deals', 'children'), Output('calc-total-fund', 'children'),
     Output('calc-target-twr', 'children'), Output('calc-dry-powder', 'children'),
     Output('calc-required-irr', 'children'), Output('calc-explanation', 'children'),
     Output('calc-waterfall', 'children')],
    [Input('deals-store', 'data'), Input('config-store', 'data')]
)
def update_calculator(deals, config):
    m = calculate_portfolio_metrics(deals)
    dry_powder = config['fund_parameters']['dry_powder']
    total_fund = m['total_nav'] + dry_powder
    target_twr = config['fund_parameters']['target_net_twr']

    req_irr = calculate_required_future_irr(m['weighted_irr'], m['total_nav'], dry_powder, config)
    gap = req_irr - m['weighted_irr'] if m['total_nav'] > 0 else 0
    explanation = f"{'Higher' if gap > 0 else 'Lower'} than current portfolio by {abs(gap):.1%}" if m[
                                                                                                        'total_nav'] > 0 else "Add deals to calculate"

    # Waterfall
    invested_ratio = max(0, 1 - config['fund_parameters']['liquidity_reserve_pct'])
    idle_ratio = config['fund_parameters']['liquidity_reserve_pct']

    waterfall_steps = [
        ("1. Target Net TWR (to LPs)", f"{target_twr:.1%}"),
        ("2. + Management Fee", f"{config['fund_parameters']['management_fee']:.2%}"),
        ("3. + Loss Drag (impairments)", f"{config['fund_parameters']['loss_drag']:.1%}"),
        ("4. - Cash Yield on reserves", f"-{idle_ratio * config['fund_parameters']['cash_yield']:.2%}"),
        ("5. ÷ Invested Ratio", f"÷ {invested_ratio:.1%}"),
        ("6. + Carry Drag (if > hurdle)", f"+Variable ({config['fund_parameters']['carry_rate']:.0%})"),
        ("═══════════════════════", "═══════════"),
        ("REQUIRED GROSS DEAL RETURN", f"{req_irr:.1%}")
    ]

    waterfall_div = dbc.Table([
        html.Tbody([
            html.Tr([
                html.Td(step, style={'fontSize': '14px', 'fontWeight': 'bold' if '═' in step else 'normal'}),
                html.Td(val, className="text-end", style={
                    'fontSize': '14px',
                    'fontWeight': 'bold',
                    'color': COLORS['success'] if '═' in step else COLORS['dark']
                })
            ]) for step, val in waterfall_steps
        ])
    ], bordered=True, striped=True, hover=True, className="shadow-sm")

    return (
        f"${m['total_nav']:.1f}M",
        f"{m['weighted_irr']:.1%}",
        str(m['num_deals']),
        f"${total_fund:.0f}M",
        f"{target_twr:.1%}",
        f"${dry_powder:.0f}M",
        f"{req_irr:.1%}",
        explanation,
        waterfall_div
    )


# Dry Powder Page
@app.callback(
    [Output('dp-current', 'children'), Output('dp-planned', 'children'),
     Output('dp-distributions', 'children'), Output('dp-forecast-12m', 'children'),
     Output('dp-forecast-chart', 'figure'), Output('dp-monthly-table', 'children')],
    [Input('deals-store', 'data'), Input('placeholder-deals-store', 'data'), Input('config-store', 'data')]
)
def update_drypowder_page(deals, placeholders, config):
    m = calculate_portfolio_metrics(deals)
    dry_powder = config['fund_parameters']['dry_powder']

    total_planned = sum(p['size'] for p in placeholders)
    annual_dist = m['total_nav'] * config['fund_parameters']['distribution_rate']

    forecast = forecast_dry_powder(m['total_nav'], dry_powder, deals, placeholders, config, 12)
    forecast_12m = forecast[-1]['dry_powder']

    # Chart
    months = [f['month'] for f in forecast]
    dp_vals = [f['dry_powder'] for f in forecast]
    calls = [f['calls'] for f in forecast]
    dists = [f['distributions'] for f in forecast]

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(go.Scatter(
        x=months, y=dp_vals, name='Dry Powder',
        line=dict(color=C['blue'], width=3),
        fill='tozeroy',
        fillcolor=rgba(C['blue'], 0.2),
        marker=dict(size=6, color=C['sky'])
    ), secondary_y=False)
    fig.add_trace(go.Bar(
        x=months, y=calls, name='Capital Calls',
        marker_color=C['red'],
        opacity=0.7
    ), secondary_y=True)
    fig.add_trace(go.Bar(
        x=months, y=dists, name='Distributions',
        marker_color=C['green'],
        opacity=0.7
    ), secondary_y=True)

    fig.update_xaxes(title_text="Month", gridcolor=C['border'])
    fig.update_yaxes(title_text="Dry Powder ($mm)", secondary_y=False, gridcolor=C['border'])
    fig.update_yaxes(title_text="Flows ($mm)", secondary_y=True, gridcolor=C['border'])
    fig.update_layout(**CHART_BASE, height=400, hovermode='x unified')

    # Table
    table_data = [{
        'Month': f['month'],
        'Dry Powder': f"${f['dry_powder']:.1f}M",
        'NAV': f"${f['nav']:.1f}M",
        'Distributions': f"${f['distributions']:.1f}M",
        'Calls': f"${f['calls']:.1f}M" if f['calls'] > 0 else "-"
    } for f in forecast]

    table = dash_table.DataTable(
        data=table_data,
        columns=[{"name": c, "id": c} for c in table_data[0].keys()],
        style_cell={'textAlign': 'left', 'padding': '10px', 'fontFamily': C['mono'], 'fontSize': '11px'},
        style_header={
            'backgroundColor': C['surface'],
            'color': C['text'],
            'fontWeight': 'bold',
            'border': f'1px solid {C["border"]}'
        },
        style_data={
            'backgroundColor': C['panel'],
            'color': C['text'],
            'border': f'1px solid {C["border"]}'
        },
        style_data_conditional=[{
            'if': {'row_index': 'odd'},
            'backgroundColor': C['surface']
        }],
        page_size=12
    )

    return (
        f"${dry_powder:.0f}M",
        f"${total_planned:.0f}M",
        f"${annual_dist:.0f}M",
        f"${forecast_12m:.0f}M",
        fig,
        table
    )


# ==================== TWR FORECASTER CALLBACKS ====================

@app.callback(
    Output('twr-current-irr', 'children'),
    Input('deals-store', 'data')
)
def update_twr_current_irr(deals):
    m = calculate_portfolio_metrics(deals)
    return f"{m['weighted_irr']:.1%}"


@app.callback(
    [Output('twr-distribution-chart', 'figure'), Output('twr-statistics', 'children'),
     Output('twr-sensitivity-chart', 'figure')],
    Input('btn-run-twr', 'n_clicks'),
    [State('deals-store', 'data'), State('config-store', 'data'),
     State('twr-future-mean', 'value'), State('twr-future-std', 'value'),
     State('twr-n-sims', 'value')],
    prevent_initial_call=True
)
def run_twr_simulation(n, deals, config, future_mean, future_std, n_sims):
    m = calculate_portfolio_metrics(deals)
    current_irr = m['weighted_irr']
    current_nav = m['total_nav']
    dry_powder = config['fund_parameters']['dry_powder']
    target_twr = config['fund_parameters']['target_net_twr']

    # Run Monte Carlo
    np.random.seed(42)
    future_irrs = np.random.normal(future_mean / 100, future_std / 100, n_sims)

    results = []
    for future_irr in future_irrs:
        # Blended portfolio IRR
        total = current_nav + dry_powder
        if total > 0:
            blended = (current_nav * current_irr + dry_powder * future_irr) / total
        else:
            blended = 0

        # Apply fees and drag
        mgmt_fee = config['fund_parameters']['management_fee']
        loss_drag = config['fund_parameters']['loss_drag']
        liq_reserve = config['fund_parameters']['liquidity_reserve_pct']
        cash_yield = config['fund_parameters']['cash_yield']
        carry_rate = config['fund_parameters']['carry_rate']
        hurdle = config['fund_parameters']['hurdle_rate']

        invested_ratio = 1 - liq_reserve
        gross_before_carry = (blended * invested_ratio) + (cash_yield * liq_reserve) - mgmt_fee - loss_drag

        # Carry drag
        if gross_before_carry > hurdle:
            carry_drag = (gross_before_carry - hurdle) * carry_rate
            net_twr = gross_before_carry - carry_drag
        else:
            net_twr = gross_before_carry

        results.append(net_twr)

    results = np.array(results)

    # Distribution chart
    fig_dist = go.Figure()
    fig_dist.add_trace(go.Histogram(
        x=results * 100, nbinsx=50,
        marker_color=C['blue'], opacity=0.7,
        name='Simulated TWR'
    ))
    fig_dist.add_vline(x=target_twr * 100, line_color=C['red'], line_width=3, line_dash='dash',
                       annotation_text=f"Target: {target_twr:.1%}", annotation_position="top right")
    fig_dist.add_vline(x=np.median(results) * 100, line_color=C['green'], line_width=3,
                       annotation_text=f"Median: {np.median(results):.1%}", annotation_position="top left")
    fig_dist.update_layout(
        **CHART_BASE,
        xaxis_title="Net TWR (%)", yaxis_title="Frequency",
        title=f"Distribution of Net TWR ({n_sims:,} Simulations)",
        height=400
    )

    # Statistics
    prob_above_target = (results >= target_twr).sum() / len(results)
    percentile_5 = np.percentile(results, 5)
    percentile_95 = np.percentile(results, 95)
    sharpe = (np.mean(results) - 0.03) / np.std(results)  # Assuming 3% risk-free

    stats_div = html.Div([
        html.Div([
            html.Strong("Probability of Hitting Target:", style={'color': C['text'], 'fontFamily': C['sans']}),
            html.H3(f"{prob_above_target:.1%}", style={
                'color': C['green'] if prob_above_target >= 0.75 else C['amber'] if prob_above_target >= 0.5 else C[
                    'red'],
                'fontFamily': C['mono'], 'marginTop': '0.5rem'
            })
        ], className="mb-3"),
        html.Hr(style={'borderColor': C['border']}),
        html.Div([
            html.P([html.Strong("Mean TWR: "),
                    html.Span(f"{np.mean(results):.2%}", style={'color': C['blue'], 'fontFamily': C['mono']})]),
            html.P([html.Strong("Median TWR: "),
                    html.Span(f"{np.median(results):.2%}", style={'color': C['green'], 'fontFamily': C['mono']})]),
            html.P([html.Strong("Std Dev: "),
                    html.Span(f"{np.std(results):.2%}", style={'color': C['muted'], 'fontFamily': C['mono']})]),
            html.P([html.Strong("5th Percentile: "),
                    html.Span(f"{percentile_5:.2%}", style={'color': C['red'], 'fontFamily': C['mono']})]),
            html.P([html.Strong("95th Percentile: "),
                    html.Span(f"{percentile_95:.2%}", style={'color': C['green'], 'fontFamily': C['mono']})]),
            html.P([html.Strong("Sharpe Ratio: "),
                    html.Span(f"{sharpe:.2f}", style={'color': C['purple'], 'fontFamily': C['mono']})]),
        ], style={'fontSize': '14px'})
    ])

    # Sensitivity Analysis
    future_irr_range = np.linspace(0.15, 0.35, 20)
    twr_sensitivity = []
    for fir in future_irr_range:
        total = current_nav + dry_powder
        blended = (current_nav * current_irr + dry_powder * fir) / total if total > 0 else 0
        invested_ratio = 1 - liq_reserve
        gross = (blended * invested_ratio) + (cash_yield * liq_reserve) - mgmt_fee - loss_drag
        if gross > hurdle:
            net = gross - (gross - hurdle) * carry_rate
        else:
            net = gross
        twr_sensitivity.append(net)

    fig_sens = go.Figure()
    fig_sens.add_trace(go.Scatter(
        x=future_irr_range * 100, y=np.array(twr_sensitivity) * 100,
        mode='lines+markers', name='Net TWR',
        line=dict(color=C['blue'], width=3),
        marker=dict(size=6, color=C['sky'])
    ))
    fig_sens.add_hline(y=target_twr * 100, line_color=C['red'], line_dash='dash',
                       annotation_text=f"Target: {target_twr:.1%}")
    fig_sens.update_layout(
        **CHART_BASE,
        xaxis_title="Future Deal IRR (%)", yaxis_title="Net TWR (%)",
        title="How Net TWR Changes with Future Deal Performance",
        height=400
    )

    return fig_dist, stats_div, fig_sens


# ==================== PORTFOLIO SEGMENTATION CALLBACKS ====================

# Segment Summary Cards
@app.callback(
    [Output('seg-total-nav', 'children'), Output('seg-total-twr', 'children'),
     Output('seg-seed-nav', 'children'), Output('seg-seed-twr', 'children'),
     Output('seg-new-nav', 'children'), Output('seg-new-twr', 'children'),
     Output('seg-mm-nav', 'children'), Output('seg-mm-twr', 'children'),
     Output('seg-pipeline-nav', 'children'), Output('seg-pipeline-twr', 'children'),
     Output('seg-future-nav', 'children'), Output('seg-future-twr', 'children')],
    [Input('deals-store', 'data'), Input('pipeline-store', 'data'),
     Input('placeholder-deals-store', 'data')]
)
def update_segment_summary(deals, pipeline, placeholders):
    """Calculate NAV and TWR for each segment"""

    # Segment deals
    seed_deals = [d for d in deals if d.get('segment', 'Seed') == 'Seed'] if deals else []
    new_deals = [d for d in deals if d.get('segment', 'Seed') == 'New'] if deals else []
    mm_deals = [d for d in deals if d.get('segment', 'Seed') == 'MoneyMarket'] if deals else []

    # Calculate NAVs
    seed_nav = sum(d['size'] for d in seed_deals)
    new_nav = sum(d['size'] for d in new_deals)
    mm_nav = sum(d['size'] for d in mm_deals)
    pipeline_nav = sum(p['size'] for p in pipeline) if pipeline else 0
    future_nav = sum(p['size'] for p in placeholders) if placeholders else 0
    total_nav = seed_nav + new_nav + mm_nav

    # Calculate TWRs (simplified - would use actual cashflows in production)
    seed_twr = sum(d['target_irr'] * d['size'] for d in seed_deals) / seed_nav if seed_nav > 0 else 0
    new_twr = sum(d['target_irr'] * d['size'] for d in new_deals) / new_nav if new_nav > 0 else 0
    mm_twr = 0.03  # Money market assumed 3%
    pipeline_twr = sum(
        p['target_irr'] * p['size'] for p in pipeline) / pipeline_nav if pipeline_nav > 0 and pipeline else 0
    future_twr = 0.25  # Placeholder assumed target
    total_twr = (seed_nav * seed_twr + new_nav * new_twr + mm_nav * mm_twr) / total_nav if total_nav > 0 else 0

    return (
        f"${total_nav:.1f}M", f"TWR: {total_twr:.1%}",
        f"${seed_nav:.1f}M", f"TWR: {seed_twr:.1%}",
        f"${new_nav:.1f}M", f"TWR: {new_twr:.1%}",
        f"${mm_nav:.1f}M", f"TWR: {mm_twr:.1%}",
        f"${pipeline_nav:.1f}M", f"Est TWR: {pipeline_twr:.1%}",
        f"${future_nav:.1f}M", f"Target TWR: {future_twr:.1%}"
    )


# TWR Forecast Chart
@app.callback(
    Output('seg-twr-forecast-chart', 'figure'),
    [Input('deals-store', 'data'), Input('config-store', 'data')]
)
def generate_twr_forecast_by_segment(deals, config):
    """Generate 12-month TWR forecast for each segment"""

    base_date = datetime(2026, 1, 1)
    months = []

    # Initialize TWR arrays for each segment
    total_twr = []
    seed_twr = []
    new_twr = []
    mm_twr = []

    target_twr = config['fund_parameters']['target_net_twr']

    for i in range(12):
        month_date = base_date + relativedelta(months=i)
        months.append(month_date.strftime('%b %Y'))

        # Cumulative TWR calculation (simplified)
        cumulative_factor = (1 + target_twr) ** (i / 12)

        # Different growth rates per segment
        total_twr.append((cumulative_factor - 1) * 100)
        seed_twr.append((cumulative_factor * 0.95 - 1) * 100)  # Slightly lower
        new_twr.append((cumulative_factor * 1.05 - 1) * 100)  # Slightly higher
        mm_twr.append((((1 + 0.03) ** (i / 12)) - 1) * 100)  # 3% fixed

    fig = go.Figure()

    fig.add_trace(go.Scatter(
        x=months, y=total_twr, name='Total Portfolio',
        line=dict(color=C['blue'], width=3), mode='lines+markers'
    ))
    fig.add_trace(go.Scatter(
        x=months, y=seed_twr, name='Seed Portfolio',
        line=dict(color=C['green'], width=2, dash='dash'), mode='lines'
    ))
    fig.add_trace(go.Scatter(
        x=months, y=new_twr, name='New Deals',
        line=dict(color=C['purple'], width=2, dash='dot'), mode='lines'
    ))
    fig.add_trace(go.Scatter(
        x=months, y=mm_twr, name='Money Market',
        line=dict(color=C['amber'], width=2, dash='dashdot'), mode='lines'
    ))

    fig.update_layout(
        **CHART_BASE,
        yaxis_title='Cumulative TWR (%)',
        height=400,
        hovermode='x unified',
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1)
    )

    return fig


# Segment Tables
@app.callback(
    [Output('seg-seed-table', 'children'), Output('seg-new-table', 'children')],
    Input('deals-store', 'data')
)
def update_segment_tables(deals):
    """Display deals by segment"""

    if not deals:
        return (html.P("No seed deals", className="text-muted"),
                html.P("No new deals", className="text-muted"))

    seed_deals = [d for d in deals if d.get('segment', 'Seed') == 'Seed']
    new_deals = [d for d in deals if d.get('segment', 'Seed') == 'New']

    def create_segment_table(segment_deals, segment_name):
        if not segment_deals:
            return html.P(f"No {segment_name.lower()} yet", className="text-muted")

        data = [{
            'Deal': d['name'],
            'NAV': f"${d['size']:.1f}M",
            'IRR': f"{d['target_irr']:.1%}",
            'Vintage': d.get('vintage', 'N/A')
        } for d in segment_deals]

        return dash_table.DataTable(
            data=data,
            columns=[{"name": c, "id": c} for c in data[0].keys()],
            style_cell={'textAlign': 'left', 'padding': '10px', 'fontFamily': C['mono'], 'fontSize': '11px'},
            style_header={'backgroundColor': C['surface'], 'color': C['text'], 'fontWeight': 'bold'},
            style_data={'backgroundColor': C['panel'], 'color': C['text']},
            style_data_conditional=[{'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']}],
            page_size=10
        )

    return (create_segment_table(seed_deals, 'Seed'),
            create_segment_table(new_deals, 'New'))


# Allocation Chart
@app.callback(
    Output('seg-allocation-chart', 'figure'),
    [Input('deals-store', 'data'), Input('pipeline-store', 'data'),
     Input('placeholder-deals-store', 'data')]
)
def generate_allocation_chart(deals, pipeline, placeholders):
    """Show NAV allocation across segments"""

    # Calculate segment NAVs
    seed_nav = sum(d['size'] for d in deals if d.get('segment', 'Seed') == 'Seed') if deals else 0
    new_nav = sum(d['size'] for d in deals if d.get('segment', 'Seed') == 'New') if deals else 0
    mm_nav = sum(d['size'] for d in deals if d.get('segment', 'Seed') == 'MoneyMarket') if deals else 0
    pipeline_nav = sum(p['size'] for p in pipeline) if pipeline else 0
    future_nav = sum(p['size'] for p in placeholders) if placeholders else 0

    labels = []
    values = []
    colors = []

    if seed_nav > 0:
        labels.append('Seed Portfolio')
        values.append(seed_nav)
        colors.append(C['green'])
    if new_nav > 0:
        labels.append('New Deals')
        values.append(new_nav)
        colors.append(C['purple'])
    if mm_nav > 0:
        labels.append('Money Market')
        values.append(mm_nav)
        colors.append(C['amber'])
    if pipeline_nav > 0:
        labels.append('Pipeline')
        values.append(pipeline_nav)
        colors.append(C['teal'])
    if future_nav > 0:
        labels.append('Future/Placeholder')
        values.append(future_nav)
        colors.append(C['pink'])

    fig = go.Figure(data=[go.Pie(
        labels=labels, values=values, hole=0.4,
        marker=dict(colors=colors),
        textfont=dict(color=C['text'], family=C['mono'])
    )])

    fig.update_layout(**CHART_BASE, height=350, margin=dict(t=20, b=0, l=0, r=0), showlegend=True)

    return fig


# Contribution Chart
@app.callback(
    Output('seg-contribution-chart', 'figure'),
    Input('deals-store', 'data')
)
def generate_contribution_chart(deals):
    """Show TWR contribution by segment"""

    if not deals:
        fig = go.Figure()
        fig.update_layout(**CHART_BASE)
        return fig

    # Calculate weighted TWR contribution
    seed_deals = [d for d in deals if d.get('segment', 'Seed') == 'Seed']
    new_deals = [d for d in deals if d.get('segment', 'Seed') == 'New']
    mm_deals = [d for d in deals if d.get('segment', 'Seed') == 'MoneyMarket']

    seed_contribution = sum(d['target_irr'] * d['size'] for d in seed_deals)
    new_contribution = sum(d['target_irr'] * d['size'] for d in new_deals)
    mm_contribution = sum(0.03 * d['size'] for d in mm_deals)

    segments = []
    contributions = []
    colors_list = []

    if seed_contribution > 0:
        segments.append('Seed')
        contributions.append(seed_contribution * 100)
        colors_list.append(C['green'])
    if new_contribution > 0:
        segments.append('New')
        contributions.append(new_contribution * 100)
        colors_list.append(C['purple'])
    if mm_contribution > 0:
        segments.append('MM')
        contributions.append(mm_contribution * 100)
        colors_list.append(C['amber'])

    fig = go.Figure(data=[go.Bar(
        x=segments, y=contributions, marker_color=colors_list,
        text=[f"{v:.1f}" for v in contributions], textposition='outside',
        textfont=dict(color=C['text'], family=C['mono'])
    )])

    fig.update_layout(**CHART_BASE, yaxis_title="TWR Contribution", height=350)

    return fig


# ==================== CASHFLOWS CALLBACKS ====================

# Upload Fund CF Excel file
@app.callback(
    [Output('fund-cf-data-store', 'data'), Output('upload-fund-cf-status', 'children')],
    Input('upload-fund-cf', 'contents'),
    State('upload-fund-cf', 'filename'),
    prevent_initial_call=True
)
def upload_fund_cf_file(contents, filename):
    """Parse uploaded Fund Level CF Excel/CSV file - stores RAW cashflow values"""
    if contents is None:
        return None, ""

    try:
        # Decode the uploaded file
        content_type, content_string = contents.split(',')
        decoded = base64.b64decode(content_string)

        # Check if CSV or Excel
        is_csv = filename.lower().endswith('.csv')

        deals_data = []

        if is_csv:
            # Handle CSV file
            import csv
            from io import StringIO

            csv_string = decoded.decode('utf-8')
            csv_reader = csv.reader(StringIO(csv_string))
            rows = list(csv_reader)

            # Find header row
            header_row_idx = None
            for idx, row in enumerate(rows):
                if any('Investment' in str(cell) or 'AIC Name' in str(cell) for cell in row):
                    header_row_idx = idx
                    break

            if header_row_idx is not None:
                headers = rows[header_row_idx]

                # Process data rows
                for row in rows[header_row_idx + 1:]:
                    if len(row) > 5 and row[0] and 'Export' not in str(row[0]):
                        deal_data = {
                            'name': str(row[0]).strip() if len(row) > 0 else 'Unknown',
                            'type': str(row[1]).strip() if len(row) > 1 else '',
                            'commitment': 0,
                            'nav': 0,
                            'monthly_cfs': {}
                        }

                        # Try to parse commitment and NAV
                        try:
                            if len(row) > 2 and row[2]:
                                deal_data['commitment'] = float(str(row[2]).replace(',', '').replace('$', ''))
                        except:
                            pass

                        try:
                            if len(row) > 3 and row[3]:
                                deal_data['nav'] = float(str(row[3]).replace(',', '').replace('$', ''))
                        except:
                            pass

                        # Store monthly cashflows as-is (no calculations!)
                        for col_idx in range(4, min(len(row), 64)):
                            try:
                                val = str(row[col_idx]).replace(',', '').replace('$', '').strip()
                                if val and val not in ['', '-', 'N/A']:
                                    deal_data['monthly_cfs'][col_idx - 4] = float(val)
                            except:
                                pass

                        if deal_data['name'] != 'Unknown':
                            deals_data.append(deal_data)
        else:
            # Handle Excel file
            from io import BytesIO
            from openpyxl import load_workbook

            wb = load_workbook(BytesIO(decoded), data_only=True)

            # Try to find Fund Level CF sheet
            sheet_names = ['Fund Level CF', 'Fund_Level_CF', 'FundLevelCF', 'Sheet1']
            ws = None
            for name in sheet_names:
                if name in wb.sheetnames:
                    ws = wb[name]
                    break

            if ws is None:
                ws = wb[wb.sheetnames[0]]

            # Find header row
            header_row = None
            for row in range(1, min(20, ws.max_row + 1)):
                for col in range(1, min(10, ws.max_column + 1)):
                    val = ws.cell(row, col).value
                    if val and ('AIC Name' in str(val) or 'Investment' in str(val) or 'Deal' in str(val)):
                        header_row = row
                        break
                if header_row:
                    break

            if header_row:
                # Parse each deal row
                for row in range(header_row + 1, min(header_row + 100, ws.max_row + 1)):
                    # Look for deal name in various columns
                    deal_name = None
                    for col in [1, 2, 6]:  # Try columns A, B, F
                        val = ws.cell(row, col).value
                        if val and str(val).strip() and 'Export' not in str(val) and 'Total' not in str(val):
                            deal_name = str(val).strip()
                            break

                    if deal_name:
                        deal_data = {
                            'name': deal_name,
                            'type': str(ws.cell(row, 8).value or ''),
                            'commitment': 0,
                            'nav': 0,
                            'monthly_cfs': {}
                        }

                        # Extract commitment and NAV (try multiple columns)
                        for col in [9, 10, 11]:
                            try:
                                val = ws.cell(row, col).value
                                if val and isinstance(val, (int, float)):
                                    if deal_data['commitment'] == 0:
                                        deal_data['commitment'] = float(val)
                                    elif deal_data['nav'] == 0:
                                        deal_data['nav'] = float(val)
                            except:
                                pass

                        # Extract monthly cashflows - STORE AS-IS, NO CALCULATIONS!
                        for col in range(14, min(ws.max_column + 1, 100)):
                            try:
                                month_val = ws.cell(row, col).value
                                if month_val is not None and isinstance(month_val, (int, float)):
                                    # Store raw value (could be positive, negative, or zero)
                                    deal_data['monthly_cfs'][col - 14] = float(month_val)
                            except:
                                pass

                        deals_data.append(deal_data)

        success_msg = dbc.Alert([
            html.I(className="fas fa-check-circle me-2"),
            f"✅ Uploaded: {filename} ({len(deals_data)} deals found)"
        ], color="success", className="mt-2")

        return deals_data, success_msg

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        error_msg = dbc.Alert([
            html.I(className="fas fa-exclamation-triangle me-2"),
            html.Div([
                f"❌ Error uploading {filename}",
                html.Br(),
                html.Small(str(e), className="text-muted"),
                html.Br(),
                html.Small("Supported: Excel (.xlsx, .xlsm) or CSV (.csv)", className="text-muted")
            ])
        ], color="danger", className="mt-2")
        print(f"Upload error: {error_details}")  # For debugging
        return None, error_msg


# Summary metrics
@app.callback(
    [Output('cf-total-commitment', 'children'), Output('cf-total-paid-in', 'children'),
     Output('cf-total-unfunded', 'children'), Output('cf-current-nav', 'children')],
    [Input('deals-store', 'data'), Input('fund-cf-data-store', 'data')]
)
def update_cf_summary(deals, uploaded_data):
    # Use uploaded data if available, otherwise use deals
    if uploaded_data:
        total_commitment = sum(d.get('commitment', 0) for d in uploaded_data)
        total_paid_in = total_commitment * 0.75  # Estimate
        total_unfunded = total_commitment - total_paid_in
        current_nav = total_paid_in
    elif deals:
        total_commitment = sum(d.get('commitment', d['size']) for d in deals)
        total_paid_in = sum(d['size'] for d in deals)
        total_unfunded = total_commitment - total_paid_in
        current_nav = sum(d['size'] for d in deals)
    else:
        return "$0.0M", "$0.0M", "$0.0M", "$0.0M"

    return (f"${total_commitment:.1f}M", f"${total_paid_in:.1f}M",
            f"${total_unfunded:.1f}M", f"${current_nav:.1f}M")


# Monthly Cashflow Table - Deal by Deal with Monthly Columns
@app.callback(
    Output('cf-monthly-table', 'children'),
    [Input('deals-store', 'data'), Input('fund-cf-data-store', 'data'),
     Input('cf-view-period', 'value'), Input('cf-type-view', 'value')]
)
def generate_monthly_cf_table(deals, uploaded_data, num_months, cf_type):
    """Generate Fund Level CF table from uploaded Excel or current deals"""

    # Use uploaded data if available
    data_source = uploaded_data if uploaded_data else deals

    if not data_source:
        return html.P("Upload Fund Level CF Excel file or add deals to see monthly cashflows",
                      className="text-muted")

    # Generate month headers (starting from Jan 2026)
    base_date = datetime(2026, 1, 1)
    months = []
    for i in range(num_months):
        month_date = base_date + relativedelta(months=i)
        months.append(month_date.strftime('%b-%y'))

    # Build table data
    table_rows = []

    for deal in data_source:
        # Base deal info
        if uploaded_data:
            row_data = {
                'Deal': deal['name'],
                'Type': deal['type'],
                'Status': 'Executed',
                'Commitment': f"${deal.get('commitment', 0):.1f}M",
            }

            # Add monthly columns from uploaded data
            for i, month in enumerate(months):
                cf_value = deal.get('monthly_cfs', {}).get(i, 0)
                row_data[month] = f"${cf_value:.2f}M" if cf_value != 0 else "-"
        else:
            row_data = {
                'Deal': deal['name'],
                'Type': deal['strategy'],
                'Status': 'Executed',
                'Commitment': f"${deal.get('commitment', deal['size']):.1f}M",
            }

            # Add placeholder monthly columns
            for month in months:
                row_data[month] = "-"  # Placeholder until real cashflows added

        table_rows.append(row_data)

    if not table_rows:
        return html.P("No data", className="text-muted")

    # Create DataTable with horizontal scroll
    all_columns = ['Deal', 'Type', 'Status', 'Commitment'] + months

    return dash_table.DataTable(
        data=table_rows,
        columns=[{"name": c, "id": c} for c in all_columns],
        style_cell={
            'textAlign': 'left',
            'padding': '10px',
            'fontFamily': C['mono'],
            'fontSize': '11px',
            'minWidth': '80px',
            'maxWidth': '150px',
            'whiteSpace': 'normal'
        },
        style_header={
            'backgroundColor': C['surface'],
            'color': C['text'],
            'fontWeight': 'bold',
            'border': f'1px solid {C["border"]}',
            'position': 'sticky',
            'top': 0
        },
        style_data={
            'backgroundColor': C['panel'],
            'color': C['text'],
            'border': f'1px solid {C["border"]}'
        },
        style_data_conditional=[
            {'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']},
            {'if': {'column_id': ['Deal', 'Type']}, 'fontWeight': 'bold'}
        ],
        style_table={
            'overflowX': 'auto',
            'maxHeight': '600px',
            'overflowY': 'auto'
        },
        fixed_columns={'headers': True, 'data': 2},
        fixed_rows={'headers': True},
        export_format='xlsx',
        export_headers='display',
    )


# Monthly Chart
@app.callback(
    Output('cf-monthly-chart', 'figure'),
    [Input('deals-store', 'data'), Input('cf-view-period', 'value')]
)
def generate_monthly_cf_chart(deals, num_months):
    """Generate monthly total cashflows chart"""

    base_date = datetime(2026, 1, 1)
    months = []
    calls = []
    dists = []
    nav = []

    for i in range(num_months):
        month_date = base_date + relativedelta(months=i)
        months.append(month_date.strftime('%b %Y'))
        # Placeholder - would calculate from actual cashflows
        calls.append(0)
        dists.append(0)
        nav.append(sum(d['size'] for d in deals) if deals else 0)

    fig = go.Figure()

    fig.add_trace(go.Bar(
        x=months, y=calls, name='Capital Calls',
        marker_color=C['red'], opacity=0.7
    ))

    fig.add_trace(go.Bar(
        x=months, y=dists, name='Distributions',
        marker_color=C['green'], opacity=0.7
    ))

    fig.add_trace(go.Scatter(
        x=months, y=nav, name='Ending NAV',
        line=dict(color=C['blue'], width=3),
        yaxis='y2'
    ))

    fig.update_layout(
        **CHART_BASE,
        barmode='group',
        height=400,
        yaxis=dict(title='Cashflows ($mm)', gridcolor=C['border']),
        yaxis2=dict(
            title='NAV ($mm)',
            overlaying='y',
            side='right',
            gridcolor=C['border']
        ),
        hovermode='x unified'
    )

    return fig
    cards = []
    for deal, data in by_deal.items():
        net = data['dists'] - data['calls']
        cards.append(dbc.Card([dbc.CardBody([
            html.H6(deal, style={'color': C['text'], 'fontFamily': C['sans']}),
            dbc.Row([
                dbc.Col([html.Small("Calls:", className="text-muted"), html.Br(),
                         html.Strong(f"${data['calls']:.1f}M", style={'color': C['red'], 'fontFamily': C['mono']})],
                        width=4),
                dbc.Col([html.Small("Dists:", className="text-muted"), html.Br(),
                         html.Strong(f"${data['dists']:.1f}M", style={'color': C['green'], 'fontFamily': C['mono']})],
                        width=4),
                dbc.Col([html.Small("Net:", className="text-muted"), html.Br(),
                         html.Strong(f"${net:.1f}M", style={'color': C['blue'], 'fontFamily': C['mono']})], width=4),
            ])
        ])], className="mb-2", style={'backgroundColor': C['surface']}))

    return html.Div(cards)


# ==================== PRO FORMA CALLBACKS ====================

# Pro Forma Tabs Content
@app.callback(
    Output('pf-tab-content', 'children'),
    [Input('pf-tabs', 'active_tab'), Input('deals-store', 'data'),
     Input('pipeline-store', 'data'), Input('pf-target-month', 'value'),
     Input('cashflows-store', 'data')]
)
def render_pf_tab_content(active_tab, deals, pipeline, target_month, cashflows):
    if active_tab == "tab-pf-portfolio":
        # COMPLETE PORTFOLIO TAB - Show ALL deals
        return render_complete_portfolio(deals, pipeline, target_month, cashflows)
    elif active_tab == "tab-pf-metrics":
        # METRICS COMPARISON TAB
        return render_metrics_comparison(deals, pipeline)
    elif active_tab == "tab-pf-charts":
        # CHARTS TAB
        return render_impact_charts(deals, pipeline)
    return html.Div()


def render_complete_portfolio(deals, pipeline, target_month, cashflows):
    """Show complete portfolio with current + pipeline deals and NAV at target month"""

    # Calculate NAV at target month using actual cashflows
    month_names = generate_month_options()
    target_month_name = month_names[target_month]['label'] if target_month < len(month_names) else "Unknown"

    # Build complete portfolio
    all_deals = []

    # Current portfolio deals
    for deal in deals:
        deal_copy = deal.copy()
        deal_copy['source'] = 'Current Portfolio'
        deal_copy['status'] = '✓ Active'

        # Calculate NAV at target month using cashflows
        nav_at_month = calculate_nav_at_month(deal['name'], target_month, cashflows, deal['size'])
        deal_copy['nav_at_target'] = nav_at_month

        all_deals.append(deal_copy)

    # Pipeline deals (assumed to close in their respective months)
    for p_deal in pipeline:
        deal_copy = {
            'name': p_deal['name'],
            'strategy': p_deal['type'],
            'size': p_deal['size'],
            'target_irr': p_deal['target_irr'],
            'hold_period': 5.0,
            'moic': (1 + p_deal['target_irr']) ** 5,
            'vintage': 2026,
            'sector': 'TBD',
            'geography': 'Global',
            'source': 'Pipeline',
            'status': '⏳ Pending',
            'nav_at_target': p_deal['size']  # Assume full deployment
        }
        all_deals.append(deal_copy)

    if not all_deals:
        return dbc.Alert("No deals to display. Add deals to Current Portfolio or Pipeline.", color="info")

    # Summary cards
    total_current_nav = sum(d['size'] for d in deals)
    total_pipeline_nav = sum(p['size'] for p in pipeline)
    total_nav_at_target = sum(d['nav_at_target'] for d in all_deals)

    summary = dbc.Row([
        dbc.Col(dbc.Card([dbc.CardBody([
            html.H6("Current Portfolio NAV", className="text-muted"),
            html.H4(f"${total_current_nav:.1f}M", style={'color': C['green'], 'fontFamily': C['mono']})
        ])], className="shadow-sm"), width=3),
        dbc.Col(dbc.Card([dbc.CardBody([
            html.H6("Pipeline NAV", className="text-muted"),
            html.H4(f"${total_pipeline_nav:.1f}M", style={'color': C['blue'], 'fontFamily': C['mono']})
        ])], className="shadow-sm"), width=3),
        dbc.Col(dbc.Card([dbc.CardBody([
            html.H6(f"Pro Forma NAV at {target_month_name}", className="text-muted"),
            html.H4(f"${total_nav_at_target:.1f}M", style={'color': C['amber'], 'fontFamily': C['mono']})
        ])], className="shadow-sm"), width=3),
        dbc.Col(dbc.Card([dbc.CardBody([
            html.H6("Total Deals", className="text-muted"),
            html.H4(f"{len(all_deals)}", style={'color': C['purple'], 'fontFamily': C['mono']})
        ])], className="shadow-sm"), width=3),
    ], className="mb-4")

    # Complete portfolio table
    table_data = []
    for d in all_deals:
        table_data.append({
            'Deal': d['name'],
            'Strategy': d['strategy'],
            'Size': f"${d['size']:.1f}M",
            'IRR': f"{d['target_irr']:.1%}",
            'MOIC': f"{d.get('moic', 0):.2f}x",
            'Vintage': str(d.get('vintage', 'N/A')),
            'Sector': d.get('sector', 'N/A'),
            'Source': d['source'],
            'Status': d['status'],
            f'NAV @ {target_month_name}': f"${d['nav_at_target']:.1f}M"
        })

    portfolio_table = dbc.Card([
        dbc.CardHeader(f"Complete Pro Forma Portfolio ({len(all_deals)} deals)",
                       style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
        dbc.CardBody([
            dash_table.DataTable(
                data=table_data,
                columns=[{"name": c, "id": c} for c in table_data[0].keys()] if table_data else [],
                style_cell={'textAlign': 'left', 'padding': '12px', 'fontFamily': C['mono'], 'fontSize': '12px'},
                style_header={
                    'backgroundColor': C['surface'],
                    'color': C['text'],
                    'fontWeight': 'bold',
                    'border': f'1px solid {C["border"]}'
                },
                style_data={
                    'backgroundColor': C['panel'],
                    'color': C['text'],
                    'border': f'1px solid {C["border"]}'
                },
                style_data_conditional=[
                    {'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']},
                    {'if': {'column_id': 'Source', 'filter_query': '{Source} = "Pipeline"'},
                     'color': C['blue'], 'fontWeight': 'bold'},
                    {'if': {'column_id': 'Source', 'filter_query': '{Source} = "Current Portfolio"'},
                     'color': C['green'], 'fontWeight': 'bold'},
                ],
                page_size=100,  # Show many deals
                sort_action='native',
                filter_action='native',
            )
        ])
    ], className="shadow-sm")

    return html.Div([summary, portfolio_table])


def calculate_nav_at_month(deal_name, target_month_idx, cashflows, initial_nav):
    """Calculate NAV of a deal at target month using actual cashflows"""
    if not cashflows:
        return initial_nav

    # Filter cashflows for this deal
    deal_cashflows = [cf for cf in cashflows if cf['deal'] == deal_name]

    if not deal_cashflows:
        return initial_nav

    # Get target month date
    base_date = datetime(2026, 1, 1)
    from dateutil.relativedelta import relativedelta
    target_date = base_date + relativedelta(months=target_month_idx)

    # Calculate NAV = Initial + Calls - Distributions up to target month
    nav = initial_nav

    for cf in deal_cashflows:
        try:
            cf_date = datetime.fromisoformat(
                cf['date'].replace('Z', '+00:00') if 'T' in cf['date'] else cf['date'] + 'T00:00:00')
        except:
            continue

        # Only include cashflows up to target month
        if cf_date <= target_date:
            if cf['type'] == 'Call':
                nav -= cf['amount']  # Calls reduce NAV
            else:  # Distribution
                nav += cf['amount']  # Distributions increase NAV

    return max(0, nav)  # NAV can't be negative


def render_metrics_comparison(deals, pipeline):
    """Render before/after metrics comparison"""
    from datetime import datetime

    current_m = calculate_portfolio_metrics(deals)

    # Pro forma (current + pipeline)
    if pipeline:
        proforma_deals = deals + [{
            'name': p['name'],
            'strategy': p['type'],
            'size': p['size'],
            'target_irr': p['target_irr'],
            'hold_period': 5.0,
            'moic': (1 + p['target_irr']) ** 5,
            'vintage': 2026,
            'sector': 'TBD',
            'geography': 'Global'
        } for p in pipeline]

        pf_m = calculate_portfolio_metrics(proforma_deals)
    else:
        pf_m = current_m

    comparison_data = [
        {"Metric": "Total NAV ($mm)", "Current": f"${current_m['total_nav']:.1f}",
         "Pro Forma": f"${pf_m['total_nav']:.1f}",
         "Change": f"+${pf_m['total_nav'] - current_m['total_nav']:.1f}"},
        {"Metric": "Weighted IRR", "Current": f"{current_m['weighted_irr']:.2%}",
         "Pro Forma": f"{pf_m['weighted_irr']:.2%}",
         "Change": f"{pf_m['weighted_irr'] - current_m['weighted_irr']:+.2%}"},
        {"Metric": "Number of Deals", "Current": str(current_m['num_deals']),
         "Pro Forma": str(pf_m['num_deals']),
         "Change": f"+{pf_m['num_deals'] - current_m['num_deals']}"},
        {"Metric": "Top 1 Concentration", "Current": f"{current_m['concentration_top1']:.1%}",
         "Pro Forma": f"{pf_m['concentration_top1']:.1%}",
         "Change": f"{pf_m['concentration_top1'] - current_m['concentration_top1']:+.1%}"},
    ]

    return dbc.Card([
        dbc.CardHeader("Before vs After Comparison", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
        dbc.CardBody([
            dash_table.DataTable(
                data=comparison_data,
                columns=[{"name": c, "id": c} for c in ["Metric", "Current", "Pro Forma", "Change"]],
                style_cell={'textAlign': 'left', 'padding': '12px', 'fontFamily': C['mono'], 'fontSize': '13px'},
                style_header={'backgroundColor': C['surface'], 'color': C['text'], 'fontWeight': 'bold'},
                style_data={'backgroundColor': C['panel'], 'color': C['text']},
                style_data_conditional=[
                    {'if': {'column_id': 'Change'}, 'fontWeight': 'bold', 'color': C['blue']}
                ]
            )
        ])
    ], className="shadow-sm")


def render_impact_charts(deals, pipeline):
    """Render impact charts"""
    current_m = calculate_portfolio_metrics(deals)

    if pipeline:
        proforma_deals = deals + [{
            'name': p['name'], 'strategy': p['type'], 'size': p['size'],
            'target_irr': p['target_irr'], 'hold_period': 5.0,
            'moic': (1 + p['target_irr']) ** 5, 'vintage': 2026,
            'sector': 'TBD', 'geography': 'Global'
        } for p in pipeline]
        pf_m = calculate_portfolio_metrics(proforma_deals)
    else:
        pf_m = current_m

    # Strategy chart
    fig_strat = go.Figure()
    if current_m['by_strategy']:
        strats = list(current_m['by_strategy'].keys())
        current_vals = [current_m['by_strategy'][s]['nav'] for s in strats]
        fig_strat.add_trace(go.Bar(name='Current', x=strats, y=current_vals, marker_color=C['blue']))

    if pf_m['by_strategy']:
        pf_strats = list(pf_m['by_strategy'].keys())
        pf_vals = [pf_m['by_strategy'][s]['nav'] for s in pf_strats]
        fig_strat.add_trace(go.Bar(name='Pro Forma', x=pf_strats, y=pf_vals, marker_color=C['green']))

    fig_strat.update_layout(**CHART_BASE, barmode='group', yaxis_title="NAV ($mm)", height=400)

    # Concentration chart
    fig_conc = go.Figure()
    cats = ['Top 1', 'Top 3', 'Top 5']
    current_conc = [current_m['concentration_top1'] * 100, current_m['concentration_top3'] * 100,
                    current_m['concentration_top5'] * 100]
    pf_conc = [pf_m['concentration_top1'] * 100, pf_m['concentration_top3'] * 100,
               pf_m['concentration_top5'] * 100]

    fig_conc.add_trace(go.Bar(name='Current', x=cats, y=current_conc, marker_color=C['blue']))
    fig_conc.add_trace(go.Bar(name='Pro Forma', x=cats, y=pf_conc, marker_color=C['green']))
    fig_conc.update_layout(**CHART_BASE, barmode='group', yaxis_title="% of NAV", height=400)

    return dbc.Row([
        dbc.Col([
            dbc.Card([
                dbc.CardHeader("Strategy Allocation", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                dbc.CardBody([dcc.Graph(figure=fig_strat, config={'displayModeBar': False})])
            ], className="shadow-sm")
        ], width=6),
        dbc.Col([
            dbc.Card([
                dbc.CardHeader("Concentration Risk", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                dbc.CardBody([dcc.Graph(figure=fig_conc, config={'displayModeBar': False})])
            ], className="shadow-sm")
        ], width=6),
    ])


# ==================== LIQUIDITY ASSUMPTIONS CALLBACKS ====================

# ==================== LIQUIDITY PULL CALLBACKS ====================

# Upload Liquidity Pull Excel file
@app.callback(
    [Output('liquidity-data-store', 'data'), Output('upload-liquidity-status', 'children')],
    Input('upload-liquidity-pull', 'contents'),
    State('upload-liquidity-pull', 'filename'),
    prevent_initial_call=True
)
def upload_liquidity_file(contents, filename):
    """Parse uploaded Liquidity Pull Excel file with all key metrics"""
    if contents is None:
        return None, ""

    try:
        # Decode the uploaded file
        content_type, content_string = contents.split(',')
        decoded = base64.b64decode(content_string)

        # Read Excel file
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(decoded), data_only=True)

        # Try to find Liquidity Pull sheet
        sheet_names = ['Liquidity Pull', 'Liquidity_Pull', 'LiquidityPull']
        ws = None
        for name in sheet_names:
            if name in wb.sheetnames:
                ws = wb[name]
                break

        if ws is None:
            # Use first sheet if can't find specific name
            ws = wb[wb.sheetnames[0]]

        # Parse the sheet structure
        liquidity_data = {
            'as_at_date': None,
            'current_quarter': None,
            'current_month': None,
            'fund_nav': 0,
            'current_cash': 0,  # Row 58
            'glf_balance': 0,  # Row 59
            'cqs_balance': 0,  # Row 60
            'total_liquidity': 0,
            'max_deployable_capital': {},  # Row 53 (monthly)
            'surplus_liquidity_post_buffer': 0,  # Row 63
            'projected_nav_existing': 0,  # Row 71
            'projected_nav_existing_pipeline': 0,  # Row 72
            'near_term_flows': {},
            'nav_projections': {}
        }

        # Extract key data points
        # As At date (typically row 4, column 2)
        as_at = ws.cell(4, 2).value
        if as_at:
            liquidity_data['as_at_date'] = str(as_at)

        # Current Quarter (row 5, column 2)
        curr_q = ws.cell(5, 2).value
        if curr_q:
            liquidity_data['current_quarter'] = str(curr_q)

        # Current Month (row 6, column 2)
        curr_m = ws.cell(6, 2).value
        if curr_m:
            liquidity_data['current_month'] = str(curr_m)

        # Fund NAV (row 10, column 3 typically)
        fund_nav = ws.cell(10, 3).value
        if fund_nav and isinstance(fund_nav, (int, float)):
            liquidity_data['fund_nav'] = float(fund_nav) / 1_000_000

        # MAX DEPLOYABLE CAPITAL - Row 53 (monthly values)
        for col in range(1, 25):  # Up to 24 months
            try:
                month_label = ws.cell(28, col).value  # Month header from row 28
                max_deploy = ws.cell(53, col).value
                if month_label and max_deploy and isinstance(max_deploy, (int, float)):
                    liquidity_data['max_deployable_capital'][str(month_label)] = float(max_deploy) / 1_000_000
            except:
                pass

        # CASH BALANCE - Row 58
        cash = ws.cell(58, 3).value
        if cash and isinstance(cash, (int, float)):
            liquidity_data['current_cash'] = float(cash) / 1_000_000

        # GLF BALANCE - Row 59
        glf = ws.cell(59, 3).value
        if glf and isinstance(glf, (int, float)):
            liquidity_data['glf_balance'] = float(glf) / 1_000_000

        # CQS BALANCE - Row 60
        cqs = ws.cell(60, 3).value
        if cqs and isinstance(cqs, (int, float)):
            liquidity_data['cqs_balance'] = float(cqs) / 1_000_000

        # SURPLUS LIQUIDITY POST BUFFER - Row 63
        surplus = ws.cell(63, 3).value
        if surplus and isinstance(surplus, (int, float)):
            liquidity_data['surplus_liquidity_post_buffer'] = float(surplus) / 1_000_000

        # PROJECTED NAV EXISTING - Row 71
        proj_nav_existing = ws.cell(71, 3).value
        if proj_nav_existing and isinstance(proj_nav_existing, (int, float)):
            liquidity_data['projected_nav_existing'] = float(proj_nav_existing) / 1_000_000

        # PROJECTED NAV EXISTING + PIPELINE - Row 72
        proj_nav_pipeline = ws.cell(72, 3).value
        if proj_nav_pipeline and isinstance(proj_nav_pipeline, (int, float)):
            liquidity_data['projected_nav_existing_pipeline'] = float(proj_nav_pipeline) / 1_000_000

        # Total Liquidity (row 16, column 3)
        total_liq = ws.cell(16, 3).value
        if total_liq and isinstance(total_liq, (int, float)):
            liquidity_data['total_liquidity'] = float(total_liq) / 1_000_000

        # Near term flows - starting around row 28
        for col in range(1, 25):  # Up to 24 months
            try:
                month_val = ws.cell(28, col).value  # Month header
                subs = ws.cell(32, col).value or 0  # Subscriptions
                reds = ws.cell(33, col).value or 0  # Redemptions
                flows = ws.cell(34, col).value or 0  # Portfolio flows

                if month_val:
                    liquidity_data['near_term_flows'][str(month_val)] = {
                        'subscriptions': float(subs) / 1_000_000 if isinstance(subs, (int, float)) else 0,
                        'redemptions': float(reds) / 1_000_000 if isinstance(reds, (int, float)) else 0,
                        'portfolio_flows': float(flows) / 1_000_000 if isinstance(flows, (int, float)) else 0
                    }
            except:
                pass

        # NAV projections (row 39)
        for col in range(1, 25):
            try:
                month_val = ws.cell(28, col).value
                nav_end = ws.cell(39, col).value
                if month_val and nav_end and isinstance(nav_end, (int, float)):
                    liquidity_data['nav_projections'][str(month_val)] = float(nav_end) / 1_000_000
            except:
                pass

        success_msg = dbc.Alert([
            html.I(className="fas fa-check-circle me-2"),
            f"✅ Uploaded: {filename} - Max Deployable, Cash ({liquidity_data['current_cash']:.1f}M), Projected NAV loaded"
        ], color="success", className="mt-2")

        return liquidity_data, success_msg

    except Exception as e:
        import traceback
        print(f"Liquidity upload error: {traceback.format_exc()}")
        error_msg = dbc.Alert([
            html.I(className="fas fa-exclamation-triangle me-2"),
            f"❌ Error: {str(e)}"
        ], color="danger", className="mt-2")
        return None, error_msg


# Date fields
@app.callback(
    [Output('liq-today-date', 'children'), Output('liq-as-at-date', 'children'),
     Output('liq-current-quarter', 'children'), Output('liq-current-month', 'children')],
    Input('liquidity-data-store', 'data')
)
def update_liq_dates(uploaded_data):
    if uploaded_data:
        # Use uploaded dates
        return (
            datetime.now().strftime('%Y-%m-%d'),
            uploaded_data.get('as_at_date', 'N/A'),
            uploaded_data.get('current_quarter', 'N/A'),
            uploaded_data.get('current_month', 'N/A')
        )
    else:
        # Use calculated dates
        today = datetime.now()
        as_at = (today.replace(day=1) - relativedelta(days=1))
        quarter_month = ((today.month - 1) // 3 + 1) * 3
        quarter_end = datetime(today.year, quarter_month, 1) + relativedelta(months=1) - relativedelta(days=1)
        month_end = today.replace(day=1) + relativedelta(months=1) - relativedelta(days=1)

        return (
            today.strftime('%Y-%m-%d'),
            as_at.strftime('%Y-%m-%d'),
            quarter_end.strftime('%Y-%m-%d'),
            month_end.strftime('%Y-%m-%d')
        )


# Liquidity Waterfall
@app.callback(
    Output('liq-waterfall-table', 'children'),
    [Input('deals-store', 'data'), Input('liquidity-data-store', 'data')]
)
def generate_liquidity_waterfall(deals, uploaded_data):
    """Generate liquidity waterfall from uploaded data or calculated"""

    if uploaded_data:
        # Use uploaded data
        total_nav = uploaded_data.get('fund_nav', 0)
        current_cash = uploaded_data.get('current_cash', 0)
        glf_balance = uploaded_data.get('glf_balance', 0)
        cqs_balance = uploaded_data.get('cqs_balance', 0)
        total_liquidity = uploaded_data.get('total_liquidity', 0)
    else:
        # Calculate from deals
        total_nav = sum(d['size'] for d in deals) if deals else 0
        current_cash = 0.88
        glf_balance = 19.95
        cqs_balance = 17.27
        total_liquidity = current_cash + glf_balance + cqs_balance

    waterfall_data = [
        {'Source': 'Fund NAV', 'Item': 'Total Current NAV', 'Amount ($mm)': f'{total_nav:.2f}'},
        {'Source': 'Current Cash', 'Item': 'Cash Balance', 'Amount ($mm)': f'{current_cash:.2f}'},
        {'Source': 'GLF', 'Item': 'GLF Balance', 'Amount ($mm)': f'{glf_balance:.2f}'},
        {'Source': 'CQS', 'Item': 'CQS Balance', 'Amount ($mm)': f'{cqs_balance:.2f}'},
        {'Source': 'Total Liquidity', 'Item': 'Total Liquidity Balance', 'Amount ($mm)': f'{total_liquidity:.2f}'},
    ]

    return dash_table.DataTable(
        data=waterfall_data,
        columns=[{"name": c, "id": c} for c in waterfall_data[0].keys()],
        style_cell={'textAlign': 'left', 'padding': '12px', 'fontFamily': C['mono'], 'fontSize': '12px'},
        style_header={'backgroundColor': C['surface'], 'color': C['text'], 'fontWeight': 'bold',
                      'border': f'1px solid {C["border"]}'},
        style_data={'backgroundColor': C['panel'], 'color': C['text'], 'border': f'1px solid {C["border"]}'},
        style_data_conditional=[
            {'if': {'column_id': 'Source', 'filter_query': '{Source} = "Total Liquidity"'},
             'fontWeight': 'bold', 'color': C['green'], 'backgroundColor': C['surface']}
        ]
    )


# Near Term Flows
@app.callback(
    Output('liq-near-term-flows', 'children'),
    [Input('deals-store', 'data'), Input('liquidity-data-store', 'data')]
)
def generate_near_term_flows(deals, uploaded_data):
    """Generate near term flows from uploaded data or calculated"""

    base_date = datetime(2026, 3, 31)
    months = []
    subscriptions = []
    redemptions = []
    portfolio_flows = []

    if uploaded_data and uploaded_data.get('near_term_flows'):
        # Use uploaded data
        for month_key, data in list(uploaded_data['near_term_flows'].items())[:12]:
            try:
                month_date = pd.to_datetime(month_key)
                months.append(month_date.strftime('%b-%y'))
            except:
                months.append(str(month_key)[:6])
            subscriptions.append(data.get('subscriptions', 0))
            redemptions.append(data.get('redemptions', 0))
            portfolio_flows.append(data.get('portfolio_flows', 0))
    else:
        # Use calculated/placeholder data
        for i in range(12):
            month_date = base_date + relativedelta(months=i)
            months.append(month_date.strftime('%b-%y'))
            subscriptions.append(50 if i < 3 else 0)
            redemptions.append(0)
            portfolio_flows.append(2.6 + (i * 0.5))

    # Create table
    table_data = [
        {'Flow Type': 'Subscriptions', **{months[i]: f'${subscriptions[i]:.1f}M' for i in range(len(months))}},
        {'Flow Type': 'Redemptions', **{months[i]: f'${redemptions[i]:.1f}M' for i in range(len(months))}},
        {'Flow Type': 'Portfolio Net Flows', **{months[i]: f'${portfolio_flows[i]:.1f}M' for i in range(len(months))}},
    ]

    all_columns = ['Flow Type'] + months

    return dash_table.DataTable(
        data=table_data,
        columns=[{"name": c, "id": c} for c in all_columns],
        style_cell={
            'textAlign': 'left',
            'padding': '10px',
            'fontFamily': C['mono'],
            'fontSize': '11px',
            'minWidth': '100px'
        },
        style_header={
            'backgroundColor': C['surface'],
            'color': C['text'],
            'fontWeight': 'bold',
            'position': 'sticky',
            'top': 0,
            'border': f'1px solid {C["border"]}'
        },
        style_data={'backgroundColor': C['panel'], 'color': C['text'], 'border': f'1px solid {C["border"]}'},
        style_data_conditional=[
            {'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']},
            {'if': {'column_id': 'Flow Type'}, 'fontWeight': 'bold'}
        ],
        style_table={'overflowX': 'auto'},
        fixed_columns={'headers': True, 'data': 1}
    )


# NAV Projection Chart
@app.callback(
    Output('liq-nav-projection-chart', 'figure'),
    [Input('deals-store', 'data'), Input('liquidity-data-store', 'data')]
)
def generate_nav_projection_chart(deals, uploaded_data):
    """Generate NAV end projections from uploaded data or calculated"""

    base_date = datetime(2026, 3, 31)
    months = []
    nav_projections = []

    if uploaded_data and uploaded_data.get('nav_projections'):
        # Use uploaded NAV projections
        for month_key, nav_val in list(uploaded_data['nav_projections'].items())[:12]:
            try:
                month_date = pd.to_datetime(month_key)
                months.append(month_date.strftime('%b %Y'))
            except:
                months.append(str(month_key)[:10])
            nav_projections.append(nav_val)
    else:
        # Calculate projections
        base_nav = sum(d['size'] for d in deals) if deals else 100

        for i in range(12):
            month_date = base_date + relativedelta(months=i)
            months.append(month_date.strftime('%b %Y'))
            nav = base_nav * (1 + (0.13 / 12)) ** i
            nav_projections.append(nav)

    fig = go.Figure()

    fig.add_trace(go.Scatter(
        x=months, y=nav_projections,
        mode='lines+markers',
        name='NAV End',
        line=dict(color=C['blue'], width=3),
        marker=dict(size=8, color=C['sky']),
        fill='tozeroy',
        fillcolor=rgba(C['blue'], 0.2)
    ))

    fig.update_layout(
        **CHART_BASE,
        height=400,
        yaxis_title='NAV ($mm)',
        hovermode='x unified'
    )

    return fig


# Tab content
@app.callback(
    Output('liq-tab-content', 'children'),
    Input('liq-tabs', 'active_tab')
)
def render_liq_tab_content(active_tab):
    if active_tab == "tab-liq-subs":
        return dbc.Card([
            dbc.CardBody([
                html.H6("Subscriptions & Redemptions", className="mb-3"),
                html.P("Expected subscriptions: $150M (confirmed)", style={'fontFamily': C['mono']}),
                html.P("Expected redemptions: $0M", style={'fontFamily': C['mono']}),
                html.P("Redemption gate limit: 5% of NAV", style={'fontFamily': C['mono']})
            ])
        ], className="shadow-sm")
    elif active_tab == "tab-liq-flows":
        return dbc.Card([
            dbc.CardBody([
                html.H6("Portfolio Net Flows", className="mb-3"),
                html.P("Seed Portfolio Net Flows calculated from deal-level cashflows",
                       style={'fontFamily': C['mono']}),
                html.P("See Fund Level CF tab for detailed monthly breakdown", style={'fontFamily': C['mono']})
            ])
        ], className="shadow-sm")
    else:  # unfunded
        return dbc.Card([
            dbc.CardBody([
                html.H6("Unfunded Commitments", className="mb-3"),
                html.P("Current unfunded commitments: $140.3M", style={'fontFamily': C['mono']}),
                html.P("Rolling unfunded exposure monitored monthly", style={'fontFamily': C['mono']})
            ])
        ], className="shadow-sm")


@app.callback(
    Output('liq-strategy-assumptions', 'children'),
    Input('liquidity-assumptions-store', 'data')
)
def render_liquidity_assumptions(liq_data):
    """Render editable liquidity assumptions by strategy"""

    cards = []
    strategies = ['GP-Led (Multi-Asset)', 'GP-Led (Single-Asset)', 'Diversified LP-Led', 'Co-Investments']
    colors = [C['blue'], C['purple'], C['teal'], C['green']]

    for idx, strategy in enumerate(strategies):
        data = liq_data.get(strategy, {'annual_dist_rate': 0.20, 'call_pattern': 'immediate'})

        cards.append(dbc.Col([
            dbc.Card([
                dbc.CardBody([
                    html.H6(strategy, style={'color': colors[idx], 'fontFamily': C['sans'], 'fontWeight': 'bold'}),
                    html.Hr(style={'borderColor': C['border']}),
                    dbc.Row([
                        dbc.Col([
                            html.Small("Annual Distribution Rate", className="text-muted d-block"),
                            html.H5(f"{data['annual_dist_rate']:.1%}",
                                    style={'color': C['green'], 'fontFamily': C['mono']})
                        ], width=6),
                        dbc.Col([
                            html.Small("Call Pattern", className="text-muted d-block"),
                            html.H6(data['call_pattern'].title(),
                                    style={'color': C['amber'], 'fontFamily': C['mono'], 'fontSize': '14px'})
                        ], width=6),
                    ])
                ], style={'padding': '1rem'})
            ], className="shadow-sm", style={'backgroundColor': C['surface'], 'border': f'1px solid {C["border"]}'})
        ], width=3))

    return dbc.Row(cards)


@app.callback(
    Output('liq-summary', 'children'),
    [Input('deals-store', 'data'), Input('liquidity-assumptions-store', 'data')]
)
def render_liquidity_summary(deals, liq_data):
    """Show expected liquidity by strategy"""

    if not deals:
        return html.P("Add deals to see liquidity forecast", className="text-muted")

    # Group by strategy
    by_strategy = {}
    for deal in deals:
        strat = deal['strategy']
        if strat not in by_strategy:
            by_strategy[strat] = {'nav': 0, 'count': 0}
        by_strategy[strat]['nav'] += deal['size']
        by_strategy[strat]['count'] += 1

    # Calculate expected annual distributions
    summary_rows = []
    total_annual_dist = 0

    for strategy, data in by_strategy.items():
        dist_rate = liq_data.get(strategy, {}).get('annual_dist_rate', 0.20)
        annual_dist = data['nav'] * dist_rate
        total_annual_dist += annual_dist

        summary_rows.append(html.Tr([
            html.Td(strategy, style={'fontFamily': C['sans'], 'color': C['text']}),
            html.Td(f"${data['nav']:.1f}M", style={'fontFamily': C['mono'], 'color': C['blue']}),
            html.Td(f"{dist_rate:.1%}", style={'fontFamily': C['mono'], 'color': C['muted']}),
            html.Td(f"${annual_dist:.1f}M", style={'fontFamily': C['mono'], 'color': C['green'], 'fontWeight': 'bold'}),
        ]))

    # Total row
    summary_rows.append(html.Tr([
        html.Td("TOTAL", style={'fontFamily': C['sans'], 'fontWeight': 'bold', 'color': C['text']}),
        html.Td(f"${sum(d['nav'] for d in by_strategy.values()):.1f}M",
                style={'fontFamily': C['mono'], 'fontWeight': 'bold', 'color': C['blue']}),
        html.Td("", style={'fontFamily': C['mono']}),
        html.Td(f"${total_annual_dist:.1f}M",
                style={'fontFamily': C['mono'], 'fontWeight': 'bold', 'color': C['green']}),
    ], style={'borderTop': f'2px solid {C["border"]}'}))

    return dbc.Table([
        html.Thead(html.Tr([
            html.Th("Strategy", style={'color': C['text'], 'fontFamily': C['sans']}),
            html.Th("Current NAV", style={'color': C['text'], 'fontFamily': C['sans']}),
            html.Th("Dist Rate", style={'color': C['text'], 'fontFamily': C['sans']}),
            html.Th("Annual Distributions", style={'color': C['text'], 'fontFamily': C['sans']}),
        ]), style={'backgroundColor': C['surface']}),
        html.Tbody(summary_rows)
    ], bordered=True, hover=True, className="shadow-sm")


# Populate deal dropdown
@app.callback(
    Output('pf-deal-select', 'options'),
    Input('pipeline-store', 'data')
)
def populate_pf_deals(pipeline):
    if not pipeline:
        return [{"label": "No pipeline deals", "value": ""}]
    return [{"label": f"{p['name']} (${p['size']:.1f}M @ {p['target_irr']:.1%})", "value": i}
            for i, p in enumerate(pipeline)]


# Add deal to scenario
@app.callback(
    Output('proforma-scenario-store', 'data', allow_duplicate=True),
    Input('btn-add-to-pf', 'n_clicks'),
    [State('pf-deal-select', 'value'), State('pipeline-store', 'data'),
     State('pf-custom-name', 'value'), State('pf-custom-size', 'value'),
     State('pf-custom-irr', 'value'), State('proforma-scenario-store', 'data')],
    prevent_initial_call=True
)
def add_to_proforma(n, selected_idx, pipeline, custom_name, custom_size, custom_irr, scenario):
    if custom_name and custom_size and custom_irr:
        # Add custom deal
        return scenario + [{
            'name': custom_name,
            'size': float(custom_size),
            'target_irr': float(custom_irr) / 100,
            'strategy': 'GP-Led (Multi-Asset)',  # Default
            'source': 'custom'
        }]
    elif selected_idx and selected_idx != '':
        # Add pipeline deal
        idx = int(selected_idx)
        if 0 <= idx < len(pipeline):
            p = pipeline[idx]
            return scenario + [{
                'name': p['name'],
                'size': p['size'],
                'target_irr': p['target_irr'],
                'strategy': p['type'],
                'source': 'pipeline'
            }]
    return scenario


# Reset scenario
@app.callback(
    Output('proforma-scenario-store', 'data', allow_duplicate=True),
    Input('btn-reset-pf', 'n_clicks'),
    prevent_initial_call=True
)
def reset_proforma(n):
    return []


# Show scenario deals
@app.callback(
    Output('pf-scenario-deals', 'children'),
    Input('proforma-scenario-store', 'data')
)
def show_pf_scenario(scenario):
    if not scenario:
        return html.P("No deals in scenario yet. Add deals above.", className="text-muted",
                      style={'fontFamily': C['mono']})

    items = []
    for i, deal in enumerate(scenario):
        items.append(html.Div([
            html.Strong(f"{i + 1}. {deal['name']}", style={'color': C['blue'], 'fontFamily': C['mono']}),
            html.Span(f" — ${deal['size']:.1f}M @ {deal['target_irr']:.1%}",
                      style={'color': C['muted'], 'fontFamily': C['mono']}),
            html.Br()
        ]))

    total_size = sum(d['size'] for d in scenario)
    items.append(html.Hr(style={'borderColor': C['border']}))
    items.append(html.P([
        html.Strong("Total Scenario Size: ", style={'fontFamily': C['sans'], 'color': C['text']}),
        html.Span(f"${total_size:.1f}M", style={'fontFamily': C['mono'], 'color': C['green']})
    ]))

    return html.Div(items)


# Comparison table
@app.callback(
    Output('pf-comparison-table', 'children'),
    [Input('proforma-scenario-store', 'data'), Input('deals-store', 'data'), Input('config-store', 'data')]
)
def update_pf_comparison(scenario, deals, config):
    # Current metrics
    current_m = calculate_portfolio_metrics(deals)
    current_nav = current_m['total_nav']
    current_irr = current_m['weighted_irr']
    current_deals_count = current_m['num_deals']
    current_top1 = current_m['concentration_top1']
    current_eff_n = current_m['effective_n']
    dry_powder = config['fund_parameters']['dry_powder']
    current_req_irr = calculate_required_future_irr(current_irr, current_nav, dry_powder, config)

    # Pro forma metrics (current + scenario)
    if scenario:
        proforma_deals = deals + scenario
        pf_m = calculate_portfolio_metrics(proforma_deals)
        pf_nav = pf_m['total_nav']
        pf_irr = pf_m['weighted_irr']
        pf_deals_count = pf_m['num_deals']
        pf_top1 = pf_m['concentration_top1']
        pf_eff_n = pf_m['effective_n']
        scenario_size = sum(d['size'] for d in scenario)
        pf_dry_powder = dry_powder - scenario_size
        pf_req_irr = calculate_required_future_irr(pf_irr, pf_nav, pf_dry_powder, config)
    else:
        pf_nav = current_nav
        pf_irr = current_irr
        pf_deals_count = current_deals_count
        pf_top1 = current_top1
        pf_eff_n = current_eff_n
        pf_req_irr = current_req_irr

    # Build comparison table
    comparison_data = [
        {"Metric": "Total NAV ($mm)", "Current": f"${current_nav:.1f}", "Pro Forma": f"${pf_nav:.1f}",
         "Change": f"+${pf_nav - current_nav:.1f}" if pf_nav > current_nav else "$0.0"},
        {"Metric": "Weighted IRR", "Current": f"{current_irr:.2%}", "Pro Forma": f"{pf_irr:.2%}",
         "Change": f"+{pf_irr - current_irr:.2%}" if pf_irr > current_irr else "0.0%"},
        {"Metric": "Required Future IRR", "Current": f"{current_req_irr:.2%}", "Pro Forma": f"{pf_req_irr:.2%}",
         "Change": f"{pf_req_irr - current_req_irr:+.2%}"},
        {"Metric": "Number of Deals", "Current": str(current_deals_count), "Pro Forma": str(pf_deals_count),
         "Change": f"+{pf_deals_count - current_deals_count}" if pf_deals_count > current_deals_count else "0"},
        {"Metric": "Top 1 Concentration", "Current": f"{current_top1:.1%}", "Pro Forma": f"{pf_top1:.1%}",
         "Change": f"{pf_top1 - current_top1:+.1%}"},
        {"Metric": "Effective # Positions", "Current": f"{current_eff_n:.1f}", "Pro Forma": f"{pf_eff_n:.1f}",
         "Change": f"+{pf_eff_n - current_eff_n:.1f}" if pf_eff_n > current_eff_n else "0.0"},
    ]

    return dash_table.DataTable(
        data=comparison_data,
        columns=[{"name": c, "id": c} for c in ["Metric", "Current", "Pro Forma", "Change"]],
        style_cell={'textAlign': 'left', 'padding': '12px', 'fontFamily': C['mono'], 'fontSize': '13px'},
        style_header={'backgroundColor': C['surface'], 'color': C['text'], 'fontWeight': 'bold',
                      'border': f'1px solid {C["border"]}'},
        style_data={'backgroundColor': C['panel'], 'color': C['text'], 'border': f'1px solid {C["border"]}'},
        style_data_conditional=[
            {'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']},
            {'if': {'column_id': 'Change'}, 'fontWeight': 'bold', 'color': C['blue']}
        ]
    )


# Pro forma charts
@app.callback(
    [Output('pf-strategy-chart', 'figure'), Output('pf-concentration-chart', 'figure')],
    [Input('proforma-scenario-store', 'data'), Input('deals-store', 'data')]
)
def update_pf_charts(scenario, deals):
    # Current
    current_m = calculate_portfolio_metrics(deals)

    # Pro forma
    if scenario:
        proforma_deals = deals + scenario
        pf_m = calculate_portfolio_metrics(proforma_deals)
    else:
        pf_m = current_m

    # Strategy comparison
    fig_strat = go.Figure()
    if current_m['by_strategy']:
        current_strats = list(current_m['by_strategy'].keys())
        current_vals = [current_m['by_strategy'][s]['nav'] for s in current_strats]
        fig_strat.add_trace(go.Bar(name='Current', x=current_strats, y=current_vals, marker_color=C['blue']))

    if pf_m['by_strategy']:
        pf_strats = list(pf_m['by_strategy'].keys())
        pf_vals = [pf_m['by_strategy'][s]['nav'] for s in pf_strats]
        fig_strat.add_trace(go.Bar(name='Pro Forma', x=pf_strats, y=pf_vals, marker_color=C['green']))

    fig_strat.update_layout(**CHART_BASE, barmode='group', yaxis_title="NAV ($mm)", height=350)

    # Concentration comparison
    fig_conc = go.Figure()
    current_conc = [current_m['concentration_top1'] * 100, current_m['concentration_top3'] * 100,
                    current_m['concentration_top5'] * 100]
    pf_conc = [pf_m['concentration_top1'] * 100, pf_m['concentration_top3'] * 100, pf_m['concentration_top5'] * 100]

    cats = ['Top 1', 'Top 3', 'Top 5']
    fig_conc.add_trace(go.Bar(name='Current', x=cats, y=current_conc, marker_color=C['blue']))
    fig_conc.add_trace(go.Bar(name='Pro Forma', x=cats, y=pf_conc, marker_color=C['green']))

    fig_conc.add_hline(y=15, line_color=C['red'], line_dash='dash', opacity=0.5)
    fig_conc.add_hline(y=40, line_color=C['red'], line_dash='dash', opacity=0.5)
    fig_conc.add_hline(y=60, line_color=C['red'], line_dash='dash', opacity=0.5)

    fig_conc.update_layout(**CHART_BASE, barmode='group', yaxis_title="% of NAV", height=350)

    return fig_strat, fig_conc


# ==================== ANALYTICS CALLBACKS WITH CURRENT/PRO FORMA TOGGLE ====================

# ==================== ANALYTICS CALLBACKS WITH 3-WAY TOGGLE ====================

def get_portfolio_for_analytics_view(view, deals, pipeline, placeholders):
    """Helper to get portfolio based on 3-way toggle: current, current+pipeline, or full"""
    if view == 'current':
        return deals if deals else []
    elif view == 'current_pipeline':
        return (deals if deals else []) + \
            [{'name': p['name'], 'size': p['size'], 'strategy': p.get('type', p.get('strategy', 'GP-Led')),
              'geography': p.get('region', 'North America'), 'sector': 'Technology',
              'vintage': 2026, 'deal_type': p.get('deal_type', 'Secondary')}
             for p in (pipeline if pipeline else [])]
    else:  # full_proforma
        return (deals if deals else []) + \
            [{'name': p['name'], 'size': p['size'], 'strategy': p.get('type', p.get('strategy', 'GP-Led')),
              'geography': p.get('region', 'North America'), 'sector': 'Technology',
              'vintage': 2026, 'deal_type': p.get('deal_type', 'Secondary')}
             for p in (pipeline if pipeline else [])] + \
            [{'name': p['name'], 'size': p['size'], 'strategy': p.get('strategy', 'GP-Led'),
              'geography': p.get('region', 'North America'), 'sector': 'Technology',
              'vintage': 2026, 'deal_type': p.get('deal_type', 'Secondary')}
             for p in (placeholders if placeholders else [])]


# Summary Cards
@app.callback(
    [Output('analytics-total-nav', 'children'), Output('analytics-num-deals', 'children'),
     Output('analytics-weighted-irr', 'children'), Output('analytics-top1-conc', 'children')],
    [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
     Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')]
)
def update_analytics_summary(view, deals, pipeline, placeholders):
    """Update summary cards based on view toggle - 3 options"""
    portfolio_deals = get_portfolio_for_analytics_view(view, deals, pipeline, placeholders)

    if not portfolio_deals:
        return "$0.0M", "0", "0.0%", "0.0%"

    total_nav = sum(d['size'] for d in portfolio_deals)
    num_deals = len(portfolio_deals)
    weighted_irr = sum(d.get('target_irr', 0) * d['size'] for d in portfolio_deals) / total_nav if total_nav > 0 else 0

    # Top 1 concentration
    sorted_deals = sorted(portfolio_deals, key=lambda x: x['size'], reverse=True)
    top1_conc = sorted_deals[0]['size'] / total_nav if len(sorted_deals) > 0 and total_nav > 0 else 0

    return f"${total_nav:.1f}M", str(num_deals), f"{weighted_irr:.1%}", f"{top1_conc:.1%}"


# Strategy Exposure
@app.callback(
    Output('analytics-strategy', 'figure'),
    [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
     Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')]
)
def update_strategy_chart(view, deals, pipeline, placeholders):
    portfolio_deals = get_portfolio_for_analytics_view(view, deals, pipeline, placeholders)

    if not portfolio_deals:
        fig = go.Figure()
        fig.update_layout(**CHART_BASE)
        return fig

    # Group by strategy
    by_strategy = {}
    for d in portfolio_deals:
        strat = d.get('strategy', 'Other')
        by_strategy[strat] = by_strategy.get(strat, 0) + d['size']

    labels = list(by_strategy.keys())
    values = list(by_strategy.values())
    colors = [C['blue'], C['purple'], C['teal'], C['green']][:len(labels)]

    fig = go.Figure(data=[go.Pie(
        labels=labels, values=values, hole=0.4,
        marker=dict(colors=colors),
        textfont=dict(color=C['text'], family=C['mono'])
    )])
    fig.update_layout(**CHART_BASE, height=350, margin=dict(t=20, b=0, l=0, r=0))

    return fig


# Regional Exposure
@app.callback(
    Output('analytics-region', 'figure'),
    [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
     Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')]
)
def update_region_chart(view, deals, pipeline, placeholders):
    if view == 'current':
        portfolio_deals = deals if deals else []
    else:
        portfolio_deals = (deals if deals else []) + \
                          [{'name': p['name'], 'size': p['size'], 'geography': p.get('region', 'North America')}
                           for p in (pipeline if pipeline else [])] + \
                          [{'name': p['name'], 'size': p['size'], 'geography': p.get('region', 'North America')}
                           for p in (placeholders if placeholders else [])]

    if not portfolio_deals:
        fig = go.Figure()
        fig.update_layout(**CHART_BASE)
        return fig

    # Group by region
    by_region = {}
    for d in portfolio_deals:
        region = d.get('geography', d.get('region', 'North America'))
        by_region[region] = by_region.get(region, 0) + d['size']

    labels = list(by_region.keys())
    values = list(by_region.values())

    fig = go.Figure(data=[go.Bar(
        x=labels, y=values, marker_color=C['blue'],
        text=[f"${v:.1f}M" for v in values], textposition='outside',
        textfont=dict(color=C['text'], family=C['mono'])
    )])
    fig.update_layout(**CHART_BASE, yaxis_title="NAV ($mm)", height=350)

    return fig


# Vintage Exposure
@app.callback(
    Output('analytics-vintage', 'figure'),
    [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
     Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')]
)
def update_vintage_chart(view, deals, pipeline, placeholders):
    if view == 'current':
        portfolio_deals = deals if deals else []
    else:
        portfolio_deals = (deals if deals else []) + \
                          [{'name': p['name'], 'size': p['size'], 'vintage': 2026} for p in
                           (pipeline if pipeline else [])] + \
                          [{'name': p['name'], 'size': p['size'], 'vintage': 2026} for p in
                           (placeholders if placeholders else [])]

    if not portfolio_deals:
        fig = go.Figure()
        fig.update_layout(**CHART_BASE)
        return fig

    # Group by vintage
    by_vintage = {}
    for d in portfolio_deals:
        vint = d.get('vintage', 2024)
        by_vintage[vint] = by_vintage.get(vint, 0) + d['size']

    labels = [str(v) for v in sorted(by_vintage.keys())]
    values = [by_vintage[int(v)] for v in labels]

    fig = go.Figure(data=[go.Bar(
        x=labels, y=values, marker_color=C['purple'],
        text=[f"${v:.1f}M" for v in values], textposition='outside',
        textfont=dict(color=C['text'], family=C['mono'])
    )])
    fig.update_layout(**CHART_BASE, yaxis_title="NAV ($mm)", height=350)

    return fig


# Deal Type (Secondary vs Co-Investment)
@app.callback(
    Output('analytics-dealtype', 'figure'),
    [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
     Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')]
)
def update_dealtype_chart(view, deals, pipeline, placeholders):
    if view == 'current':
        portfolio_deals = deals if deals else []
    else:
        portfolio_deals = (deals if deals else []) + \
                          [{'name': p['name'], 'size': p['size'], 'deal_type': 'Secondary'} for p in
                           (pipeline if pipeline else [])] + \
                          [{'name': p['name'], 'size': p['size'], 'deal_type': p.get('deal_type', 'Secondary')}
                           for p in (placeholders if placeholders else [])]

    if not portfolio_deals:
        fig = go.Figure()
        fig.update_layout(**CHART_BASE)
        return fig

    # Group by deal type
    secondary = sum(d['size'] for d in portfolio_deals if d.get('deal_type', 'Secondary') == 'Secondary')
    coinvest = sum(d['size'] for d in portfolio_deals if d.get('deal_type', 'Secondary') == 'Co-Investment')

    labels = ['Secondary', 'Co-Investment']
    values = [secondary, coinvest]
    colors_dt = [C['blue'], C['green']]

    fig = go.Figure(data=[go.Pie(
        labels=labels, values=values, hole=0.4,
        marker=dict(colors=colors_dt),
        textfont=dict(color=C['text'], family=C['mono'])
    )])
    fig.update_layout(**CHART_BASE, height=350, margin=dict(t=20, b=0, l=0, r=0))

    return fig


# Sector Exposure
@app.callback(
    Output('analytics-sector', 'figure'),
    [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
     Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')]
)
def update_sector_chart(view, deals, pipeline, placeholders):
    if view == 'current':
        portfolio_deals = deals if deals else []
    else:
        portfolio_deals = (deals if deals else []) + \
                          [{'name': p['name'], 'size': p['size'], 'sector': 'Technology'} for p in
                           (pipeline if pipeline else [])] + \
                          [{'name': p['name'], 'size': p['size'], 'sector': 'Technology'} for p in
                           (placeholders if placeholders else [])]

    if not portfolio_deals:
        fig = go.Figure()
        fig.update_layout(**CHART_BASE)
        return fig

    # Group by sector
    by_sector = {}
    for d in portfolio_deals:
        sector = d.get('sector', 'Technology')
        by_sector[sector] = by_sector.get(sector, 0) + d['size']

    labels = list(by_sector.keys())
    values = list(by_sector.values())
    colors_sec = [C['teal'], C['purple'], C['amber'], C['green'], C['pink']][:len(labels)]

    fig = go.Figure(data=[go.Pie(
        labels=labels, values=values, hole=0.4,
        marker=dict(colors=colors_sec),
        textfont=dict(color=C['text'], family=C['mono'])
    )])
    fig.update_layout(**CHART_BASE, height=350, margin=dict(t=20, b=0, l=0, r=0))

    return fig


# Concentration Risk
@app.callback(
    Output('analytics-concentration', 'figure'),
    [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
     Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')]
)
def update_concentration_chart(view, deals, pipeline, placeholders):
    if view == 'current':
        portfolio_deals = deals if deals else []
    else:
        portfolio_deals = (deals if deals else []) + \
                          [{'name': p['name'], 'size': p['size']} for p in (pipeline if pipeline else [])] + \
                          [{'name': p['name'], 'size': p['size']} for p in (placeholders if placeholders else [])]

    if not portfolio_deals:
        fig = go.Figure()
        fig.update_layout(**CHART_BASE)
        return fig

    total_nav = sum(d['size'] for d in portfolio_deals)
    sorted_deals = sorted(portfolio_deals, key=lambda x: x['size'], reverse=True)

    top1 = (sorted_deals[0]['size'] / total_nav * 100) if len(sorted_deals) >= 1 else 0
    top3 = (sum(d['size'] for d in sorted_deals[:3]) / total_nav * 100) if len(sorted_deals) >= 3 else 0
    top5 = (sum(d['size'] for d in sorted_deals[:5]) / total_nav * 100) if len(sorted_deals) >= 5 else 0

    cats = ['Top 1', 'Top 3', 'Top 5']
    vals = [top1, top3, top5]
    limits = [15, 40, 60]
    colors_conc = [C['red'] if v > l else C['green'] for v, l in zip(vals, limits)]

    fig = go.Figure(data=[go.Bar(
        x=cats, y=vals, marker_color=colors_conc,
        text=[f"{v:.1f}%" for v in vals], textposition='outside',
        textfont=dict(color=C['text'], family=C['mono'])
    )])
    for limit in limits:
        fig.add_hline(y=limit, line_dash='dash', line_color=C['red'], opacity=0.5,
                      annotation_text=f"Limit: {limit}%", annotation_position="right")
    fig.update_layout(**CHART_BASE, yaxis_title="% of NAV", height=350)

    return fig


# Dashboard Recent Activity & Pipeline Status callbacks
@app.callback(
    [Output('dash-recent-activity', 'children'), Output('dash-pipeline-status', 'children')],
    [Input('deals-store', 'data'), Input('pipeline-store', 'data')]
)
def update_dashboard_activity(deals, pipeline):
    # Recent Activity
    if not deals:
        recent = html.P("No deals yet", className="text-muted", style={'fontFamily': C['mono']})
    else:
        recent_deals = sorted(deals, key=lambda x: x.get('date_added', ''), reverse=True)[:5]
        activity_items = []
        for d in recent_deals:
            activity_items.append(html.Div([
                html.Strong(d['name'], style={'color': C['blue'], 'fontFamily': C['mono']}),
                html.Br(),
                html.Small(f"${d['size']:.1f}M • {d['target_irr']:.1%} IRR",
                           style={'color': C['muted'], 'fontFamily': C['mono']}),
                html.Hr(style={'borderColor': C['border'], 'margin': '0.5rem 0'})
            ]))
        recent = html.Div(activity_items)

    # Pipeline Status
    if not pipeline:
        pipe_status = html.P("No pipeline deals", className="text-muted", style={'fontFamily': C['mono']})
    else:
        stage_counts = {}
        for p in pipeline:
            stage = p.get('stage', 'Screening')
            stage_counts[stage] = stage_counts.get(stage, 0) + 1

        pipe_items = []
        stage_colors = {
            'Screening': C['muted'],
            'Due Diligence': C['amber'],
            'Term Sheet': C['blue'],
            'Final Docs': C['green']
        }
        for stage, count in stage_counts.items():
            pipe_items.append(html.Div([
                html.Span("●", style={'color': stage_colors.get(stage, C['muted']), 'marginRight': '0.5rem'}),
                html.Strong(f"{stage}: ", style={'fontFamily': C['sans'], 'color': C['text']}),
                html.Span(f"{count} deals", style={'fontFamily': C['mono'], 'color': C['muted']})
            ], style={'marginBottom': '0.5rem'}))
        pipe_status = html.Div(pipe_items)

    return recent, pipe_status


# Export CSV
@app.callback(
    Output("download-csv", "data"),
    Input("btn-export", "n_clicks"),
    [State("deals-store", "data"), State("placeholder-deals-store", "data"), State("pipeline-store", "data")],
    prevent_initial_call=True
)
def export_all(n, deals, placeholders, pipeline):
    if deals:
        df_deals = pd.DataFrame(deals)
        df_deals.to_csv('current_portfolio.csv', index=False)
    if placeholders:
        df_ph = pd.DataFrame(placeholders)
        df_ph.to_csv('placeholder_deals.csv', index=False)
    if pipeline:
        df_pipe = pd.DataFrame(pipeline)
        df_pipe.to_csv('pipeline.csv', index=False)

    return dcc.send_data_frame(pd.DataFrame(deals).to_csv if deals else pd.DataFrame().to_csv,
                               f"portfolio_{date.today()}.csv", index=False)


# Auto-save callback - saves data whenever it changes
@app.callback(
    Output('save-status', 'children'),
    [Input('deals-store', 'data'), Input('pipeline-store', 'data'),
     Input('placeholder-deals-store', 'data'), Input('config-store', 'data')],
    prevent_initial_call=True
)
def auto_save_data(deals, pipeline, placeholders, config):
    """Automatically save data when anything changes"""
    save_data(deals or [], pipeline or [], placeholders or [], config or DEFAULT_CONFIG)
    return ""


# For production deployment (Render, Heroku, AWS, etc.)
server = app.server

if __name__ == '__main__':
    print("\n" + "=" * 80)
    print("HORIZON PORTFOLIO TOOL - INSTITUTIONAL GRADE LP MANAGEMENT")
    print("=" * 80)
    print("\n✅ Starting on http://localhost:8050")
    print("\n🎯 Features:")
    print("   • Current Portfolio Management (17 fields per deal)")
    print("   • Portfolio Segmentation & TWR Tracking")
    print("   • Pro Forma Portfolio (Placeholder Deals)")
    print("   • Custom Bite Sizes & Weights")
    print("   • 12-Month Dry Powder Forecasting")
    print("   • Deal Bite Sizing (Min/Desired/Max)")
    print("   • Pipeline Management")
    print("   • Return Calculator with Waterfall")
    print("   • Analytics with 3-Way Toggle (Current/Pipeline/Full)")
    print("   • Excel File Uploads (Fund CF + Liquidity Pull)")
    print("   • Comprehensive Exposure Analysis")
    print("   • Auto-Save Data Persistence")
    print("\nPress CTRL+C to stop\n")

    app.run(debug=True, host='0.0.0.0', port=8050)
