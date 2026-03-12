"""
HORIZON PORTFOLIO TOOL - INSTITUTIONAL GRADE LP MANAGEMENT
Comprehensive evergreen fund portfolio management and analytics platform

Features:
✓ Current Portfolio Management (add/delete deals)
✓ Pro Forma Portfolio (placeholder/future deals)
✓ Dry Powder Forecasting & Bite Sizing
✓ Pipeline Management
✓ Return Calculator with Waterfall
✓ Pacing Model with Multi-Year Forecasts
✓ Portfolio Construction Analytics
✓ Deal Sizing Recommendations (Min/Desired/Max)

Run: python ULTIMATE_Evergreen_Fund_Dashboard.py
Open: http://localhost:8050
"""

import dash
from dash import dcc, html, dash_table, Input, Output, State, callback_context, ALL
import dash_bootstrap_components as dbc
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import pandas as pd
import numpy as np
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import json
import base64
from io import BytesIO
import pickle
import os

# ==================== DATA PERSISTENCE ====================

DATA_FILE = 'portfolio_data.pkl'


def save_data(deals, pipeline, placeholders, config):
    """Save all data to disk"""
    data = {
        'deals': deals,
        'pipeline': pipeline,
        'placeholders': placeholders,
        'config': config,
        'saved_at': datetime.now().isoformat()
    }
    try:
        with open(DATA_FILE, 'wb') as f:
            pickle.dump(data, f)
        print(f"✅ Data saved successfully at {datetime.now().strftime('%H:%M:%S')}")
    except Exception as e:
        print(f"⚠️ Error saving data: {e}")


def load_data():
    """Load data from disk"""
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'rb') as f:
                data = pickle.load(f)
            print(f"✅ Data loaded from {data.get('saved_at', 'unknown time')}")
            return data
        except Exception as e:
            print(f"⚠️ Error loading data: {e}")
            return None
    return None



def get_last_month_end(reference_date=None):
    """Return last month end date."""
    reference_date = reference_date or datetime.now()
    month_start = reference_date.replace(day=1)
    return (month_start - relativedelta(days=1)).date()


def normalize_deal_name(name):
    return str(name or '').strip().lower()


def get_effective_dry_powder(config, uploaded_dry_powder=None):
    """Use uploaded liquidity dry powder when available, else config default."""
    if uploaded_dry_powder not in (None, '', 0):
        try:
            return float(uploaded_dry_powder)
        except Exception:
            pass
    try:
        return float(config.get('fund_parameters', {}).get('dry_powder', 0))
    except Exception:
        return 0.0


def get_latest_available_nav_for_deal(fund_cf_data, deal_name, as_of_date=None):
    """Get latest available NAV for a deal from uploaded cashflow data up to last month end."""
    if not fund_cf_data or not deal_name:
        return None

    as_of_date = as_of_date or get_last_month_end()
    target_name = normalize_deal_name(deal_name)

    for cf_deal in fund_cf_data:
        if normalize_deal_name(cf_deal.get('name')) != target_name:
            continue

        nav_series = cf_deal.get('nav_series') or {}
        dated_values = []

        for key, value in nav_series.items():
            try:
                dt = pd.to_datetime(key).date()
            except Exception:
                continue

            try:
                num = float(value)
            except Exception:
                continue

            if num != 0:
                dated_values.append((dt, num))

        dated_values.sort(key=lambda x: x[0])

        for dt, num in reversed(dated_values):
            if dt <= as_of_date:
                return num

        if dated_values:
            return dated_values[-1][1]

        fallback_nav = cf_deal.get('nav')
        if fallback_nav not in (None, ''):
            try:
                return float(fallback_nav)
            except Exception:
                pass

    return None


def enrich_deals_with_latest_nav(deals, fund_cf_data):
    """Overlay uploaded latest NAVs onto deal records for display and calculations."""
    enriched = []
    for deal in deals or []:
        updated = dict(deal)
        latest_nav = get_latest_available_nav_for_deal(fund_cf_data, deal.get('name'))
        if latest_nav is not None:
            updated['nav'] = latest_nav
            updated['size'] = latest_nav
        else:
            nav = updated.get('nav', updated.get('size', 0))
            updated['nav'] = nav
            updated['size'] = nav
        enriched.append(updated)
    return enriched


def build_cashflow_template_table(title, data_rows, columns):
    """Render a cashflow table section in the uploaded template style."""
    if not data_rows:
        return dbc.Alert(f"No {title.lower()} data available.", color="light")

    return dbc.Card([
        dbc.CardHeader(title, style={
            'fontWeight': 'bold',
            'backgroundColor': C['surface'],
            'borderBottom': f'1px solid {C["border"]}'
        }),
        dbc.CardBody([
            dash_table.DataTable(
                data=data_rows,
                columns=[{"name": c, "id": c} for c in columns],
                style_cell={
                    'textAlign': 'left',
                    'padding': '8px 10px',
                    'fontFamily': C['mono'],
                    'fontSize': '11px',
                    'minWidth': '90px',
                    'maxWidth': '180px',
                    'whiteSpace': 'normal'
                },
                style_header={
                    'backgroundColor': C['surface'],
                    'color': C['text'],
                    'fontWeight': 'bold',
                    'border': f'1px solid {C["border"]}',
                    'position': 'sticky',
                    'top': 0,
                    'zIndex': 1
                },
                style_data={
                    'backgroundColor': C['panel'],
                    'color': C['text'],
                    'border': f'1px solid {C["border"]}'
                },
                style_data_conditional=[
                    {'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']},
                    {'if': {'column_id': 'AIC Name'}, 'fontWeight': 'bold', 'minWidth': '180px'},
                    {'if': {'column_id': 'Investments'}, 'fontWeight': 'bold', 'minWidth': '180px'},
                ],
                style_table={
                    'overflowX': 'auto',
                    'maxHeight': '480px',
                    'overflowY': 'auto'
                },
                fixed_columns={'headers': True, 'data': 5},
                fixed_rows={'headers': True},
                page_action='none',
                export_format='xlsx',
                export_headers='display',
            )
        ])
    ], className="shadow-sm mb-4")

# ==================== COLORS ====================

C = dict(
    bg="#07090f", panel="#0d1117", surface="#111822",
    border="#1c2a3a", border2="#243649",
    blue="#2979c8", sky="#56b4f5", teal="#2ec4b6",
    green="#2ecc71", red="#e5493a", amber="#f0a500",
    purple="#9b72cf", pink="#e05c8a",
    text="#d4e6f5", muted="#5e82a0", dim="#324d63",
    mono="'JetBrains Mono', 'Fira Mono', monospace",
    sans="'IBM Plex Sans', 'Segoe UI', sans-serif",
)


def rgba(hex_color, alpha):
    r, g, b = int(hex_color[1:3], 16), int(hex_color[3:5], 16), int(hex_color[5:7], 16)
    return f"rgba({r},{g},{b},{alpha})"


CHART_BASE = dict(
    paper_bgcolor=C["panel"], plot_bgcolor=C["surface"],
    font=dict(family=C["mono"], color=C["text"], size=11),
    margin=dict(l=52, r=24, t=40, b=40),
    xaxis=dict(gridcolor=C["border"], zerolinecolor=C["border"]),
    yaxis=dict(gridcolor=C["border"], zerolinecolor=C["border"]),
    legend=dict(bgcolor=C["panel"], bordercolor=C["border"], borderwidth=1),
)

# Legacy color mappings for compatibility
COLORS = {
    'primary': C['blue'],
    'secondary': C['purple'],
    'accent': C['amber'],
    'success': C['green'],
    'danger': C['red'],
    'dark': C['bg'],
    'light': C['panel'],
    'warning': C['amber'],
    'info': C['sky']
}


# ==================== HELPER FUNCTIONS ====================

# ==================== HELPER FUNCTIONS ====================

def generate_month_options():
    """Generate month options from Jan 2026 to Dec 2040 (180 months)"""
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    options = []
    for i in range(180):  # 15 years * 12 months
        month_name = month_names[i % 12]
        year = 2026 + (i // 12)
        options.append({"label": f"{month_name} {year}", "value": i})
    return options


def calculate_required_future_irr(current_portfolio_irr, current_nav, dry_powder, config):
    """Calculate IRR needed on future deals to hit target net TWR"""
    target_twr = config['fund_parameters']['target_net_twr']
    mgmt_fee = config['fund_parameters']['management_fee']
    carry_rate = config['fund_parameters']['carry_rate']
    hurdle = config['fund_parameters']['hurdle_rate']
    loss_drag = config['fund_parameters']['loss_drag']
    cash_reserve = config['fund_parameters']['liquidity_reserve_pct']
    cash_yield = config['fund_parameters'].get('cash_yield', 0.03)

    # Calculate invested ratio
    total_fund = current_nav + dry_powder
    if total_fund == 0:
        return 0.25

    invested_ratio = max(0, 1 - cash_reserve)
    idle_ratio = cash_reserve

    # Gross return needed before carry
    gross_needed = (target_twr + mgmt_fee + loss_drag - (idle_ratio * cash_yield)) / invested_ratio

    # Add carry drag if above hurdle
    if gross_needed > hurdle:
        carry_drag = (gross_needed - hurdle) * carry_rate
        gross_needed += carry_drag

    # Calculate weighted average
    if dry_powder == 0:
        return gross_needed

    current_weight = current_nav / total_fund
    future_weight = dry_powder / total_fund

    required_future = (gross_needed - (current_weight * current_portfolio_irr)) / future_weight

    return max(0, min(1.0, required_future))


def calculate_portfolio_metrics(deals):
    """Calculate comprehensive portfolio metrics"""
    if not deals:
        return {
            'total_nav': 0,
            'num_deals': 0,
            'weighted_irr': 0,
            'by_strategy': {},
            'by_vintage': {},
            'by_sector': {},
            'concentration_top1': 0,
            'concentration_top3': 0,
            'concentration_top5': 0,
            'effective_n': 0
        }

    total_nav = sum(d['size'] for d in deals)
    num_deals = len(deals)
    weighted_irr = sum(d['size'] * d['target_irr'] for d in deals) / total_nav if total_nav > 0 else 0

    # Effective number of positions (Herfindahl)
    weights = [d['size'] / total_nav for d in deals] if total_nav > 0 else []
    effective_n = 1 / sum(w ** 2 for w in weights) if weights else 0

    # By strategy
    by_strategy = {}
    for deal in deals:
        strategy = deal['strategy']
        if strategy not in by_strategy:
            by_strategy[strategy] = {'nav': 0, 'count': 0, 'deals': []}
        by_strategy[strategy]['nav'] += deal['size']
        by_strategy[strategy]['count'] += 1
        by_strategy[strategy]['deals'].append(deal)

    for strategy in by_strategy:
        nav = by_strategy[strategy]['nav']
        by_strategy[strategy]['weighted_irr'] = sum(
            d['size'] * d['target_irr'] for d in by_strategy[strategy]['deals']) / nav if nav > 0 else 0
        by_strategy[strategy]['allocation'] = nav / total_nav if total_nav > 0 else 0

    # By vintage
    by_vintage = {}
    for deal in deals:
        vintage = deal.get('vintage', 2024)
        if vintage not in by_vintage:
            by_vintage[vintage] = {'nav': 0, 'count': 0}
        by_vintage[vintage]['nav'] += deal['size']
        by_vintage[vintage]['count'] += 1

    # By sector
    by_sector = {}
    for deal in deals:
        sector = deal.get('sector', 'Other')
        if sector not in by_sector:
            by_sector[sector] = {'nav': 0, 'count': 0}
        by_sector[sector]['nav'] += deal['size']
        by_sector[sector]['count'] += 1

    # Concentration
    sorted_deals = sorted(deals, key=lambda x: x['size'], reverse=True)
    concentration_top1 = sorted_deals[0]['size'] / total_nav if len(sorted_deals) > 0 and total_nav > 0 else 0
    concentration_top3 = sum(d['size'] for d in sorted_deals[:3]) / total_nav if len(
        sorted_deals) >= 3 and total_nav > 0 else 0
    concentration_top5 = sum(d['size'] for d in sorted_deals[:5]) / total_nav if len(
        sorted_deals) >= 5 and total_nav > 0 else 0

    return {
        'total_nav': total_nav,
        'num_deals': num_deals,
        'weighted_irr': weighted_irr,
        'by_strategy': by_strategy,
        'by_vintage': by_vintage,
        'by_sector': by_sector,
        'concentration_top1': concentration_top1,
        'concentration_top3': concentration_top3,
        'concentration_top5': concentration_top5,
        'effective_n': effective_n
    }


def calculate_bite_sizes(dry_powder, config):
    """Calculate recommended deal sizes (min/desired/max) for each strategy"""
    custom = calculate_custom_bite_sizes(dry_powder, config)
    if custom:
        return custom

    bite_sizes = {}

    for strategy_name in ['GP-Led (Multi-Asset)', 'GP-Led (Single-Asset)', 'Diversified LP-Led', 'Co-Investments']:
        if 'Multi-Asset' in strategy_name:
            min_pct, desired_pct, max_pct = 0.005, 0.0275, 0.05
        elif 'Single-Asset' in strategy_name:
            min_pct, desired_pct, max_pct = 0.005, 0.0225, 0.04
        elif 'LP-Led' in strategy_name:
            min_pct, desired_pct, max_pct = 0.005, 0.0275, 0.05
        else:  # Co-Investments
            min_pct, desired_pct, max_pct = 0.005, 0.0175, 0.03

        bite_sizes[strategy_name] = {
            'min': dry_powder * min_pct,
            'desired': dry_powder * desired_pct,
            'max': dry_powder * max_pct,
            'min_pct': min_pct,
            'desired_pct': desired_pct,
            'max_pct': max_pct
        }

    return bite_sizes




def get_fund_overview_config(config):
    config = config or DEFAULT_CONFIG
    overview = json.loads(json.dumps(DEFAULT_FUND_OVERVIEW))
    stored = config.get('fund_overview', {}) or {}
    for key, value in stored.items():
        overview[key] = value
    return overview


def calculate_current_nav_with_liquidity(deals, liquidity_data=None, fund_cf_data=None):
    deals = deals or []
    liquidity_data = liquidity_data or {}
    enriched_deals = enrich_deals_with_latest_nav(deals, fund_cf_data)
    investment_nav = sum(float(d.get('nav', d.get('size', 0)) or 0) for d in enriched_deals)
    cash = float(liquidity_data.get('cash_balance', 0) or 0)
    glf = float(liquidity_data.get('glf_balance', 0) or 0)
    cqs = float(liquidity_data.get('cqs_balance', 0) or 0)
    return investment_nav + cash + glf + cqs


def calculate_custom_bite_sizes(dry_powder, config):
    overview = get_fund_overview_config(config)
    bite_targets = overview.get('bite_size_targets', []) or []
    bite_sizes = {}
    for row in bite_targets:
        asset_type = row.get('asset_type')
        if not asset_type:
            continue
        min_pct = float(row.get('min_pct', 0) or 0)
        desired_pct = float(row.get('desired_pct', 0) or 0)
        max_pct = float(row.get('max_pct', 0) or 0)
        bite_sizes[asset_type] = {
            'min': dry_powder * min_pct,
            'desired': dry_powder * desired_pct,
            'max': dry_powder * max_pct,
            'min_pct': min_pct,
            'desired_pct': desired_pct,
            'max_pct': max_pct,
        }
    return bite_sizes


def build_exposure_map(deals, dimension):
    exposures = {}
    total_nav = sum(float(d.get('nav', d.get('size', 0)) or 0) for d in (deals or []))
    if total_nav <= 0:
        return exposures, 0

    for deal in deals or []:
        nav = float(deal.get('nav', deal.get('size', 0)) or 0)
        if dimension == 'Strategy':
            key = deal.get('strategy', 'Unknown')
        elif dimension == 'Region':
            key = deal.get('geography', deal.get('region', 'Unknown'))
        elif dimension == 'Sector':
            key = deal.get('sector', 'Unknown')
        elif dimension == 'Vintage':
            key = str(deal.get('vintage', 'Unknown'))
        elif dimension == 'Manager':
            key = deal.get('manager', 'Unknown')
        elif dimension == 'Single Asset':
            key = deal.get('name', 'Unknown')
        else:
            key = deal.get(dimension.lower(), 'Unknown')
        exposures[key] = exposures.get(key, 0) + nav
    return exposures, total_nav


def build_exposure_comparison_rows(deals, config):
    overview = get_fund_overview_config(config)
    limits = overview.get('exposure_limits', []) or []
    rows = []
    for row in limits:
        dimension = row.get('dimension', 'Unknown')
        category = str(row.get('category', 'Unknown'))
        exposures, total_nav = build_exposure_map(deals, dimension)
        current_value = exposures.get(category, 0.0)
        current_pct = current_value / total_nav if total_nav > 0 else 0.0
        min_pct = float(row.get('min_pct', 0) or 0)
        target_pct = float(row.get('target_pct', 0) or 0)
        max_pct = float(row.get('max_pct', 0) or 0)
        status = 'Within Range'
        if current_pct < min_pct:
            status = 'Below Min'
        elif current_pct > max_pct and max_pct > 0:
            status = 'Above Max'
        rows.append({
            'Dimension': dimension,
            'Category': category,
            'Current NAV ($m)': round(current_value, 2),
            'Current %': round(current_pct * 100, 2),
            'Min %': round(min_pct * 100, 2),
            'Target %': round(target_pct * 100, 2),
            'Max %': round(max_pct * 100, 2),
            'Gap to Target ($m)': round((target_pct - current_pct) * total_nav, 2),
            'Status': status,
        })
    return rows


def build_forward_bite_rows(pipeline, placeholders, dry_powder, config):
    bite_sizes = calculate_custom_bite_sizes(dry_powder, config)
    rows = []

    def add_row(item, source):
        strategy = item.get('strategy') or item.get('type') or 'Unknown'
        bite = bite_sizes.get(strategy, {'min': 0, 'desired': 0, 'max': 0})
        override = item.get('bite_override')
        try:
            override = float(override) if override not in (None, '') else None
        except Exception:
            override = None
        current_size = float(item.get('size', 0) or 0)
        effective = override if override is not None else (current_size if current_size > 0 else bite.get('desired', 0))
        rows.append({
            'Source': source,
            'Name': item.get('name', ''),
            'Asset Type': strategy,
            'Current Size ($m)': round(current_size, 2),
            'Base Min ($m)': round(bite.get('min', 0), 2),
            'Base Desired ($m)': round(bite.get('desired', 0), 2),
            'Base Max ($m)': round(bite.get('max', 0), 2),
            'Override Bite ($m)': override,
            'Effective Bite ($m)': round(effective, 2),
        })

    for p in pipeline or []:
        add_row(p, 'Pipeline')
    for p in placeholders or []:
        add_row(p, 'Future Deals')
    return rows

def forecast_dry_powder(current_nav, dry_powder, deals, placeholder_deals, config, months=12):
    """Forecast dry powder availability over next N months starting from Jan 2026"""
    forecast = []
    nav = current_nav
    powder = dry_powder

    base_date = datetime(2026, 1, 1)

    for month in range(months):
        month_date = base_date + relativedelta(months=month)

        # Simple monthly growth on NAV
        monthly_return = config['fund_parameters']['target_net_twr'] / 12
        nav_growth = nav * monthly_return

        # Distributions (20% annually = 1.67% monthly)
        monthly_dist = nav * (config['fund_parameters']['distribution_rate'] / 12)

        # Capital calls from placeholder deals in this month
        calls_this_month = 0
        for pd_deal in placeholder_deals:
            if pd_deal.get('expected_month') == month:
                calls_this_month += pd_deal['size']

        # Update
        powder = powder + monthly_dist - calls_this_month
        nav = nav + nav_growth + calls_this_month - monthly_dist

        forecast.append({
            'month': month_date.strftime('%b %Y'),
            'dry_powder': powder,
            'nav': nav,
            'distributions': monthly_dist,
            'calls': calls_this_month
        })

    return forecast


# ==================== DASH APP ====================

app = dash.Dash(__name__,
                external_stylesheets=[dbc.themes.BOOTSTRAP, dbc.icons.FONT_AWESOME],
                suppress_callback_exceptions=True)

app.title = "Ultimate Evergreen Fund Manager"
server = app.server

# Custom CSS for dark theme
app.index_string = '''
<!DOCTYPE html>
<html>
    <head>
        {%metas%}
        <title>{%title%}</title>
        {%favicon%}
        {%css%}
        <style>
            @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500;600&display=swap');

            body {
                background-color: ''' + C['bg'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
                font-family: ''' + C['sans'] + ''' !important;
            }

            .card {
                background-color: ''' + C['panel'] + ''' !important;
                border: 1px solid ''' + C['border'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
            }

            .card-header {
                background-color: ''' + C['surface'] + ''' !important;
                border-bottom: 1px solid ''' + C['border'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
            }

            .card-body {
                background-color: ''' + C['panel'] + ''' !important;
            }

            .nav-pills .nav-link {
                color: ''' + C['muted'] + ''' !important;
                background-color: transparent !important;
                border-radius: 8px !important;
                font-family: ''' + C['sans'] + ''' !important;
            }

            .nav-pills .nav-link:hover {
                background-color: ''' + C['surface'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
            }

            .nav-pills .nav-link.active {
                background: linear-gradient(135deg, ''' + C['blue'] + ''' 0%, ''' + C['sky'] + ''' 100%) !important;
                color: white !important;
                font-weight: 600 !important;
            }

            .navbar {
                background: linear-gradient(135deg, ''' + C['bg'] + ''' 0%, ''' + C['surface'] + ''' 100%) !important;
                border-bottom: 2px solid ''' + C['border'] + ''' !important;
            }

            .btn-primary {
                background: linear-gradient(135deg, ''' + C['blue'] + ''' 0%, ''' + C['sky'] + ''' 100%) !important;
                border: none !important;
                font-weight: 600 !important;
                font-family: ''' + C['sans'] + ''' !important;
            }

            .btn-primary:hover {
                background: linear-gradient(135deg, ''' + C['sky'] + ''' 0%, ''' + C['teal'] + ''' 100%) !important;
                transform: translateY(-1px);
                box-shadow: 0 4px 12px ''' + rgba(C['blue'], 0.3) + ''' !important;
            }

            .btn-success {
                background: linear-gradient(135deg, ''' + C['green'] + ''' 0%, ''' + C['teal'] + ''' 100%) !important;
                border: none !important;
            }

            .btn-danger {
                background-color: ''' + C['red'] + ''' !important;
                border: none !important;
            }

            .btn-secondary {
                background-color: ''' + C['surface'] + ''' !important;
                border: 1px solid ''' + C['border'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
            }

            .btn-info {
                background: linear-gradient(135deg, ''' + C['purple'] + ''' 0%, ''' + C['pink'] + ''' 100%) !important;
                border: none !important;
            }

            .modal-content {
                background-color: ''' + C['panel'] + ''' !important;
                border: 1px solid ''' + C['border'] + ''' !important;
            }

            .modal-header {
                background-color: ''' + C['surface'] + ''' !important;
                border-bottom: 1px solid ''' + C['border'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
            }

            .modal-body {
                background-color: ''' + C['panel'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
            }

            .modal-footer {
                background-color: ''' + C['surface'] + ''' !important;
                border-top: 1px solid ''' + C['border'] + ''' !important;
            }

            .form-control, .form-select {
                background-color: ''' + C['surface'] + ''' !important;
                border: 1px solid ''' + C['border'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
                font-family: ''' + C['mono'] + ''' !important;
            }

            .form-control:focus, .form-select:focus {
                background-color: ''' + C['surface'] + ''' !important;
                border-color: ''' + C['blue'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
                box-shadow: 0 0 0 0.2rem ''' + rgba(C['blue'], 0.25) + ''' !important;
            }

            .form-label {
                color: ''' + C['text'] + ''' !important;
                font-weight: 600 !important;
                font-family: ''' + C['sans'] + ''' !important;
            }

            .text-muted {
                color: ''' + C['muted'] + ''' !important;
            }

            .alert-info {
                background-color: ''' + rgba(C['blue'], 0.1) + ''' !important;
                border-color: ''' + C['blue'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
            }

            h1, h2, h3, h4, h5, h6 {
                color: ''' + C['text'] + ''' !important;
                font-family: ''' + C['sans'] + ''' !important;
            }

            .shadow-sm {
                box-shadow: 0 2px 8px ''' + rgba(C['bg'], 0.4) + ''' !important;
            }

            .border-0 {
                border: 1px solid ''' + C['border'] + ''' !important;
            }

            small {
                color: ''' + C['muted'] + ''' !important;
            }

            /* Dash DataTable */
            .dash-table-container {
                font-family: ''' + C['mono'] + ''' !important;
            }

            .dash-spreadsheet {
                background-color: ''' + C['panel'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
            }

            .dash-spreadsheet td {
                background-color: ''' + C['panel'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
                border: 1px solid ''' + C['border'] + ''' !important;
            }

            .dash-spreadsheet th {
                background-color: ''' + C['surface'] + ''' !important;
                color: ''' + C['text'] + ''' !important;
                border: 1px solid ''' + C['border'] + ''' !important;
            }

            /* Tab styling */
            .nav-tabs .nav-link {
                color: ''' + C['muted'] + ''' !important;
                background-color: transparent !important;
                border: none !important;
                border-bottom: 2px solid transparent !important;
            }

            .nav-tabs .nav-link.active {
                color: ''' + C['blue'] + ''' !important;
                border-bottom: 2px solid ''' + C['blue'] + ''' !important;
                background-color: transparent !important;
            }

            /* Scrollbar */
            ::-webkit-scrollbar {
                width: 10px;
                height: 10px;
            }

            ::-webkit-scrollbar-track {
                background: ''' + C['surface'] + ''';
            }

            ::-webkit-scrollbar-thumb {
                background: ''' + C['border2'] + ''';
                border-radius: 5px;
            }

            ::-webkit-scrollbar-thumb:hover {
                background: ''' + C['muted'] + ''';
            }
        </style>
    </head>
    <body>
        {%app_entry%}
        <footer>
            {%config%}
            {%scripts%}
            {%renderer%}
        </footer>
    </body>
</html>
'''

DEFAULT_FUND_OVERVIEW = {
    'target_fund_size': 1000.0,
    'target_current_nav': None,
    'exposure_limits': [
        {'dimension': 'Strategy', 'category': 'GP-Led (Single-Asset)', 'min_pct': 0.0, 'target_pct': 0.25, 'max_pct': 0.40},
        {'dimension': 'Strategy', 'category': 'GP-Led (Multi-Asset)', 'min_pct': 0.0, 'target_pct': 0.30, 'max_pct': 0.45},
        {'dimension': 'Strategy', 'category': 'Diversified LP-Led', 'min_pct': 0.0, 'target_pct': 0.30, 'max_pct': 0.45},
        {'dimension': 'Strategy', 'category': 'Co-Investments', 'min_pct': 0.0, 'target_pct': 0.15, 'max_pct': 0.25},
        {'dimension': 'Region', 'category': 'North America', 'min_pct': 0.0, 'target_pct': 0.40, 'max_pct': 0.70},
        {'dimension': 'Region', 'category': 'Europe', 'min_pct': 0.0, 'target_pct': 0.35, 'max_pct': 0.60},
        {'dimension': 'Sector', 'category': 'Technology', 'min_pct': 0.0, 'target_pct': 0.20, 'max_pct': 0.35},
        {'dimension': 'Vintage', 'category': '2026', 'min_pct': 0.0, 'target_pct': 0.20, 'max_pct': 0.35},
        {'dimension': 'Manager', 'category': 'Single Manager', 'min_pct': 0.0, 'target_pct': 0.10, 'max_pct': 0.20},
        {'dimension': 'Single Asset', 'category': 'Single Deal', 'min_pct': 0.0, 'target_pct': 0.05, 'max_pct': 0.10},
    ],
    'bite_size_targets': [
        {'asset_type': 'GP-Led (Single-Asset)', 'min_pct': 0.005, 'desired_pct': 0.0225, 'max_pct': 0.040},
        {'asset_type': 'GP-Led (Multi-Asset)', 'min_pct': 0.005, 'desired_pct': 0.0275, 'max_pct': 0.050},
        {'asset_type': 'Co-Investments', 'min_pct': 0.005, 'desired_pct': 0.0175, 'max_pct': 0.030},
        {'asset_type': 'Diversified LP-Led', 'min_pct': 0.005, 'desired_pct': 0.0275, 'max_pct': 0.050},
    ]
}

DEFAULT_CONFIG = {
    'fund_parameters': {
        'dry_powder': 300,
        'target_net_twr': 0.13,
        'management_fee': 0.0125,
        'carry_rate': 0.125,
        'hurdle_rate': 0.10,
        'loss_drag': 0.01,
        'liquidity_reserve_pct': 0.05,
        'distribution_rate': 0.20,
        'cash_yield': 0.03,
        'avg_hold_period': 5.0
    },
    'bite_sizing': {
        'enabled': True
    },
    'fund_overview': DEFAULT_FUND_OVERVIEW
}

# ==================== LAYOUT ====================

navbar = dbc.Navbar(
    dbc.Container([
        html.Div([
            html.I(className="fas fa-chart-line me-3", style={'fontSize': '32px', 'color': C['blue']}),
            dbc.NavbarBrand("Horizon Portfolio Tool", style={
                'fontSize': '24px',
                'fontWeight': 'bold',
                'fontFamily': C['sans'],
                'background': f'linear-gradient(135deg, {C["blue"]} 0%, {C["sky"]} 100%)',
                '-webkit-background-clip': 'text',
                '-webkit-text-fill-color': 'transparent'
            })
        ]),
        html.Div(id='live-clock', style={
            'fontSize': '14px',
            'color': C['muted'],
            'fontFamily': C['mono']
        })
    ], fluid=True),
    style={
        'background': f'linear-gradient(135deg, {C["bg"]} 0%, {C["surface"]} 100%)',
        'borderBottom': f'2px solid {C["border"]}',
        'padding': '1.2rem 2rem'
    },
    dark=True,
    className="mb-4"
)

sidebar = dbc.Card([
    dbc.CardBody([
        html.H5("📊 Navigation", className="mb-4", style={
            'fontWeight': 'bold',
            'color': C['text'],
            'fontFamily': C['sans']
        }),
        dbc.Nav([
            dbc.NavLink([html.I(className="fas fa-home me-2"), "Dashboard"], href="/", active="exact"),
            dbc.NavLink([html.I(className="fas fa-compass me-2"), "Fund Overview"], href="/overview", active="exact"),
            dbc.NavLink([html.I(className="fas fa-briefcase me-2"), "Current Portfolio"], href="/portfolio",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-layer-group me-2"), "Portfolio Segments & TWR"], href="/segmentation",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-calendar-plus me-2"), "Future Deals"], href="/future",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-water me-2"), "Dry Powder Forecast"], href="/drypowder",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-tint me-2"), "Liquidity Assumptions"], href="/liquidity",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-bullseye me-2"), "Return Calculator"], href="/calculator",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-chart-area me-2"), "TWR Forecaster"], href="/twr", active="exact"),
            dbc.NavLink([html.I(className="fas fa-dollar-sign me-2"), "Deal Cashflows"], href="/cashflows",
                        active="exact"),
            dbc.NavLink([html.I(className="fas fa-magic me-2"), "Pro Forma"], href="/proforma", active="exact"),
            dbc.NavLink([html.I(className="fas fa-funnel-dollar me-2"), "Pipeline"], href="/pipeline", active="exact"),
            dbc.NavLink([html.I(className="fas fa-chart-bar me-2"), "Analytics"], href="/analytics", active="exact"),
            dbc.NavLink([html.I(className="fas fa-cog me-2"), "Settings"], href="/settings", active="exact"),
        ], vertical=True, pills=True),
        html.Hr(style={'borderColor': C['border']}),
        html.H6("📈 Quick Stats", className="mb-3", style={
            'fontWeight': 'bold',
            'color': C['text'],
            'fontFamily': C['sans']
        }),
        html.Div(id='sidebar-stats'),
        html.Hr(style={'borderColor': C['border']}),
        dbc.Button("📥 Export All", id="btn-export", color="secondary", size="sm", className="w-100"),
        dcc.Download(id="download-csv"),
    ], style={'backgroundColor': C['panel']})
], style={
    'position': 'sticky',
    'top': '20px',
    'backgroundColor': C['panel'],
    'border': f'1px solid {C["border"]}'
})

# Load saved data at startup
LOADED_DATA = load_data()

app.layout = dbc.Container([
    dcc.Location(id='url', refresh=False),
    html.Div(id='save-status', style={'display': 'none'}),  # Hidden save status
    dcc.Store(id='deals-store', data=LOADED_DATA.get('deals', []) if LOADED_DATA else []),
    dcc.Store(id='placeholder-deals-store', data=LOADED_DATA.get('placeholders', []) if LOADED_DATA else []),
    dcc.Store(id='pipeline-store', data=LOADED_DATA.get('pipeline', []) if LOADED_DATA else []),
    dcc.Store(id='cashflows-store', data=[]),
    dcc.Store(id='fund-cf-data-store', data=None),   # Fund Level CF upload — global
    dcc.Store(id='liquidity-data-store', data=None),  # Liquidity Pull upload
    dcc.Store(id='discount-store', data={}),           # Per-deal discount / price-to-NAV entries
    dcc.Store(id='liquidity-dry-powder-store', data=None),  # Dry powder from Liquidity file
    dcc.Store(id='proforma-scenario-store', data=[]),
    dcc.Store(id='liquidity-assumptions-store', data={
        'GP-Led (Multi-Asset)': {'annual_dist_rate': 0.20, 'call_pattern': 'immediate'},
        'GP-Led (Single-Asset)': {'annual_dist_rate': 0.25, 'call_pattern': 'staged'},
        'Diversified LP-Led': {'annual_dist_rate': 0.18, 'call_pattern': 'staged'},
        'Co-Investments': {'annual_dist_rate': 0.15, 'call_pattern': 'delayed'},
    }),
    dcc.Store(id='config-store', data=LOADED_DATA.get('config', DEFAULT_CONFIG) if LOADED_DATA else DEFAULT_CONFIG),
    dcc.Store(id='restrictions-store', data=LOADED_DATA.get('restrictions', {}) if LOADED_DATA else {}),
    dcc.Interval(id='clock-interval', interval=1000),
    navbar,
    dbc.Row([
        dbc.Col(sidebar, width=2),
        dbc.Col(html.Div(id="page-content"), width=10)
    ])
], fluid=True, style={'backgroundColor': C['bg'], 'minHeight': '100vh'})


# ==================== PAGE LAYOUTS ====================

def dashboard_page():
    return html.Div([
        html.H2("📊 Fund Dashboard", className="mb-4", style={
            'fontWeight': 'bold',
            'fontFamily': C['sans'],
            'color': C['text']
        }),

        # Top KPIs
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("💰 Current NAV", className="text-muted mb-2"),
                html.H3(id="dash-nav", style={'color': C['green'], 'fontWeight': 'bold', 'fontFamily': C['mono']}),
                html.Small(id="dash-num-deals", className="text-muted")
            ])], className="shadow-sm border-0",
                style={'background': f'linear-gradient(135deg, {rgba(C["green"], 0.1)} 0%, {C["panel"]} 100%)'}),
                width=2),

            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("💵 Dry Powder", className="text-muted mb-2"),
                html.H3(id="dash-dry-powder",
                        style={'color': C['blue'], 'fontWeight': 'bold', 'fontFamily': C['mono']}),
                html.Small("Available", className="text-muted")
            ])], className="shadow-sm border-0",
                style={'background': f'linear-gradient(135deg, {rgba(C["blue"], 0.1)} 0%, {C["panel"]} 100%)'}),
                width=2),

            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("📊 Total Fund", className="text-muted mb-2"),
                html.H3(id="dash-total", style={'color': C['sky'], 'fontWeight': 'bold', 'fontFamily': C['mono']}),
                html.Small("NAV + Powder", className="text-muted")
            ])], className="shadow-sm border-0",
                style={'background': f'linear-gradient(135deg, {rgba(C["sky"], 0.1)} 0%, {C["panel"]} 100%)'}),
                width=2),

            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("🎯 Portfolio IRR", className="text-muted mb-2"),
                html.H3(id="dash-current-irr",
                        style={'color': C['teal'], 'fontWeight': 'bold', 'fontFamily': C['mono']}),
                html.Small("Weighted", className="text-muted")
            ])], className="shadow-sm border-0",
                style={'background': f'linear-gradient(135deg, {rgba(C["teal"], 0.1)} 0%, {C["panel"]} 100%)'}),
                width=2),

            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("🚀 Required IRR", className="text-muted mb-2"),
                html.H3(id="dash-req-irr", style={'color': C['amber'], 'fontWeight': 'bold', 'fontFamily': C['mono']}),
                html.Small("Future Deals", className="text-muted")
            ])], className="shadow-sm border-0",
                style={'background': f'linear-gradient(135deg, {rgba(C["amber"], 0.1)} 0%, {C["panel"]} 100%)'}),
                width=2),

            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("📅 Placeholders", className="text-muted mb-2"),
                html.H3(id="dash-placeholders",
                        style={'color': C['purple'], 'fontWeight': 'bold', 'fontFamily': C['mono']}),
                html.Small(id="dash-placeholder-value", className="text-muted")
            ])], className="shadow-sm border-0",
                style={'background': f'linear-gradient(135deg, {rgba(C["purple"], 0.1)} 0%, {C["panel"]} 100%)'}),
                width=2),
        ], className="mb-4"),

        # Charts Row
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Portfolio Allocation", style={
                        'fontWeight': 'bold',
                        'backgroundColor': C['surface'],
                        'borderBottom': f'1px solid {C["border"]}'
                    }),
                    dbc.CardBody([dcc.Graph(id='dash-allocation-chart', config={'displayModeBar': False})])
                ], className="shadow-sm border-0")
            ], width=4),

            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Dry Powder Forecast (12 Months)", style={
                        'fontWeight': 'bold',
                        'backgroundColor': C['surface'],
                        'borderBottom': f'1px solid {C["border"]}'
                    }),
                    dbc.CardBody([dcc.Graph(id='dash-forecast-chart', config={'displayModeBar': False})])
                ], className="shadow-sm border-0")
            ], width=4),

            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Deal Bite Sizing Guide", style={
                        'fontWeight': 'bold',
                        'backgroundColor': C['surface'],
                        'borderBottom': f'1px solid {C["border"]}'
                    }),
                    dbc.CardBody([html.Div(id='dash-bite-sizing')])
                ], className="shadow-sm border-0")
            ], width=4),
        ], className="mb-4"),

        # Recent Activity
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Recent Portfolio Activity", style={
                        'fontWeight': 'bold',
                        'backgroundColor': C['surface'],
                        'borderBottom': f'1px solid {C["border"]}'
                    }),
                    dbc.CardBody([html.Div(id='dash-recent-activity')])
                ], className="shadow-sm border-0")
            ], width=6),

            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Pipeline Status", style={
                        'fontWeight': 'bold',
                        'backgroundColor': C['surface'],
                        'borderBottom': f'1px solid {C["border"]}'
                    }),
                    dbc.CardBody([html.Div(id='dash-pipeline-status')])
                ], className="shadow-sm border-0")
            ], width=6),
        ])
    ])



def fund_overview_page():
    # ── Shared dropdown option lists ───────────────────────────────────────
    STRAT_OPTS  = ['GP-Led (Multi-Asset CV)', 'GP-Led (Single-Asset CV)',
                   'LP-Led Secondary', 'Co-Investment', 'Primary']
    REGION_OPTS = ['North America', 'Europe', 'Asia & ROW']
    SECTOR_OPTS = ['Technology', 'Healthcare', 'Financial Services', 'Consumer',
                   'Industrials', 'Energy', 'Real Estate', 'Other']
    STAGE_OPTS  = ['Buyout', 'Growth', 'Venture', 'Liquidity']
    VINTAGE_OPTS = [str(y) for y in range(2026, 2014, -1)]
    DIM_OPTS    = ['Investment Type', 'Region', 'Sector', 'Stage', 'Vintage']

    def _tbl_style():
        return dict(
            style_cell={'textAlign': 'left', 'padding': '8px 12px',
                        'fontFamily': C['mono'], 'fontSize': '12px',
                        'border': f'1px solid {C["border"]}'},
            style_header={'backgroundColor': C['surface'], 'color': C['text'],
                          'fontWeight': 'bold', 'fontSize': '12px',
                          'border': f'1px solid {C["border"]}'},
            style_data={'backgroundColor': C['panel'], 'color': C['text']},
            style_data_conditional=[
                {'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']},
                {'if': {'column_editable': True}, 'backgroundColor': '#fff9e6'},
            ],
        )

    # Default targets table rows (Investment Type / Region / Sector / Stage / Vintage)
    DEFAULT_TARGETS = [
        {'dimension': 'Investment Type', 'category': 'GP-Led (Multi-Asset CV)',    'min_pct': 0,  'target_pct': 30, 'max_pct': 45},
        {'dimension': 'Investment Type', 'category': 'GP-Led (Single-Asset CV)',   'min_pct': 0,  'target_pct': 25, 'max_pct': 40},
        {'dimension': 'Investment Type', 'category': 'LP-Led Secondary',           'min_pct': 0,  'target_pct': 30, 'max_pct': 45},
        {'dimension': 'Investment Type', 'category': 'Co-Investment',              'min_pct': 0,  'target_pct': 15, 'max_pct': 25},
        {'dimension': 'Investment Type', 'category': 'Primary',                    'min_pct': 0,  'target_pct': 5,  'max_pct': 15},
        {'dimension': 'Region',          'category': 'North America',              'min_pct': 0,  'target_pct': 40, 'max_pct': 70},
        {'dimension': 'Region',          'category': 'Europe',                     'min_pct': 0,  'target_pct': 35, 'max_pct': 60},
        {'dimension': 'Region',          'category': 'Asia & ROW',                 'min_pct': 0,  'target_pct': 10, 'max_pct': 30},
        {'dimension': 'Sector',          'category': 'Technology',                 'min_pct': 0,  'target_pct': 20, 'max_pct': 35},
        {'dimension': 'Sector',          'category': 'Healthcare',                 'min_pct': 0,  'target_pct': 15, 'max_pct': 30},
        {'dimension': 'Stage',           'category': 'Buyout',                     'min_pct': 0,  'target_pct': 50, 'max_pct': 75},
        {'dimension': 'Stage',           'category': 'Growth',                     'min_pct': 0,  'target_pct': 20, 'max_pct': 40},
        {'dimension': 'Stage',           'category': 'Venture',                    'min_pct': 0,  'target_pct': 10, 'max_pct': 25},
        {'dimension': 'Vintage',         'category': '2025',                       'min_pct': 0,  'target_pct': 20, 'max_pct': 35},
        {'dimension': 'Vintage',         'category': '2026',                       'min_pct': 0,  'target_pct': 20, 'max_pct': 35},
    ]

    # Default legal restrictions rows
    DEFAULT_RESTRICTIONS = [
        {'restriction': 'Single Asset Exposure',          'legal_min': None, 'legal_max': 10,  'internal_min': None, 'internal_max': 8,  'target': 5},
        {'restriction': 'Single Manager Exposure',        'legal_min': None, 'legal_max': 20,  'internal_min': None, 'internal_max': 15, 'target': 10},
        {'restriction': 'Listed Securities',              'legal_min': None, 'legal_max': 10,  'internal_min': None, 'internal_max': 5,  'target': 0},
        {'restriction': 'North America + Europe Combined','legal_min': None, 'legal_max': 100, 'internal_min': None, 'internal_max': 90, 'target': 75},
        {'restriction': 'Leverage (Portfolio Level)',     'legal_min': None, 'legal_max': 30,  'internal_min': None, 'internal_max': 20, 'target': 10},
        {'restriction': 'Leverage (Deal Level)',          'legal_min': None, 'legal_max': 50,  'internal_min': None, 'internal_max': 40, 'target': 30},
    ]

    return html.Div([
        dbc.Row([
            dbc.Col(html.H2("🧭 Fund Overview", style={'fontWeight': 'bold'}), width=8),
            dbc.Col(dbc.Button([html.I(className="fas fa-save me-2"), "Save All Overview Inputs"],
                               id="btn-save-overview", color="primary",
                               className="float-end", size="lg"), width=4)
        ], className="mb-3"),

        dbc.Alert([
            html.I(className="fas fa-info-circle me-2"),
            "Set investment targets, legal restrictions and exposure guardrails. "
            "Current levels are calculated from the live portfolio; Pro Forma adds pipeline deals."
        ], color="info", className="mb-4"),

        # ── KPI cards ────────────────────────────────────────────────────────
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Current NAV", className="text-muted"),
                html.H3(id='overview-current-nav', style={'color': C['green'], 'fontFamily': C['mono']}),
                html.Small("Current portfolio (deals + liquidity)", className="text-muted")
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Dry Powder", className="text-muted"),
                html.H3(id='overview-dry-powder', style={'color': C['blue'], 'fontFamily': C['mono']}),
                html.Small("Available for deployment", className="text-muted")
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Target Fund Size ($m)", className="text-muted"),
                dbc.Input(id='overview-target-fund-size', type='number', step=10, value=1000,
                          style={'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Deployment Headroom", className="text-muted"),
                html.H3(id='overview-headroom', style={'color': C['amber'], 'fontFamily': C['mono']}),
                html.Small("Target size less current NAV", className="text-muted")
            ])], className="shadow-sm"), width=3),
        ], className="mb-4"),

        # ── SECTION 1: Targets & Investment Restrictions ─────────────────────
        html.H4([html.I(className="fas fa-bullseye me-2", style={'color': C['blue']}),
                 "Investment Targets & Restrictions"],
                className="mb-3 mt-2",
                style={'fontWeight': 'bold', 'fontFamily': C['sans'], 'color': C['text']}),

        dbc.Tabs([
            dbc.Tab(label="Investment Type", tab_id="ov-tab-invtype"),
            dbc.Tab(label="Region",          tab_id="ov-tab-region"),
            dbc.Tab(label="Sector",          tab_id="ov-tab-sector"),
            dbc.Tab(label="Stage",           tab_id="ov-tab-stage"),
            dbc.Tab(label="Vintage",         tab_id="ov-tab-vintage"),
        ], id="ov-targets-tabs", active_tab="ov-tab-invtype", className="mb-0"),

        dbc.Card([
            dbc.CardBody([
                html.Div(id='ov-targets-tab-content'),
                dbc.Row([
                    dbc.Col(dbc.Button([html.I(className="fas fa-plus me-1"), "Add Row"],
                                       id='btn-add-target-row', color='secondary',
                                       size='sm', className='mt-3'), width="auto"),
                ]),
            ])
        ], className="shadow-sm mb-4",
           style={'borderTop': 'none', 'borderRadius': '0 0 0.375rem 0.375rem'}),

        # ── SECTION 2: Legal Restrictions ─────────────────────────────────
        html.H4([html.I(className="fas fa-gavel me-2", style={'color': C['red']}),
                 "Legal & Internal Investment Restrictions"],
                className="mb-1 mt-2",
                style={'fontWeight': 'bold', 'fontFamily': C['sans'], 'color': C['text']}),
        html.P("All values are % of portfolio NAV unless noted. "
               "Current and Pro Forma columns are calculated automatically.",
               className="text-muted mb-3", style={'fontSize': '13px'}),

        dbc.Card([
            dbc.CardHeader([
                dbc.Row([
                    dbc.Col("⚖️ Restriction Limits", width=8,
                            style={'fontWeight': 'bold', 'paddingTop': '6px'}),
                    dbc.Col(dbc.Button([html.I(className="fas fa-plus me-1"), "Add Restriction"],
                                       id='btn-add-restriction-row', color='secondary',
                                       size='sm', className='float-end'), width=4,
                            className="text-end")
                ])
            ], style={'backgroundColor': C['surface']}),
            dbc.CardBody([
                html.Small([
                    html.I(className="fas fa-pencil-alt me-1"),
                    "Edit Legal Min/Max, Internal Min/Max and Target columns directly. "
                    "Current and Pro Forma are read-only (auto-calculated)."
                ], className="text-muted d-block mb-2"),
                html.Div(id='restrictions-table-container'),
            ])
        ], className="shadow-sm mb-4"),

        # ── SECTION 3: Exposure vs Targets (live read-only) ───────────────
        html.H4([html.I(className="fas fa-chart-bar me-2", style={'color': C['purple']}),
                 "Live Exposure Monitor"],
                className="mb-3 mt-2",
                style={'fontWeight': 'bold', 'fontFamily': C['sans'], 'color': C['text']}),

        dbc.Tabs([
            dbc.Tab(label="📋 Current Portfolio", tab_id="ov-exp-current"),
            dbc.Tab(label="📋 + Pipeline (Pro Forma)", tab_id="ov-exp-proforma"),
        ], id="ov-exposure-view-tabs", active_tab="ov-exp-current", className="mb-0"),

        dbc.Card([
            dbc.CardBody([html.Div(id='ov-exposure-monitor')])
        ], className="shadow-sm mb-4",
           style={'borderTop': 'none', 'borderRadius': '0 0 0.375rem 0.375rem'}),

        # ── SECTION 4: Bite sizes + charts ────────────────────────────────
        dbc.Row([
            dbc.Col(dbc.Card([
                dbc.CardHeader("Exposure Mix (Current)",
                               style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                dbc.CardBody([dcc.Graph(id='overview-exposure-chart',
                                        config={'displayModeBar': False})])
            ], className="shadow-sm"), width=5),
            dbc.Col(dbc.Card([
                dbc.CardHeader("Target Bite Sizes by Asset Type",
                               style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                dbc.CardBody([
                    dash_table.DataTable(
                        id='overview-bite-table',
                        columns=[
                            {'name': 'Asset Type', 'id': 'asset_type',
                             'presentation': 'dropdown'},
                            {'name': 'Min % DP', 'id': 'min_pct', 'type': 'numeric',
                             'format': {'specifier': '.3f'}},
                            {'name': 'Desired % DP', 'id': 'desired_pct', 'type': 'numeric',
                             'format': {'specifier': '.3f'}},
                            {'name': 'Max % DP', 'id': 'max_pct', 'type': 'numeric',
                             'format': {'specifier': '.3f'}},
                        ],
                        data=[],
                        editable=True,
                        row_deletable=True,
                        dropdown={'asset_type': {'options': [
                            {'label': x, 'value': x} for x in STRAT_OPTS
                        ]}},
                        **_tbl_style(),
                    ),
                    dbc.Button("Add Bite Size Row", id='btn-add-bite-row',
                               color='secondary', size='sm', className='mt-3')
                ])
            ], className="shadow-sm"), width=7),
        ], className="mb-4"),

        # ── Legacy hidden table (kept for backward compat with save callback) ─
        html.Div([
            dash_table.DataTable(
                id='overview-exposure-table',
                columns=[
                    {'name': 'Dimension',  'id': 'dimension',  'presentation': 'dropdown'},
                    {'name': 'Category',   'id': 'category'},
                    {'name': 'Min %',      'id': 'min_pct',    'type': 'numeric'},
                    {'name': 'Target %',   'id': 'target_pct', 'type': 'numeric'},
                    {'name': 'Max %',      'id': 'max_pct',    'type': 'numeric'},
                ],
                data=DEFAULT_TARGETS,
                editable=True,
                row_deletable=True,
                dropdown={'dimension': {'options': [
                    {'label': x, 'value': x} for x in DIM_OPTS
                ]}},
                **_tbl_style(),
            )
        ], style={'display': 'none'}),

        dbc.Card([
            dbc.CardHeader("Pipeline and Future Deal Bite Sizing",
                           style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
            dbc.CardBody([
                html.Div(id='overview-forward-bite-table'),
                html.Div(id='overview-save-status', className='mt-3')
            ])
        ], className="shadow-sm"),

        # Store defaults inline so callbacks can access them before first save
        dcc.Store(id='ov-targets-store',
                  data=DEFAULT_TARGETS),
        dcc.Store(id='ov-restrictions-store',
                  data=DEFAULT_RESTRICTIONS),
    ])


def portfolio_page():
    return html.Div([
        dbc.Row([
            dbc.Col(html.H2("💼 Current Portfolio", style={'fontWeight': 'bold'}), width=8),
            dbc.Col(dbc.Button("➕ Add Deal", id="btn-open-deal", color="primary", className="float-end", size="lg"),
                    width=4)
        ], className="mb-3"),

        # Fund Status Cards
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("💰 Current NAV", className="text-muted"),
                html.H4(id="port-nav", style={'color': COLORS['success']}),
                html.Small("Latest month-end NAV", className="text-muted")
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("📥 Current Commitment", className="text-muted"),
                html.H4(id="port-current-commitment", style={'color': COLORS['info']}),
                html.Small("Capital called to date", className="text-muted")
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("📊 Total Commitment", className="text-muted"),
                html.H4(id="port-total-commitment", style={'color': COLORS['primary']}),
                html.Small("Across current portfolio", className="text-muted")
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("💵 Dry Powder", className="text-muted"),
                html.H4(id="port-dry-powder", style={'color': COLORS['primary']}),
                html.Small("From liquidity upload when available", className="text-muted")
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("📦 Total Fund Size", className="text-muted"),
                html.H4(id="port-total-fund", style={'color': COLORS['dark']}),
                html.Small("NAV + Dry Powder", className="text-muted")
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("🎯 Target Return", className="text-muted"),
                html.H4(id="port-target-return", style={'color': COLORS['warning']}),
                html.Small("Required deal IRR", className="text-muted")
            ])], className="shadow-sm"), width=2),
        ], className="mb-4"),

        # Add Deal Modal
        dbc.Modal([
            dbc.ModalHeader("Add Deal to Current Portfolio"),
            dbc.ModalBody([
                dbc.Alert([
                    html.I(className="fas fa-info-circle me-2"),
                    html.Strong("Current Portfolio: "),
                    "Enter ACTUAL NAV of deals already in the portfolio. These do NOT reduce dry powder."
                ], color="info", className="mb-3", style={'fontSize': '13px'}),

                # Row 1: Deal Name, Fund Manager
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Deal Name *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-name", type="text", placeholder="e.g., Coller VII")
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Fund Manager *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-manager", type="text", placeholder="e.g., Coller Capital")
                    ], width=6)
                ], className="mb-3"),

                # Row 2: Strategy, Stage
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Strategy Type *", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-strategy", options=[
                            {"label": "GP-Led (Multi-Asset)", "value": "GP-Led (Multi-Asset)"},
                            {"label": "GP-Led (Single-Asset)", "value": "GP-Led (Single-Asset)"},
                            {"label": "Diversified LP-Led", "value": "Diversified LP-Led"},
                            {"label": "Co-Investments", "value": "Co-Investments"},
                        ])
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Stage *", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-stage", options=[
                            {"label": "Buyout", "value": "Buyout"},
                            {"label": "Venture", "value": "Venture"},
                            {"label": "Growth", "value": "Growth"},
                            {"label": "Liquidity", "value": "Liquidity"},
                        ], value="Buyout")
                    ], width=6)
                ], className="mb-3"),

                # Row 3: Total Commitment, Current Commitment
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Total Commitment ($mm) *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-total-commitment", type="number", step=0.1, placeholder="70.0"),
                        html.Small("Total committed capital", className="text-muted")
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Current Commitment ($mm) *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-current-commitment", type="number", step=0.1, placeholder="52.0"),
                        html.Small("Capital called to date", className="text-muted")
                    ], width=6),
                ], className="mb-3"),

                # Row 4: Current NAV (MAIN DRIVER), Unfunded (auto-calculated)
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Current NAV ($mm) *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-nav", type="number", step=0.1, placeholder="48.0"),
                        html.Small("Current Net Asset Value (MAIN DRIVER of calcs)", className="text-muted")
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Unfunded Commitment ($mm)", style={'fontWeight': 'bold'}),
                        html.Div(id="unfunded-display", style={
                            'marginTop': '8px',
                            'padding': '8px',
                            'backgroundColor': C['surface'],
                            'borderRadius': '4px',
                            'fontFamily': C['mono'],
                            'fontSize': '16px',
                            'fontWeight': 'bold',
                            'color': C['amber']
                        }),
                        html.Small("Auto: Total - Current Commitment", className="text-muted")
                    ], width=6),
                ], className="mb-3"),

                # Row 5: IRR, Hold Period, Currency
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Target Gross IRR (%) *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-irr", type="number", step=0.5, placeholder="22.0")
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Hold Period (years)", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-hold", type="number", value=5, step=0.5)
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Currency", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-currency", options=[
                            {"label": "USD", "value": "USD"},
                            {"label": "EUR", "value": "EUR"},
                            {"label": "GBP", "value": "GBP"},
                        ], value="USD")
                    ], width=4),
                ], className="mb-3"),

                # Row 5: Vintage, Segment, Allocation Status
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Vintage Year", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-vintage", options=[
                            {"label": str(y), "value": y} for y in range(2026, 2014, -1)
                        ], value=2024)
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Portfolio Segment *", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-segment", options=[
                            {"label": "Seed Portfolio", "value": "Seed"},
                            {"label": "New Deals", "value": "New"},
                            {"label": "Money Market", "value": "MoneyMarket"},
                        ], value="Seed"),
                        html.Small("For TWR tracking", className="text-muted")
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Allocation Status *", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-allocation-status", options=[
                            {"label": "Closed", "value": "Closed"},
                            {"label": "Pending Close", "value": "Pending Close"},
                            {"label": "Pending Allocation", "value": "Pending Allocation"},
                        ], value="Closed")
                    ], width=4),
                ], className="mb-3"),

                # Row 6: Sector, Geography
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Sector", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-sector", options=[
                            {"label": "Diversified", "value": "Diversified"},
                            {"label": "Technology", "value": "Technology"},
                            {"label": "Healthcare", "value": "Healthcare"},
                            {"label": "Consumer", "value": "Consumer"},
                            {"label": "Industrials", "value": "Industrials"},
                            {"label": "Financials", "value": "Financials"},
                            {"label": "Other", "value": "Other"}
                        ], value="Diversified")
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Geography", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-geo", options=[
                            {"label": "North America", "value": "North America"},
                            {"label": "Europe", "value": "Europe"},
                            {"label": "Asia", "value": "Asia"},
                            {"label": "Global", "value": "Global"}
                        ], value="North America")
                    ], width=6),
                ])
            ]),
            dbc.ModalFooter([
                dbc.Button("Cancel", id="btn-cancel", color="secondary"),
                dbc.Button("Add Deal", id="btn-submit", color="primary")
            ])
        ], id="modal-deal", size="lg", is_open=False),

        # Edit Deal Modal
        dbc.Modal([
            dbc.ModalHeader("Edit Deal"),
            dbc.ModalBody([
                dbc.Alert([
                    html.I(className="fas fa-edit me-2"),
                    html.Strong("Editing Deal: "),
                    "Update deal details. NAV will auto-pull from latest cashflows if available."
                ], color="warning", className="mb-3", style={'fontSize': '13px'}),

                dcc.Store(id='edit-deal-index', data=None),

                # Row 1: Deal Name, Fund Manager
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Deal Name *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="edit-name", type="text")
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Fund Manager *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="edit-manager", type="text")
                    ], width=6)
                ], className="mb-3"),

                # Row 2: Strategy, Stage
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Strategy Type *", style={'fontWeight': 'bold'}),
                        dbc.Select(id="edit-strategy", options=[
                            {"label": "GP-Led (Multi-Asset)", "value": "GP-Led (Multi-Asset)"},
                            {"label": "GP-Led (Single-Asset)", "value": "GP-Led (Single-Asset)"},
                            {"label": "Diversified LP-Led", "value": "Diversified LP-Led"},
                            {"label": "Co-Investments", "value": "Co-Investments"},
                        ])
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Stage *", style={'fontWeight': 'bold'}),
                        dbc.Select(id="edit-stage", options=[
                            {"label": "Buyout", "value": "Buyout"},
                            {"label": "Venture", "value": "Venture"},
                            {"label": "Growth", "value": "Growth"},
                            {"label": "Liquidity", "value": "Liquidity"},
                        ])
                    ], width=6)
                ], className="mb-3"),

                # Row 3: Total Commitment, Current Commitment
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Total Commitment ($mm) *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="edit-total-commitment", type="number", step=0.1),
                        html.Small("Total committed capital", className="text-muted")
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Current Commitment ($mm) *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="edit-current-commitment", type="number", step=0.1),
                        html.Small("Capital called to date", className="text-muted")
                    ], width=6),
                ], className="mb-3"),

                # Row 4: Current NAV (auto-pulls from cashflows), Unfunded
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Current NAV ($mm) *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="edit-nav", type="number", step=0.1, disabled=True),
                        html.Small("Auto-pulled from latest cashflows (MAIN DRIVER)", className="text-success"),
                        html.Br(),
                        html.Small(id="edit-nav-source", className="text-muted")
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Unfunded Commitment ($mm)", style={'fontWeight': 'bold'}),
                        html.Div(id="edit-unfunded-display", style={
                            'marginTop': '8px',
                            'padding': '8px',
                            'backgroundColor': C['surface'],
                            'borderRadius': '4px',
                            'fontFamily': C['mono'],
                            'fontSize': '16px',
                            'fontWeight': 'bold',
                            'color': C['amber']
                        }),
                        html.Small("Auto: Total - Current Commitment", className="text-muted")
                    ], width=6),
                ], className="mb-3"),

                # Row 5: IRR, Hold Period
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Target Gross IRR (%) *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="edit-irr", type="number", step=0.5)
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Hold Period (years)", style={'fontWeight': 'bold'}),
                        dbc.Input(id="edit-hold", type="number", step=0.5)
                    ], width=6),
                ])
            ]),
            dbc.ModalFooter([
                dbc.Button("Cancel", id="btn-cancel-edit", color="secondary"),
                dbc.Button("Save Changes", id="btn-save-edit", color="warning")
            ])
        ], id="modal-edit-deal", size="lg", is_open=False),

        # Edit Pipeline Modal
        dbc.Modal([
            dbc.ModalHeader("Edit Pipeline Deal"),
            dbc.ModalBody([
                dcc.Store(id='edit-pipeline-index', data=None),

                dbc.Row([
                    dbc.Col([
                        dbc.Label("Deal Name *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="edit-pipeline-name", type="text")
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Type *", style={'fontWeight': 'bold'}),
                        dbc.Select(id="edit-pipeline-type", options=[
                            {"label": "GP-Led", "value": "GP-Led"},
                            {"label": "LP-Led", "value": "LP-Led"},
                            {"label": "Co-Investment", "value": "Co-Investment"},
                        ])
                    ], width=6)
                ], className="mb-3"),

                dbc.Row([
                    dbc.Col([
                        dbc.Label("Stage *", style={'fontWeight': 'bold'}),
                        dbc.Select(id="edit-pipeline-stage", options=[
                            {"label": "Screening", "value": "Screening"},
                            {"label": "Diligence", "value": "Diligence"},
                            {"label": "IC Review", "value": "IC Review"},
                            {"label": "Negotiation", "value": "Negotiation"},
                        ])
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Size ($mm)", style={'fontWeight': 'bold'}),
                        dbc.Input(id="edit-pipeline-size", type="number", step=0.1)
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Target IRR (%)", style={'fontWeight': 'bold'}),
                        dbc.Input(id="edit-pipeline-irr", type="number", step=0.5)
                    ], width=4),
                ])
            ]),
            dbc.ModalFooter([
                dbc.Button("Cancel", id="btn-cancel-edit-pipeline", color="secondary"),
                dbc.Button("Save Changes", id="btn-save-edit-pipeline", color="warning")
            ])
        ], id="modal-edit-pipeline", size="lg", is_open=False),

        # Edit Placeholder Modal
        dbc.Modal([
            dbc.ModalHeader("Edit Placeholder Deal"),
            dbc.ModalBody([
                dcc.Store(id='edit-placeholder-index', data=None),

                dbc.Row([
                    dbc.Col([
                        dbc.Label("Deal Name *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="edit-placeholder-name", type="text")
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Strategy *", style={'fontWeight': 'bold'}),
                        dbc.Select(id="edit-placeholder-strategy", options=[
                            {"label": "GP-Led (Multi-Asset)", "value": "GP-Led (Multi-Asset)"},
                            {"label": "GP-Led (Single-Asset)", "value": "GP-Led (Single-Asset)"},
                            {"label": "Diversified LP-Led", "value": "Diversified LP-Led"},
                            {"label": "Co-Investments", "value": "Co-Investments"},
                        ])
                    ], width=6)
                ], className="mb-3"),

                dbc.Row([
                    dbc.Col([
                        dbc.Label("Size ($mm) *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="edit-placeholder-size", type="number", step=0.1)
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Expected Month (0-11) *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="edit-placeholder-month", type="number", min=0, max=11, step=1),
                        html.Small("0 = Jan 2026, 11 = Dec 2026", className="text-muted")
                    ], width=6),
                ])
            ]),
            dbc.ModalFooter([
                dbc.Button("Cancel", id="btn-cancel-edit-placeholder", color="secondary"),
                dbc.Button("Save Changes", id="btn-save-edit-placeholder", color="warning")
            ])
        ], id="modal-edit-placeholder", size="lg", is_open=False),

        html.Div(id='portfolio-table')
    ])


def future_deals_page():
    return html.Div([
        dbc.Row([
            dbc.Col(html.H2("📅 Pro Forma Portfolio (Future/Placeholder Deals)", style={'fontWeight': 'bold'}), width=8),
            dbc.Col(dbc.Button("➕ Add Placeholder", id="btn-open-placeholder", color="success", className="float-end",
                               size="lg"), width=4)
        ], className="mb-3"),

        dbc.Alert([
            html.I(className="fas fa-lightbulb me-2"),
            "Plan future deals to forecast dry powder usage and ensure you stay within bite size limits. These are NOT in your current portfolio."
        ], color="info", className="mb-4"),

        # Add Placeholder Modal
        dbc.Modal([
            dbc.ModalHeader("Add Placeholder Deal (Future Commitment)"),
            dbc.ModalBody([
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Deal Name *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-ph-name", type="text", placeholder="e.g., GP-Led (Multi-Asset) 1")
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Strategy Type *", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-ph-strategy", options=[
                            {"label": "GP-Led (Multi-Asset)", "value": "GP-Led (Multi-Asset)"},
                            {"label": "GP-Led (Single-Asset)", "value": "GP-Led (Single-Asset)"},
                            {"label": "Diversified LP-Led", "value": "Diversified LP-Led"},
                            {"label": "Co-Investments", "value": "Co-Investments"},
                        ])
                    ], width=6)
                ], className="mb-3"),
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Expected Size ($mm) *", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-ph-size", type="number", step=0.1, placeholder="15.0"),
                        html.Small(id="ph-bite-warning", className="text-muted")
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Custom Bite Size ($mm)", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-ph-bite", type="number", step=0.1, placeholder="Auto-calculate"),
                        html.Small("Override default bite size limits", className="text-muted")
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Expected Month *", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-ph-month", options=generate_month_options(), value=0,
                                   style={'fontFamily': C['mono']})
                    ], width=4),
                ], className="mb-3"),
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Target Weight (% of Portfolio)", style={'fontWeight': 'bold'}),
                        dbc.Input(id="in-ph-weight", type="number", step=0.5, placeholder="5.0"),
                        html.Small("Desired portfolio allocation %", className="text-muted")
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Region", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-ph-region", options=[
                            {"label": "North America", "value": "North America"},
                            {"label": "Europe", "value": "Europe"},
                            {"label": "Asia", "value": "Asia"},
                            {"label": "Global", "value": "Global"},
                        ], value="North America")
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Deal Type", style={'fontWeight': 'bold'}),
                        dbc.Select(id="in-ph-type", options=[
                            {"label": "Secondary", "value": "Secondary"},
                            {"label": "Co-Investment", "value": "Co-Investment"},
                        ], value="Secondary")
                    ], width=4),
                ])
            ]),
            dbc.ModalFooter([
                dbc.Button("Cancel", id="btn-cancel-ph", color="secondary"),
                dbc.Button("Add Placeholder", id="btn-submit-ph", color="success")
            ])
        ], id="modal-placeholder", size="lg", is_open=False),

        html.Div(id='placeholder-table')
    ])


def drypowder_page():
    return html.Div([
        html.H2("💧 Dry Powder Forecaster", className="mb-4", style={'fontWeight': 'bold'}),

        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Current Dry Powder", className="text-muted"),
                html.H3(id="dp-current", style={'color': COLORS['primary']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Planned Deployments", className="text-muted"),
                html.H3(id="dp-planned", style={'color': COLORS['warning']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Expected Distributions", className="text-muted"),
                html.H3(id="dp-distributions", style={'color': COLORS['success']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("12-Month Forecast", className="text-muted"),
                html.H3(id="dp-forecast-12m", style={'color': COLORS['info']})
            ])], className="shadow-sm"), width=3),
        ], className="mb-4"),

        dbc.Card([
            dbc.CardHeader("12-Month Dry Powder Forecast",
                           style={'fontWeight': 'bold', 'backgroundColor': COLORS['light']}),
            dbc.CardBody([
                dcc.Graph(id='dp-forecast-chart', config={'displayModeBar': True})
            ])
        ], className="shadow-sm mb-4"),

        dbc.Card([
            dbc.CardHeader("Monthly Breakdown", style={'fontWeight': 'bold', 'backgroundColor': COLORS['light']}),
            dbc.CardBody([
                html.Div(id='dp-monthly-table')
            ])
        ], className="shadow-sm")
    ])


def calculator_page():
    return html.Div([
        html.H2("🧮 Return Calculator & Waterfall", className="mb-4", style={'fontWeight': 'bold'}),

        dbc.Card([dbc.CardBody([
            html.H5("Calculate required deal IRR to achieve target net TWR", className="text-muted mb-4"),

            html.Hr(),
            html.H4("📊 Current Portfolio", className="mt-4", style={'color': COLORS['dark'], 'fontWeight': 'bold'}),
            dbc.Row([
                dbc.Col([
                    html.H6("Current NAV:", className="text-muted"),
                    html.H3(id="calc-current-nav", style={'color': COLORS['success']})
                ], width=4),
                dbc.Col([
                    html.H6("Weighted IRR:", className="text-muted"),
                    html.H3(id="calc-current-irr", style={'color': COLORS['success']})
                ], width=4),
                dbc.Col([
                    html.H6("Number of Deals:", className="text-muted"),
                    html.H3(id="calc-num-deals", style={'color': COLORS['success']})
                ], width=4),
            ], className="mb-4"),

            html.Hr(),
            html.H4("🎯 Evergreen Fund Structure", className="mt-4",
                    style={'color': COLORS['dark'], 'fontWeight': 'bold'}),
            dbc.Row([
                dbc.Col([
                    html.H6("Total Fund Size:", className="text-muted"),
                    html.H3(id="calc-total-fund", style={'color': COLORS['primary']}),
                    html.Small("NAV + Dry Powder", className="text-muted")
                ], width=4),
                dbc.Col([
                    html.H6("Target Net TWR:", className="text-muted"),
                    html.H3(id="calc-target-twr", style={'color': COLORS['warning']})
                ], width=4),
                dbc.Col([
                    html.H6("Dry Powder:", className="text-muted"),
                    html.H3(id="calc-dry-powder", style={'color': COLORS['info']}),
                    html.Small("For Pipeline/Future Only", className="text-muted")
                ], width=4),
            ], className="mb-4"),

            html.Hr(),
            html.Div([
                html.H3("✅ REQUIRED FUTURE DEAL IRR", className="text-center mb-3",
                        style={'color': COLORS['success'], 'fontWeight': 'bold'}),
                html.H1(id="calc-required-irr", className="text-center mb-2",
                        style={'fontSize': '80px', 'fontWeight': 'bold', 'color': COLORS['success']}),
                html.P(id="calc-explanation", className="text-center text-muted mb-4", style={'fontSize': '18px'}),
            ], style={'backgroundColor': '#F0F8F0', 'padding': '2rem', 'borderRadius': '10px',
                      'border': f'3px solid {COLORS["success"]}'}),

            html.Hr(className="mt-4"),
            html.H5("💡 Return Waterfall Breakdown", className="mb-3",
                    style={'color': COLORS['dark'], 'fontWeight': 'bold'}),
            html.Div(id="calc-waterfall")
        ])], className="shadow-sm")
    ])


def twr_forecaster_page():
    return html.Div([
        html.H2("📈 TWR Returns Forecaster", className="mb-4",
                style={'fontWeight': 'bold', 'fontFamily': C['sans'], 'color': C['text']}),

        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Monte Carlo Simulation Parameters",
                                   style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([
                        dbc.Row([
                            dbc.Col([
                                dbc.Label("Current Portfolio IRR",
                                          style={'fontWeight': 'bold', 'fontFamily': C['sans']}),
                                html.H4(id="twr-current-irr", style={'color': C['green'], 'fontFamily': C['mono']})
                            ], width=4),
                            dbc.Col([
                                dbc.Label("Future Deals Mean IRR (%)", style={'fontWeight': 'bold'}),
                                dbc.Input(id="twr-future-mean", type="number", value=25, step=0.5,
                                          style={'fontFamily': C['mono']})
                            ], width=4),
                            dbc.Col([
                                dbc.Label("Future Deals Std Dev (%)", style={'fontWeight': 'bold'}),
                                dbc.Input(id="twr-future-std", type="number", value=5, step=0.5,
                                          style={'fontFamily': C['mono']})
                            ], width=4),
                        ], className="mb-3"),
                        dbc.Row([
                            dbc.Col([
                                dbc.Label("# Simulations", style={'fontWeight': 'bold'}),
                                dbc.Select(id="twr-n-sims", options=[
                                    {"label": "1,000", "value": 1000},
                                    {"label": "5,000", "value": 5000},
                                    {"label": "10,000", "value": 10000},
                                ], value=5000, style={'fontFamily': C['mono']})
                            ], width=6),
                            dbc.Col([
                                dbc.Button("🎲 Run Simulation", id="btn-run-twr", color="primary", size="lg",
                                           className="w-100")
                            ], width=6)
                        ])
                    ])
                ], className="shadow-sm mb-4")
            ], width=12),
        ]),

        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Probability Distribution",
                                   style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='twr-distribution-chart', config={'displayModeBar': True})])
                ], className="shadow-sm")
            ], width=8),
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Key Statistics", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([html.Div(id='twr-statistics')])
                ], className="shadow-sm")
            ], width=4),
        ], className="mb-4"),

        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Sensitivity Analysis",
                                   style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='twr-sensitivity-chart', config={'displayModeBar': True})])
                ], className="shadow-sm")
            ], width=12),
        ])
    ])


def cashflows_page():
    # Pre-build month options from Sep-2024 → Dec-2040 for the start-month picker
    from dateutil.relativedelta import relativedelta as _rd
    _start = datetime(2024, 9, 30)
    _end   = datetime(2040, 12, 31)
    _month_opts = []
    _cur = _start
    while _cur <= _end:
        _month_opts.append({
            "label": _cur.strftime("%b %Y"),
            "value": _cur.strftime("%Y-%m-%d")
        })
        _cur = (_cur.replace(day=1) + _rd(months=1))
        _cur = _cur.replace(day=pd.Timestamp(_cur).days_in_month)

    # Forecast horizon options  (label, months)
    HORIZONS = [
        ("12 Months",  12),
        ("18 Months",  18),
        ("24 Months",  24),
        ("36 Months",  36),
        ("5 Years",    60),
    ]

    return html.Div([
        html.H2("💰 Fund Level Cashflows", className="mb-4",
                style={'fontWeight': 'bold', 'fontFamily': C['sans'], 'color': C['text']}),

        dbc.Alert([
            html.I(className="fas fa-info-circle me-2"),
            "Upload your Excel Fund Level CF file. Select a start month, then choose a forecast horizon — "
            "the table and chart show Net CF, Calls, Distributions and NAV for that window."
        ], color="info", className="mb-4"),

        # ── Upload ───────────────────────────────────────────────────────────
        dbc.Card([
            dbc.CardHeader("📥 Upload Fund Level CF Excel File",
                           style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
            dbc.CardBody([
                dbc.Row([
                    dbc.Col([
                        dcc.Upload(
                            id='upload-fund-cf',
                            children=dbc.Button([
                                html.I(className="fas fa-file-excel me-2"),
                                'Upload Fund Level CF Excel File'
                            ], color="primary", size="lg", className="w-100"),
                            multiple=False
                        ),
                        html.Small("Upload the 'Fund Level CF' tab from your Excel file (.xlsx, .xlsm)",
                                   className="text-muted d-block mt-2")
                    ], width=6),
                    dbc.Col([
                        html.Div(id='upload-fund-cf-status', className="mt-2")
                    ], width=6)
                ])
            ])
        ], className="shadow-sm mb-4"),

        # ── Summary KPI cards ────────────────────────────────────────────────
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Total Commitment", className="text-muted"),
                html.H4(id="cf-total-commitment", style={'color': C['blue'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Total Paid In", className="text-muted"),
                html.H4(id="cf-total-paid-in", style={'color': C['green'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Total Unfunded", className="text-muted"),
                html.H4(id="cf-total-unfunded", style={'color': C['amber'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Current Total NAV", className="text-muted"),
                html.H4(id="cf-current-nav", style={'color': C['purple'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
        ], className="mb-4"),

        # ── Master controls card ─────────────────────────────────────────────
        dbc.Card([
            dbc.CardHeader("🔭 View Controls",
                           style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
            dbc.CardBody([
                dbc.Row([
                    # Start month
                    dbc.Col([
                        dbc.Label([
                            html.I(className="fas fa-calendar-alt me-1"),
                            html.Strong("Start Month"),
                            html.Small("  (scroll to any month from Sep 2024 → Dec 2040)",
                                       className="text-muted ms-1")
                        ], style={'fontWeight': 'bold'}),
                        dbc.Select(
                            id="cf-start-month",
                            options=_month_opts,
                            value=datetime.now().replace(day=1).strftime("%Y-%m-%d"),
                            style={'fontFamily': C['mono']}
                        ),
                    ], width=4),

                    # Forecast horizon tabs
                    dbc.Col([
                        dbc.Label([
                            html.I(className="fas fa-forward me-1"),
                            html.Strong("Forecast Horizon"),
                        ], style={'fontWeight': 'bold'}),
                        html.Div(
                            dbc.RadioItems(
                                id="cf-horizon",
                                options=[{"label": lbl, "value": mo}
                                         for lbl, mo in HORIZONS],
                                value=12,
                                inline=True,
                                input_class_name="btn-check",
                                label_class_name="btn btn-outline-primary btn-sm me-1",
                                label_checked_class_name="active",
                            ),
                            className="mt-1"
                        ),
                    ], width=5),

                    # Section toggles + export
                    dbc.Col([
                        dbc.Label([
                            html.I(className="fas fa-filter me-1"),
                            html.Strong("Sections"),
                        ], style={'fontWeight': 'bold'}),
                        dbc.Checklist(
                            id="cf-sections",
                            options=[
                                {"label": "Net CF",       "value": "net"},
                                {"label": "Calls",        "value": "calls"},
                                {"label": "Distributions","value": "dists"},
                                {"label": "NAV",          "value": "nav"},
                            ],
                            value=["net", "calls", "dists", "nav"],
                            inline=True,
                            switch=True,
                        ),
                        dbc.Button("📥 Export CSV", id="btn-export-cf", color="success",
                                   size="sm", className="mt-2 d-block"),
                    ], width=3),
                ]),

                # Selected window summary badge
                html.Div(id='cf-window-badge', className="mt-3"),
            ])
        ], className="shadow-sm mb-4"),

        # ── Main cashflow table ──────────────────────────────────────────────
        dbc.Card([
            dbc.CardHeader(id='cf-table-header',
                           style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
            dbc.CardBody([
                html.Div(id='cf-monthly-table', style={'overflowX': 'auto'})
            ])
        ], className="shadow-sm mb-4"),

        # ── Forecast chart ───────────────────────────────────────────────────
        dbc.Card([
            dbc.CardHeader(id='cf-chart-header',
                           style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
            dbc.CardBody([
                dcc.Graph(id='cf-monthly-chart', config={'displayModeBar': True})
            ])
        ], className="shadow-sm mb-4"),

        # ── Discount & Price-to-NAV Tracker ──────────────────────────────────
        html.Hr(className="my-4"),
        html.H4([
            html.I(className="fas fa-tags me-2", style={'color': C['amber']}),
            "Discount & Price-to-NAV Tracker"
        ], className="mb-1", style={'fontWeight': 'bold', 'fontFamily': C['sans'], 'color': C['text']}),
        html.P(
            "Enter the discount % negotiated per deal. The tool computes implied purchase price and "
            "Price-to-NAV ratio, and rolls up a portfolio-level weighted discount forecast.",
            className="text-muted mb-3", style={'fontSize': '13px'}
        ),

        # Portfolio summary cards
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Portfolio NAV (basis)", className="text-muted mb-1", style={'fontSize':'12px'}),
                html.H4(id="disc-total-nav", style={'color': C['blue'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Wtd. Avg Discount", className="text-muted mb-1", style={'fontSize':'12px'}),
                html.H4(id="disc-wtd-avg", style={'color': C['amber'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Total Purchase Price", className="text-muted mb-1", style={'fontSize':'12px'}),
                html.H4(id="disc-total-price", style={'color': C['green'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Portfolio P/NAV", className="text-muted mb-1", style={'fontSize':'12px'}),
                html.H4(id="disc-portfolio-ptbnav", style={'color': C['purple'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
        ], className="mb-3"),

        # Editable per-deal discount table
        dbc.Card([
            dbc.CardHeader([
                dbc.Row([
                    dbc.Col("📊 Per-Deal Discount Entry", width=8,
                            style={'fontWeight': 'bold', 'paddingTop': '6px'}),
                    dbc.Col([
                        dbc.Button([html.I(className="fas fa-save me-1"), "Save Discounts"],
                                   id="btn-save-discounts", color="warning", size="sm", className="me-2"),
                        dbc.Button([html.I(className="fas fa-sync me-1"), "Refresh from NAV"],
                                   id="btn-refresh-discounts", color="secondary", size="sm"),
                    ], width=4, className="text-end")
                ])
            ], style={'backgroundColor': C['surface']}),
            dbc.CardBody([
                dbc.Alert(id="discount-save-status", is_open=False, dismissable=True,
                          duration=3000, className="mb-2"),
                html.Div([
                    html.Small([
                        html.I(className="fas fa-info-circle me-1"),
                        "Edit the ",
                        html.Strong("Discount (%)"),
                        " column directly. All other columns auto-calculate. "
                        "NAV sourced from uploaded cashflows (or manual entry if not uploaded)."
                    ], className="text-muted d-block mb-2"),
                ]),
                html.Div(id='discount-table-container')
            ])
        ], className="shadow-sm mb-4"),

        # Discount visualisation
        dbc.Card([
            dbc.CardHeader("Discount Distribution by Deal",
                           style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
            dbc.CardBody([
                dcc.Graph(id='discount-chart', config={'displayModeBar': True})
            ])
        ], className="shadow-sm"),

    ])


def proforma_page():
    return html.Div([
        html.H2("🔮 Pro Forma Analyzer", className="mb-4",
                style={'fontWeight': 'bold', 'fontFamily': C['sans'], 'color': C['text']}),

        # Month selector for NAV calculation
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Pro Forma Settings", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([
                        dbc.Row([
                            dbc.Col([
                                dbc.Label("Target Month for NAV Calculation *", style={'fontWeight': 'bold'}),
                                dbc.Select(id="pf-target-month", options=generate_month_options(), value=12,
                                           style={'fontFamily': C['mono']}),
                                html.Small("Calculate pro forma NAV at this month using actual cashflows",
                                           className="text-muted")
                            ], width=6),
                            dbc.Col([
                                dbc.Button("📊 Calculate Pro Forma NAV", id="btn-calc-pf-nav",
                                           color="primary", size="lg", className="w-100 mt-4")
                            ], width=6)
                        ])
                    ])
                ], className="shadow-sm mb-4")
            ], width=12)
        ]),

        # Tabs
        dbc.Tabs([
            dbc.Tab(label="📋 Complete Portfolio", tab_id="tab-pf-portfolio"),
            dbc.Tab(label="📊 Metrics Comparison", tab_id="tab-pf-metrics"),
            dbc.Tab(label="📈 Impact Charts", tab_id="tab-pf-charts"),
        ], id="pf-tabs", active_tab="tab-pf-portfolio"),

        html.Div(id='pf-tab-content', className="mt-4")
    ])


def liquidity_assumptions_page():
    return html.Div([
        html.H2("💧 Liquidity Pull", className="mb-4",
                style={'fontWeight': 'bold', 'fontFamily': C['sans'], 'color': C['text']}),

        dbc.Alert([
            html.I(className="fas fa-database me-2"),
            html.Strong("Data Pull From Liquidity Model - "),
            "Upload your Excel Liquidity Pull file or view current liquidity position and projections"
        ], color="info", className="mb-4"),

        # File Upload Section
        dbc.Card([
            dbc.CardHeader("📥 Upload Liquidity Pull Excel File",
                           style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
            dbc.CardBody([
                dbc.Row([
                    dbc.Col([
                        dcc.Upload(
                            id='upload-liquidity-pull',
                            children=dbc.Button([
                                html.I(className="fas fa-file-excel me-2"),
                                'Upload Liquidity Pull Excel File'
                            ], color="info", size="lg", className="w-100"),
                            multiple=False
                        ),
                        html.Small("Upload the 'Liquidity Pull' tab from your Excel file (.xlsx, .xlsm)",
                                   className="text-muted d-block mt-2")
                    ], width=6),
                    dbc.Col([
                        html.Div(id='upload-liquidity-status', className="mt-2")
                    ], width=6)
                ])
            ])
        ], className="shadow-sm mb-4"),

        # Current Date Info
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Today's Date", className="text-muted"),
                html.H5(id="liq-today-date", style={'color': C['text'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("As At", className="text-muted"),
                html.H5(id="liq-as-at-date", style={'color': C['text'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Current Quarter", className="text-muted"),
                html.H5(id="liq-current-quarter", style={'color': C['text'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Current Month", className="text-muted"),
                html.H5(id="liq-current-month", style={'color': C['text'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
        ], className="mb-4"),

        # Liquidity Waterfall
        dbc.Card([
            dbc.CardHeader("Liquidity Waterfall", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
            dbc.CardBody([
                html.Div(id='liq-waterfall-table')
            ])
        ], className="shadow-sm mb-4"),

        # Near Term Flows (Monthly Projections)
        dbc.Card([
            dbc.CardHeader("Near Term Flows (Next 12 Months)",
                           style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
            dbc.CardBody([
                html.Div(id='liq-near-term-flows')
            ])
        ], className="shadow-sm mb-4"),

        # NAV Projections
        dbc.Card([
            dbc.CardHeader("NAV End Projections", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
            dbc.CardBody([
                dcc.Graph(id='liq-nav-projection-chart', config={'displayModeBar': True})
            ])
        ], className="shadow-sm mb-4"),

        # Detailed Breakdown Tabs
        dbc.Tabs([
            dbc.Tab(label="Subscriptions & Redemptions", tab_id="tab-liq-subs"),
            dbc.Tab(label="Portfolio Net Flows", tab_id="tab-liq-flows"),
            dbc.Tab(label="Unfunded Commitments", tab_id="tab-liq-unfunded"),
        ], id="liq-tabs", active_tab="tab-liq-subs"),

        html.Div(id='liq-tab-content', className="mt-4"),
    ])


def segmentation_page():
    return html.Div([
        html.H2("📊 Portfolio Segmentation & TWR Analysis", className="mb-4",
                style={'fontWeight': 'bold', 'fontFamily': C['sans'], 'color': C['text']}),

        dbc.Alert([
            html.I(className="fas fa-layer-group me-2"),
            "Track TWR performance across portfolio segments: Seed Portfolio, New Deals, Money Market, Pipeline, and Future Deals"
        ], color="info", className="mb-4"),

        # Segment Summary Cards
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Total Portfolio", className="text-muted"),
                html.H4(id="seg-total-nav", style={'color': C['blue'], 'fontFamily': C['mono']}),
                html.Small(id="seg-total-twr", className="text-muted", style={'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Seed Portfolio", className="text-muted"),
                html.H4(id="seg-seed-nav", style={'color': C['green'], 'fontFamily': C['mono']}),
                html.Small(id="seg-seed-twr", className="text-muted", style={'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("New Deals", className="text-muted"),
                html.H4(id="seg-new-nav", style={'color': C['purple'], 'fontFamily': C['mono']}),
                html.Small(id="seg-new-twr", className="text-muted", style={'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Money Market", className="text-muted"),
                html.H4(id="seg-mm-nav", style={'color': C['amber'], 'fontFamily': C['mono']}),
                html.Small(id="seg-mm-twr", className="text-muted", style={'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Pipeline", className="text-muted"),
                html.H4(id="seg-pipeline-nav", style={'color': C['teal'], 'fontFamily': C['mono']}),
                html.Small(id="seg-pipeline-twr", className="text-muted", style={'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=2),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Future/Placeholder", className="text-muted"),
                html.H4(id="seg-future-nav", style={'color': C['pink'], 'fontFamily': C['mono']}),
                html.Small(id="seg-future-twr", className="text-muted", style={'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=2),
        ], className="mb-4"),

        # TWR Forecast by Segment
        dbc.Card([
            dbc.CardHeader("TWR Forecast by Segment (Next 12 Months)",
                           style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
            dbc.CardBody([
                dcc.Graph(id='seg-twr-forecast-chart', config={'displayModeBar': True})
            ])
        ], className="shadow-sm mb-4"),

        # Segment Breakdown Tables
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Seed Portfolio Deals",
                                   style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([html.Div(id='seg-seed-table')])
                ], className="shadow-sm")
            ], width=6),
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("New Deals", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([html.Div(id='seg-new-table')])
                ], className="shadow-sm")
            ], width=6),
        ], className="mb-4"),

        # Allocation and Contribution Charts
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Segment Allocation", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='seg-allocation-chart', config={'displayModeBar': False})])
                ], className="shadow-sm")
            ], width=6),
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("TWR Contribution by Segment",
                                   style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='seg-contribution-chart', config={'displayModeBar': False})])
                ], className="shadow-sm")
            ], width=6),
        ])
    ])


def analytics_page():
    return html.Div([
        html.H2("📊 Portfolio Analytics", className="mb-4",
                style={'fontWeight': 'bold', 'fontFamily': C['sans'], 'color': C['text']}),

        # Toggle between Current and Pro Forma
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardBody([
                        dbc.Row([
                            dbc.Col([
                                html.H6("Portfolio View:",
                                        style={'color': C['text'], 'fontFamily': C['sans'], 'marginBottom': 0}),
                            ], width=2),
                            dbc.Col([
                                dbc.RadioItems(
                                    id="analytics-view-toggle",
                                    options=[
                                        {"label": " Current Portfolio", "value": "current"},
                                        {"label": " Current + Pipeline", "value": "current_pipeline"},
                                        {"label": " Current + Pipeline + Placeholder", "value": "full_proforma"},
                                    ],
                                    value="current",
                                    inline=True,
                                    style={'fontFamily': C['sans'], 'fontSize': '13px'}
                                )
                            ], width=10)
                        ])
                    ])
                ], className="shadow-sm")
            ], width=12)
        ], className="mb-4"),

        # Summary Cards
        dbc.Row([
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Total NAV", className="text-muted"),
                html.H4(id="analytics-total-nav", style={'color': C['blue'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Number of Deals", className="text-muted"),
                html.H4(id="analytics-num-deals", style={'color': C['green'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Weighted IRR", className="text-muted"),
                html.H4(id="analytics-weighted-irr", style={'color': C['purple'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
            dbc.Col(dbc.Card([dbc.CardBody([
                html.H6("Top 1 Concentration", className="text-muted"),
                html.H4(id="analytics-top1-conc", style={'color': C['amber'], 'fontFamily': C['mono']})
            ])], className="shadow-sm"), width=3),
        ], className="mb-4"),

        # Exposure Charts Row 1
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Strategy Exposure", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='analytics-strategy', config={'displayModeBar': False})])
                ], className="shadow-sm")
            ], width=6),
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Regional Exposure", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='analytics-region', config={'displayModeBar': False})])
                ], className="shadow-sm")
            ], width=6),
        ], className="mb-4"),

        # Exposure Charts Row 2
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Vintage Year Exposure",
                                   style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='analytics-vintage', config={'displayModeBar': False})])
                ], className="shadow-sm")
            ], width=6),
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Secondary vs Co-Investment",
                                   style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='analytics-dealtype', config={'displayModeBar': False})])
                ], className="shadow-sm")
            ], width=6),
        ], className="mb-4"),

        # Exposure Charts Row 3
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Sector Exposure", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='analytics-sector', config={'displayModeBar': False})])
                ], className="shadow-sm")
            ], width=6),
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader("Concentration Risk", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                    dbc.CardBody([dcc.Graph(id='analytics-concentration', config={'displayModeBar': False})])
                ], className="shadow-sm")
            ], width=6),
        ])
    ])


def pipeline_page():
    return html.Div([
        dbc.Row([
            dbc.Col(html.H2("🔍 Deal Pipeline", style={'fontWeight': 'bold'}), width=8),
            dbc.Col(
                dbc.Button("➕ Add to Pipeline", id="btn-open-pipeline", color="info", className="float-end", size="lg"),
                width=4)
        ], className="mb-3"),

        # Add Pipeline Modal
        dbc.Modal([
            dbc.ModalHeader("Add Deal to Pipeline"),
            dbc.ModalBody([
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Deal Name *"),
                        dbc.Input(id="in-pipe-name", type="text", placeholder="Project Sigil")
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Deal Type *"),
                        dbc.Select(id="in-pipe-type", options=[
                            {"label": "GP-Led (Multi-Asset)", "value": "GP-Led (Multi-Asset)"},
                            {"label": "GP-Led (Single-Asset)", "value": "GP-Led (Single-Asset)"},
                            {"label": "Diversified LP-Led", "value": "Diversified LP-Led"},
                            {"label": "Co-Investments", "value": "Co-Investments"},
                        ])
                    ], width=6)
                ], className="mb-3"),
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Stage"),
                        dbc.Select(id="in-pipe-stage", options=[
                            {"label": "Screening", "value": "Screening"},
                            {"label": "Due Diligence", "value": "Due Diligence"},
                            {"label": "Term Sheet", "value": "Term Sheet"},
                            {"label": "Final Docs", "value": "Final Docs"},
                        ], value="Screening")
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Target Size ($mm)"),
                        dbc.Input(id="in-pipe-size", type="number", step=0.1, placeholder="7.5")
                    ], width=4),
                    dbc.Col([
                        dbc.Label("Target IRR (%)"),
                        dbc.Input(id="in-pipe-irr", type="number", step=0.5, placeholder="16")
                    ], width=4),
                ])
            ]),
            dbc.ModalFooter([
                dbc.Button("Cancel", id="btn-cancel-pipe", color="secondary"),
                dbc.Button("Add to Pipeline", id="btn-submit-pipe", color="info")
            ])
        ], id="modal-pipeline", size="lg", is_open=False),

        html.Div(id='pipeline-table')
    ])


def settings_page():
    return html.Div([
        html.H2("⚙️ Fund Settings", className="mb-4", style={'fontWeight': 'bold'}),

        dbc.Card([dbc.CardBody([
            html.H5("💰 Fund Parameters", className="mb-4", style={'color': COLORS['dark'], 'fontWeight': 'bold'}),
            dbc.Row([
                dbc.Col([
                    dbc.Label("💵 Dry Powder ($mm)", style={'fontWeight': 'bold'}),
                    dbc.Input(id="set-dry-powder", type="number", value=300, step=10),
                    html.Small("Available capital to deploy", className="text-muted")
                ], width=4),
                dbc.Col([
                    dbc.Label("🎯 Target Net TWR (%)", style={'fontWeight': 'bold'}),
                    dbc.Input(id="set-twr", type="number", value=13, step=0.5),
                    html.Small("Net return to LPs", className="text-muted")
                ], width=4),
                dbc.Col([
                    dbc.Label("📊 Average Hold Period (years)", style={'fontWeight': 'bold'}),
                    dbc.Input(id="set-hold", type="number", value=5.0, step=0.5),
                    html.Small("For MOIC translation", className="text-muted")
                ], width=4),
            ], className="mb-4"),

            html.Hr(),
            html.H5("💼 Fee Structure", className="mb-4", style={'color': COLORS['dark'], 'fontWeight': 'bold'}),
            dbc.Row([
                dbc.Col([
                    dbc.Label("Management Fee (% p.a.)", style={'fontWeight': 'bold'}),
                    dbc.Input(id="set-fee", type="number", value=1.25, step=0.05),
                    html.Small("Annual fee on NAV", className="text-muted")
                ], width=4),
                dbc.Col([
                    dbc.Label("Carry Rate (%)", style={'fontWeight': 'bold'}),
                    dbc.Input(id="set-carry", type="number", value=12.5, step=0.5),
                    html.Small("Performance fee", className="text-muted")
                ], width=4),
                dbc.Col([
                    dbc.Label("Hurdle Rate (%)", style={'fontWeight': 'bold'}),
                    dbc.Input(id="set-hurdle", type="number", value=10, step=0.5),
                    html.Small("Return before carry", className="text-muted")
                ], width=4),
            ], className="mb-4"),

            html.Hr(),
            html.H5("🔧 Portfolio Assumptions", className="mb-4", style={'color': COLORS['dark'], 'fontWeight': 'bold'}),
            dbc.Row([
                dbc.Col([
                    dbc.Label("Liquidity Reserve (%)", style={'fontWeight': 'bold'}),
                    dbc.Input(id="set-liq", type="number", value=5, step=1),
                    html.Small("Cash buffer", className="text-muted")
                ], width=4),
                dbc.Col([
                    dbc.Label("Loss Drag (%)", style={'fontWeight': 'bold'}),
                    dbc.Input(id="set-loss", type="number", value=1, step=0.5),
                    html.Small("Expected annual impairment", className="text-muted")
                ], width=4),
                dbc.Col([
                    dbc.Label("Cash Yield (%)", style={'fontWeight': 'bold'}),
                    dbc.Input(id="set-cash-yield", type="number", value=3, step=0.5),
                    html.Small("Return on uninvested cash", className="text-muted")
                ], width=4),
            ], className="mb-4"),

            html.Hr(),
            dbc.Row([
                dbc.Col([
                    dbc.Button("💾 Save Settings", id="btn-save-settings", color="primary", size="lg", className="me-2"),
                    dbc.Button("🔄 Reset to Default", id="btn-reset-settings", color="secondary", size="lg")
                ])
            ], className="mb-3"),
            html.Div(id='settings-save-status')
        ])], className="shadow-sm")
    ])



# ==================== CALLBACKS ====================

@app.callback(
    [Output('set-dry-powder', 'value'), Output('set-twr', 'value'), Output('set-hold', 'value'),
     Output('set-fee', 'value'), Output('set-carry', 'value'), Output('set-hurdle', 'value'),
     Output('set-liq', 'value'), Output('set-loss', 'value'), Output('set-cash-yield', 'value')],
    Input('config-store', 'data')
)
def populate_settings_inputs(config):
    config = config or DEFAULT_CONFIG
    fund = config.get('fund_parameters', {})
    return (
        fund.get('dry_powder', 450),
        fund.get('target_net_twr', 0.13) * 100,
        fund.get('hold_years', 5.0),
        fund.get('management_fee', 0.0125) * 100,
        fund.get('carry_rate', 0.125) * 100,
        fund.get('hurdle_rate', 0.10) * 100,
        fund.get('liquidity_reserve_pct', 0.05) * 100,
        fund.get('loss_drag', 0.01) * 100,
        fund.get('cash_yield', 0.03) * 100
    )


@app.callback(
    [Output('config-store', 'data', allow_duplicate=True), Output('settings-save-status', 'children')],
    [Input('btn-save-settings', 'n_clicks'), Input('btn-reset-settings', 'n_clicks')],
    [State('set-dry-powder', 'value'), State('set-twr', 'value'), State('set-hold', 'value'),
     State('set-fee', 'value'), State('set-carry', 'value'), State('set-hurdle', 'value'),
     State('set-liq', 'value'), State('set-loss', 'value'), State('set-cash-yield', 'value'),
     State('config-store', 'data')],
    prevent_initial_call=True
)
def save_or_reset_settings(save_clicks, reset_clicks, dry_powder, twr, hold, fee, carry, hurdle, liq, loss,
                           cash_yield, config):
    ctx = callback_context
    if not ctx.triggered:
        raise dash.exceptions.PreventUpdate

    trigger = ctx.triggered[0]['prop_id'].split('.')[0]

    if trigger == 'btn-reset-settings':
        return DEFAULT_CONFIG, dbc.Alert("Settings reset to defaults.", color="secondary", className="mb-0")

    config = config or DEFAULT_CONFIG
    updated = json.loads(json.dumps(config))
    updated.setdefault('fund_parameters', {})
    updated['fund_parameters'].update({
        'dry_powder': float(dry_powder or 0),
        'target_net_twr': float(twr or 0) / 100,
        'hold_years': float(hold or 0),
        'management_fee': float(fee or 0) / 100,
        'carry_rate': float(carry or 0) / 100,
        'hurdle_rate': float(hurdle or 0) / 100,
        'liquidity_reserve_pct': float(liq or 0) / 100,
        'loss_drag': float(loss or 0) / 100,
        'cash_yield': float(cash_yield or 0) / 100,
    })

    return updated, dbc.Alert("Settings saved.", color="success", className="mb-0")


@app.callback(Output('live-clock', 'children'), Input('clock-interval', 'n_intervals'))

def update_clock(n):
    return datetime.now().strftime('%H:%M:%S')


@app.callback(Output('sidebar-stats', 'children'),
              [Input('deals-store', 'data'), Input('config-store', 'data')])
def update_sidebar(deals, config):
    m = calculate_portfolio_metrics(deals)
    dry_powder = config['fund_parameters']['dry_powder']
    return [
        html.Small("💰 NAV", className="text-muted"),
        html.H6(f"${m['total_nav']:.0f}M", className="mb-2", style={'fontWeight': 'bold'}),
        html.Small("💵 Dry Powder", className="text-muted"),
        html.H6(f"${dry_powder:.0f}M", className="mb-2", style={'fontWeight': 'bold'}),
        html.Small("📊 Deals", className="text-muted"),
        html.H6(str(m['num_deals']), className="mb-2", style={'fontWeight': 'bold'}),
        html.Small("📈 Avg IRR", className="text-muted"),
        html.H6(f"{m['weighted_irr']:.1%}", style={'fontWeight': 'bold'})
    ]


@app.callback(Output('page-content', 'children'), Input('url', 'pathname'))
def display_page(pathname):
    if pathname == '/overview':
        return fund_overview_page()
    elif pathname == '/portfolio':
        return portfolio_page()
    elif pathname == '/segmentation':
        return segmentation_page()
    elif pathname == '/future':
        return future_deals_page()
    elif pathname == '/drypowder':
        return drypowder_page()
    elif pathname == '/liquidity':
        return liquidity_assumptions_page()
    elif pathname == '/calculator':
        return calculator_page()
    elif pathname == '/twr':
        return twr_forecaster_page()
    elif pathname == '/cashflows':
        return cashflows_page()
    elif pathname == '/proforma':
        return proforma_page()
    elif pathname == '/pipeline':
        return pipeline_page()
    elif pathname == '/analytics':
        return analytics_page()
    elif pathname == '/settings':
        return settings_page()
    else:
        return dashboard_page()


# Deal Modal
@app.callback(
    Output('modal-deal', 'is_open'),
    [Input('btn-open-deal', 'n_clicks'), Input('btn-cancel', 'n_clicks'), Input('btn-submit', 'n_clicks')],
    State('modal-deal', 'is_open'), prevent_initial_call=True
)
def toggle_deal_modal(o, c, s, is_open):
    return not is_open


# Placeholder Modal
@app.callback(
    Output('modal-placeholder', 'is_open'),
    [Input('btn-open-placeholder', 'n_clicks'), Input('btn-cancel-ph', 'n_clicks'), Input('btn-submit-ph', 'n_clicks')],
    State('modal-placeholder', 'is_open'), prevent_initial_call=True
)
def toggle_placeholder_modal(o, c, s, is_open):
    return not is_open


# Pipeline Modal
@app.callback(
    Output('modal-pipeline', 'is_open'),
    [Input('btn-open-pipeline', 'n_clicks'), Input('btn-cancel-pipe', 'n_clicks'),
     Input('btn-submit-pipe', 'n_clicks')],
    State('modal-pipeline', 'is_open'), prevent_initial_call=True
)
def toggle_pipeline_modal(o, c, s, is_open):
    return not is_open


# Edit Deal Modal Toggle
@app.callback(
    Output('modal-edit-deal', 'is_open'),
    [Input({'type': 'edit-deal', 'index': ALL}, 'n_clicks'),
     Input('btn-cancel-edit', 'n_clicks'), Input('btn-save-edit', 'n_clicks')],
    State('modal-edit-deal', 'is_open'),
    prevent_initial_call=True
)
def toggle_edit_modal(edit_clicks, cancel, save, is_open):
    return not is_open


# Populate Edit Modal with Deal Data
@app.callback(
    [Output('edit-deal-index', 'data'),
     Output('edit-name', 'value'), Output('edit-manager', 'value'),
     Output('edit-strategy', 'value'), Output('edit-stage', 'value'),
     Output('edit-total-commitment', 'value'), Output('edit-current-commitment', 'value'),
     Output('edit-nav', 'value'), Output('edit-irr', 'value'), Output('edit-hold', 'value'),
     Output('edit-nav-source', 'children')],
    Input({'type': 'edit-deal', 'index': ALL}, 'n_clicks'),
    [State('deals-store', 'data'), State('fund-cf-data-store', 'data')],
    prevent_initial_call=True
)
def populate_edit_modal(n_clicks, deals, fund_cf_data):
    """Populate edit modal and auto-pull NAV from cashflows"""
    print(f"\n=== POPULATE EDIT MODAL DEBUG ===")
    print(f"n_clicks received: {n_clicks}")
    print(f"Number of deals: {len(deals) if deals else 0}")

    if not any(n_clicks) or not deals:
        print("No clicks or no deals - returning empty")
        return None, "", "", "", "", 0, 0, 0, 0, 5, ""

    ctx = callback_context
    if not ctx.triggered:
        print("No trigger context - returning empty")
        return None, "", "", "", "", 0, 0, 0, 0, 5, ""

    # Get clicked button
    button_id = ctx.triggered[0]['prop_id'].split('.')[0]
    print(f"Button ID: {button_id}")
    button_info = json.loads(button_id)
    idx = button_info['index']
    print(f"Editing deal at index: {idx}")

    if idx < 0 or idx >= len(deals):
        print(f"Invalid index {idx}")
        return None, "", "", "", "", 0, 0, 0, 0, 5, ""

    deal = deals[idx]
    print(f"Deal name: {deal['name']}")

    # Get NAV from fund CF data if available
    nav = deal.get('nav', deal.get('size', 0))
    nav_source = "Current NAV"

    if fund_cf_data:
        print(f"Checking {len(fund_cf_data)} cashflow deals for NAV")
        latest_nav = get_latest_available_nav_for_deal(fund_cf_data, deal['name'])
        if latest_nav is not None:
            nav = latest_nav
            nav_source = f"✓ Auto-pulled from latest month-end NAV ({get_last_month_end().isoformat()})"
            print(f"✅ Found NAV in cashflows: {nav}")
    else:
        print("No cashflow data available")

    print(f"Returning NAV: {nav}, Source: {nav_source}")

    return (
        idx,
        deal['name'],
        deal.get('manager', ''),
        deal.get('strategy', ''),
        deal.get('stage', 'Buyout'),
        deal.get('total_commitment', deal.get('commitment', 0)),
        deal.get('current_commitment', deal.get('commitment', 0)),
        nav,
        deal.get('target_irr', 0) * 100,
        deal.get('hold_period', 5),
        nav_source
    )


# Auto-calculate Unfunded
@app.callback(
    Output('unfunded-display', 'children'),
    [Input('in-total-commitment', 'value'), Input('in-current-commitment', 'value')]
)
def calculate_unfunded(total_commitment, current_commitment):
    if total_commitment is None or current_commitment is None:
        return "$0.0M"
    unfunded = float(total_commitment) - float(current_commitment)
    return f"${unfunded:.1f}M"


# Auto-calculate Unfunded for Edit Modal
@app.callback(
    Output('edit-unfunded-display', 'children'),
    [Input('edit-total-commitment', 'value'), Input('edit-current-commitment', 'value')]
)
def calculate_edit_unfunded(total_commitment, current_commitment):
    if total_commitment is None or current_commitment is None:
        return "$0.0M"
    unfunded = float(total_commitment) - float(current_commitment)
    return f"${unfunded:.1f}M"


# Save Edited Deal
@app.callback(
    Output('deals-store', 'data', allow_duplicate=True),
    Input('btn-save-edit', 'n_clicks'),
    [State('edit-deal-index', 'data'),
     State('edit-name', 'value'), State('edit-manager', 'value'),
     State('edit-strategy', 'value'), State('edit-stage', 'value'),
     State('edit-total-commitment', 'value'), State('edit-current-commitment', 'value'),
     State('edit-nav', 'value'), State('edit-irr', 'value'), State('edit-hold', 'value'),
     State('deals-store', 'data')],
    prevent_initial_call='initial_duplicate'
)
def save_edited_deal(n, idx, name, manager, strategy, stage, total_comm, current_comm,
                     nav, irr, hold, deals):
    """Save changes to edited deal"""
    if not n or idx is None or not deals:
        return deals or []

    if idx < 0 or idx >= len(deals):
        return deals

    print(f"\n=== EDIT DEAL DEBUG ===")
    print(f"Editing deal index: {idx}")
    print(f"Name: {name}")
    print(f"NAV: {nav}")

    try:
        # Update the deal
        unfunded = float(total_comm) - float(current_comm)

        deals[idx].update({
            'name': name,
            'manager': manager,
            'strategy': strategy,
            'stage': stage,
            'total_commitment': float(total_comm),
            'current_commitment': float(current_comm),
            'nav': float(nav),
            'size': float(nav),  # Keep for compatibility
            'commitment': float(total_comm),  # Keep for compatibility
            'unfunded': unfunded,
            'target_irr': float(irr) / 100,
            'hold_period': float(hold),
            'moic': (1 + float(irr) / 100) ** float(hold),
        })

        print(f"✅ Deal updated successfully: {name}")
        return deals

    except Exception as e:
        print(f"ERROR updating deal: {e}")
        import traceback
        traceback.print_exc()
        return deals


# Pipeline Edit Modal Toggle
@app.callback(
    Output('modal-edit-pipeline', 'is_open'),
    [Input({'type': 'edit-pipeline', 'index': ALL}, 'n_clicks'),
     Input('btn-cancel-edit-pipeline', 'n_clicks'), Input('btn-save-edit-pipeline', 'n_clicks')],
    State('modal-edit-pipeline', 'is_open'),
    prevent_initial_call=True
)
def toggle_edit_pipeline_modal(edit_clicks, cancel, save, is_open):
    return not is_open


# Populate Pipeline Edit Modal
@app.callback(
    [Output('edit-pipeline-index', 'data'),
     Output('edit-pipeline-name', 'value'), Output('edit-pipeline-type', 'value'),
     Output('edit-pipeline-stage', 'value'), Output('edit-pipeline-size', 'value'),
     Output('edit-pipeline-irr', 'value')],
    Input({'type': 'edit-pipeline', 'index': ALL}, 'n_clicks'),
    State('pipeline-store', 'data'),
    prevent_initial_call=True
)
def populate_edit_pipeline_modal(n_clicks, pipeline):
    """Populate pipeline edit modal with current values"""
    if not any(n_clicks) or not pipeline:
        return None, "", "", "", 0, 0

    ctx = callback_context
    if not ctx.triggered:
        return None, "", "", "", 0, 0

    button_id = ctx.triggered[0]['prop_id'].split('.')[0]
    button_info = json.loads(button_id)
    idx = button_info['index']

    if idx < 0 or idx >= len(pipeline):
        return None, "", "", "", 0, 0

    p = pipeline[idx]

    return (
        idx,
        p['name'],
        p['type'],
        p['stage'],
        p['size'],
        p['target_irr'] * 100 if p['target_irr'] < 1 else p['target_irr']
    )


# Save Pipeline Edit
@app.callback(
    Output('pipeline-store', 'data', allow_duplicate=True),
    Input('btn-save-edit-pipeline', 'n_clicks'),
    [State('edit-pipeline-index', 'data'),
     State('edit-pipeline-name', 'value'), State('edit-pipeline-type', 'value'),
     State('edit-pipeline-stage', 'value'), State('edit-pipeline-size', 'value'),
     State('edit-pipeline-irr', 'value'),
     State('pipeline-store', 'data')],
    prevent_initial_call='initial_duplicate'
)
def save_edited_pipeline(n, idx, name, ptype, stage, size, irr, pipeline):
    """Save changes to edited pipeline deal"""
    if not n or idx is None or not pipeline:
        return pipeline or []

    if idx < 0 or idx >= len(pipeline):
        return pipeline

    try:
        pipeline[idx].update({
            'name': name,
            'type': ptype,
            'stage': stage,
            'size': float(size) if size else 0,
            'target_irr': float(irr) / 100 if irr and irr > 1 else float(irr) if irr else 0,
        })

        print(f"✅ Pipeline deal updated: {name}")
        return pipeline

    except Exception as e:
        print(f"ERROR updating pipeline deal: {e}")
        return pipeline


# Placeholder Edit Modal Toggle
@app.callback(
    Output('modal-edit-placeholder', 'is_open'),
    [Input({'type': 'edit-placeholder', 'index': ALL}, 'n_clicks'),
     Input('btn-cancel-edit-placeholder', 'n_clicks'), Input('btn-save-edit-placeholder', 'n_clicks')],
    State('modal-edit-placeholder', 'is_open'),
    prevent_initial_call=True
)
def toggle_edit_placeholder_modal(edit_clicks, cancel, save, is_open):
    return not is_open


# Populate Placeholder Edit Modal
@app.callback(
    [Output('edit-placeholder-index', 'data'),
     Output('edit-placeholder-name', 'value'), Output('edit-placeholder-strategy', 'value'),
     Output('edit-placeholder-size', 'value'), Output('edit-placeholder-month', 'value')],
    Input({'type': 'edit-placeholder', 'index': ALL}, 'n_clicks'),
    State('placeholder-deals-store', 'data'),
    prevent_initial_call=True
)
def populate_edit_placeholder_modal(n_clicks, placeholders):
    """Populate placeholder edit modal with current values"""
    if not any(n_clicks) or not placeholders:
        return None, "", "", 0, 0

    ctx = callback_context
    if not ctx.triggered:
        return None, "", "", 0, 0

    button_id = ctx.triggered[0]['prop_id'].split('.')[0]
    button_info = json.loads(button_id)
    idx = button_info['index']

    if idx < 0 or idx >= len(placeholders):
        return None, "", "", 0, 0

    ph = placeholders[idx]

    return (
        idx,
        ph['name'],
        ph['strategy'],
        ph['size'],
        ph['expected_month']
    )


# Save Placeholder Edit
@app.callback(
    Output('placeholder-deals-store', 'data', allow_duplicate=True),
    Input('btn-save-edit-placeholder', 'n_clicks'),
    [State('edit-placeholder-index', 'data'),
     State('edit-placeholder-name', 'value'), State('edit-placeholder-strategy', 'value'),
     State('edit-placeholder-size', 'value'), State('edit-placeholder-month', 'value'),
     State('placeholder-deals-store', 'data')],
    prevent_initial_call='initial_duplicate'
)
def save_edited_placeholder(n, idx, name, strategy, size, month, placeholders):
    """Save changes to edited placeholder deal"""
    if not n or idx is None or not placeholders:
        return placeholders or []

    if idx < 0 or idx >= len(placeholders):
        return placeholders

    try:
        placeholders[idx].update({
            'name': name,
            'strategy': strategy,
            'size': float(size) if size else 0,
            'expected_month': int(month) if month is not None else 0,
        })

        print(f"✅ Placeholder deal updated: {name}")
        return placeholders

    except Exception as e:
        print(f"ERROR updating placeholder deal: {e}")
        return placeholders


# Extract Dry Powder from Liquidity Upload
@app.callback(
    Output('liquidity-dry-powder-store', 'data'),
    Input('liquidity-data-store', 'data')
)
def extract_dry_powder(liquidity_data):
    """Extract Dec 2026 dry powder value from Row 53, Column 9"""
    if not liquidity_data:
        return None

    # Priority 1: Direct Dec 2026 value from Row 53, Col 9
    if 'dec_2026_dry_powder' in liquidity_data:
        return liquidity_data['dec_2026_dry_powder']

    # Priority 2: Look in max_deployable_capital dict
    max_deploy = liquidity_data.get('max_deployable_capital', {})

    # Try to find Dec 2026 value
    for key, value in max_deploy.items():
        if 'Dec' in str(key) and '2026' in str(key):
            return value
        if '12' in str(key) and '2026' in str(key):
            return value

    # Priority 3: If we have at least 10 months, get the 10th (approx Dec 2026)
    if len(max_deploy) >= 10:
        values_list = list(max_deploy.values())
        return values_list[9] if len(values_list) > 9 else None

    return None


# Add Deal
@app.callback(
    Output('deals-store', 'data', allow_duplicate=True),
    Input('btn-submit', 'n_clicks'),
    [State('in-name', 'value'), State('in-manager', 'value'), State('in-strategy', 'value'),
     State('in-stage', 'value'), State('in-total-commitment', 'value'), State('in-current-commitment', 'value'),
     State('in-nav', 'value'), State('in-irr', 'value'), State('in-hold', 'value'), State('in-currency', 'value'),
     State('in-vintage', 'value'), State('in-segment', 'value'), State('in-allocation-status', 'value'),
     State('in-sector', 'value'), State('in-geo', 'value'),
     State('deals-store', 'data'), State('config-store', 'data')],
    prevent_initial_call='initial_duplicate'
)
def add_deal(n, name, manager, strat, stage, total_commitment, current_commitment, nav, irr, hold, currency, vint,
             segment,
             alloc_status, sec, geo, deals, config):
    """Add deal with comprehensive error handling"""

    # Debug logging
    print(f"\n=== ADD DEAL DEBUG ===")
    print(f"Button clicks: {n}")
    print(f"Name: {name}")
    print(f"Manager: {manager}")
    print(f"Strategy: {strat}")
    print(f"Total Commitment: {total_commitment}")
    print(f"Current Commitment: {current_commitment}")
    print(f"NAV: {nav}")
    print(f"IRR: {irr}")

    # Validation
    if not n or n == 0:
        print("No button click detected")
        return deals or []

    if not name:
        print("ERROR: Missing deal name")
        return deals or []

    if not manager:
        print("ERROR: Missing fund manager")
        return deals or []

    if not strat:
        print("ERROR: Missing strategy")
        return deals or []

    if total_commitment is None or total_commitment == '':
        print("ERROR: Missing total commitment")
        return deals or []

    if current_commitment is None or current_commitment == '':
        print("ERROR: Missing current commitment")
        return deals or []

    if nav is None or nav == '':
        print("ERROR: Missing NAV")
        return deals or []

    if irr is None or irr == '':
        print("ERROR: Missing IRR")
        return deals or []

    try:
        # Calculate unfunded
        unfunded = float(total_commitment) - float(current_commitment)

        # Create new deal
        new_deal = {
            'name': name,
            'manager': manager,
            'strategy': strat,
            'stage': stage or 'Buyout',
            'total_commitment': float(total_commitment),
            'current_commitment': float(current_commitment),
            'nav': float(nav),
            'size': float(nav),  # Keep for backward compatibility
            'commitment': float(total_commitment),  # Keep for backward compatibility
            'unfunded': unfunded,
            'target_irr': float(irr) / 100,
            'hold_period': float(hold) if hold else 5.0,
            'moic': (1 + float(irr) / 100) ** (float(hold) if hold else 5.0),
            'currency': currency or 'USD',
            'vintage': int(vint) if vint else 2024,
            'segment': segment or 'Seed',
            'allocation_status': alloc_status or 'Closed',
            'sector': sec or 'Diversified',
            'geography': geo or 'North America',
            'date_added': datetime.now().isoformat(),
            'is_actual': True
        }

        print(f"✅ Deal created successfully: {name}")
        print(f"New deal data: {new_deal}")

        updated_deals = (deals or []) + [new_deal]
        print(f"Total deals after add: {len(updated_deals)}")

        return updated_deals

    except Exception as e:
        print(f"ERROR creating deal: {e}")
        import traceback
        traceback.print_exc()
        return deals or []


# Add Placeholder
@app.callback(
    Output('placeholder-deals-store', 'data', allow_duplicate=True),
    Input('btn-submit-ph', 'n_clicks'),
    [State('in-ph-name', 'value'), State('in-ph-strategy', 'value'),
     State('in-ph-size', 'value'), State('in-ph-bite', 'value'),
     State('in-ph-weight', 'value'), State('in-ph-month', 'value'),
     State('in-ph-region', 'value'), State('in-ph-type', 'value'),
     State('placeholder-deals-store', 'data')],
    prevent_initial_call='initial_duplicate'
)
def add_placeholder(n, name, strat, size, bite, weight, month, region, deal_type, placeholders):
    if not n or n == 0:
        return placeholders or []
    if not all([name, strat, size is not None, month is not None]):
        return placeholders or []

    return placeholders + [{
        'name': name, 'strategy': strat, 'size': float(size),
        'custom_bite_size': float(bite) if bite else None,
        'target_weight': float(weight) if weight else None,
        'expected_month': int(month),
        'region': region or 'North America',
        'deal_type': deal_type or 'Secondary',
        'date_added': datetime.now().isoformat()
    }]


# Add Pipeline Deal
@app.callback(
    Output('pipeline-store', 'data', allow_duplicate=True),
    Input('btn-submit-pipe', 'n_clicks'),
    [State('in-pipe-name', 'value'), State('in-pipe-type', 'value'),
     State('in-pipe-stage', 'value'), State('in-pipe-size', 'value'),
     State('in-pipe-irr', 'value'), State('pipeline-store', 'data')],
    prevent_initial_call='initial_duplicate'
)
def add_pipeline(n, name, ptype, stage, size, irr, pipeline):
    if not all([name, ptype]):
        return pipeline

    return pipeline + [{
        'name': name, 'type': ptype, 'stage': stage or 'Screening',
        'size': float(size) if size else 0, 'target_irr': float(irr) / 100 if irr else 0,
        'date_added': datetime.now().isoformat()
    }]


# Delete Deal
@app.callback(
    Output('deals-store', 'data', allow_duplicate=True),
    Input({'type': 'delete-deal', 'index': ALL}, 'n_clicks'),
    State('deals-store', 'data'),
    prevent_initial_call='initial_duplicate'
)
def delete_deal(n_clicks, deals):
    if not any(n_clicks):
        return deals
    ctx = callback_context
    if not ctx.triggered:
        return deals
    button_id = ctx.triggered[0]['prop_id'].split('.')[0]
    if button_id == '':
        return deals
    button_info = json.loads(button_id)
    delete_idx = button_info['index']
    if 0 <= delete_idx < len(deals):
        deals.pop(delete_idx)
    return deals


# Delete Placeholder
@app.callback(
    Output('placeholder-deals-store', 'data', allow_duplicate=True),
    Input({'type': 'delete-placeholder', 'index': ALL}, 'n_clicks'),
    State('placeholder-deals-store', 'data'),
    prevent_initial_call='initial_duplicate'
)
def delete_placeholder(n_clicks, placeholders):
    if not any(n_clicks):
        return placeholders
    ctx = callback_context
    if not ctx.triggered:
        return placeholders
    button_id = ctx.triggered[0]['prop_id'].split('.')[0]
    if button_id == '':
        return placeholders
    button_info = json.loads(button_id)
    delete_idx = button_info['index']
    if 0 <= delete_idx < len(placeholders):
        placeholders.pop(delete_idx)
    return placeholders


# Delete Pipeline
@app.callback(
    Output('pipeline-store', 'data', allow_duplicate=True),
    Input({'type': 'delete-pipeline', 'index': ALL}, 'n_clicks'),
    State('pipeline-store', 'data'),
    prevent_initial_call='initial_duplicate'
)
def delete_pipeline(n_clicks, pipeline):
    if not any(n_clicks):
        return pipeline
    ctx = callback_context
    if not ctx.triggered:
        return pipeline
    button_id = ctx.triggered[0]['prop_id'].split('.')[0]
    if button_id == '':
        return pipeline
    button_info = json.loads(button_id)
    delete_idx = button_info['index']
    if 0 <= delete_idx < len(pipeline):
        pipeline.pop(delete_idx)
    return pipeline


# Portfolio Table
@app.callback(Output('portfolio-table', 'children'), [Input('deals-store', 'data'), Input('fund-cf-data-store', 'data')])
def update_portfolio_table(deals, fund_cf_data):
    deals = enrich_deals_with_latest_nav(deals, fund_cf_data)
    if not deals:
        return dbc.Alert("📝 No deals yet. Click 'Add Deal' to start.", color="info")
    rows = []
    for idx, d in enumerate(deals):
        # Get new structure or fall back to old
        total_commitment = d.get('total_commitment', d.get('commitment', d.get('size', 0)))
        current_commitment = d.get('current_commitment', d.get('commitment', d.get('size', 0)))
        nav = d.get('nav', d.get('size', 0))
        unfunded = d.get('unfunded', total_commitment - current_commitment)

        # Allocation status badge color
        status = d.get('allocation_status', 'Closed')
        if status == 'Closed':
            status_color = 'success'
        elif status == 'Pending Close':
            status_color = 'warning'
        else:
            status_color = 'info'

        rows.append(dbc.Card([dbc.CardBody([
            dbc.Row([
                dbc.Col([
                    html.H5([
                        d['name'],
                        dbc.Badge(status, color=status_color, className="ms-2", style={'fontSize': '10px'})
                    ], className="mb-1", style={'fontFamily': C['sans'], 'color': C['text']}),
                    html.Small([
                        html.Strong(d.get('manager', 'N/A'), style={'color': C['blue']}),
                        f" • {d['strategy']} • {d.get('stage', 'Buyout')} • {d['sector']} • {d['geography']}"
                    ], className="text-muted", style={'fontFamily': C['mono'], 'fontSize': '12px'})
                ], width=3),
                dbc.Col([
                    html.Div([
                        html.Strong("Total Commitment: ",
                                    style={'fontFamily': C['sans'], 'color': C['muted'], 'fontSize': '11px'}),
                        html.Span(f"${total_commitment:.1f}M {d.get('currency', 'USD')}",
                                  style={'fontFamily': C['mono'], 'color': C['blue'], 'fontSize': '13px'}),
                        html.Br(),
                        html.Strong("Current Commitment: ",
                                    style={'fontFamily': C['sans'], 'color': C['muted'], 'fontSize': '11px'}),
                        html.Span(f"${current_commitment:.1f}M {d.get('currency', 'USD')}",
                                  style={'fontFamily': C['mono'], 'color': C['purple'], 'fontSize': '13px'}),
                        html.Br(),
                        html.Strong("Current NAV: ", style={'fontFamily': C['sans'], 'color': C['muted'], 'fontSize': '11px'}),
                        html.Span(f"${nav:.1f}M {d.get('currency', 'USD')}",
                                  style={'fontFamily': C['mono'], 'color': C['green'], 'fontSize': '13px'}),
                        html.Br(),
                        html.Strong("Unfunded: ",
                                    style={'fontFamily': C['sans'], 'color': C['muted'], 'fontSize': '11px'}),
                        html.Span(f"${unfunded:.1f}M",
                                  style={'fontFamily': C['mono'], 'color': C['amber'], 'fontSize': '13px'})
                    ])
                ], width=4),
                dbc.Col([
                    html.Div([
                        html.Strong("IRR: ", style={'fontFamily': C['sans'], 'color': C['muted'], 'fontSize': '11px'}),
                        html.Span(f"{d['target_irr']:.1%}",
                                  style={'fontFamily': C['mono'], 'color': C['text'], 'fontSize': '13px'}),
                        html.Br(),
                        html.Strong("MOIC: ", style={'fontFamily': C['sans'], 'color': C['muted'], 'fontSize': '11px'}),
                        html.Span(f"{d.get('moic', 0):.2f}x",
                                  style={'fontFamily': C['mono'], 'color': C['text'], 'fontSize': '13px'}),
                        html.Br(),
                        html.Strong("Hold: ", style={'fontFamily': C['sans'], 'color': C['muted'], 'fontSize': '11px'}),
                        html.Span(f"{d.get('hold_period', 5):.1f}y",
                                  style={'fontFamily': C['mono'], 'color': C['text'], 'fontSize': '13px'}),
                    ])
                ], width=2),
                dbc.Col([
                    html.Div([
                        html.Strong("Vintage: ",
                                    style={'fontFamily': C['sans'], 'color': C['muted'], 'fontSize': '11px'}),
                        html.Span(str(d.get('vintage', 2024)),
                                  style={'fontFamily': C['mono'], 'color': C['text'], 'fontSize': '13px'}),
                        html.Br(),
                        html.Strong("Segment: ",
                                    style={'fontFamily': C['sans'], 'color': C['muted'], 'fontSize': '11px'}),
                        html.Span(d.get('segment', 'Seed'),
                                  style={'fontFamily': C['mono'], 'color': C['purple'], 'fontSize': '11px'}),
                        html.Br(),
                        html.Strong("Added: ",
                                    style={'fontFamily': C['sans'], 'color': C['muted'], 'fontSize': '11px'}),
                        html.Span(d.get('date_added', '')[:10] if d.get('date_added') else 'N/A',
                                  style={'fontFamily': C['mono'], 'color': C['text'], 'fontSize': '11px'})
                    ])
                ], width=2),
                dbc.Col([
                    dbc.ButtonGroup([
                        dbc.Button("✏️", id={'type': 'edit-deal', 'index': idx},
                                   color="warning", size="sm", outline=True, className="me-1"),
                        dbc.Button("🗑️", id={'type': 'delete-deal', 'index': idx},
                                   color="danger", size="sm", outline=True)
                    ])
                ], width=1, className="text-end")
            ])
        ])], className="mb-2 shadow-sm"))
    return html.Div(rows)


# Placeholder Table
@app.callback(Output('placeholder-table', 'children'), Input('placeholder-deals-store', 'data'))
def update_placeholder_table(placeholders):
    if not placeholders:
        return dbc.Alert("📝 No placeholder deals yet. Add future deals to forecast dry powder.", color="info")
    rows = []
    for idx, pd in enumerate(placeholders):
        # Month index starts from Jan 2026
        base_date = datetime(2026, 1, 1)
        month_date = base_date + relativedelta(months=pd['expected_month'])
        month_name = month_date.strftime('%b %Y')
        rows.append(dbc.Card([dbc.CardBody([
            dbc.Row([
                dbc.Col([
                    html.H5(pd['name'], className="mb-1", style={'fontFamily': C['sans'], 'color': C['text']}),
                    html.Small(f"{pd['strategy']} • Expected: {month_name}",
                               className="text-muted", style={'fontFamily': C['mono']})
                ], width=8),
                dbc.Col([
                    html.H5(f"${pd['size']:.1f}M", className="text-end mb-1",
                            style={'color': C['green'], 'fontFamily': C['mono']})
                ], width=2),
                dbc.Col([
                    dbc.ButtonGroup([
                        dbc.Button("✏️", id={'type': 'edit-placeholder', 'index': idx},
                                   color="warning", size="sm", outline=True, className="me-1"),
                        dbc.Button("🗑️", id={'type': 'delete-placeholder', 'index': idx},
                                   color="danger", size="sm", outline=True)
                    ])
                ], width=2, className="text-end")
            ])
        ])], className="mb-2 shadow-sm", style={'backgroundColor': C['panel'], 'border': f'1px solid {C["border"]}'}))
    return html.Div(rows)


# Pipeline Table
@app.callback(Output('pipeline-table', 'children'), Input('pipeline-store', 'data'))
def update_pipeline_table(pipeline):
    if not pipeline:
        return dbc.Alert("📝 No pipeline deals yet.", color="info")

    rows = []
    for idx, p in enumerate(pipeline):
        rows.append(dbc.Card([dbc.CardBody([
            dbc.Row([
                dbc.Col([
                    html.H5(p['name'], className="mb-1", style={'fontFamily': C['sans'], 'color': C['text']}),
                    html.Small(f"{p['type']} • {p['stage']}",
                               className="text-muted", style={'fontFamily': C['mono']})
                ], width=6),
                dbc.Col([
                    html.Div([
                        html.Strong("Size: ", style={'fontSize': '11px', 'color': C['muted']}),
                        html.Span(f"${p['size']:.1f}M" if p['size'] > 0 else "TBD",
                                  style={'color': C['green'], 'fontFamily': C['mono'], 'fontSize': '13px'})
                    ])
                ], width=2),
                dbc.Col([
                    html.Div([
                        html.Strong("IRR: ", style={'fontSize': '11px', 'color': C['muted']}),
                        html.Span(f"{p['target_irr']:.1%}" if p['target_irr'] > 0 else "TBD",
                                  style={'fontFamily': C['mono'], 'fontSize': '13px'})
                    ])
                ], width=2),
                dbc.Col([
                    dbc.ButtonGroup([
                        dbc.Button("✏️", id={'type': 'edit-pipeline', 'index': idx},
                                   color="warning", size="sm", outline=True, className="me-1"),
                        dbc.Button("🗑️", id={'type': 'delete-pipeline', 'index': idx},
                                   color="danger", size="sm", outline=True)
                    ])
                ], width=2, className="text-end")
            ])
        ])], className="mb-2 shadow-sm", style={'backgroundColor': C['panel'], 'border': f'1px solid {C["border"]}'}))
    return html.Div(rows)


# Dashboard KPIs
@app.callback(
    [Output('dash-nav', 'children'), Output('dash-num-deals', 'children'),
     Output('dash-dry-powder', 'children'), Output('dash-total', 'children'),
     Output('dash-current-irr', 'children'), Output('dash-req-irr', 'children'),
     Output('dash-placeholders', 'children'), Output('dash-placeholder-value', 'children')],
    [Input('deals-store', 'data'), Input('placeholder-deals-store', 'data'),
     Input('config-store', 'data'), Input('liquidity-data-store', 'data'),
     Input('fund-cf-data-store', 'data'), Input('liquidity-dry-powder-store', 'data')]
)
def update_dashboard_kpis(deals, placeholders, config, liquidity_data, fund_cf_data, dry_powder_uploaded):
    """
    Total NAV = Investment NAV + Cash + GLF + CQS

    Sources:
    - Investment NAV: From fund_cf_data (monthly) OR deals manually entered
    - Cash, GLF, CQS: From liquidity_data (Rows 58, 59, 60)
    - Dry Powder: From liquidity_dry_powder_store (Row 53, Col 9 - Dec 2026)
    """
    # Calculate investment NAV from deals
    m = calculate_portfolio_metrics(deals)
    investment_nav = m['total_nav']  # NAV from manually entered deals

    # Get Cash, GLF, CQS from liquidity upload (if available)
    cash = 0
    glf = 0
    cqs = 0

    if liquidity_data:
        cash = liquidity_data.get('current_cash', 0)
        glf = liquidity_data.get('glf_balance', 0)
        cqs = liquidity_data.get('cqs_balance', 0)

    # If fund_cf_data is available, use it for investment NAV
    # (This would override manually entered deals if both exist)
    if fund_cf_data and len(fund_cf_data) > 0:
        # Sum up latest NAV from fund CF data
        # This represents the most recent monthly snapshot
        fund_cf_nav = sum(deal.get('nav', deal.get('size', 0)) for deal in fund_cf_data)
        investment_nav = fund_cf_nav

    # TOTAL NAV = Investment NAV + Cash + GLF + CQS
    total_nav = investment_nav + cash + glf + cqs

    # Dry powder calculation
    # Priority: 1) Uploaded from Liquidity file (Dec 2026 value), 2) Config setting
    if dry_powder_uploaded and dry_powder_uploaded > 0:
        dry_powder = dry_powder_uploaded
    else:
        dry_powder = config['fund_parameters']['dry_powder']

    # Required IRR calculation (based on investment NAV only, not cash/GLF/CQS)
    req_irr = calculate_required_future_irr(m['weighted_irr'], investment_nav, dry_powder, config)

    # Placeholder stats
    num_placeholders = len(placeholders)
    total_placeholder_value = sum(p['size'] for p in placeholders)

    return (
        f"${total_nav:.1f}M",  # Shows Total NAV (investments + cash + GLF + CQS)
        f"{m['num_deals']} deals",
        f"${dry_powder:.0f}M",  # Now pulls from Liquidity file Dec 2026!
        f"${total_nav + dry_powder:.0f}M",  # Total Fund = Total NAV + Dry Powder
        f"{m['weighted_irr']:.1%}",
        f"{req_irr:.1%}",
        str(num_placeholders),
        f"${total_placeholder_value:.0f}M"
    )


# Dashboard Charts
@app.callback(
    [Output('dash-allocation-chart', 'figure'), Output('dash-forecast-chart', 'figure'),
     Output('dash-bite-sizing', 'children')],
    [Input('deals-store', 'data'), Input('placeholder-deals-store', 'data'), Input('config-store', 'data')]
)
def update_dashboard_charts(deals, placeholders, config):
    m = calculate_portfolio_metrics(deals)
    dry_powder = config['fund_parameters']['dry_powder']

    # Allocation Pie
    if m['by_strategy']:
        labels = list(m['by_strategy'].keys())
        values = [m['by_strategy'][s]['nav'] for s in labels]
        colors_pie = [C['blue'], C['purple'], C['teal'], C['green']][:len(labels)]

        fig_alloc = go.Figure(data=[go.Pie(
            labels=labels, values=values, hole=0.4,
            marker=dict(colors=colors_pie, line=dict(color=C['border'], width=1)),
            textfont=dict(color=C['text'], family=C['mono'])
        )])
        fig_alloc.update_layout(**CHART_BASE, height=300, margin=dict(t=20, b=0, l=0, r=0), showlegend=False)
    else:
        fig_alloc = go.Figure()
        fig_alloc.update_layout(**CHART_BASE)

    # Forecast
    forecast = forecast_dry_powder(m['total_nav'], dry_powder, deals, placeholders, config, 12)
    months = [f['month'] for f in forecast]
    dp_values = [f['dry_powder'] for f in forecast]

    fig_forecast = go.Figure()
    fig_forecast.add_trace(go.Scatter(
        x=months, y=dp_values, mode='lines+markers', name='Dry Powder',
        line=dict(color=C['blue'], width=3),
        fill='tozeroy',
        fillcolor=rgba(C['blue'], 0.2),
        marker=dict(size=6, color=C['sky'], line=dict(color=C['blue'], width=2))
    ))
    fig_forecast.update_layout(
        **CHART_BASE,
        height=300,
        xaxis_title="Month",
        yaxis_title="Dry Powder ($mm)",
        hovermode='x unified'
    )

    # Bite Sizing
    bite_sizes = calculate_bite_sizes(dry_powder, config)
    bite_cards = []
    for strategy, sizes in bite_sizes.items():
        bite_cards.append(html.Div([
            html.H6(strategy, style={'fontSize': '12px', 'fontWeight': 'bold'}),
            html.Small(f"Min: ${sizes['min']:.1f}M ({sizes['min_pct']:.1%})", className="d-block text-muted"),
            html.Small(f"Desired: ${sizes['desired']:.1f}M ({sizes['desired_pct']:.1%})",
                       className="d-block text-success"),
            html.Small(f"Max: ${sizes['max']:.1f}M ({sizes['max_pct']:.1%})", className="d-block text-danger"),
            html.Hr(className="my-2")
        ]))

    return fig_alloc, fig_forecast, html.Div(bite_cards)


# Portfolio Page Metrics
@app.callback(
    [Output('port-nav', 'children'), Output('port-current-commitment', 'children'),
     Output('port-total-commitment', 'children'), Output('port-dry-powder', 'children'),
     Output('port-total-fund', 'children'), Output('port-target-return', 'children')],
    [Input('deals-store', 'data'), Input('config-store', 'data'),
     Input('fund-cf-data-store', 'data'), Input('liquidity-dry-powder-store', 'data')]
)
def update_portfolio_metrics(deals, config, fund_cf_data, dry_powder_uploaded):
    deals_enriched = enrich_deals_with_latest_nav(deals, fund_cf_data)
    m = calculate_portfolio_metrics(deals_enriched)
    total_commitment = sum(d.get('total_commitment', d.get('commitment', d.get('size', 0))) for d in deals_enriched or [])
    current_commitment = sum(d.get('current_commitment', d.get('commitment', d.get('size', 0))) for d in deals_enriched or [])
    dry_powder = get_effective_dry_powder(config, dry_powder_uploaded)
    req_irr = calculate_required_future_irr(m['weighted_irr'], m['total_nav'], dry_powder, config)
    return (
        f"${m['total_nav']:.1f}M",
        f"${current_commitment:.1f}M",
        f"${total_commitment:.1f}M",
        f"${dry_powder:.1f}M",
        f"${m['total_nav'] + dry_powder:.1f}M",
        f"{req_irr:.1%}"
    )


# Calculator Page
@app.callback(
    [Output('calc-current-nav', 'children'), Output('calc-current-irr', 'children'),
     Output('calc-num-deals', 'children'), Output('calc-total-fund', 'children'),
     Output('calc-target-twr', 'children'), Output('calc-dry-powder', 'children'),
     Output('calc-required-irr', 'children'), Output('calc-explanation', 'children'),
     Output('calc-waterfall', 'children')],
    [Input('deals-store', 'data'), Input('config-store', 'data')]
)
def update_calculator(deals, config):
    m = calculate_portfolio_metrics(deals)
    dry_powder = config['fund_parameters']['dry_powder']
    total_fund = m['total_nav'] + dry_powder
    target_twr = config['fund_parameters']['target_net_twr']

    req_irr = calculate_required_future_irr(m['weighted_irr'], m['total_nav'], dry_powder, config)
    gap = req_irr - m['weighted_irr'] if m['total_nav'] > 0 else 0
    explanation = f"{'Higher' if gap > 0 else 'Lower'} than current portfolio by {abs(gap):.1%}" if m[
                                                                                                        'total_nav'] > 0 else "Add deals to calculate"

    # Waterfall
    invested_ratio = max(0, 1 - config['fund_parameters']['liquidity_reserve_pct'])
    idle_ratio = config['fund_parameters']['liquidity_reserve_pct']

    waterfall_steps = [
        ("1. Target Net TWR (to LPs)", f"{target_twr:.1%}"),
        ("2. + Management Fee", f"{config['fund_parameters']['management_fee']:.2%}"),
        ("3. + Loss Drag (impairments)", f"{config['fund_parameters']['loss_drag']:.1%}"),
        ("4. - Cash Yield on reserves", f"-{idle_ratio * config['fund_parameters']['cash_yield']:.2%}"),
        ("5. ÷ Invested Ratio", f"÷ {invested_ratio:.1%}"),
        ("6. + Carry Drag (if > hurdle)", f"+Variable ({config['fund_parameters']['carry_rate']:.0%})"),
        ("═══════════════════════", "═══════════"),
        ("REQUIRED GROSS DEAL RETURN", f"{req_irr:.1%}")
    ]

    waterfall_div = dbc.Table([
        html.Tbody([
            html.Tr([
                html.Td(step, style={'fontSize': '14px', 'fontWeight': 'bold' if '═' in step else 'normal'}),
                html.Td(val, className="text-end", style={
                    'fontSize': '14px',
                    'fontWeight': 'bold',
                    'color': COLORS['success'] if '═' in step else COLORS['dark']
                })
            ]) for step, val in waterfall_steps
        ])
    ], bordered=True, striped=True, hover=True, className="shadow-sm")

    return (
        f"${m['total_nav']:.1f}M",
        f"{m['weighted_irr']:.1%}",
        str(m['num_deals']),
        f"${total_fund:.0f}M",
        f"{target_twr:.1%}",
        f"${dry_powder:.0f}M",
        f"{req_irr:.1%}",
        explanation,
        waterfall_div
    )


# Dry Powder Page
@app.callback(
    [Output('dp-current', 'children'), Output('dp-planned', 'children'),
     Output('dp-distributions', 'children'), Output('dp-forecast-12m', 'children'),
     Output('dp-forecast-chart', 'figure'), Output('dp-monthly-table', 'children')],
    [Input('deals-store', 'data'), Input('placeholder-deals-store', 'data'), Input('config-store', 'data')]
)
def update_drypowder_page(deals, placeholders, config):
    m = calculate_portfolio_metrics(deals)
    dry_powder = config['fund_parameters']['dry_powder']

    total_planned = sum(p['size'] for p in placeholders)
    annual_dist = m['total_nav'] * config['fund_parameters']['distribution_rate']

    forecast = forecast_dry_powder(m['total_nav'], dry_powder, deals, placeholders, config, 12)
    forecast_12m = forecast[-1]['dry_powder']

    # Chart
    months = [f['month'] for f in forecast]
    dp_vals = [f['dry_powder'] for f in forecast]
    calls = [f['calls'] for f in forecast]
    dists = [f['distributions'] for f in forecast]

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(go.Scatter(
        x=months, y=dp_vals, name='Dry Powder',
        line=dict(color=C['blue'], width=3),
        fill='tozeroy',
        fillcolor=rgba(C['blue'], 0.2),
        marker=dict(size=6, color=C['sky'])
    ), secondary_y=False)
    fig.add_trace(go.Bar(
        x=months, y=calls, name='Capital Calls',
        marker_color=C['red'],
        opacity=0.7
    ), secondary_y=True)
    fig.add_trace(go.Bar(
        x=months, y=dists, name='Distributions',
        marker_color=C['green'],
        opacity=0.7
    ), secondary_y=True)

    fig.update_xaxes(title_text="Month", gridcolor=C['border'])
    fig.update_yaxes(title_text="Dry Powder ($mm)", secondary_y=False, gridcolor=C['border'])
    fig.update_yaxes(title_text="Flows ($mm)", secondary_y=True, gridcolor=C['border'])
    fig.update_layout(**CHART_BASE, height=400, hovermode='x unified')

    # Table
    table_data = [{
        'Month': f['month'],
        'Dry Powder': f"${f['dry_powder']:.1f}M",
        'NAV': f"${f['nav']:.1f}M",
        'Distributions': f"${f['distributions']:.1f}M",
        'Calls': f"${f['calls']:.1f}M" if f['calls'] > 0 else "-"
    } for f in forecast]

    table = dash_table.DataTable(
        data=table_data,
        columns=[{"name": c, "id": c} for c in table_data[0].keys()],
        style_cell={'textAlign': 'left', 'padding': '10px', 'fontFamily': C['mono'], 'fontSize': '11px'},
        style_header={
            'backgroundColor': C['surface'],
            'color': C['text'],
            'fontWeight': 'bold',
            'border': f'1px solid {C["border"]}'
        },
        style_data={
            'backgroundColor': C['panel'],
            'color': C['text'],
            'border': f'1px solid {C["border"]}'
        },
        style_data_conditional=[{
            'if': {'row_index': 'odd'},
            'backgroundColor': C['surface']
        }],
        page_size=12
    )

    return (
        f"${dry_powder:.0f}M",
        f"${total_planned:.0f}M",
        f"${annual_dist:.0f}M",
        f"${forecast_12m:.0f}M",
        fig,
        table
    )


# ==================== TWR FORECASTER CALLBACKS ====================

@app.callback(
    Output('twr-current-irr', 'children'),
    Input('deals-store', 'data')
)
def update_twr_current_irr(deals):
    m = calculate_portfolio_metrics(deals)
    return f"{m['weighted_irr']:.1%}"


@app.callback(
    [Output('twr-distribution-chart', 'figure'), Output('twr-statistics', 'children'),
     Output('twr-sensitivity-chart', 'figure')],
    Input('btn-run-twr', 'n_clicks'),
    [State('deals-store', 'data'), State('config-store', 'data'),
     State('twr-future-mean', 'value'), State('twr-future-std', 'value'),
     State('twr-n-sims', 'value')],
    prevent_initial_call=True
)
def run_twr_simulation(n, deals, config, future_mean, future_std, n_sims):
    m = calculate_portfolio_metrics(deals)
    current_irr = m['weighted_irr']
    current_nav = m['total_nav']
    dry_powder = config['fund_parameters']['dry_powder']
    target_twr = config['fund_parameters']['target_net_twr']

    # Run Monte Carlo
    np.random.seed(42)
    future_irrs = np.random.normal(future_mean / 100, future_std / 100, n_sims)

    results = []
    for future_irr in future_irrs:
        # Blended portfolio IRR
        total = current_nav + dry_powder
        if total > 0:
            blended = (current_nav * current_irr + dry_powder * future_irr) / total
        else:
            blended = 0

        # Apply fees and drag
        mgmt_fee = config['fund_parameters']['management_fee']
        loss_drag = config['fund_parameters']['loss_drag']
        liq_reserve = config['fund_parameters']['liquidity_reserve_pct']
        cash_yield = config['fund_parameters']['cash_yield']
        carry_rate = config['fund_parameters']['carry_rate']
        hurdle = config['fund_parameters']['hurdle_rate']

        invested_ratio = 1 - liq_reserve
        gross_before_carry = (blended * invested_ratio) + (cash_yield * liq_reserve) - mgmt_fee - loss_drag

        # Carry drag
        if gross_before_carry > hurdle:
            carry_drag = (gross_before_carry - hurdle) * carry_rate
            net_twr = gross_before_carry - carry_drag
        else:
            net_twr = gross_before_carry

        results.append(net_twr)

    results = np.array(results)

    # Distribution chart
    fig_dist = go.Figure()
    fig_dist.add_trace(go.Histogram(
        x=results * 100, nbinsx=50,
        marker_color=C['blue'], opacity=0.7,
        name='Simulated TWR'
    ))
    fig_dist.add_vline(x=target_twr * 100, line_color=C['red'], line_width=3, line_dash='dash',
                       annotation_text=f"Target: {target_twr:.1%}", annotation_position="top right")
    fig_dist.add_vline(x=np.median(results) * 100, line_color=C['green'], line_width=3,
                       annotation_text=f"Median: {np.median(results):.1%}", annotation_position="top left")
    fig_dist.update_layout(
        **CHART_BASE,
        xaxis_title="Net TWR (%)", yaxis_title="Frequency",
        title=f"Distribution of Net TWR ({n_sims:,} Simulations)",
        height=400
    )

    # Statistics
    prob_above_target = (results >= target_twr).sum() / len(results)
    percentile_5 = np.percentile(results, 5)
    percentile_95 = np.percentile(results, 95)
    sharpe = (np.mean(results) - 0.03) / np.std(results)  # Assuming 3% risk-free

    stats_div = html.Div([
        html.Div([
            html.Strong("Probability of Hitting Target:", style={'color': C['text'], 'fontFamily': C['sans']}),
            html.H3(f"{prob_above_target:.1%}", style={
                'color': C['green'] if prob_above_target >= 0.75 else C['amber'] if prob_above_target >= 0.5 else C[
                    'red'],
                'fontFamily': C['mono'], 'marginTop': '0.5rem'
            })
        ], className="mb-3"),
        html.Hr(style={'borderColor': C['border']}),
        html.Div([
            html.P([html.Strong("Mean TWR: "),
                    html.Span(f"{np.mean(results):.2%}", style={'color': C['blue'], 'fontFamily': C['mono']})]),
            html.P([html.Strong("Median TWR: "),
                    html.Span(f"{np.median(results):.2%}", style={'color': C['green'], 'fontFamily': C['mono']})]),
            html.P([html.Strong("Std Dev: "),
                    html.Span(f"{np.std(results):.2%}", style={'color': C['muted'], 'fontFamily': C['mono']})]),
            html.P([html.Strong("5th Percentile: "),
                    html.Span(f"{percentile_5:.2%}", style={'color': C['red'], 'fontFamily': C['mono']})]),
            html.P([html.Strong("95th Percentile: "),
                    html.Span(f"{percentile_95:.2%}", style={'color': C['green'], 'fontFamily': C['mono']})]),
            html.P([html.Strong("Sharpe Ratio: "),
                    html.Span(f"{sharpe:.2f}", style={'color': C['purple'], 'fontFamily': C['mono']})]),
        ], style={'fontSize': '14px'})
    ])

    # Sensitivity Analysis
    future_irr_range = np.linspace(0.15, 0.35, 20)
    twr_sensitivity = []
    for fir in future_irr_range:
        total = current_nav + dry_powder
        blended = (current_nav * current_irr + dry_powder * fir) / total if total > 0 else 0
        invested_ratio = 1 - liq_reserve
        gross = (blended * invested_ratio) + (cash_yield * liq_reserve) - mgmt_fee - loss_drag
        if gross > hurdle:
            net = gross - (gross - hurdle) * carry_rate
        else:
            net = gross
        twr_sensitivity.append(net)

    fig_sens = go.Figure()
    fig_sens.add_trace(go.Scatter(
        x=future_irr_range * 100, y=np.array(twr_sensitivity) * 100,
        mode='lines+markers', name='Net TWR',
        line=dict(color=C['blue'], width=3),
        marker=dict(size=6, color=C['sky'])
    ))
    fig_sens.add_hline(y=target_twr * 100, line_color=C['red'], line_dash='dash',
                       annotation_text=f"Target: {target_twr:.1%}")
    fig_sens.update_layout(
        **CHART_BASE,
        xaxis_title="Future Deal IRR (%)", yaxis_title="Net TWR (%)",
        title="How Net TWR Changes with Future Deal Performance",
        height=400
    )

    return fig_dist, stats_div, fig_sens


# ==================== PORTFOLIO SEGMENTATION CALLBACKS ====================

# Segment Summary Cards
@app.callback(
    [Output('seg-total-nav', 'children'), Output('seg-total-twr', 'children'),
     Output('seg-seed-nav', 'children'), Output('seg-seed-twr', 'children'),
     Output('seg-new-nav', 'children'), Output('seg-new-twr', 'children'),
     Output('seg-mm-nav', 'children'), Output('seg-mm-twr', 'children'),
     Output('seg-pipeline-nav', 'children'), Output('seg-pipeline-twr', 'children'),
     Output('seg-future-nav', 'children'), Output('seg-future-twr', 'children')],
    [Input('deals-store', 'data'), Input('pipeline-store', 'data'),
     Input('placeholder-deals-store', 'data')]
)
def update_segment_summary(deals, pipeline, placeholders):
    """Calculate NAV and TWR for each segment"""

    # Segment deals
    seed_deals = [d for d in deals if d.get('segment', 'Seed') == 'Seed'] if deals else []
    new_deals = [d for d in deals if d.get('segment', 'Seed') == 'New'] if deals else []
    mm_deals = [d for d in deals if d.get('segment', 'Seed') == 'MoneyMarket'] if deals else []

    # Calculate NAVs
    seed_nav = sum(d['size'] for d in seed_deals)
    new_nav = sum(d['size'] for d in new_deals)
    mm_nav = sum(d['size'] for d in mm_deals)
    pipeline_nav = sum(p['size'] for p in pipeline) if pipeline else 0
    future_nav = sum(p['size'] for p in placeholders) if placeholders else 0
    total_nav = seed_nav + new_nav + mm_nav

    # Calculate TWRs (simplified - would use actual cashflows in production)
    seed_twr = sum(d['target_irr'] * d['size'] for d in seed_deals) / seed_nav if seed_nav > 0 else 0
    new_twr = sum(d['target_irr'] * d['size'] for d in new_deals) / new_nav if new_nav > 0 else 0
    mm_twr = 0.03  # Money market assumed 3%
    pipeline_twr = sum(
        p['target_irr'] * p['size'] for p in pipeline) / pipeline_nav if pipeline_nav > 0 and pipeline else 0
    future_twr = 0.25  # Placeholder assumed target
    total_twr = (seed_nav * seed_twr + new_nav * new_twr + mm_nav * mm_twr) / total_nav if total_nav > 0 else 0

    return (
        f"${total_nav:.1f}M", f"TWR: {total_twr:.1%}",
        f"${seed_nav:.1f}M", f"TWR: {seed_twr:.1%}",
        f"${new_nav:.1f}M", f"TWR: {new_twr:.1%}",
        f"${mm_nav:.1f}M", f"TWR: {mm_twr:.1%}",
        f"${pipeline_nav:.1f}M", f"Est TWR: {pipeline_twr:.1%}",
        f"${future_nav:.1f}M", f"Target TWR: {future_twr:.1%}"
    )


# TWR Forecast Chart
@app.callback(
    Output('seg-twr-forecast-chart', 'figure'),
    [Input('deals-store', 'data'), Input('config-store', 'data')]
)
def generate_twr_forecast_by_segment(deals, config):
    """Generate 12-month TWR forecast for each segment"""

    base_date = datetime(2026, 1, 1)
    months = []

    # Initialize TWR arrays for each segment
    total_twr = []
    seed_twr = []
    new_twr = []
    mm_twr = []

    target_twr = config['fund_parameters']['target_net_twr']

    for i in range(12):
        month_date = base_date + relativedelta(months=i)
        months.append(month_date.strftime('%b %Y'))

        # Cumulative TWR calculation (simplified)
        cumulative_factor = (1 + target_twr) ** (i / 12)

        # Different growth rates per segment
        total_twr.append((cumulative_factor - 1) * 100)
        seed_twr.append((cumulative_factor * 0.95 - 1) * 100)  # Slightly lower
        new_twr.append((cumulative_factor * 1.05 - 1) * 100)  # Slightly higher
        mm_twr.append((((1 + 0.03) ** (i / 12)) - 1) * 100)  # 3% fixed

    fig = go.Figure()

    fig.add_trace(go.Scatter(
        x=months, y=total_twr, name='Total Portfolio',
        line=dict(color=C['blue'], width=3), mode='lines+markers'
    ))
    fig.add_trace(go.Scatter(
        x=months, y=seed_twr, name='Seed Portfolio',
        line=dict(color=C['green'], width=2, dash='dash'), mode='lines'
    ))
    fig.add_trace(go.Scatter(
        x=months, y=new_twr, name='New Deals',
        line=dict(color=C['purple'], width=2, dash='dot'), mode='lines'
    ))
    fig.add_trace(go.Scatter(
        x=months, y=mm_twr, name='Money Market',
        line=dict(color=C['amber'], width=2, dash='dashdot'), mode='lines'
    ))

    fig.update_layout(
        **CHART_BASE,
        yaxis_title='Cumulative TWR (%)',
        height=400,
        hovermode='x unified',
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1)
    )

    return fig


# Segment Tables
@app.callback(
    [Output('seg-seed-table', 'children'), Output('seg-new-table', 'children')],
    Input('deals-store', 'data')
)
def update_segment_tables(deals):
    """Display deals by segment"""

    if not deals:
        return (html.P("No seed deals", className="text-muted"),
                html.P("No new deals", className="text-muted"))

    seed_deals = [d for d in deals if d.get('segment', 'Seed') == 'Seed']
    new_deals = [d for d in deals if d.get('segment', 'Seed') == 'New']

    def create_segment_table(segment_deals, segment_name):
        if not segment_deals:
            return html.P(f"No {segment_name.lower()} yet", className="text-muted")

        data = [{
            'Deal': d['name'],
            'NAV': f"${d['size']:.1f}M",
            'IRR': f"{d['target_irr']:.1%}",
            'Vintage': d.get('vintage', 'N/A')
        } for d in segment_deals]

        return dash_table.DataTable(
            data=data,
            columns=[{"name": c, "id": c} for c in data[0].keys()],
            style_cell={'textAlign': 'left', 'padding': '10px', 'fontFamily': C['mono'], 'fontSize': '11px'},
            style_header={'backgroundColor': C['surface'], 'color': C['text'], 'fontWeight': 'bold'},
            style_data={'backgroundColor': C['panel'], 'color': C['text']},
            style_data_conditional=[{'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']}],
            page_size=10
        )

    return (create_segment_table(seed_deals, 'Seed'),
            create_segment_table(new_deals, 'New'))


# Allocation Chart
@app.callback(
    Output('seg-allocation-chart', 'figure'),
    [Input('deals-store', 'data'), Input('pipeline-store', 'data'),
     Input('placeholder-deals-store', 'data')]
)
def generate_allocation_chart(deals, pipeline, placeholders):
    """Show NAV allocation across segments"""

    # Calculate segment NAVs
    seed_nav = sum(d['size'] for d in deals if d.get('segment', 'Seed') == 'Seed') if deals else 0
    new_nav = sum(d['size'] for d in deals if d.get('segment', 'Seed') == 'New') if deals else 0
    mm_nav = sum(d['size'] for d in deals if d.get('segment', 'Seed') == 'MoneyMarket') if deals else 0
    pipeline_nav = sum(p['size'] for p in pipeline) if pipeline else 0
    future_nav = sum(p['size'] for p in placeholders) if placeholders else 0

    labels = []
    values = []
    colors = []

    if seed_nav > 0:
        labels.append('Seed Portfolio')
        values.append(seed_nav)
        colors.append(C['green'])
    if new_nav > 0:
        labels.append('New Deals')
        values.append(new_nav)
        colors.append(C['purple'])
    if mm_nav > 0:
        labels.append('Money Market')
        values.append(mm_nav)
        colors.append(C['amber'])
    if pipeline_nav > 0:
        labels.append('Pipeline')
        values.append(pipeline_nav)
        colors.append(C['teal'])
    if future_nav > 0:
        labels.append('Future/Placeholder')
        values.append(future_nav)
        colors.append(C['pink'])

    fig = go.Figure(data=[go.Pie(
        labels=labels, values=values, hole=0.4,
        marker=dict(colors=colors),
        textfont=dict(color=C['text'], family=C['mono'])
    )])

    fig.update_layout(**CHART_BASE, height=350, margin=dict(t=20, b=0, l=0, r=0), showlegend=True)

    return fig


# Contribution Chart
@app.callback(
    Output('seg-contribution-chart', 'figure'),
    Input('deals-store', 'data')
)
def generate_contribution_chart(deals):
    """Show TWR contribution by segment"""

    if not deals:
        fig = go.Figure()
        fig.update_layout(**CHART_BASE)
        return fig

    # Calculate weighted TWR contribution
    seed_deals = [d for d in deals if d.get('segment', 'Seed') == 'Seed']
    new_deals = [d for d in deals if d.get('segment', 'Seed') == 'New']
    mm_deals = [d for d in deals if d.get('segment', 'Seed') == 'MoneyMarket']

    seed_contribution = sum(d['target_irr'] * d['size'] for d in seed_deals)
    new_contribution = sum(d['target_irr'] * d['size'] for d in new_deals)
    mm_contribution = sum(0.03 * d['size'] for d in mm_deals)

    segments = []
    contributions = []
    colors_list = []

    if seed_contribution > 0:
        segments.append('Seed')
        contributions.append(seed_contribution * 100)
        colors_list.append(C['green'])
    if new_contribution > 0:
        segments.append('New')
        contributions.append(new_contribution * 100)
        colors_list.append(C['purple'])
    if mm_contribution > 0:
        segments.append('MM')
        contributions.append(mm_contribution * 100)
        colors_list.append(C['amber'])

    fig = go.Figure(data=[go.Bar(
        x=segments, y=contributions, marker_color=colors_list,
        text=[f"{v:.1f}" for v in contributions], textposition='outside',
        textfont=dict(color=C['text'], family=C['mono'])
    )])

    fig.update_layout(**CHART_BASE, yaxis_title="TWR Contribution", height=350)

    return fig



# ==================== CASHFLOWS CALLBACKS ====================

# Upload Fund CF Excel file
@app.callback(
    [Output('fund-cf-data-store', 'data'), Output('upload-fund-cf-status', 'children')],
    Input('upload-fund-cf', 'contents'),
    State('upload-fund-cf', 'filename'),
    prevent_initial_call=True
)
def upload_fund_cf_file(contents, filename):
    """
    Parse uploaded Fund Level CF Excel file matching exact template structure:
    - Row 1: Section labels at fixed cols — Net CF (col 14), Called Capital (col 190),
              Distributions (col 364), NAV (col 538)
    - Row 2: Month-end dates (datetime) across all sections
    - Row 3: Header row — AIC Name(5), Investments(6), Portfolio Type(7),
              Investment Type(8), Investment Status(9), Commitment Year(10),
              Commitment($m)(11), Paid In(12), Unfunded(13)
    - Data rows 7–29: Actual investments (skip rows 30+ which are placeholders/totals)
    - Rows 49+: Total rows — excluded
    - Rows 59–64: Liquidity sleeve — excluded
    Unfunded values in col 13 are raw dollars; convert to millions.
    """
    if contents is None:
        return None, ""

    try:
        content_type, content_string = contents.split(',')
        decoded = base64.b64decode(content_string)
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(decoded), data_only=True)

        # Find the Fund Level CF sheet
        target_sheets = ['Fund Level CF', 'Fund_Level_CF', 'FundLevelCF', 'Sheet1']
        ws = None
        for sname in target_sheets:
            if sname in wb.sheetnames:
                ws = wb[sname]
                break
        if ws is None:
            ws = wb[wb.sheetnames[0]]

        # ── Section column boundaries ──────────────────────────────────────────
        # Detect from Row 1 labels; fall back to known template offsets
        SECTION_LABELS = {
            'Net CF': 'net_cf',
            'Called Capital': 'calls',
            'Distributions': 'distributions',
            'NAV': 'nav',
        }
        section_start_cols = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(1, col).value
            if val in SECTION_LABELS:
                section_start_cols[SECTION_LABELS[val]] = col

        # If detection fails, use known template offsets
        if not section_start_cols:
            section_start_cols = {
                'net_cf': 14,
                'calls': 190,
                'distributions': 364,
                'nav': 538,
            }

        section_order = ['net_cf', 'calls', 'distributions', 'nav']
        section_bounds = {}
        for i, key in enumerate(section_order):
            start = section_start_cols.get(key)
            if start is None:
                continue
            next_starts = [section_start_cols[k] for k in section_order[i+1:] if k in section_start_cols]
            end = (min(next_starts) - 1) if next_starts else ws.max_column
            section_bounds[key] = (start, end)

        # ── Build month-date lookup from Row 2 (use Net CF section range) ─────
        net_start, net_end = section_bounds.get('net_cf', (14, 189))
        month_col_to_key = {}   # col → 'YYYY-MM-DD' using Net CF section cols
        all_month_keys_ordered = []
        for col in range(net_start, net_end + 1):
            val = ws.cell(2, col).value
            if val is None:
                continue
            try:
                mk = pd.to_datetime(val).strftime('%Y-%m-%d')
                month_col_to_key[col] = mk
                if mk not in all_month_keys_ordered:
                    all_month_keys_ordered.append(mk)
            except Exception:
                pass

        # For non-Net-CF sections the col offset to the month date row is the same;
        # we compute each section's col → month key using the same row-2 dates
        def build_section_month_map(start_col, end_col):
            """Map absolute col → 'YYYY-MM-DD' within a section."""
            result = {}
            for col in range(start_col, end_col + 1):
                val = ws.cell(2, col).value
                if val is None:
                    continue
                try:
                    result[col] = pd.to_datetime(val).strftime('%Y-%m-%d')
                except Exception:
                    pass
            return result

        section_month_maps = {}
        for key, (s, e) in section_bounds.items():
            section_month_maps[key] = build_section_month_map(s, e)

        # ── Row classification ────────────────────────────────────────────────
        SKIP_PREFIXES = ('total ', 'cash', 'liquidity')
        SKIP_NAMES = {'AIC Name', 'Investments', 'Investment Type', 'Investment Status',
                      'Commitment Year', 'Commitment ($m)', 'Total Liquidity'}

        def is_investment_row(row_idx):
            name = ws.cell(row_idx, 5).value
            if not name:
                return False
            name_str = str(name).strip()
            if not name_str:
                return False
            if name_str in SKIP_NAMES:
                return False
            if name_str.lower().startswith(SKIP_PREFIXES):
                return False
            # Skip pure-placeholder rows (portfolio_type == 'Placeholder') if desired
            # Actually we INCLUDE placeholders as they appear in the template
            return True

        # ── Parse each investment row ─────────────────────────────────────────
        deals_data = []
        for row in range(7, ws.max_row + 1):
            if not is_investment_row(row):
                continue

            deal_name = str(ws.cell(row, 5).value).strip()
            portfolio_type = str(ws.cell(row, 7).value or '').strip()
            investment_type = str(ws.cell(row, 8).value or '').strip()
            investment_status = str(ws.cell(row, 9).value or '').strip()
            commitment_year = ws.cell(row, 10).value
            commitment_raw = ws.cell(row, 11).value
            paid_in_raw = ws.cell(row, 12).value
            unfunded_raw = ws.cell(row, 13).value

            # Commitment is already in millions in the template
            commitment_m = float(commitment_raw) if isinstance(commitment_raw, (int, float)) else 0.0
            # Paid-in may be absent; default to 0
            paid_in_m = float(paid_in_raw) if isinstance(paid_in_raw, (int, float)) else 0.0
            # Unfunded is stored in raw dollars in col 13 — convert to millions
            if isinstance(unfunded_raw, (int, float)):
                unfunded_m = float(unfunded_raw) / 1_000_000 if abs(float(unfunded_raw)) > 1000 else float(unfunded_raw)
            else:
                unfunded_m = 0.0

            deal_data = {
                'name': deal_name,
                'investments': str(ws.cell(row, 6).value or deal_name).strip(),
                'portfolio_type': portfolio_type,
                'type': investment_type,
                'investment_status': investment_status,
                'commitment_year': commitment_year,
                'commitment': commitment_m,
                'paid_in': paid_in_m,
                'unfunded': unfunded_m,
                'current_commitment': paid_in_m if paid_in_m > 0 else max(commitment_m - unfunded_m, 0),
                'nav': 0.0,
                'net_cf_series': {},
                'calls_series': {},
                'distributions_series': {},
                'nav_series': {},
            }

            # Read each cashflow section
            series_map = {
                'net_cf': 'net_cf_series',
                'calls': 'calls_series',
                'distributions': 'distributions_series',
                'nav': 'nav_series',
            }
            for section_key, store_key in series_map.items():
                if section_key not in section_bounds:
                    continue
                s_col, e_col = section_bounds[section_key]
                month_map = section_month_maps[section_key]
                for col in range(s_col, e_col + 1):
                    mk = month_map.get(col)
                    if not mk:
                        continue
                    cell_val = ws.cell(row, col).value
                    if isinstance(cell_val, (int, float)):
                        num = float(cell_val)
                    else:
                        num = 0.0
                    deal_data[store_key][mk] = num

            # Derive latest NAV (last non-zero NAV up to last month-end)
            nav_series = deal_data['nav_series']
            latest_nav = get_latest_available_nav_for_deal([deal_data], deal_name)
            if latest_nav is None:
                non_zero = [(mk, v) for mk, v in sorted(nav_series.items()) if v not in (None, 0, 0.0)]
                latest_nav = non_zero[-1][1] if non_zero else commitment_m
            deal_data['nav'] = float(latest_nav or 0)

            deals_data.append(deal_data)

        # Store ordered month keys globally for table rendering
        # (attach to first deal so the table callback can use them)
        if deals_data:
            deals_data[0]['_all_month_keys'] = all_month_keys_ordered

        n_investments = sum(1 for d in deals_data if d.get('portfolio_type', '').lower() != 'placeholder')
        n_placeholders = sum(1 for d in deals_data if d.get('portfolio_type', '').lower() == 'placeholder')

        success_msg = dbc.Alert([
            html.I(className="fas fa-check-circle me-2"),
            f"✅ Uploaded {filename} — {n_investments} investments + {n_placeholders} placeholders parsed "
            f"({len(all_month_keys_ordered)} months, "
            f"{pd.to_datetime(all_month_keys_ordered[0]).strftime('%b %Y') if all_month_keys_ordered else '?'} – "
            f"{pd.to_datetime(all_month_keys_ordered[-1]).strftime('%b %Y') if all_month_keys_ordered else '?'})"
        ], color="success", className="mt-2")
        return deals_data, success_msg

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Fund CF upload error:\n{error_details}")
        error_msg = dbc.Alert([
            html.I(className="fas fa-exclamation-triangle me-2"),
            html.Div([
                f"❌ Error uploading {filename}: {str(e)}",
                html.Br(),
                html.Small("Expected: Excel (.xlsx/.xlsm) with 'Fund Level CF' sheet matching the Horizon template.",
                           className="text-muted"),
            ])
        ], color="danger", className="mt-2")
        return None, error_msg


# Summary metrics
@app.callback(
    [Output('cf-total-commitment', 'children'), Output('cf-total-paid-in', 'children'),
     Output('cf-total-unfunded', 'children'), Output('cf-current-nav', 'children')],
    [Input('deals-store', 'data'), Input('fund-cf-data-store', 'data')]
)
def update_cf_summary(deals, uploaded_data):
    if uploaded_data:
        total_commitment = sum(float(d.get('commitment', 0) or 0) for d in uploaded_data)
        total_paid_in = sum(float(d.get('paid_in', d.get('current_commitment', 0)) or 0) for d in uploaded_data)
        total_unfunded = sum(float(d.get('unfunded', 0) or 0) for d in uploaded_data)
        current_nav = sum(float(get_latest_available_nav_for_deal(uploaded_data, d.get('name')) or d.get('nav', 0) or 0) for d in uploaded_data)
    elif deals:
        deals_enriched = enrich_deals_with_latest_nav(deals, None)
        total_commitment = sum(d.get('total_commitment', d.get('commitment', d['size'])) for d in deals_enriched)
        total_paid_in = sum(d.get('current_commitment', d.get('commitment', d['size'])) for d in deals_enriched)
        total_unfunded = sum(d.get('unfunded', max(d.get('total_commitment', d.get('commitment', d['size'])) - d.get('current_commitment', d.get('commitment', d['size'])), 0)) for d in deals_enriched)
        current_nav = sum(d.get('nav', d.get('size', 0)) for d in deals_enriched)
    else:
        return "$0.0M", "$0.0M", "$0.0M", "$0.0M"

    return (f"${total_commitment:.1f}M", f"${total_paid_in:.1f}M",
            f"${total_unfunded:.1f}M", f"${current_nav:.1f}M")


def _resolve_display_window(uploaded_data, start_month_val, num_months):
    """
    Given uploaded deal data, a start-month ISO string, and a horizon in months,
    return (display_months, month_labels, start_idx) where:
      display_months  = list of 'YYYY-MM-DD' keys in that window
      month_labels    = list of 'Mon-YY' strings for column headers
      start_idx       = index into all_months_ordered where the window begins
    Falls back gracefully when the start month isn't in the data.
    """
    all_months = (
        uploaded_data[0].get('_all_month_keys') or
        sorted({mk for d in uploaded_data
                for mk in (d.get('net_cf_series') or {}).keys()})
    ) if uploaded_data else []

    if not all_months:
        return [], [], 0

    # Find the closest month >= start_month_val
    start_idx = 0
    if start_month_val:
        try:
            target = pd.to_datetime(start_month_val)
            for i, mk in enumerate(all_months):
                if pd.to_datetime(mk) >= target:
                    start_idx = i
                    break
            else:
                start_idx = max(0, len(all_months) - num_months)
        except Exception:
            start_idx = 0

    window = all_months[start_idx: start_idx + num_months]
    labels = [pd.to_datetime(m).strftime("%b-%y") for m in window]
    return window, labels, start_idx


# Window badge callback
@app.callback(
    [Output('cf-window-badge', 'children'),
     Output('cf-table-header', 'children'),
     Output('cf-chart-header', 'children')],
    [Input('cf-start-month', 'value'),
     Input('cf-horizon', 'value'),
     Input('fund-cf-data-store', 'data')],
    prevent_initial_call=False
)
def update_cf_window_info(start_month, horizon, uploaded_data):
    horizon = int(horizon or 12)
    window, labels, _ = _resolve_display_window(uploaded_data, start_month, horizon)

    if window:
        first = pd.to_datetime(window[0]).strftime("%b %Y")
        last  = pd.to_datetime(window[-1]).strftime("%b %Y")
        n     = len(window)
        badge = dbc.Alert([
            html.I(className="fas fa-calendar-check me-2"),
            html.Strong(f"Window: {first} → {last}  "),
            dbc.Badge(f"{n} months", color="primary", className="me-2"),
            dbc.Badge(f"{horizon}m horizon", color="secondary"),
        ], color="light", className="mb-0 py-2",
           style={'border': f'1px solid {C["border"]}', 'fontSize': '13px'})
        tbl_hdr = f"Fund Level Cashflows — {first} to {last} ({n} months)"
        cht_hdr = f"Monthly Totals — {first} to {last}"
    else:
        badge = dbc.Alert("Upload a Fund Level CF file to activate the view window.",
                          color="warning", className="mb-0 py-2")
        tbl_hdr = "Fund Level Cashflows"
        cht_hdr = "Monthly Cashflow Totals"

    return badge, tbl_hdr, cht_hdr


# Monthly Cashflow Table
@app.callback(
    Output('cf-monthly-table', 'children'),
    [Input('deals-store', 'data'), Input('fund-cf-data-store', 'data'),
     Input('cf-start-month', 'value'), Input('cf-horizon', 'value'),
     Input('cf-sections', 'value')]
)
def generate_monthly_cf_table(deals, uploaded_data, start_month, horizon, sections):
    """
    Mirror the exact 4-section Fund Level CF template layout:
      • Section 1 – Net CF
      • Section 2 – Called Capital
      • Section 3 – Distributions
      • Section 4 – NAV
    Window: from cf-start-month for cf-horizon months.
    Sections toggled by cf-sections checklist.
    """
    num_months = int(horizon or 12)
    sections = sections or ["net", "calls", "dists", "nav"]

    META_COLS = [
        'AIC Name', 'Investments', 'Portfolio Type', 'Investment Type',
        'Investment Status', 'Commitment Year', 'Commitment ($m)', 'Paid In', 'Unfunded',
    ]

    def fmt(val, decimals=2):
        try:
            v = float(val)
        except (TypeError, ValueError):
            return "-"
        if v == 0.0:
            return "-"
        return f"{v:+.{decimals}f}" if v < 0 else f"{v:.{decimals}f}"

    def fmt_nav(val):
        try:
            v = float(val)
        except (TypeError, ValueError):
            return "-"
        return "-" if v == 0.0 else f"{v:.2f}"

    def build_section_table(title, series_key, display_months, month_labels, formatter):
        all_cols = META_COLS + month_labels
        rows = []
        col_totals = {lbl: 0.0 for lbl in month_labels}

        for deal in uploaded_data:
            if deal.get('name') in ('_all_month_keys',):
                continue
            series = deal.get(series_key) or {}
            row = {
                'AIC Name':          deal.get('name', ''),
                'Investments':       deal.get('investments', deal.get('name', '')),
                'Portfolio Type':    deal.get('portfolio_type', ''),
                'Investment Type':   deal.get('type', ''),
                'Investment Status': deal.get('investment_status', ''),
                'Commitment Year':   str(deal.get('commitment_year', '') or ''),
                'Commitment ($m)':   f"{float(deal.get('commitment', 0) or 0):.2f}",
                'Paid In':           f"{float(deal.get('paid_in', deal.get('current_commitment', 0)) or 0):.2f}",
                'Unfunded':          f"{float(deal.get('unfunded', 0) or 0):.2f}",
            }
            for mk, lbl in zip(display_months, month_labels):
                v = float(series.get(mk) or 0)
                row[lbl] = formatter(v)
                col_totals[lbl] += v
            rows.append(row)

        # Totals row
        total_commitment = sum(float(d.get('commitment', 0) or 0) for d in uploaded_data)
        total_paid_in    = sum(float(d.get('paid_in', d.get('current_commitment', 0)) or 0)
                               for d in uploaded_data)
        total_unfunded   = sum(float(d.get('unfunded', 0) or 0) for d in uploaded_data)
        totals_row = {
            'AIC Name': 'TOTAL', 'Investments': '', 'Portfolio Type': '',
            'Investment Type': '', 'Investment Status': '', 'Commitment Year': '',
            'Commitment ($m)': f"{total_commitment:.2f}",
            'Paid In': f"{total_paid_in:.2f}",
            'Unfunded': f"{total_unfunded:.2f}",
        }
        for lbl in month_labels:
            totals_row[lbl] = formatter(col_totals[lbl])
        rows.append(totals_row)

        n_rows = len(rows)
        style_cond = [
            {'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']},
            {'if': {'row_index': n_rows - 1},
             'backgroundColor': C['surface'], 'fontWeight': 'bold', 'color': C['sky']},
            {'if': {'column_id': 'AIC Name'}, 'fontWeight': 'bold', 'minWidth': '160px'},
            {'if': {'column_id': 'Investments'}, 'minWidth': '160px'},
        ]
        if series_key != 'nav_series':
            style_cond += [
                {'if': {'filter_query': f'{{{lbl}}} contains "-"', 'column_id': lbl},
                 'color': C['red']}
                for lbl in month_labels
            ]

        section_colors = {
            'Net CF':         C['blue'],
            'Called Capital': C['red'],
            'Distributions':  C['green'],
            'NAV':            C['purple'],
        }
        hdr_color = section_colors.get(title, C['text'])

        columns = [{'name': c, 'id': c} for c in all_cols]

        return dbc.Card([
            dbc.CardHeader(
                html.Div([
                    html.Span("■ ", style={'color': hdr_color, 'fontSize': '16px'}),
                    html.Span(title, style={'fontWeight': 'bold', 'fontSize': '14px',
                                           'color': hdr_color}),
                    html.Span(
                        f"  ({len(rows)-1} deals · {len(month_labels)} months)",
                        style={'fontSize': '11px', 'color': C['muted'], 'marginLeft': '10px'}
                    ),
                ]),
                style={'backgroundColor': C['surface'],
                       'borderBottom': f'2px solid {hdr_color}'}
            ),
            dbc.CardBody([
                dash_table.DataTable(
                    data=rows,
                    columns=columns,
                    style_cell={
                        'textAlign': 'right', 'padding': '6px 10px',
                        'fontFamily': C['mono'], 'fontSize': '11px',
                        'minWidth': '80px', 'maxWidth': '140px',
                        'whiteSpace': 'normal',
                    },
                    style_cell_conditional=[
                        {'if': {'column_id': c},
                         'textAlign': 'left', 'minWidth': '140px'}
                        for c in ['AIC Name', 'Investments', 'Portfolio Type',
                                  'Investment Type', 'Investment Status']
                    ],
                    style_header={
                        'backgroundColor': C['surface'], 'color': C['text'],
                        'fontWeight': 'bold', 'border': f'1px solid {C["border"]}',
                        'position': 'sticky', 'top': 0, 'zIndex': 1,
                        'textAlign': 'center',
                    },
                    style_data={
                        'backgroundColor': C['panel'], 'color': C['text'],
                        'border': f'1px solid {C["border"]}',
                    },
                    style_data_conditional=style_cond,
                    style_table={'overflowX': 'auto', 'maxHeight': '460px',
                                 'overflowY': 'auto'},
                    fixed_columns={'headers': True, 'data': 2},
                    fixed_rows={'headers': True},
                    page_action='none',
                    export_format='xlsx',
                    export_headers='display',
                )
            ], style={'padding': '0'})
        ], className="shadow-sm mb-3",
           style={'borderTop': f'2px solid {hdr_color}'})

    # ── UPLOADED DATA PATH ────────────────────────────────────────────────
    if uploaded_data:
        display_months, month_labels, _ = _resolve_display_window(
            uploaded_data, start_month, num_months)

        if not display_months:
            return dbc.Alert("No data found in the selected window. "
                             "Try a different start month.", color="warning")

        section_map = [
            ("net",   "Net CF",         'net_cf_series',       fmt),
            ("calls", "Called Capital", 'calls_series',         fmt),
            ("dists", "Distributions",  'distributions_series', fmt),
            ("nav",   "NAV",            'nav_series',           fmt_nav),
        ]
        built = [build_section_table(title, key, display_months, month_labels, formatter)
                 for tag, title, key, formatter in section_map
                 if tag in sections]

        return html.Div([
            dbc.Alert([
                html.I(className="fas fa-info-circle me-2"),
                html.Strong("Tip: "),
                "Values in $m. Negative cashflows in red. "
                "Each section has its own Excel export button."
            ], color="info", className="mb-3", style={"fontSize": "12px"}),
            *built
        ])

    # ── FALLBACK ─────────────────────────────────────────────────────────
    if not deals:
        return dbc.Alert(
            "Upload your Fund Level CF Excel file above to see the full cashflow matrix, "
            "or add deals on the Current Portfolio page.",
            color="info"
        )

    deals_enriched = enrich_deals_with_latest_nav(deals, None)
    fallback_rows = []
    for d in deals_enriched:
        fallback_rows.append({
            'AIC Name':          d.get('name', ''),
            'Investments':       d.get('name', ''),
            'Portfolio Type':    d.get('segment', ''),
            'Investment Type':   d.get('strategy', ''),
            'Investment Status': d.get('allocation_status', 'Executed'),
            'Commitment Year':   str(d.get('vintage', '') or ''),
            'Commitment ($m)':   f"{float(d.get('total_commitment', d.get('commitment', d.get('size', 0))) or 0):.2f}",
            'Paid In':           f"{float(d.get('current_commitment', d.get('commitment', d.get('size', 0))) or 0):.2f}",
            'Unfunded':          f"{float(d.get('unfunded', 0) or 0):.2f}",
            'Current NAV':       f"{float(d.get('nav', d.get('size', 0)) or 0):.2f}",
        })
    return build_cashflow_template_table(
        "Current Portfolio Snapshot (no CF file uploaded)",
        fallback_rows,
        list(fallback_rows[0].keys()) if fallback_rows else []
    )


# Monthly Forecast Chart
@app.callback(
    Output('cf-monthly-chart', 'figure'),
    [Input('deals-store', 'data'), Input('fund-cf-data-store', 'data'),
     Input('cf-start-month', 'value'), Input('cf-horizon', 'value'),
     Input('cf-sections', 'value')]
)
def generate_monthly_cf_chart(deals, uploaded_data, start_month, horizon, sections):
    """Stacked bar chart: Calls (red) + Distributions (green) + Net CF line (amber) + NAV line (blue)."""
    import plotly.graph_objects as go

    num_months = int(horizon or 12)
    sections   = sections or ["net", "calls", "dists", "nav"]
    fig = go.Figure()

    if uploaded_data:
        display_months, month_labels, _ = _resolve_display_window(
            uploaded_data, start_month, num_months)
        if not display_months:
            fig.update_layout(**CHART_BASE, height=320)
            return fig

        month_labels_long = [pd.to_datetime(m).strftime("%b %Y") for m in display_months]

        calls  = [sum(float((d.get('calls_series') or {}).get(mk, 0) or 0)
                      for d in uploaded_data) for mk in display_months]
        dists  = [sum(float((d.get('distributions_series') or {}).get(mk, 0) or 0)
                      for d in uploaded_data) for mk in display_months]
        net_cf = [sum(float((d.get('net_cf_series') or {}).get(mk, 0) or 0)
                      for d in uploaded_data) for mk in display_months]
        nav    = [sum(float((d.get('nav_series') or {}).get(mk, 0) or 0)
                      for d in uploaded_data) for mk in display_months]
    else:
        base_date = datetime.now().replace(day=1)
        month_labels_long = [
            (base_date + relativedelta(months=i)).strftime("%b %Y")
            for i in range(num_months)
        ]
        total_nav = sum(d.get('nav', d.get('size', 0)) for d in (deals or []))
        calls  = [0] * num_months
        dists  = [0] * num_months
        net_cf = [0] * num_months
        nav    = [total_nav] * num_months

    if "calls" in sections:
        fig.add_trace(go.Bar(
            x=month_labels_long, y=calls, name='Called Capital',
            marker_color=C['red'], opacity=0.85,
            hovertemplate='%{x}<br>Calls: $%{y:.2f}M<extra></extra>'
        ))
    if "dists" in sections:
        fig.add_trace(go.Bar(
            x=month_labels_long, y=dists, name='Distributions',
            marker_color=C['green'], opacity=0.85,
            hovertemplate='%{x}<br>Dists: $%{y:.2f}M<extra></extra>'
        ))
    if "net" in sections:
        fig.add_trace(go.Scatter(
            x=month_labels_long, y=net_cf, name='Net CF',
            line=dict(color=C['amber'], width=2, dash='dot'),
            mode='lines+markers', marker=dict(size=4),
            hovertemplate='%{x}<br>Net CF: $%{y:.2f}M<extra></extra>'
        ))
    if "nav" in sections:
        fig.add_trace(go.Scatter(
            x=month_labels_long, y=nav, name='Ending NAV',
            line=dict(color=C['blue'], width=3),
            marker=dict(size=4, color=C['sky']),
            yaxis='y2',
            hovertemplate='%{x}<br>NAV: $%{y:.2f}M<extra></extra>'
        ))

    fig.update_layout(
        **CHART_BASE,
        barmode='group',
        height=440,
        yaxis=dict(
            title='Cashflows ($m)', gridcolor=C['border'],
            zeroline=True, zerolinecolor=C['border2']
        ),
        yaxis2=dict(
            title='NAV ($m)', overlaying='y', side='right',
            gridcolor=C['border'], showgrid=False
        ),
        hovermode='x unified',
        legend=dict(orientation='h', yanchor='bottom', y=1.02,
                    xanchor='right', x=1),
        xaxis=dict(
            tickangle=-45,
            tickfont=dict(size=10, family=C['mono']),
            gridcolor=C['border'],
        ),
    )
    return fig
    """
    Mirror the exact 4-section Fund Level CF template layout:
      • Section 1 – Net CF
      • Section 2 – Called Capital
      • Section 3 – Distributions
      • Section 4 – NAV
    Each section shows the 9 meta columns followed by up to num_months date columns.
    A totals row is appended at the bottom of every section.
    """

    META_COLS = [
        'AIC Name', 'Investments', 'Portfolio Type', 'Investment Type',
        'Investment Status', 'Commitment Year', 'Commitment ($m)', 'Paid In', 'Unfunded',
    ]

    def fmt(val, decimals=2):
        """Format a numeric cashflow value; show dash for zero."""
        try:
            v = float(val)
        except (TypeError, ValueError):
            return "-"
        if v == 0.0:
            return "-"
        return f"{v:+.{decimals}f}" if v < 0 else f"{v:.{decimals}f}"

    def fmt_nav(val):
        """NAV is always positive – no sign prefix needed."""
        try:
            v = float(val)
        except (TypeError, ValueError):
            return "-"
        return "-" if v == 0.0 else f"{v:.2f}"

    def build_section_table(title, series_key, display_months, month_labels, formatter):
        """Build one DataTable section that mirrors a spreadsheet block."""
        all_cols = META_COLS + month_labels

        rows = []
        col_totals = {lbl: 0.0 for lbl in month_labels}

        for deal in uploaded_data:
            if deal.get('name') in ('_all_month_keys',):
                continue
            series = deal.get(series_key) or {}
            row = {
                'AIC Name':          deal.get('name', ''),
                'Investments':       deal.get('investments', deal.get('name', '')),
                'Portfolio Type':    deal.get('portfolio_type', ''),
                'Investment Type':   deal.get('type', ''),
                'Investment Status': deal.get('investment_status', ''),
                'Commitment Year':   str(deal.get('commitment_year', '') or ''),
                'Commitment ($m)':   f"{float(deal.get('commitment', 0) or 0):.2f}",
                'Paid In':           f"{float(deal.get('paid_in', deal.get('current_commitment', 0)) or 0):.2f}",
                'Unfunded':          f"{float(deal.get('unfunded', 0) or 0):.2f}",
            }
            for mk, lbl in zip(display_months, month_labels):
                v = float(series.get(mk) or 0)
                row[lbl] = formatter(v)
                col_totals[lbl] += v
            rows.append(row)

        # Totals row
        total_commitment = sum(float(d.get('commitment', 0) or 0) for d in uploaded_data)
        total_paid_in    = sum(float(d.get('paid_in', d.get('current_commitment', 0)) or 0) for d in uploaded_data)
        total_unfunded   = sum(float(d.get('unfunded', 0) or 0) for d in uploaded_data)
        totals_row = {
            'AIC Name':          'TOTAL',
            'Investments':       '',
            'Portfolio Type':    '',
            'Investment Type':   '',
            'Investment Status': '',
            'Commitment Year':   '',
            'Commitment ($m)':   f"{total_commitment:.2f}",
            'Paid In':           f"{total_paid_in:.2f}",
            'Unfunded':          f"{total_unfunded:.2f}",
        }
        for lbl in month_labels:
            totals_row[lbl] = formatter(col_totals[lbl])
        rows.append(totals_row)

        n_meta = len(META_COLS)
        style_cond = [
            {'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']},
            {'if': {'row_index': len(rows) - 1},   # Totals row
             'backgroundColor': C['surface'], 'fontWeight': 'bold', 'color': C['sky']},
            {'if': {'column_id': 'AIC Name'}, 'fontWeight': 'bold', 'minWidth': '160px'},
            {'if': {'column_id': 'Investments'}, 'minWidth': '160px'},
        ]

        # Colour negative cashflow values red, positive green (for CF sections)
        if series_key != 'nav_series':
            style_cond += [
                {'if': {'filter_query': f'{{{lbl}}} contains "-"', 'column_id': lbl},
                 'color': C['red']}
                for lbl in month_labels
            ]

        columns = [{'name': c, 'id': c} for c in all_cols]

        return dbc.Card([
            dbc.CardHeader(
                html.Div([
                    html.Span(title, style={'fontWeight': 'bold', 'fontSize': '14px'}),
                    html.Span(
                        f"  ({len(rows)-1} investments, {len(month_labels)} months shown)",
                        style={'fontSize': '11px', 'color': C['muted'], 'marginLeft': '12px'}
                    ),
                ]),
                style={'backgroundColor': C['surface'], 'borderBottom': f'1px solid {C["border"]}'}
            ),
            dbc.CardBody([
                dash_table.DataTable(
                    data=rows,
                    columns=columns,
                    style_cell={
                        'textAlign': 'right',
                        'padding': '6px 10px',
                        'fontFamily': C['mono'],
                        'fontSize': '11px',
                        'minWidth': '80px',
                        'maxWidth': '140px',
                        'whiteSpace': 'normal',
                    },
                    style_cell_conditional=[
                        {'if': {'column_id': c}, 'textAlign': 'left', 'minWidth': '140px'}
                        for c in ['AIC Name', 'Investments', 'Portfolio Type',
                                  'Investment Type', 'Investment Status']
                    ],
                    style_header={
                        'backgroundColor': C['surface'],
                        'color': C['text'],
                        'fontWeight': 'bold',
                        'border': f'1px solid {C["border"]}',
                        'position': 'sticky',
                        'top': 0,
                        'zIndex': 1,
                        'textAlign': 'center',
                    },
                    style_data={
                        'backgroundColor': C['panel'],
                        'color': C['text'],
                        'border': f'1px solid {C["border"]}',
                    },
                    style_data_conditional=style_cond,
                    style_table={
                        'overflowX': 'auto',
                        'maxHeight': '500px',
                        'overflowY': 'auto',
                    },
                    fixed_columns={'headers': True, 'data': 2},
                    fixed_rows={'headers': True},
                    page_action='none',
                    export_format='xlsx',
                    export_headers='display',
                    tooltip_data=[
                        {c: {'value': str(row.get(c, '')), 'type': 'markdown'}
                         for c in ['AIC Name', 'Investment Type']}
                        for row in rows
                    ],
                    tooltip_duration=None,
                )
            ], style={'padding': '0'})
        ], className="shadow-sm mb-4")

    # ── UPLOADED DATA PATH ─────────────────────────────────────────────────────
    if uploaded_data:
        # Recover ordered month keys stored during upload
        all_months_ordered = (
            uploaded_data[0].get('_all_month_keys') or
            sorted({mk for d in uploaded_data for mk in (d.get('net_cf_series') or {}).keys()})
        )
        display_months = all_months_ordered[:num_months]
        month_labels   = [pd.to_datetime(m).strftime("%b-%y") for m in display_months]

        sections = []
        if cf_type in ('all', 'net'):
            sections.append(build_section_table('Net CF',         'net_cf_series',       display_months, month_labels, fmt))
        if cf_type == 'all':
            sections.append(build_section_table('Called Capital', 'calls_series',         display_months, month_labels, fmt))
            sections.append(build_section_table('Distributions',  'distributions_series', display_months, month_labels, fmt))
        if cf_type in ('all', 'nav'):
            sections.append(build_section_table('NAV',            'nav_series',           display_months, month_labels, fmt_nav))

        return html.Div([
            dbc.Alert([
                html.I(className="fas fa-info-circle me-2"),
                html.Strong("Tip: "),
                "Values in millions USD. Negative cashflows shown in red. "
                "Use the export button on each table to download as Excel."
            ], color="info", className="mb-3", style={"fontSize": "12px"}),
            *sections
        ])

    # ── FALLBACK: no upload, use manually added deals ──────────────────────────
    if not deals:
        return dbc.Alert(
            "Upload your Fund Level CF Excel file above to see the full cashflow matrix, "
            "or add deals on the Current Portfolio page.",
            color="info"
        )

    deals_enriched = enrich_deals_with_latest_nav(deals, None)
    fallback_rows = []
    for d in deals_enriched:
        fallback_rows.append({
            'AIC Name':          d.get('name', ''),
            'Investments':       d.get('name', ''),
            'Portfolio Type':    d.get('segment', ''),
            'Investment Type':   d.get('strategy', ''),
            'Investment Status': d.get('allocation_status', 'Executed'),
            'Commitment Year':   str(d.get('vintage', '') or ''),
            'Commitment ($m)':   f"{float(d.get('total_commitment', d.get('commitment', d.get('size', 0))) or 0):.2f}",
            'Paid In':           f"{float(d.get('current_commitment', d.get('commitment', d.get('size', 0))) or 0):.2f}",
            'Unfunded':          f"{float(d.get('unfunded', 0) or 0):.2f}",
            'Current NAV':       f"{float(d.get('nav', d.get('size', 0)) or 0):.2f}",
        })
    return build_cashflow_template_table(
        "Current Portfolio Snapshot (no CF file uploaded)",
        fallback_rows,
        list(fallback_rows[0].keys()) if fallback_rows else []
    )


# ==================== PRO FORMA CALLBACKS ====================

# Pro Forma Tabs Content
@app.callback(
    Output('pf-tab-content', 'children'),
    [Input('pf-tabs', 'active_tab'), Input('deals-store', 'data'),
     Input('pipeline-store', 'data'), Input('pf-target-month', 'value'),
     Input('cashflows-store', 'data')]
)
def render_pf_tab_content(active_tab, deals, pipeline, target_month, cashflows):
    if active_tab == "tab-pf-portfolio":
        # COMPLETE PORTFOLIO TAB - Show ALL deals
        return render_complete_portfolio(deals, pipeline, target_month, cashflows)
    elif active_tab == "tab-pf-metrics":
        # METRICS COMPARISON TAB
        return render_metrics_comparison(deals, pipeline)
    elif active_tab == "tab-pf-charts":
        # CHARTS TAB
        return render_impact_charts(deals, pipeline)
    return html.Div()


def render_complete_portfolio(deals, pipeline, target_month, cashflows):
    """Show complete portfolio with current + pipeline deals and NAV at target month"""

    # Calculate NAV at target month using actual cashflows
    month_names = generate_month_options()
    target_month_name = month_names[target_month]['label'] if target_month < len(month_names) else "Unknown"

    # Build complete portfolio
    all_deals = []

    # Current portfolio deals
    for deal in deals:
        deal_copy = deal.copy()
        deal_copy['source'] = 'Current Portfolio'
        deal_copy['status'] = '✓ Active'

        # Calculate NAV at target month using cashflows
        nav_at_month = calculate_nav_at_month(deal['name'], target_month, cashflows, deal['size'])
        deal_copy['nav_at_target'] = nav_at_month

        all_deals.append(deal_copy)

    # Pipeline deals (assumed to close in their respective months)
    for p_deal in pipeline:
        deal_copy = {
            'name': p_deal['name'],
            'strategy': p_deal['type'],
            'size': p_deal['size'],
            'target_irr': p_deal['target_irr'],
            'hold_period': 5.0,
            'moic': (1 + p_deal['target_irr']) ** 5,
            'vintage': 2026,
            'sector': 'TBD',
            'geography': 'Global',
            'source': 'Pipeline',
            'status': '⏳ Pending',
            'nav_at_target': p_deal['size']  # Assume full deployment
        }
        all_deals.append(deal_copy)

    if not all_deals:
        return dbc.Alert("No deals to display. Add deals to Current Portfolio or Pipeline.", color="info")

    # Summary cards
    total_current_nav = sum(d['size'] for d in deals)
    total_pipeline_nav = sum(p['size'] for p in pipeline)
    total_nav_at_target = sum(d['nav_at_target'] for d in all_deals)

    summary = dbc.Row([
        dbc.Col(dbc.Card([dbc.CardBody([
            html.H6("Current Portfolio NAV", className="text-muted"),
            html.H4(f"${total_current_nav:.1f}M", style={'color': C['green'], 'fontFamily': C['mono']})
        ])], className="shadow-sm"), width=3),
        dbc.Col(dbc.Card([dbc.CardBody([
            html.H6("Pipeline NAV", className="text-muted"),
            html.H4(f"${total_pipeline_nav:.1f}M", style={'color': C['blue'], 'fontFamily': C['mono']})
        ])], className="shadow-sm"), width=3),
        dbc.Col(dbc.Card([dbc.CardBody([
            html.H6(f"Pro Forma NAV at {target_month_name}", className="text-muted"),
            html.H4(f"${total_nav_at_target:.1f}M", style={'color': C['amber'], 'fontFamily': C['mono']})
        ])], className="shadow-sm"), width=3),
        dbc.Col(dbc.Card([dbc.CardBody([
            html.H6("Total Deals", className="text-muted"),
            html.H4(f"{len(all_deals)}", style={'color': C['purple'], 'fontFamily': C['mono']})
        ])], className="shadow-sm"), width=3),
    ], className="mb-4")

    # Complete portfolio table
    table_data = []
    for d in all_deals:
        table_data.append({
            'Deal': d['name'],
            'Strategy': d['strategy'],
            'Size': f"${d['size']:.1f}M",
            'IRR': f"{d['target_irr']:.1%}",
            'MOIC': f"{d.get('moic', 0):.2f}x",
            'Vintage': str(d.get('vintage', 'N/A')),
            'Sector': d.get('sector', 'N/A'),
            'Source': d['source'],
            'Status': d['status'],
            f'NAV @ {target_month_name}': f"${d['nav_at_target']:.1f}M"
        })

    portfolio_table = dbc.Card([
        dbc.CardHeader(f"Complete Pro Forma Portfolio ({len(all_deals)} deals)",
                       style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
        dbc.CardBody([
            dash_table.DataTable(
                data=table_data,
                columns=[{"name": c, "id": c} for c in table_data[0].keys()] if table_data else [],
                style_cell={'textAlign': 'left', 'padding': '12px', 'fontFamily': C['mono'], 'fontSize': '12px'},
                style_header={
                    'backgroundColor': C['surface'],
                    'color': C['text'],
                    'fontWeight': 'bold',
                    'border': f'1px solid {C["border"]}'
                },
                style_data={
                    'backgroundColor': C['panel'],
                    'color': C['text'],
                    'border': f'1px solid {C["border"]}'
                },
                style_data_conditional=[
                    {'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']},
                    {'if': {'column_id': 'Source', 'filter_query': '{Source} = "Pipeline"'},
                     'color': C['blue'], 'fontWeight': 'bold'},
                    {'if': {'column_id': 'Source', 'filter_query': '{Source} = "Current Portfolio"'},
                     'color': C['green'], 'fontWeight': 'bold'},
                ],
                page_size=100,  # Show many deals
                sort_action='native',
                filter_action='native',
            )
        ])
    ], className="shadow-sm")

    return html.Div([summary, portfolio_table])


def calculate_nav_at_month(deal_name, target_month_idx, cashflows, initial_nav):
    """Calculate NAV of a deal at target month using actual cashflows"""
    if not cashflows:
        return initial_nav

    # Filter cashflows for this deal
    deal_cashflows = [cf for cf in cashflows if cf['deal'] == deal_name]

    if not deal_cashflows:
        return initial_nav

    # Get target month date
    base_date = datetime(2026, 1, 1)
    from dateutil.relativedelta import relativedelta
    target_date = base_date + relativedelta(months=target_month_idx)

    # Calculate NAV = Initial + Calls - Distributions up to target month
    nav = initial_nav

    for cf in deal_cashflows:
        try:
            cf_date = datetime.fromisoformat(
                cf['date'].replace('Z', '+00:00') if 'T' in cf['date'] else cf['date'] + 'T00:00:00')
        except:
            continue

        # Only include cashflows up to target month
        if cf_date <= target_date:
            if cf['type'] == 'Call':
                nav -= cf['amount']  # Calls reduce NAV
            else:  # Distribution
                nav += cf['amount']  # Distributions increase NAV

    return max(0, nav)  # NAV can't be negative


def render_metrics_comparison(deals, pipeline):
    """Render before/after metrics comparison"""
    from datetime import datetime

    current_m = calculate_portfolio_metrics(deals)

    # Pro forma (current + pipeline)
    if pipeline:
        proforma_deals = deals + [{
            'name': p['name'],
            'strategy': p['type'],
            'size': p['size'],
            'target_irr': p['target_irr'],
            'hold_period': 5.0,
            'moic': (1 + p['target_irr']) ** 5,
            'vintage': 2026,
            'sector': 'TBD',
            'geography': 'Global'
        } for p in pipeline]

        pf_m = calculate_portfolio_metrics(proforma_deals)
    else:
        pf_m = current_m

    comparison_data = [
        {"Metric": "Total NAV ($mm)", "Current": f"${current_m['total_nav']:.1f}",
         "Pro Forma": f"${pf_m['total_nav']:.1f}",
         "Change": f"+${pf_m['total_nav'] - current_m['total_nav']:.1f}"},
        {"Metric": "Weighted IRR", "Current": f"{current_m['weighted_irr']:.2%}",
         "Pro Forma": f"{pf_m['weighted_irr']:.2%}",
         "Change": f"{pf_m['weighted_irr'] - current_m['weighted_irr']:+.2%}"},
        {"Metric": "Number of Deals", "Current": str(current_m['num_deals']),
         "Pro Forma": str(pf_m['num_deals']),
         "Change": f"+{pf_m['num_deals'] - current_m['num_deals']}"},
        {"Metric": "Top 1 Concentration", "Current": f"{current_m['concentration_top1']:.1%}",
         "Pro Forma": f"{pf_m['concentration_top1']:.1%}",
         "Change": f"{pf_m['concentration_top1'] - current_m['concentration_top1']:+.1%}"},
    ]

    return dbc.Card([
        dbc.CardHeader("Before vs After Comparison", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
        dbc.CardBody([
            dash_table.DataTable(
                data=comparison_data,
                columns=[{"name": c, "id": c} for c in ["Metric", "Current", "Pro Forma", "Change"]],
                style_cell={'textAlign': 'left', 'padding': '12px', 'fontFamily': C['mono'], 'fontSize': '13px'},
                style_header={'backgroundColor': C['surface'], 'color': C['text'], 'fontWeight': 'bold'},
                style_data={'backgroundColor': C['panel'], 'color': C['text']},
                style_data_conditional=[
                    {'if': {'column_id': 'Change'}, 'fontWeight': 'bold', 'color': C['blue']}
                ]
            )
        ])
    ], className="shadow-sm")


def render_impact_charts(deals, pipeline):
    """Render impact charts"""
    current_m = calculate_portfolio_metrics(deals)

    if pipeline:
        proforma_deals = deals + [{
            'name': p['name'], 'strategy': p['type'], 'size': p['size'],
            'target_irr': p['target_irr'], 'hold_period': 5.0,
            'moic': (1 + p['target_irr']) ** 5, 'vintage': 2026,
            'sector': 'TBD', 'geography': 'Global'
        } for p in pipeline]
        pf_m = calculate_portfolio_metrics(proforma_deals)
    else:
        pf_m = current_m

    # Strategy chart
    fig_strat = go.Figure()
    if current_m['by_strategy']:
        strats = list(current_m['by_strategy'].keys())
        current_vals = [current_m['by_strategy'][s]['nav'] for s in strats]
        fig_strat.add_trace(go.Bar(name='Current', x=strats, y=current_vals, marker_color=C['blue']))

    if pf_m['by_strategy']:
        pf_strats = list(pf_m['by_strategy'].keys())
        pf_vals = [pf_m['by_strategy'][s]['nav'] for s in pf_strats]
        fig_strat.add_trace(go.Bar(name='Pro Forma', x=pf_strats, y=pf_vals, marker_color=C['green']))

    fig_strat.update_layout(**CHART_BASE, barmode='group', yaxis_title="NAV ($mm)", height=400)

    # Concentration chart
    fig_conc = go.Figure()
    cats = ['Top 1', 'Top 3', 'Top 5']
    current_conc = [current_m['concentration_top1'] * 100, current_m['concentration_top3'] * 100,
                    current_m['concentration_top5'] * 100]
    pf_conc = [pf_m['concentration_top1'] * 100, pf_m['concentration_top3'] * 100,
               pf_m['concentration_top5'] * 100]

    fig_conc.add_trace(go.Bar(name='Current', x=cats, y=current_conc, marker_color=C['blue']))
    fig_conc.add_trace(go.Bar(name='Pro Forma', x=cats, y=pf_conc, marker_color=C['green']))
    fig_conc.update_layout(**CHART_BASE, barmode='group', yaxis_title="% of NAV", height=400)

    return dbc.Row([
        dbc.Col([
            dbc.Card([
                dbc.CardHeader("Strategy Allocation", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                dbc.CardBody([dcc.Graph(figure=fig_strat, config={'displayModeBar': False})])
            ], className="shadow-sm")
        ], width=6),
        dbc.Col([
            dbc.Card([
                dbc.CardHeader("Concentration Risk", style={'fontWeight': 'bold', 'backgroundColor': C['surface']}),
                dbc.CardBody([dcc.Graph(figure=fig_conc, config={'displayModeBar': False})])
            ], className="shadow-sm")
        ], width=6),
    ])


# ==================== FUND OVERVIEW CALLBACKS ====================

@app.callback(
    [Output('overview-current-nav', 'children'), Output('overview-dry-powder', 'children'),
     Output('overview-headroom', 'children')],
    [Input('deals-store', 'data'), Input('config-store', 'data'),
     Input('fund-cf-data-store', 'data'), Input('liquidity-dry-powder-store', 'data'),
     Input('overview-target-fund-size', 'value')]
)
def update_overview_kpis(deals, config, fund_cf_data, dry_powder_uploaded, target_fund_size):
    deals_enriched = enrich_deals_with_latest_nav(deals, fund_cf_data)
    total_nav = sum(d.get('nav', d.get('size', 0)) for d in (deals_enriched or []))
    dry_powder = get_effective_dry_powder(config, dry_powder_uploaded)
    target = float(target_fund_size) if target_fund_size else 1000.0
    headroom = max(0, target - total_nav)
    return f"${total_nav:.1f}M", f"${dry_powder:.1f}M", f"${headroom:.1f}M"


# ── Targets tab content ────────────────────────────────────────────────────
OVERVIEW_TAB_DIM_MAP = {
    'ov-tab-invtype': ('Investment Type', ['GP-Led (Multi-Asset CV)', 'GP-Led (Single-Asset CV)',
                                           'LP-Led Secondary', 'Co-Investment', 'Primary']),
    'ov-tab-region':  ('Region',  ['North America', 'Europe', 'Asia & ROW']),
    'ov-tab-sector':  ('Sector',  ['Technology', 'Healthcare', 'Financial Services',
                                   'Consumer', 'Industrials', 'Energy', 'Real Estate', 'Other']),
    'ov-tab-stage':   ('Stage',   ['Buyout', 'Growth', 'Venture', 'Liquidity']),
    'ov-tab-vintage': ('Vintage', [str(y) for y in range(2026, 2014, -1)]),
}

def _tbl_sty(C):
    return dict(
        style_cell={'textAlign': 'left', 'padding': '8px 12px',
                    'fontFamily': C['mono'], 'fontSize': '12px',
                    'border': f'1px solid {C["border"]}'},
        style_header={'backgroundColor': C['surface'], 'color': C['text'],
                      'fontWeight': 'bold', 'fontSize': '12px',
                      'border': f'1px solid {C["border"]}'},
        style_data={'backgroundColor': C['panel'], 'color': C['text']},
        style_data_conditional=[
            {'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']},
            {'if': {'column_id': ['target_pct', 'min_pct', 'max_pct']},
             'backgroundColor': '#fff9e6'},
        ],
    )


@app.callback(
    Output('ov-targets-tab-content', 'children'),
    [Input('ov-targets-tabs', 'active_tab'),
     Input('ov-targets-store', 'data'),
     Input('btn-add-target-row', 'n_clicks')],
    prevent_initial_call=False
)
def render_targets_tab(active_tab, store_data, _add):
    """Render the editable targets DataTable for the selected dimension tab."""
    dim, categories = OVERVIEW_TAB_DIM_MAP.get(active_tab, ('Investment Type', []))
    rows = [r for r in (store_data or []) if r.get('dimension') == dim]
    if not rows:
        rows = [{'dimension': dim, 'category': c,
                 'min_pct': 0, 'target_pct': 0, 'max_pct': 0}
                for c in categories]

    ctx = callback_context
    triggered = ctx.triggered[0]['prop_id'].split('.')[0] if ctx.triggered else ''
    if triggered == 'btn-add-target-row':
        rows.append({'dimension': dim, 'category': '', 'min_pct': 0,
                     'target_pct': 0, 'max_pct': 0})

    columns = [
        {'name': 'Category',  'id': 'category',   'editable': True,
         'presentation': 'dropdown'},
        {'name': 'Min %',     'id': 'min_pct',    'editable': True, 'type': 'numeric',
         'format': {'specifier': '.1f'}},
        {'name': 'Target %',  'id': 'target_pct', 'editable': True, 'type': 'numeric',
         'format': {'specifier': '.1f'}},
        {'name': 'Max %',     'id': 'max_pct',    'editable': True, 'type': 'numeric',
         'format': {'specifier': '.1f'}},
    ]

    return dash_table.DataTable(
        id={'type': 'ov-dim-table', 'dim': dim},
        columns=columns,
        data=rows,
        editable=True,
        row_deletable=True,
        dropdown={'category': {'options': [{'label': c, 'value': c} for c in categories]}},
        tooltip_header={
            'min_pct':    'Minimum allowed % of portfolio NAV',
            'target_pct': 'Desired target % of portfolio NAV',
            'max_pct':    'Maximum allowed % of portfolio NAV',
        },
        **_tbl_sty(C),
    )


@app.callback(
    Output('ov-targets-store', 'data', allow_duplicate=True),
    Input({'type': 'ov-dim-table', 'dim': ALL}, 'data'),
    State('ov-targets-store', 'data'),
    prevent_initial_call=True
)
def sync_targets_store(all_table_data, store):
    """Merge any edited dimension table back into the central targets store."""
    if not all_table_data:
        return store or []
    store = list(store or [])
    for tbl_rows in all_table_data:
        if not tbl_rows:
            continue
        dim = tbl_rows[0].get('dimension', '')
        if not dim:
            continue
        store = [r for r in store if r.get('dimension') != dim]
        store.extend(tbl_rows)
    return store


# ── Restrictions table ─────────────────────────────────────────────────────
@app.callback(
    Output('restrictions-table-container', 'children'),
    [Input('ov-restrictions-store', 'data'),
     Input('btn-add-restriction-row', 'n_clicks')],
    prevent_initial_call=False
)
def render_restrictions_table(store_data, _add):
    """Render the legal/internal restrictions DataTable."""
    ctx = callback_context
    triggered = ctx.triggered[0]['prop_id'].split('.')[0] if ctx.triggered else ''

    rows = list(store_data or [])
    if triggered == 'btn-add-restriction-row':
        rows.append({'restriction': 'New Restriction',
                     'legal_min': None, 'legal_max': None,
                     'internal_min': None, 'internal_max': None,
                     'target': None})

    # Current and Pro Forma are placeholder columns — populated by separate callback
    for r in rows:
        r.setdefault('current_pct', '—')
        r.setdefault('proforma_pct', '—')

    columns = [
        {'name': 'Restriction',      'id': 'restriction',   'editable': True},
        {'name': 'Legal Min %',      'id': 'legal_min',     'editable': True, 'type': 'numeric'},
        {'name': 'Legal Max %',      'id': 'legal_max',     'editable': True, 'type': 'numeric'},
        {'name': 'Internal Min %',   'id': 'internal_min',  'editable': True, 'type': 'numeric'},
        {'name': 'Internal Max %',   'id': 'internal_max',  'editable': True, 'type': 'numeric'},
        {'name': 'Target %',         'id': 'target',        'editable': True, 'type': 'numeric'},
        {'name': 'Current %',        'id': 'current_pct',   'editable': False},
        {'name': 'Pro Forma %',      'id': 'proforma_pct',  'editable': False},
    ]

    style_data_conditional = [
        {'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']},
        # Editable input columns
        {'if': {'column_id': ['legal_min','legal_max','internal_min','internal_max','target']},
         'backgroundColor': '#fff9e6'},
        # Read-only columns
        {'if': {'column_id': 'current_pct'},  'backgroundColor': '#e8f5e9', 'fontWeight': 'bold'},
        {'if': {'column_id': 'proforma_pct'}, 'backgroundColor': '#e3f2fd', 'fontWeight': 'bold'},
        {'if': {'column_id': 'restriction'},  'fontWeight': 'bold'},
    ]

    return dash_table.DataTable(
        id='restrictions-datatable',
        columns=columns,
        data=rows,
        editable=True,
        row_deletable=True,
        style_table={'overflowX': 'auto'},
        style_cell_conditional=[
            {'if': {'column_id': 'restriction'}, 'minWidth': '220px', 'textAlign': 'left'},
            {'if': {'column_id': ['legal_min','legal_max','internal_min',
                                  'internal_max','target','current_pct','proforma_pct']},
             'minWidth': '90px', 'textAlign': 'right'},
        ],
        style_data_conditional=style_data_conditional,
        tooltip_header={
            'legal_min':    'Minimum % mandated by fund legal documents',
            'legal_max':    'Maximum % mandated by fund legal documents',
            'internal_min': 'Stricter internal minimum',
            'internal_max': 'Stricter internal maximum',
            'target':       'Desired / optimal level',
            'current_pct':  'Calculated from current portfolio (read-only)',
            'proforma_pct': 'Calculated including pipeline deals (read-only)',
        },
        tooltip_delay=400,
        **{k: v for k, v in _tbl_sty(C).items()
           if k not in ('style_data_conditional',)},
        **{'style_data_conditional': style_data_conditional},
    )


@app.callback(
    Output('restrictions-datatable', 'data'),
    [Input('restrictions-datatable', 'data_timestamp'),
     Input('deals-store', 'data'),
     Input('pipeline-store', 'data'),
     Input('fund-cf-data-store', 'data')],
    State('restrictions-datatable', 'data'),
    prevent_initial_call=True
)
def update_restrictions_current_proforma(_, deals, pipeline, fund_cf_data, rows):
    """
    Populate Current % and Pro Forma % columns in the restrictions table.
    Supports: Single Asset, Single Manager, Listed Securities,
              North America + Europe Combined, Leverage (placeholder — manual).
    """
    if not rows:
        return rows

    deals_enriched = enrich_deals_with_latest_nav(deals, fund_cf_data)
    total_nav_current = sum(d.get('nav', d.get('size', 0)) for d in (deals_enriched or []))

    pipeline_nav = sum(float(p.get('size', 0) or 0) for p in (pipeline or []))
    total_nav_pf = total_nav_current + pipeline_nav

    def calc_restriction(restriction_name, portfolio, total_nav):
        """Return % value for a given restriction against a portfolio list."""
        if total_nav == 0:
            return None
        name_l = restriction_name.lower()

        if 'single asset' in name_l:
            # Largest single deal as % of NAV
            navs = [d.get('nav', d.get('size', 0)) for d in portfolio]
            return (max(navs) / total_nav * 100) if navs else 0

        elif 'single manager' in name_l:
            # Largest single manager group as % of NAV
            from collections import defaultdict
            mgr_nav = defaultdict(float)
            for d in portfolio:
                mgr = d.get('manager', d.get('name', 'Unknown'))
                mgr_nav[mgr] += d.get('nav', d.get('size', 0))
            return (max(mgr_nav.values()) / total_nav * 100) if mgr_nav else 0

        elif 'listed' in name_l:
            # Deals tagged as listed/public equity
            listed = sum(d.get('nav', d.get('size', 0)) for d in portfolio
                         if 'listed' in str(d.get('stage', '')).lower()
                         or 'public' in str(d.get('sector', '')).lower())
            return listed / total_nav * 100

        elif 'north america' in name_l and 'europe' in name_l:
            # Combined NA + EU exposure
            combined = sum(d.get('nav', d.get('size', 0)) for d in portfolio
                           if d.get('geography', d.get('region', '')) in ('North America', 'Europe'))
            return combined / total_nav * 100

        elif 'leverage' in name_l:
            # Leverage is manually entered — keep existing value
            return None

        return None

    # Build proforma portfolio = current deals + pipeline as synthetic entries
    pipeline_as_deals = []
    for p in (pipeline or []):
        pipeline_as_deals.append({
            'nav': float(p.get('size', 0) or 0),
            'size': float(p.get('size', 0) or 0),
            'manager': p.get('name', 'Pipeline'),
            'geography': p.get('region', 'North America'),
            'stage': p.get('stage', 'Buyout'),
            'sector': p.get('sector', 'Technology'),
        })
    pf_portfolio = list(deals_enriched or []) + pipeline_as_deals

    updated = []
    for row in rows:
        restriction = row.get('restriction', '')
        current_val = calc_restriction(restriction, deals_enriched or [], total_nav_current)
        pf_val = calc_restriction(restriction, pf_portfolio, total_nav_pf)

        row = dict(row)
        if current_val is not None:
            row['current_pct'] = f"{current_val:.1f}%"
        else:
            row['current_pct'] = row.get('current_pct', '—')

        if pf_val is not None:
            row['proforma_pct'] = f"{pf_val:.1f}%"
        else:
            row['proforma_pct'] = row.get('proforma_pct', '—')

        # Colour-code status in the cell string (append flag)
        def _flag(val_str, legal_max, internal_max):
            try:
                val = float(val_str.replace('%', ''))
                if legal_max and val > float(legal_max):
                    return val_str + ' ⚠️'
                elif internal_max and val > float(internal_max):
                    return val_str + ' ⚡'
            except Exception:
                pass
            return val_str

        row['current_pct'] = _flag(row['current_pct'], row.get('legal_max'), row.get('internal_max'))
        row['proforma_pct'] = _flag(row['proforma_pct'], row.get('legal_max'), row.get('internal_max'))
        updated.append(row)

    return updated


@app.callback(
    Output('ov-restrictions-store', 'data', allow_duplicate=True),
    Input('restrictions-datatable', 'data'),
    prevent_initial_call=True
)
def persist_restrictions(rows):
    """Keep restrictions store in sync with table edits."""
    return rows or []


# ── Live Exposure Monitor (Current vs Pro Forma tabs) ─────────────────────
@app.callback(
    Output('ov-exposure-monitor', 'children'),
    [Input('ov-exposure-view-tabs', 'active_tab'),
     Input('deals-store', 'data'),
     Input('pipeline-store', 'data'),
     Input('ov-targets-store', 'data'),
     Input('fund-cf-data-store', 'data')],
    prevent_initial_call=False
)
def render_exposure_monitor(active_tab, deals, pipeline, targets, fund_cf_data):
    """Render exposure vs targets table for current or pro forma view."""
    deals_enriched = enrich_deals_with_latest_nav(deals, fund_cf_data)
    current_nav = sum(d.get('nav', d.get('size', 0)) for d in (deals_enriched or []))

    if active_tab == 'ov-exp-proforma':
        pipeline_deals = [{
            'nav': float(p.get('size', 0) or 0),
            'size': float(p.get('size', 0) or 0),
            'strategy': p.get('type', p.get('strategy', 'Other')),
            'geography': p.get('region', 'North America'),
            'sector': p.get('sector', 'Technology'),
            'stage': p.get('stage', 'Buyout'),
            'vintage': str(p.get('vintage', 2026)),
        } for p in (pipeline or [])]
        portfolio = list(deals_enriched or []) + pipeline_deals
        total_nav = current_nav + sum(p['nav'] for p in pipeline_deals)
        view_label = 'Pro Forma (Current + Pipeline)'
    else:
        portfolio = list(deals_enriched or [])
        total_nav = current_nav
        view_label = 'Current Portfolio'

    if not targets or total_nav == 0:
        return dbc.Alert(
            f"No data available for {view_label}. Add deals and set targets to see exposure.",
            color="info"
        )

    # Strategy/InvType mapping: deal.strategy values → target dimension categories
    STRAT_ALIAS = {
        'GP-Led (Multi-Asset)':  'GP-Led (Multi-Asset CV)',
        'GP-Led (Single-Asset)': 'GP-Led (Single-Asset CV)',
        'Diversified LP-Led':    'LP-Led Secondary',
        'Co-Investments':        'Co-Investment',
    }

    def get_nav_for_dim(dim, cat):
        if dim == 'Investment Type':
            return sum(d.get('nav', d.get('size', 0)) for d in portfolio
                       if STRAT_ALIAS.get(d.get('strategy', ''), d.get('strategy', '')) == cat
                       or d.get('strategy', '') == cat)
        elif dim == 'Region':
            return sum(d.get('nav', d.get('size', 0)) for d in portfolio
                       if d.get('geography', d.get('region', '')) == cat)
        elif dim == 'Sector':
            return sum(d.get('nav', d.get('size', 0)) for d in portfolio
                       if d.get('sector', '') == cat)
        elif dim == 'Stage':
            return sum(d.get('nav', d.get('size', 0)) for d in portfolio
                       if d.get('stage', '') == cat)
        elif dim == 'Vintage':
            return sum(d.get('nav', d.get('size', 0)) for d in portfolio
                       if str(d.get('vintage', '')) == str(cat))
        return 0

    rows_by_dim = {}
    for t in (targets or []):
        dim = t.get('dimension', '')
        if dim not in rows_by_dim:
            rows_by_dim[dim] = []
        rows_by_dim[dim].append(t)

    sections = []
    dim_colors = {
        'Investment Type': C['blue'],
        'Region': C['teal'],
        'Sector': C['purple'],
        'Stage': C['green'],
        'Vintage': C['amber'],
    }

    for dim, dim_rows in rows_by_dim.items():
        trs = []
        for t in dim_rows:
            cat       = t.get('category', '')
            min_pct   = float(t.get('min_pct', 0) or 0)
            tgt_pct   = float(t.get('target_pct', 0) or 0)
            max_pct   = float(t.get('max_pct', 0) or 0)
            cur_nav   = get_nav_for_dim(dim, cat)
            cur_pct   = cur_nav / total_nav * 100 if total_nav else 0

            if max_pct > 0 and cur_pct > max_pct:
                status, s_color, badge = '⚠️ Over',  C['red'],   'danger'
            elif min_pct > 0 and cur_pct < min_pct:
                status, s_color, badge = '↓ Under', C['amber'], 'warning'
            elif tgt_pct > 0:
                status, s_color, badge = '✓ On Track', C['green'], 'success'
            else:
                status, s_color, badge = '—', C['muted'], 'secondary'

            # Progress bar fill
            bar_pct = min(cur_pct / max_pct * 100, 100) if max_pct else 0
            bar_color = 'danger' if cur_pct > max_pct else ('warning' if cur_pct < min_pct else 'success')

            trs.append(html.Tr([
                html.Td(cat, style={'fontFamily': C['mono'], 'fontSize': '12px',
                                    'fontWeight': 'bold', 'paddingLeft': '16px'}),
                html.Td(f"${cur_nav:,.1f}M",
                        style={'fontFamily': C['mono'], 'fontSize': '12px', 'textAlign': 'right'}),
                html.Td([
                    html.Div([
                        dbc.Progress(value=bar_pct, color=bar_color,
                                     style={'height': '8px', 'borderRadius': '4px'}),
                        html.Small(f"{cur_pct:.1f}%",
                                   style={'fontFamily': C['mono'], 'color': C['text']})
                    ])
                ], style={'minWidth': '120px'}),
                html.Td(f"{tgt_pct:.1f}%",
                        style={'fontFamily': C['mono'], 'fontSize': '12px',
                               'textAlign': 'right', 'color': C['blue']}),
                html.Td(f"{min_pct:.1f}%–{max_pct:.1f}%",
                        style={'fontFamily': C['mono'], 'fontSize': '11px',
                               'textAlign': 'right', 'color': C['muted']}),
                html.Td(dbc.Badge(status, color=badge, className="px-2"),
                        style={'textAlign': 'center'}),
            ]))

        sections.append(html.Div([
            html.H6(dim, className="mb-0 mt-3",
                    style={'color': dim_colors.get(dim, C['text']),
                           'fontFamily': C['sans'], 'fontWeight': 'bold',
                           'fontSize': '13px', 'textTransform': 'uppercase',
                           'letterSpacing': '0.05em'}),
            dbc.Table([
                html.Thead(html.Tr([
                    html.Th('Category', style={'width': '25%'}),
                    html.Th('NAV ($m)',  style={'textAlign': 'right'}),
                    html.Th('Current %', style={'minWidth': '140px'}),
                    html.Th('Target %',  style={'textAlign': 'right'}),
                    html.Th('Range',     style={'textAlign': 'right'}),
                    html.Th('Status',    style={'textAlign': 'center'}),
                ]), style={'backgroundColor': C['surface'], 'color': C['text'],
                           'fontSize': '11px'}),
                html.Tbody(trs)
            ], bordered=True, hover=True, size='sm', className='mb-0',
               style={'fontFamily': C['mono'], 'fontSize': '12px'})
        ]))

    header_badge = dbc.Badge(view_label, color="primary", className="ms-2")
    return html.Div([
        html.H6(["📊 Exposure Breakdown ", header_badge],
                className="mb-2",
                style={'fontFamily': C['sans'], 'fontWeight': 'bold'}),
    ] + sections)


@app.callback(
    Output('overview-exposure-table', 'data'),
    [Input('config-store', 'data'), Input('btn-add-exposure-row', 'n_clicks')],
    State('overview-exposure-table', 'data'),
    prevent_initial_call=False
)
def update_exposure_table(config, n_clicks, current_data):
    ctx = callback_context
    triggered = ctx.triggered[0]['prop_id'].split('.')[0] if ctx.triggered else ''

    if triggered == 'btn-add-exposure-row':
        # Add a blank row
        rows = current_data or []
        rows = list(rows)  # make mutable copy
        rows.append({'dimension': 'Strategy', 'category': '', 'min_pct': 0.0, 'target_pct': 0.0, 'max_pct': 0.0})
        return rows

    # Initial load — populate from config
    overview = get_fund_overview_config(config)
    limits = overview.get('exposure_limits', []) or []
    return [
        {'dimension': row.get('dimension', ''), 'category': row.get('category', ''),
         'min_pct': round(row.get('min_pct', 0) * 100, 2),
         'target_pct': round(row.get('target_pct', 0) * 100, 2),
         'max_pct': round(row.get('max_pct', 0) * 100, 2)}
        for row in limits
    ]


@app.callback(
    Output('overview-bite-table', 'data'),
    [Input('config-store', 'data'), Input('btn-add-bite-row', 'n_clicks')],
    State('overview-bite-table', 'data'),
    prevent_initial_call=False
)
def update_bite_table(config, n_clicks, current_data):
    ctx = callback_context
    triggered = ctx.triggered[0]['prop_id'].split('.')[0] if ctx.triggered else ''

    if triggered == 'btn-add-bite-row':
        rows = current_data or []
        rows = list(rows)
        rows.append({'asset_type': 'GP-Led (Single-Asset)', 'min_pct': 0.5, 'desired_pct': 2.25, 'max_pct': 4.0})
        return rows

    # Initial load — populate from config
    overview = get_fund_overview_config(config)
    targets = overview.get('bite_size_targets', []) or []
    return [
        {'asset_type': row.get('asset_type', ''),
         'min_pct': round(row.get('min_pct', 0) * 100, 3),
         'desired_pct': round(row.get('desired_pct', 0) * 100, 3),
         'max_pct': round(row.get('max_pct', 0) * 100, 3)}
        for row in targets
    ]


@app.callback(
    [Output('config-store', 'data', allow_duplicate=True),
     Output('overview-save-status', 'children')],
    Input('btn-save-overview', 'n_clicks'),
    [State('overview-exposure-table', 'data'), State('overview-bite-table', 'data'),
     State('overview-target-fund-size', 'value'), State('config-store', 'data')],
    prevent_initial_call=True
)
def save_overview(n_clicks, exposure_rows, bite_rows, target_fund_size, config):
    if not n_clicks:
        return config, ""
    try:
        config = config or DEFAULT_CONFIG
        # Convert percentage columns back to decimals for storage
        exposure_limits = []
        for row in (exposure_rows or []):
            try:
                exposure_limits.append({
                    'dimension': row.get('dimension', ''),
                    'category': row.get('category', ''),
                    'min_pct': float(row.get('min_pct', 0)) / 100,
                    'target_pct': float(row.get('target_pct', 0)) / 100,
                    'max_pct': float(row.get('max_pct', 0)) / 100,
                })
            except Exception:
                pass

        bite_targets = []
        for row in (bite_rows or []):
            try:
                bite_targets.append({
                    'asset_type': row.get('asset_type', ''),
                    'min_pct': float(row.get('min_pct', 0)) / 100,
                    'desired_pct': float(row.get('desired_pct', 0)) / 100,
                    'max_pct': float(row.get('max_pct', 0)) / 100,
                })
            except Exception:
                pass

        overview = get_fund_overview_config(config)
        overview['exposure_limits'] = exposure_limits
        overview['bite_size_targets'] = bite_targets
        overview['target_fund_size'] = float(target_fund_size) if target_fund_size else 1000.0

        import copy
        new_config = copy.deepcopy(config)
        new_config['fund_overview'] = overview

        return new_config, dbc.Alert("✅ Overview saved successfully!", color="success", duration=3000)
    except Exception as e:
        return config, dbc.Alert(f"❌ Save error: {e}", color="danger")


@app.callback(
    Output('overview-exposure-chart', 'figure'),
    [Input('deals-store', 'data'), Input('fund-cf-data-store', 'data')]
)
def update_overview_exposure_chart(deals, fund_cf_data):
    deals_enriched = enrich_deals_with_latest_nav(deals, fund_cf_data)
    m = calculate_portfolio_metrics(deals_enriched)
    fig = go.Figure()
    if m['by_strategy']:
        labels = list(m['by_strategy'].keys())
        values = [m['by_strategy'][s]['nav'] for s in labels]
        colors_pie = [C['blue'], C['purple'], C['teal'], C['green']][:len(labels)]
        fig = go.Figure(data=[go.Pie(
            labels=labels, values=values, hole=0.4,
            marker=dict(colors=colors_pie),
            textfont=dict(color=C['text'], family=C['mono'])
        )])
    fig.update_layout(**CHART_BASE, height=300, margin=dict(t=20, b=0, l=0, r=0))
    return fig


@app.callback(
    Output('overview-forward-bite-table', 'children'),
    [Input('config-store', 'data'), Input('overview-bite-table', 'data'),
     Input('liquidity-dry-powder-store', 'data')]
)
def update_forward_bite_table(config, bite_rows, dry_powder_uploaded):
    dry_powder = get_effective_dry_powder(config, dry_powder_uploaded)
    rows = bite_rows or []

    if not rows:
        rows = [
            {'asset_type': 'GP-Led (Single-Asset)', 'min_pct': 0.5, 'desired_pct': 2.25, 'max_pct': 4.0},
            {'asset_type': 'GP-Led (Multi-Asset)', 'min_pct': 0.5, 'desired_pct': 2.75, 'max_pct': 5.0},
            {'asset_type': 'Co-Investments', 'min_pct': 0.5, 'desired_pct': 1.75, 'max_pct': 3.0},
            {'asset_type': 'Diversified LP-Led', 'min_pct': 0.5, 'desired_pct': 2.75, 'max_pct': 5.0},
        ]

    table_rows = []
    for row in rows:
        asset_type = row.get('asset_type', '')
        min_pct = float(row.get('min_pct', 0)) / 100
        desired_pct = float(row.get('desired_pct', 0)) / 100
        max_pct = float(row.get('max_pct', 0)) / 100

        min_size = dry_powder * min_pct
        desired_size = dry_powder * desired_pct
        max_size = dry_powder * max_pct

        table_rows.append(html.Tr([
            html.Td(asset_type, style={'fontWeight': 'bold', 'fontFamily': C['mono'], 'fontSize': '12px'}),
            html.Td(f"${min_size:.1f}M ({min_pct:.2%})", style={'color': C['muted'], 'fontFamily': C['mono'], 'fontSize': '12px'}),
            html.Td(f"${desired_size:.1f}M ({desired_pct:.2%})", style={'color': C['green'], 'fontFamily': C['mono'], 'fontWeight': 'bold', 'fontSize': '12px'}),
            html.Td(f"${max_size:.1f}M ({max_pct:.2%})", style={'color': C['red'], 'fontFamily': C['mono'], 'fontSize': '12px'}),
        ]))

    return dbc.Table([
        html.Thead(html.Tr([
            html.Th('Asset Type'), html.Th('Min Size'), html.Th('Desired Size'), html.Th('Max Size')
        ]), style={'backgroundColor': C['surface'], 'color': C['text']}),
        html.Tbody(table_rows)
    ], bordered=True, hover=True, size='sm', className='shadow-sm',
        style={'fontFamily': C['mono']})


# ==================== DISCOUNT & PRICE-TO-NAV CALLBACKS ====================

def _build_discount_rows(deals, fund_cf_data, discount_store):
    """
    Build the list of dicts used by the discount DataTable.
    NAV comes from uploaded cashflows first, then deal['nav'], then 0.
    discount_store keyed by deal name → {discount_pct, nav_override}.
    """
    rows = []
    for d in (deals or []):
        name = d.get('name', '')
        # Prefer latest uploaded NAV
        nav = get_latest_available_nav_for_deal(fund_cf_data, name)
        if nav is None or nav == 0:
            nav = float(d.get('nav', d.get('size', 0)) or 0)

        saved = (discount_store or {}).get(name, {})
        nav_override = saved.get('nav_override')
        if nav_override not in (None, '', 0):
            nav = float(nav_override)

        discount_pct = float(saved.get('discount_pct', 0) or 0)
        purchase_price = nav * (1 - discount_pct / 100) if nav else 0
        ptbnav = purchase_price / nav if nav else None

        rows.append({
            'Deal': name,
            'Strategy': d.get('strategy', ''),
            'NAV ($m)': round(nav, 2),
            'Discount (%)': round(discount_pct, 2),
            'Purchase Price ($m)': round(purchase_price, 2),
            'P/NAV': round(ptbnav, 4) if ptbnav is not None else None,
        })
    return rows


@app.callback(
    Output('discount-table-container', 'children'),
    [Input('deals-store', 'data'),
     Input('fund-cf-data-store', 'data'),
     Input('discount-store', 'data')],
    prevent_initial_call=False
)
def render_discount_table(deals, fund_cf_data, discount_store):
    """Render editable discount DataTable."""
    rows = _build_discount_rows(deals, fund_cf_data, discount_store)
    if not rows:
        return dbc.Alert(
            "No deals found. Add deals on the Portfolio page or upload a Fund Level CF file.",
            color="info"
        )

    columns = [
        {'name': 'Deal', 'id': 'Deal', 'editable': False},
        {'name': 'Strategy', 'id': 'Strategy', 'editable': False},
        {'name': 'NAV ($m)', 'id': 'NAV ($m)', 'editable': True, 'type': 'numeric',
         'format': {'specifier': ',.2f'}},
        {'name': 'Discount (%)', 'id': 'Discount (%)', 'editable': True, 'type': 'numeric',
         'format': {'specifier': '.2f'}},
        {'name': 'Purchase Price ($m)', 'id': 'Purchase Price ($m)', 'editable': False,
         'type': 'numeric', 'format': {'specifier': ',.2f'}},
        {'name': 'P/NAV', 'id': 'P/NAV', 'editable': False, 'type': 'numeric',
         'format': {'specifier': '.4f'}},
    ]

    # Style rules: highlight editable columns, colour-code P/NAV
    style_data_conditional = [
        # Editable columns highlighted
        {'if': {'column_id': 'Discount (%)'}, 'backgroundColor': '#fff9e6', 'fontWeight': 'bold',
         'border': f'2px solid {C["amber"]}'},
        {'if': {'column_id': 'NAV ($m)'}, 'backgroundColor': '#f0f4ff',
         'border': f'1px solid {C["blue"]}'},
        # P/NAV < 0.85 → green (good discount), > 0.95 → amber (thin discount)
        {'if': {'filter_query': '{P/NAV} < 0.85', 'column_id': 'P/NAV'},
         'color': C['green'], 'fontWeight': 'bold'},
        {'if': {'filter_query': '{P/NAV} >= 0.85 && {P/NAV} < 0.95', 'column_id': 'P/NAV'},
         'color': C['amber'], 'fontWeight': 'bold'},
        {'if': {'filter_query': '{P/NAV} >= 0.95', 'column_id': 'P/NAV'},
         'color': C['red'], 'fontWeight': 'bold'},
        # Zero discount rows dimmed
        {'if': {'filter_query': '{Discount (%)} = 0'}, 'opacity': '0.6'},
        # Alternating rows
        {'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']},
    ]

    return dash_table.DataTable(
        id='discount-datatable',
        columns=columns,
        data=rows,
        editable=True,
        row_selectable=False,
        sort_action='native',
        style_table={'overflowX': 'auto'},
        style_header={
            'backgroundColor': C['surface'],
            'fontWeight': 'bold',
            'fontFamily': C['sans'],
            'fontSize': '12px',
            'border': f'1px solid {C["border"]}',
            'textAlign': 'center',
        },
        style_cell={
            'fontFamily': C['mono'],
            'fontSize': '13px',
            'padding': '8px 12px',
            'textAlign': 'right',
            'border': f'1px solid {C["border"]}',
            'minWidth': '100px',
        },
        style_cell_conditional=[
            {'if': {'column_id': 'Deal'}, 'textAlign': 'left', 'minWidth': '180px',
             'fontFamily': C['sans'], 'fontWeight': 'bold'},
            {'if': {'column_id': 'Strategy'}, 'textAlign': 'left', 'minWidth': '160px'},
        ],
        style_data_conditional=style_data_conditional,
        tooltip_header={
            'Discount (%)': 'Enter the % discount negotiated on the NAV. Edit inline.',
            'NAV ($m)': 'Override NAV here if different from uploaded cashflows.',
            'Purchase Price ($m)': 'NAV × (1 − Discount %). Auto-calculated.',
            'P/NAV': 'Purchase Price ÷ NAV. Green < 0.85, Amber 0.85–0.95, Red ≥ 0.95.',
        },
        tooltip_delay=400,
        tooltip_duration=None,
    )


@app.callback(
    [Output('discount-store', 'data', allow_duplicate=True),
     Output('discount-save-status', 'children'),
     Output('discount-save-status', 'is_open'),
     Output('discount-save-status', 'color')],
    Input('btn-save-discounts', 'n_clicks'),
    [State('discount-datatable', 'data'),
     State('discount-store', 'data')],
    prevent_initial_call=True
)
def save_discounts(n, table_data, existing_store):
    """Persist edited discount and NAV override values from the DataTable."""
    if not n or not table_data:
        return existing_store or {}, "Nothing to save.", False, "info"

    store = dict(existing_store or {})
    for row in table_data:
        name = row.get('Deal', '')
        if not name:
            continue
        try:
            d_pct = float(row.get('Discount (%)', 0) or 0)
            nav_ov = float(row.get('NAV ($m)', 0) or 0)
            store[name] = {'discount_pct': d_pct, 'nav_override': nav_ov}
        except (TypeError, ValueError):
            pass

    return store, f"✅ Discounts saved for {len(store)} deal(s).", True, "success"


@app.callback(
    Output('discount-store', 'data', allow_duplicate=True),
    Input('btn-refresh-discounts', 'n_clicks'),
    [State('discount-store', 'data')],
    prevent_initial_call=True
)
def refresh_discounts(n, store):
    """Clear NAV overrides so they re-pull from uploaded cashflows."""
    if not n:
        return store or {}
    refreshed = {}
    for name, vals in (store or {}).items():
        refreshed[name] = {'discount_pct': vals.get('discount_pct', 0), 'nav_override': 0}
    return refreshed


# Live recalc: update Purchase Price & P/NAV as user edits the DataTable
@app.callback(
    Output('discount-datatable', 'data'),
    Input('discount-datatable', 'data_timestamp'),
    State('discount-datatable', 'data'),
    prevent_initial_call=True
)
def recalc_discount_table(_, rows):
    """Recalculate derived columns whenever the user edits a cell."""
    if not rows:
        return rows
    updated = []
    for row in rows:
        try:
            nav = float(row.get('NAV ($m)', 0) or 0)
            disc = float(row.get('Discount (%)', 0) or 0)
            pp = nav * (1 - disc / 100) if nav else 0
            ptbnav = pp / nav if nav else None
            row['Purchase Price ($m)'] = round(pp, 2)
            row['P/NAV'] = round(ptbnav, 4) if ptbnav is not None else None
        except (TypeError, ValueError):
            pass
        updated.append(row)
    return updated


@app.callback(
    [Output('disc-total-nav', 'children'),
     Output('disc-wtd-avg', 'children'),
     Output('disc-total-price', 'children'),
     Output('disc-portfolio-ptbnav', 'children')],
    [Input('discount-datatable', 'data')],
    prevent_initial_call=False
)
def update_discount_summary(rows):
    """Compute portfolio-level discount summary cards from table data."""
    if not rows:
        return "$—", "—", "$—", "—"

    total_nav = sum(float(r.get('NAV ($m)', 0) or 0) for r in rows)
    total_price = sum(float(r.get('Purchase Price ($m)', 0) or 0) for r in rows)

    # Weighted average discount (weight by NAV)
    wtd_num = sum(float(r.get('NAV ($m)', 0) or 0) * float(r.get('Discount (%)', 0) or 0)
                  for r in rows)
    wtd_avg_disc = wtd_num / total_nav if total_nav else 0
    portfolio_ptbnav = total_price / total_nav if total_nav else None

    nav_str = f"${total_nav:,.1f}M"
    disc_str = f"{wtd_avg_disc:.1f}%"
    price_str = f"${total_price:,.1f}M"
    ptbnav_str = f"{portfolio_ptbnav:.3f}x" if portfolio_ptbnav is not None else "—"

    return nav_str, disc_str, price_str, ptbnav_str


@app.callback(
    Output('discount-chart', 'figure'),
    Input('discount-datatable', 'data'),
    prevent_initial_call=False
)
def update_discount_chart(rows):
    """Horizontal bar chart showing discount % and P/NAV per deal."""
    import plotly.graph_objects as go

    if not rows:
        return go.Figure()

    # Sort by discount descending
    rows_sorted = sorted(
        [r for r in rows if float(r.get('NAV ($m)', 0) or 0) > 0],
        key=lambda x: float(x.get('Discount (%)', 0) or 0),
        reverse=True
    )

    if not rows_sorted:
        return go.Figure()

    deals = [r['Deal'] for r in rows_sorted]
    discounts = [float(r.get('Discount (%)', 0) or 0) for r in rows_sorted]
    ptbnavs = [float(r.get('P/NAV', 1) or 1) for r in rows_sorted]
    navs = [float(r.get('NAV ($m)', 0) or 0) for r in rows_sorted]
    prices = [float(r.get('Purchase Price ($m)', 0) or 0) for r in rows_sorted]

    # Bar colours: green if disc > 15, amber if 5–15, red if < 5
    bar_colors = []
    for d in discounts:
        if d >= 15:
            bar_colors.append(C['green'])
        elif d >= 5:
            bar_colors.append(C['amber'])
        else:
            bar_colors.append(C['red'])

    fig = go.Figure()

    fig.add_trace(go.Bar(
        name='Discount (%)',
        y=deals,
        x=discounts,
        orientation='h',
        marker_color=bar_colors,
        customdata=list(zip(navs, prices, ptbnavs)),
        hovertemplate=(
            '<b>%{y}</b><br>'
            'Discount: %{x:.1f}%<br>'
            'NAV: $%{customdata[0]:,.1f}M<br>'
            'Purchase Price: $%{customdata[1]:,.1f}M<br>'
            'P/NAV: %{customdata[2]:.3f}x'
            '<extra></extra>'
        ),
        text=[f"{d:.1f}%" for d in discounts],
        textposition='outside',
    ))

    # Overlay P/NAV as a scatter on secondary x-axis
    fig.add_trace(go.Scatter(
        name='P/NAV (right)',
        y=deals,
        x=ptbnavs,
        mode='markers',
        marker=dict(symbol='diamond', size=10, color=C['purple'], line=dict(width=1, color='white')),
        xaxis='x2',
        hovertemplate='<b>%{y}</b><br>P/NAV: %{x:.3f}x<extra></extra>',
    ))

    fig.update_layout(
        height=max(280, 40 * len(deals) + 100),
        margin=dict(l=20, r=60, t=40, b=40),
        paper_bgcolor=C['bg'],
        plot_bgcolor=C['bg'],
        font=dict(family=C['mono'], color=C['text'], size=12),
        legend=dict(orientation='h', y=1.08),
        xaxis=dict(title='Discount (%)', gridcolor=C['border'], zeroline=True,
                   zerolinecolor=C['border']),
        xaxis2=dict(title='P/NAV', overlaying='x', side='top',
                    range=[0, 1.1], gridcolor=C['border'], showgrid=False),
        yaxis=dict(autorange='reversed', gridcolor=C['border']),
        bargap=0.25,
    )

    return fig


# ==================== LIQUIDITY ASSUMPTIONS CALLBACKS ====================

# ==================== LIQUIDITY PULL CALLBACKS ====================

# Upload Liquidity Pull Excel file
@app.callback(
    [Output('liquidity-data-store', 'data'), Output('upload-liquidity-status', 'children')],
    Input('upload-liquidity-pull', 'contents'),
    State('upload-liquidity-pull', 'filename'),
    prevent_initial_call=True
)
def upload_liquidity_file(contents, filename):
    """Parse uploaded Liquidity Pull Excel file with all key metrics"""
    print(f"\n=== LIQUIDITY UPLOAD DEBUG ===")
    print(f"Contents received: {contents is not None}")
    print(f"Filename: {filename}")

    if contents is None:
        print("No contents - returning early")
        return None, ""

    try:
        # Decode the uploaded file
        content_type, content_string = contents.split(',')
        decoded = base64.b64decode(content_string)

        # Read Excel file
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(decoded), data_only=True)

        # Try to find Liquidity Pull sheet
        sheet_names = ['Liquidity Pull', 'Liquidity_Pull', 'LiquidityPull']
        ws = None
        for name in sheet_names:
            if name in wb.sheetnames:
                ws = wb[name]
                break

        if ws is None:
            # Use first sheet if can't find specific name
            ws = wb[wb.sheetnames[0]]

        # Parse the sheet structure
        liquidity_data = {
            'as_at_date': None,
            'current_quarter': None,
            'current_month': None,
            'fund_nav': 0,
            'current_cash': 0,  # Row 58
            'glf_balance': 0,  # Row 59
            'cqs_balance': 0,  # Row 60
            'total_liquidity': 0,
            'max_deployable_capital': {},  # Row 53 (monthly)
            'surplus_liquidity_post_buffer': 0,  # Row 63
            'projected_nav_existing': 0,  # Row 71
            'projected_nav_existing_pipeline': 0,  # Row 72
            'near_term_flows': {},
            'nav_projections': {}
        }

        # Extract key data points
        # As At date (typically row 4, column 2)
        as_at = ws.cell(4, 2).value
        if as_at:
            liquidity_data['as_at_date'] = str(as_at)

        # Current Quarter (row 5, column 2)
        curr_q = ws.cell(5, 2).value
        if curr_q:
            liquidity_data['current_quarter'] = str(curr_q)

        # Current Month (row 6, column 2)
        curr_m = ws.cell(6, 2).value
        if curr_m:
            liquidity_data['current_month'] = str(curr_m)

        # Fund NAV (row 10, column 3 typically)
        fund_nav = ws.cell(10, 3).value
        if fund_nav and isinstance(fund_nav, (int, float)):
            liquidity_data['fund_nav'] = float(fund_nav) / 1_000_000

        # MAX DEPLOYABLE CAPITAL - Row 53 (monthly values)
        # Special focus on Column 9 which should be Dec 2026
        dec_2026_dry_powder = None

        try:
            # Row 53, Column 9 (index 8) = Dec 2026 dry powder
            dec_val = ws.cell(53, 9).value
            if dec_val and isinstance(dec_val, (int, float)):
                dec_2026_dry_powder = float(dec_val) / 1_000_000
                liquidity_data['dec_2026_dry_powder'] = dec_2026_dry_powder
                print(f"✅ Dec 2026 Dry Powder (Row 53, Col 9): ${dec_2026_dry_powder:.1f}M")
        except:
            pass

        for col in range(1, 25):  # Up to 24 months
            try:
                month_label = ws.cell(28, col).value  # Month header from row 28
                max_deploy = ws.cell(53, col).value
                if month_label and max_deploy and isinstance(max_deploy, (int, float)):
                    liquidity_data['max_deployable_capital'][str(month_label)] = float(max_deploy) / 1_000_000
            except:
                pass

        # CASH BALANCE - Row 58
        cash = ws.cell(58, 3).value
        if cash and isinstance(cash, (int, float)):
            liquidity_data['current_cash'] = float(cash) / 1_000_000

        # GLF BALANCE - Row 59
        glf = ws.cell(59, 3).value
        if glf and isinstance(glf, (int, float)):
            liquidity_data['glf_balance'] = float(glf) / 1_000_000

        # CQS BALANCE - Row 60
        cqs = ws.cell(60, 3).value
        if cqs and isinstance(cqs, (int, float)):
            liquidity_data['cqs_balance'] = float(cqs) / 1_000_000

        # SURPLUS LIQUIDITY POST BUFFER - Row 63
        surplus = ws.cell(63, 3).value
        if surplus and isinstance(surplus, (int, float)):
            liquidity_data['surplus_liquidity_post_buffer'] = float(surplus) / 1_000_000

        # PROJECTED NAV EXISTING - Row 71
        proj_nav_existing = ws.cell(71, 3).value
        if proj_nav_existing and isinstance(proj_nav_existing, (int, float)):
            liquidity_data['projected_nav_existing'] = float(proj_nav_existing) / 1_000_000

        # PROJECTED NAV EXISTING + PIPELINE - Row 72
        proj_nav_pipeline = ws.cell(72, 3).value
        if proj_nav_pipeline and isinstance(proj_nav_pipeline, (int, float)):
            liquidity_data['projected_nav_existing_pipeline'] = float(proj_nav_pipeline) / 1_000_000

        # Total Liquidity (row 16, column 3)
        total_liq = ws.cell(16, 3).value
        if total_liq and isinstance(total_liq, (int, float)):
            liquidity_data['total_liquidity'] = float(total_liq) / 1_000_000

        # Near term flows - starting around row 28
        for col in range(1, 25):  # Up to 24 months
            try:
                month_val = ws.cell(28, col).value  # Month header
                subs = ws.cell(32, col).value or 0  # Subscriptions
                reds = ws.cell(33, col).value or 0  # Redemptions
                flows = ws.cell(34, col).value or 0  # Portfolio flows

                if month_val:
                    liquidity_data['near_term_flows'][str(month_val)] = {
                        'subscriptions': float(subs) / 1_000_000 if isinstance(subs, (int, float)) else 0,
                        'redemptions': float(reds) / 1_000_000 if isinstance(reds, (int, float)) else 0,
                        'portfolio_flows': float(flows) / 1_000_000 if isinstance(flows, (int, float)) else 0
                    }
            except:
                pass

        # NAV projections (row 39)
        for col in range(1, 25):
            try:
                month_val = ws.cell(28, col).value
                nav_end = ws.cell(39, col).value
                if month_val and nav_end and isinstance(nav_end, (int, float)):
                    liquidity_data['nav_projections'][str(month_val)] = float(nav_end) / 1_000_000
            except:
                pass

        success_msg = dbc.Alert([
            html.I(className="fas fa-check-circle me-2"),
            f"✅ Uploaded: {filename} - Max Deployable, Cash ({liquidity_data['current_cash']:.1f}M), Projected NAV loaded"
        ], color="success", className="mt-2")

        return liquidity_data, success_msg

    except Exception as e:
        import traceback
        print(f"Liquidity upload error: {traceback.format_exc()}")
        error_msg = dbc.Alert([
            html.I(className="fas fa-exclamation-triangle me-2"),
            f"❌ Error: {str(e)}"
        ], color="danger", className="mt-2")
        return None, error_msg


# Date fields
@app.callback(
    [Output('liq-today-date', 'children'), Output('liq-as-at-date', 'children'),
     Output('liq-current-quarter', 'children'), Output('liq-current-month', 'children')],
    Input('liquidity-data-store', 'data')
)
def update_liq_dates(uploaded_data):
    if uploaded_data:
        # Use uploaded dates
        return (
            datetime.now().strftime('%Y-%m-%d'),
            uploaded_data.get('as_at_date', 'N/A'),
            uploaded_data.get('current_quarter', 'N/A'),
            uploaded_data.get('current_month', 'N/A')
        )
    else:
        # Use calculated dates
        today = datetime.now()
        as_at = (today.replace(day=1) - relativedelta(days=1))
        quarter_month = ((today.month - 1) // 3 + 1) * 3
        quarter_end = datetime(today.year, quarter_month, 1) + relativedelta(months=1) - relativedelta(days=1)
        month_end = today.replace(day=1) + relativedelta(months=1) - relativedelta(days=1)

        return (
            today.strftime('%Y-%m-%d'),
            as_at.strftime('%Y-%m-%d'),
            quarter_end.strftime('%Y-%m-%d'),
            month_end.strftime('%Y-%m-%d')
        )


# Liquidity Waterfall
@app.callback(
    Output('liq-waterfall-table', 'children'),
    [Input('deals-store', 'data'), Input('liquidity-data-store', 'data')]
)
def generate_liquidity_waterfall(deals, uploaded_data):
    """Generate liquidity waterfall from uploaded data or calculated"""

    if uploaded_data:
        # Use uploaded data
        total_nav = uploaded_data.get('fund_nav', 0)
        current_cash = uploaded_data.get('current_cash', 0)
        glf_balance = uploaded_data.get('glf_balance', 0)
        cqs_balance = uploaded_data.get('cqs_balance', 0)
        total_liquidity = uploaded_data.get('total_liquidity', 0)
    else:
        # Calculate from deals
        total_nav = sum(d['size'] for d in deals) if deals else 0
        current_cash = 0.88
        glf_balance = 19.95
        cqs_balance = 17.27
        total_liquidity = current_cash + glf_balance + cqs_balance

    waterfall_data = [
        {'Source': 'Fund NAV', 'Item': 'Total Current NAV', 'Amount ($mm)': f'{total_nav:.2f}'},
        {'Source': 'Current Cash', 'Item': 'Cash Balance', 'Amount ($mm)': f'{current_cash:.2f}'},
        {'Source': 'GLF', 'Item': 'GLF Balance', 'Amount ($mm)': f'{glf_balance:.2f}'},
        {'Source': 'CQS', 'Item': 'CQS Balance', 'Amount ($mm)': f'{cqs_balance:.2f}'},
        {'Source': 'Total Liquidity', 'Item': 'Total Liquidity Balance', 'Amount ($mm)': f'{total_liquidity:.2f}'},
    ]

    return dash_table.DataTable(
        data=waterfall_data,
        columns=[{"name": c, "id": c} for c in waterfall_data[0].keys()],
        style_cell={'textAlign': 'left', 'padding': '12px', 'fontFamily': C['mono'], 'fontSize': '12px'},
        style_header={'backgroundColor': C['surface'], 'color': C['text'], 'fontWeight': 'bold',
                      'border': f'1px solid {C["border"]}'},
        style_data={'backgroundColor': C['panel'], 'color': C['text'], 'border': f'1px solid {C["border"]}'},
        style_data_conditional=[
            {'if': {'column_id': 'Source', 'filter_query': '{Source} = "Total Liquidity"'},
             'fontWeight': 'bold', 'color': C['green'], 'backgroundColor': C['surface']}
        ]
    )


# Near Term Flows
@app.callback(
    Output('liq-near-term-flows', 'children'),
    [Input('deals-store', 'data'), Input('liquidity-data-store', 'data')]
)
def generate_near_term_flows(deals, uploaded_data):
    """Generate near term flows from uploaded data or calculated"""

    base_date = datetime(2026, 3, 31)
    months = []
    subscriptions = []
    redemptions = []
    portfolio_flows = []

    if uploaded_data and uploaded_data.get('near_term_flows'):
        # Use uploaded data
        for month_key, data in list(uploaded_data['near_term_flows'].items())[:12]:
            try:
                month_date = pd.to_datetime(month_key)
                months.append(month_date.strftime('%b-%y'))
            except:
                months.append(str(month_key)[:6])
            subscriptions.append(data.get('subscriptions', 0))
            redemptions.append(data.get('redemptions', 0))
            portfolio_flows.append(data.get('portfolio_flows', 0))
    else:
        # Use calculated/placeholder data
        for i in range(12):
            month_date = base_date + relativedelta(months=i)
            months.append(month_date.strftime('%b-%y'))
            subscriptions.append(50 if i < 3 else 0)
            redemptions.append(0)
            portfolio_flows.append(2.6 + (i * 0.5))

    # Create table
    table_data = [
        {'Flow Type': 'Subscriptions', **{months[i]: f'${subscriptions[i]:.1f}M' for i in range(len(months))}},
        {'Flow Type': 'Redemptions', **{months[i]: f'${redemptions[i]:.1f}M' for i in range(len(months))}},
        {'Flow Type': 'Portfolio Net Flows', **{months[i]: f'${portfolio_flows[i]:.1f}M' for i in range(len(months))}},
    ]

    all_columns = ['Flow Type'] + months

    return dash_table.DataTable(
        data=table_data,
        columns=[{"name": c, "id": c} for c in all_columns],
        style_cell={
            'textAlign': 'left',
            'padding': '10px',
            'fontFamily': C['mono'],
            'fontSize': '11px',
            'minWidth': '100px'
        },
        style_header={
            'backgroundColor': C['surface'],
            'color': C['text'],
            'fontWeight': 'bold',
            'position': 'sticky',
            'top': 0,
            'border': f'1px solid {C["border"]}'
        },
        style_data={'backgroundColor': C['panel'], 'color': C['text'], 'border': f'1px solid {C["border"]}'},
        style_data_conditional=[
            {'if': {'row_index': 'odd'}, 'backgroundColor': C['surface']},
            {'if': {'column_id': 'Flow Type'}, 'fontWeight': 'bold'}
        ],
        style_table={'overflowX': 'auto'},
        fixed_columns={'headers': True, 'data': 1}
    )


# NAV Projection Chart
@app.callback(
    Output('liq-nav-projection-chart', 'figure'),
    [Input('deals-store', 'data'), Input('liquidity-data-store', 'data')]
)
def generate_nav_projection_chart(deals, uploaded_data):
    """Generate NAV end projections from uploaded data or calculated"""

    base_date = datetime(2026, 3, 31)
    months = []
    nav_projections = []

    if uploaded_data and uploaded_data.get('nav_projections'):
        # Use uploaded NAV projections
        for month_key, nav_val in list(uploaded_data['nav_projections'].items())[:12]:
            try:
                month_date = pd.to_datetime(month_key)
                months.append(month_date.strftime('%b %Y'))
            except:
                months.append(str(month_key)[:10])
            nav_projections.append(nav_val)
    else:
        # Calculate projections
        base_nav = sum(d['size'] for d in deals) if deals else 100

        for i in range(12):
            month_date = base_date + relativedelta(months=i)
            months.append(month_date.strftime('%b %Y'))
            nav = base_nav * (1 + (0.13 / 12)) ** i
            nav_projections.append(nav)

    fig = go.Figure()

    fig.add_trace(go.Scatter(
        x=months, y=nav_projections,
        mode='lines+markers',
        name='NAV End',
        line=dict(color=C['blue'], width=3),
        marker=dict(size=8, color=C['sky']),
        fill='tozeroy',
        fillcolor=rgba(C['blue'], 0.2)
    ))

    fig.update_layout(
        **CHART_BASE,
        height=400,
        yaxis_title='NAV ($mm)',
        hovermode='x unified'
    )

    return fig


# Tab content
@app.callback(
    Output('liq-tab-content', 'children'),
    Input('liq-tabs', 'active_tab')
)
def render_liq_tab_content(active_tab):
    if active_tab == "tab-liq-subs":
        return dbc.Card([
            dbc.CardBody([
                html.H6("Subscriptions & Redemptions", className="mb-3"),
                html.P("Expected subscriptions: $150M (confirmed)", style={'fontFamily': C['mono']}),
                html.P("Expected redemptions: $0M", style={'fontFamily': C['mono']}),
                html.P("Redemption gate limit: 5% of NAV", style={'fontFamily': C['mono']})
            ])
        ], className="shadow-sm")
    elif active_tab == "tab-liq-flows":
        return dbc.Card([
            dbc.CardBody([
                html.H6("Portfolio Net Flows", className="mb-3"),
                html.P("Seed Portfolio Net Flows calculated from deal-level cashflows",
                       style={'fontFamily': C['mono']}),
                html.P("See Fund Level CF tab for detailed monthly breakdown", style={'fontFamily': C['mono']})
            ])
        ], className="shadow-sm")
    else:  # unfunded
        return dbc.Card([
            dbc.CardBody([
                html.H6("Unfunded Commitments", className="mb-3"),
                html.P("Current unfunded commitments: $140.3M", style={'fontFamily': C['mono']}),
                html.P("Rolling unfunded exposure monitored monthly", style={'fontFamily': C['mono']})
            ])
        ], className="shadow-sm")


@app.callback(
    [Output('analytics-total-nav', 'children'), Output('analytics-num-deals', 'children'),
     Output('analytics-weighted-irr', 'children'), Output('analytics-top1-conc', 'children')],
    [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
     Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')]
)
def update_analytics_summary(view, deals, pipeline, placeholders):
    """Update summary cards based on view toggle - 3 options"""
    portfolio_deals = get_portfolio_for_analytics_view(view, deals, pipeline, placeholders)

    if not portfolio_deals:
        return "$0.0M", "0", "0.0%", "0.0%"

    total_nav = sum(d['size'] for d in portfolio_deals)
    num_deals = len(portfolio_deals)
    weighted_irr = sum(d.get('target_irr', 0) * d['size'] for d in portfolio_deals) / total_nav if total_nav > 0 else 0

    # Top 1 concentration
    sorted_deals = sorted(portfolio_deals, key=lambda x: x['size'], reverse=True)
    top1_conc = sorted_deals[0]['size'] / total_nav if len(sorted_deals) > 0 and total_nav > 0 else 0

    return f"${total_nav:.1f}M", str(num_deals), f"{weighted_irr:.1%}", f"{top1_conc:.1%}"


# Strategy Exposure
@app.callback(
    Output('analytics-strategy', 'figure'),
    [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
     Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')]
)
def update_strategy_chart(view, deals, pipeline, placeholders):
    portfolio_deals = get_portfolio_for_analytics_view(view, deals, pipeline, placeholders)

    if not portfolio_deals:
        fig = go.Figure()
        fig.update_layout(**CHART_BASE)
        return fig

    # Group by strategy
    by_strategy = {}
    for d in portfolio_deals:
        strat = d.get('strategy', 'Other')
        by_strategy[strat] = by_strategy.get(strat, 0) + d['size']

    labels = list(by_strategy.keys())
    values = list(by_strategy.values())
    colors = [C['blue'], C['purple'], C['teal'], C['green']][:len(labels)]

    fig = go.Figure(data=[go.Pie(
        labels=labels, values=values, hole=0.4,
        marker=dict(colors=colors),
        textfont=dict(color=C['text'], family=C['mono'])
    )])
    fig.update_layout(**CHART_BASE, height=350, margin=dict(t=20, b=0, l=0, r=0))

    return fig


# Regional Exposure
@app.callback(
    Output('analytics-region', 'figure'),
    [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
     Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')]
)
def update_region_chart(view, deals, pipeline, placeholders):
    if view == 'current':
        portfolio_deals = deals if deals else []
    else:
        portfolio_deals = (deals if deals else []) + \
                          [{'name': p['name'], 'size': p['size'], 'geography': p.get('region', 'North America')}
                           for p in (pipeline if pipeline else [])] + \
                          [{'name': p['name'], 'size': p['size'], 'geography': p.get('region', 'North America')}
                           for p in (placeholders if placeholders else [])]

    if not portfolio_deals:
        fig = go.Figure()
        fig.update_layout(**CHART_BASE)
        return fig

    # Group by region
    by_region = {}
    for d in portfolio_deals:
        region = d.get('geography', d.get('region', 'North America'))
        by_region[region] = by_region.get(region, 0) + d['size']

    labels = list(by_region.keys())
    values = list(by_region.values())

    fig = go.Figure(data=[go.Bar(
        x=labels, y=values, marker_color=C['blue'],
        text=[f"${v:.1f}M" for v in values], textposition='outside',
        textfont=dict(color=C['text'], family=C['mono'])
    )])
    fig.update_layout(**CHART_BASE, yaxis_title="NAV ($mm)", height=350)

    return fig


# Vintage Exposure
@app.callback(
    Output('analytics-vintage', 'figure'),
    [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
     Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')]
)
def update_vintage_chart(view, deals, pipeline, placeholders):
    if view == 'current':
        portfolio_deals = deals if deals else []
    else:
        portfolio_deals = (deals if deals else []) + \
                          [{'name': p['name'], 'size': p['size'], 'vintage': 2026} for p in
                           (pipeline if pipeline else [])] + \
                          [{'name': p['name'], 'size': p['size'], 'vintage': 2026} for p in
                           (placeholders if placeholders else [])]

    if not portfolio_deals:
        fig = go.Figure()
        fig.update_layout(**CHART_BASE)
        return fig

    # Group by vintage
    by_vintage = {}
    for d in portfolio_deals:
        vint = d.get('vintage', 2024)
        by_vintage[vint] = by_vintage.get(vint, 0) + d['size']

    labels = [str(v) for v in sorted(by_vintage.keys())]
    values = [by_vintage[int(v)] for v in labels]

    fig = go.Figure(data=[go.Bar(
        x=labels, y=values, marker_color=C['purple'],
        text=[f"${v:.1f}M" for v in values], textposition='outside',
        textfont=dict(color=C['text'], family=C['mono'])
    )])
    fig.update_layout(**CHART_BASE, yaxis_title="NAV ($mm)", height=350)

    return fig


# Deal Type (Secondary vs Co-Investment)
@app.callback(
    Output('analytics-dealtype', 'figure'),
    [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
     Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')]
)
def update_dealtype_chart(view, deals, pipeline, placeholders):
    if view == 'current':
        portfolio_deals = deals if deals else []
    else:
        portfolio_deals = (deals if deals else []) + \
                          [{'name': p['name'], 'size': p['size'], 'deal_type': 'Secondary'} for p in
                           (pipeline if pipeline else [])] + \
                          [{'name': p['name'], 'size': p['size'], 'deal_type': p.get('deal_type', 'Secondary')}
                           for p in (placeholders if placeholders else [])]

    if not portfolio_deals:
        fig = go.Figure()
        fig.update_layout(**CHART_BASE)
        return fig

    # Group by deal type
    secondary = sum(d['size'] for d in portfolio_deals if d.get('deal_type', 'Secondary') == 'Secondary')
    coinvest = sum(d['size'] for d in portfolio_deals if d.get('deal_type', 'Secondary') == 'Co-Investment')

    labels = ['Secondary', 'Co-Investment']
    values = [secondary, coinvest]
    colors_dt = [C['blue'], C['green']]

    fig = go.Figure(data=[go.Pie(
        labels=labels, values=values, hole=0.4,
        marker=dict(colors=colors_dt),
        textfont=dict(color=C['text'], family=C['mono'])
    )])
    fig.update_layout(**CHART_BASE, height=350, margin=dict(t=20, b=0, l=0, r=0))

    return fig


# Sector Exposure
@app.callback(
    Output('analytics-sector', 'figure'),
    [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
     Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')]
)
def update_sector_chart(view, deals, pipeline, placeholders):
    if view == 'current':
        portfolio_deals = deals if deals else []
    else:
        portfolio_deals = (deals if deals else []) + \
                          [{'name': p['name'], 'size': p['size'], 'sector': 'Technology'} for p in
                           (pipeline if pipeline else [])] + \
                          [{'name': p['name'], 'size': p['size'], 'sector': 'Technology'} for p in
                           (placeholders if placeholders else [])]

    if not portfolio_deals:
        fig = go.Figure()
        fig.update_layout(**CHART_BASE)
        return fig

    # Group by sector
    by_sector = {}
    for d in portfolio_deals:
        sector = d.get('sector', 'Technology')
        by_sector[sector] = by_sector.get(sector, 0) + d['size']

    labels = list(by_sector.keys())
    values = list(by_sector.values())
    colors_sec = [C['teal'], C['purple'], C['amber'], C['green'], C['pink']][:len(labels)]

    fig = go.Figure(data=[go.Pie(
        labels=labels, values=values, hole=0.4,
        marker=dict(colors=colors_sec),
        textfont=dict(color=C['text'], family=C['mono'])
    )])
    fig.update_layout(**CHART_BASE, height=350, margin=dict(t=20, b=0, l=0, r=0))

    return fig


# Concentration Risk
@app.callback(
    Output('analytics-concentration', 'figure'),
    [Input('analytics-view-toggle', 'value'), Input('deals-store', 'data'),
     Input('pipeline-store', 'data'), Input('placeholder-deals-store', 'data')]
)
def update_concentration_chart(view, deals, pipeline, placeholders):
    if view == 'current':
        portfolio_deals = deals if deals else []
    else:
        portfolio_deals = (deals if deals else []) + \
                          [{'name': p['name'], 'size': p['size']} for p in (pipeline if pipeline else [])] + \
                          [{'name': p['name'], 'size': p['size']} for p in (placeholders if placeholders else [])]

    if not portfolio_deals:
        fig = go.Figure()
        fig.update_layout(**CHART_BASE)
        return fig

    total_nav = sum(d['size'] for d in portfolio_deals)
    sorted_deals = sorted(portfolio_deals, key=lambda x: x['size'], reverse=True)

    top1 = (sorted_deals[0]['size'] / total_nav * 100) if len(sorted_deals) >= 1 else 0
    top3 = (sum(d['size'] for d in sorted_deals[:3]) / total_nav * 100) if len(sorted_deals) >= 3 else 0
    top5 = (sum(d['size'] for d in sorted_deals[:5]) / total_nav * 100) if len(sorted_deals) >= 5 else 0

    cats = ['Top 1', 'Top 3', 'Top 5']
    vals = [top1, top3, top5]
    limits = [15, 40, 60]
    colors_conc = [C['red'] if v > l else C['green'] for v, l in zip(vals, limits)]

    fig = go.Figure(data=[go.Bar(
        x=cats, y=vals, marker_color=colors_conc,
        text=[f"{v:.1f}%" for v in vals], textposition='outside',
        textfont=dict(color=C['text'], family=C['mono'])
    )])
    for limit in limits:
        fig.add_hline(y=limit, line_dash='dash', line_color=C['red'], opacity=0.5,
                      annotation_text=f"Limit: {limit}%", annotation_position="right")
    fig.update_layout(**CHART_BASE, yaxis_title="% of NAV", height=350)

    return fig


# Dashboard Recent Activity & Pipeline Status callbacks
@app.callback(
    [Output('dash-recent-activity', 'children'), Output('dash-pipeline-status', 'children')],
    [Input('deals-store', 'data'), Input('pipeline-store', 'data')]
)
def update_dashboard_activity(deals, pipeline):
    # Recent Activity
    if not deals:
        recent = html.P("No deals yet", className="text-muted", style={'fontFamily': C['mono']})
    else:
        recent_deals = sorted(deals, key=lambda x: x.get('date_added', ''), reverse=True)[:5]
        activity_items = []
        for d in recent_deals:
            activity_items.append(html.Div([
                html.Strong(d['name'], style={'color': C['blue'], 'fontFamily': C['mono']}),
                html.Br(),
                html.Small(f"${d['size']:.1f}M • {d['target_irr']:.1%} IRR",
                           style={'color': C['muted'], 'fontFamily': C['mono']}),
                html.Hr(style={'borderColor': C['border'], 'margin': '0.5rem 0'})
            ]))
        recent = html.Div(activity_items)

    # Pipeline Status
    if not pipeline:
        pipe_status = html.P("No pipeline deals", className="text-muted", style={'fontFamily': C['mono']})
    else:
        stage_counts = {}
        for p in pipeline:
            stage = p.get('stage', 'Screening')
            stage_counts[stage] = stage_counts.get(stage, 0) + 1

        pipe_items = []
        stage_colors = {
            'Screening': C['muted'],
            'Due Diligence': C['amber'],
            'Term Sheet': C['blue'],
            'Final Docs': C['green']
        }
        for stage, count in stage_counts.items():
            pipe_items.append(html.Div([
                html.Span("●", style={'color': stage_colors.get(stage, C['muted']), 'marginRight': '0.5rem'}),
                html.Strong(f"{stage}: ", style={'fontFamily': C['sans'], 'color': C['text']}),
                html.Span(f"{count} deals", style={'fontFamily': C['mono'], 'color': C['muted']})
            ], style={'marginBottom': '0.5rem'}))
        pipe_status = html.Div(pipe_items)

    return recent, pipe_status


# Export CSV
@app.callback(
    Output("download-csv", "data"),
    Input("btn-export", "n_clicks"),
    [State("deals-store", "data"), State("placeholder-deals-store", "data"), State("pipeline-store", "data")],
    prevent_initial_call=True
)
def export_all(n, deals, placeholders, pipeline):
    if deals:
        df_deals = pd.DataFrame(deals)
        df_deals.to_csv('current_portfolio.csv', index=False)
    if placeholders:
        df_ph = pd.DataFrame(placeholders)
        df_ph.to_csv('placeholder_deals.csv', index=False)
    if pipeline:
        df_pipe = pd.DataFrame(pipeline)
        df_pipe.to_csv('pipeline.csv', index=False)

    return dcc.send_data_frame(pd.DataFrame(deals).to_csv if deals else pd.DataFrame().to_csv,
                               f"portfolio_{date.today()}.csv", index=False)


# Auto-save callback - saves data whenever it changes
@app.callback(
    Output('save-status', 'children'),
    [Input('deals-store', 'data'), Input('pipeline-store', 'data'),
     Input('placeholder-deals-store', 'data'), Input('config-store', 'data')],
    prevent_initial_call=True
)
def auto_save_data(deals, pipeline, placeholders, config):
    """Automatically save data when anything changes"""
    save_data(deals or [], pipeline or [], placeholders or [], config or DEFAULT_CONFIG)
    return ""


# For production deployment (Render, Heroku, AWS, etc.)
server = app.server

if __name__ == '__main__':
    print("\n" + "=" * 80)
    print("HORIZON PORTFOLIO TOOL - INSTITUTIONAL GRADE LP MANAGEMENT")
    print("=" * 80)
    print("\n✅ Starting on http://localhost:8050")
    print("\n🎯 Features:")
    print("   • Current Portfolio Management (17 fields per deal)")
    print("   • Portfolio Segmentation & TWR Tracking")
    print("   • Pro Forma Portfolio (Placeholder Deals)")
    print("   • Custom Bite Sizes & Weights")
    print("   • 12-Month Dry Powder Forecasting")
    print("   • Deal Bite Sizing (Min/Desired/Max)")
    print("   • Pipeline Management")
    print("   • Return Calculator with Waterfall")
    print("   • Analytics with 3-Way Toggle (Current/Pipeline/Full)")
    print("   • Excel File Uploads (Fund CF + Liquidity Pull)")
    print("   • Comprehensive Exposure Analysis")
    print("   • Auto-Save Data Persistence")
    print("\nPress CTRL+C to stop\n")

    app.run(debug=True, host='0.0.0.0', port=8050)
